import streamlit as st
import pandas as pd
import hashlib
import secrets
import time
import importlib
import logging
import os
import io
from datetime import datetime, date
from functools import lru_cache
from PIL import Image
from sqlalchemy import text
import plotly.express as px
from database_mysql import Database
from data_manager import DataManager

# desktop.flag file → desktop app; absent → web/cloud (clients only)
IS_DESKTOP = os.path.exists(os.path.join(os.path.dirname(__file__), "desktop.flag"))

# Configure logging to console only (no file I/O on cloud)
logging.basicConfig(
    level=logging.WARNING,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Configure page
_logo_path = "assets/logo.png"
_page_icon = Image.open(_logo_path) if os.path.exists(_logo_path) else "📊"
st.set_page_config(
    page_title="EIMS",
    page_icon=_page_icon,
    layout="wide",
    initial_sidebar_state="expanded"
)

# Language translations dictionary
TRANSLATIONS = {
    'en': {
        'title': 'EIMS',
        'login': 'Login',
        'signup': 'Sign Up',
        'forgot_password': 'Forgot Password',
        'dashboard': 'Dashboard',
        'view': 'View Data',
        'add': 'Add Data',
        'edit': 'Edit Data',
        'delete': 'Delete Data',
        'analytics': 'Analytics & Charts',
        'export': 'Export Data',
        'request_leave': 'Request Leave',
        'manage_leaves': 'Manage Leave Requests',
        'manage_users': 'Manage Users',
        'profile': 'Profile',
        'logout': 'Logout',
        'select_page': 'Select Page:',
        'project_info': 'Information',
        'graduation_project': 'Software Engineering\nCompany Data Management System',
        'email': 'Email',
        'password': 'Password',
        'confirm_password': 'Confirm Password',
        'role': 'Role',
        'create_account': 'Create Account',
        'login_button': 'Login',
        'start_date': 'Start Date',
        'end_date': 'End Date',
        'leave_type': 'Leave Type',
        'reason': 'Reason for leave',
        'attachment': 'Optional attachment (medical note, etc.)',
        'submit': 'Submit Request',
        'my_requests': 'My Leave Requests',
        'from': 'From',
        'to': 'To',
        'type': 'Type',
        'manager_response': 'Manager response',
        'status': 'Status',
        'pending': 'Pending',
        'approved': 'Approved',
        'rejected': 'Rejected',
        'language': 'Language',
        'no_requests': 'You have no leave requests.',
        'set_status': 'Set status:',
        'response_user': 'Response to user:',
        'save': 'Save',
        'user': 'User',
        'employee': 'Employee',
        'client': 'Client',
        'manager': 'Manager',
        'shipments': 'Shipments',
        'add_shipment': 'Add Shipment',
        'export_employees': 'Export Employees Data',
        'export_shipments': 'Export Shipments Data',
        'client_dashboard': 'Overview',
        'my_shipments': 'My Shipments',
        'my_tickets': 'My Tickets',
        'my_invoices': 'My Invoices',
        'my_offers': 'My Offers',
        'accept_offer': 'Accept Offer',
        'reject_offer': 'Reject Offer',
        'offer_accepted': 'Offer accepted successfully.',
        'offer_rejected': 'Offer rejected.',
        'no_offers_client': 'No offers have been sent to you yet.',
        # Client Dashboard
        'welcome_back': 'Welcome back',
        'active_shipments': 'Active Shipments',
        'open_tickets': 'Open Tickets',
        'unpaid_invoices': 'Unpaid Invoices',
        'new_offers': 'New Offers',
        'unread_messages': 'Unread Messages',
        'recent_shipments': 'Recent Shipments',
        'recent_tickets': 'Recent Tickets',
        'recent_offers': 'Recent Offers',
        'no_recent_activity': 'No recent activity.',
        'view_all': 'View All',
        'due_lbl': 'Due',
        'dash_origin': 'Origin',
        'dash_dest': 'Destination',
        'track_shipment': 'Track Shipment',
        'manage_shipments': 'Manage Shipments',
        'shipment_analytics': 'Shipment Analytics',
        'import': 'Import',
        'export': 'Export',
        'shipment_number': 'Shipment Number',
        'origin': 'Origin Country',
        'destination': 'Destination Country',
        'departure_date': 'Departure Date',
        'expected_arrival': 'Expected Arrival',
        'actual_arrival': 'Actual Arrival',
        'shipment_status': 'Shipment Status',
        'cargo_items': 'Cargo Items',
        'add_cargo': 'Add Cargo Item',
        'tracking': 'Tracking',
        'documents': 'Documents',
        'total_weight': 'Total Weight (kg)',
        'total_value': 'Total Value',
        'customs_cleared': 'Customs Cleared',
        'in_transit': 'In Transit',
        'delivered': 'Delivered',
        'cancelled': 'Cancelled',
        'item_name': 'Item Name',
        'quantity': 'Quantity',
        'weight': 'Weight (kg)',
        'value': 'Value',
        'description': 'Description',
        'location': 'Location',
        'update_tracking': 'Update Tracking',
        'upload_document': 'Upload Document',
        'cargo_requests': 'Cargo Requests',
        'request_cargo_change': 'Request Cargo Change',
        'manage_cargo_requests': 'Manage Cargo Requests',
        'request_type': 'Request Type',
        'modify': 'Modify',
        'remove': 'Remove',
        'request_reason': 'Reason for Request',
        'my_cargo_requests': 'My Cargo Requests',
        'approve': 'Approve',
        'reject': 'Reject',
        'request_status': 'Request Status',
        'client_name': 'Client Name',
        'edit_shipment': 'Edit Shipment',
        'delete_shipment': 'Delete Shipment',
        'messages': 'Messages',
        'send_message': 'Send Message',
        'message_subject': 'Subject',
        'message_content': 'Message',
        'to': 'To',
        'from': 'From',
        'date': 'Date',
        'reply': 'Reply',
        'finance_dashboard': 'Finance Dashboard',
        'invoices': 'Invoices',
        'payments': 'Payments',
        'expenses': 'Expenses',
        'financial_reports': 'Financial Reports',
        'marketing_dashboard': 'Marketing Dashboard',
        'campaigns': 'Campaigns',
        'marketing_analytics': 'Analytics',
        'social_media': 'Social Media',
        'it_dashboard': 'IT Dashboard',
        'system_management': 'System Management',
        'support_tickets': 'Support Tickets',
        'security': 'Security',
        'logistics_dashboard': 'Logistics Dashboard',
        'routes': 'Routes',
        'delivery_assignments': 'Delivery Assignments',
        'customer_service_dashboard': 'Customer Service Dashboard',
        'client_tickets': 'Client Tickets',
        'customer_feedback': 'Customer Feedback',
        'administration_dashboard': 'Administration Dashboard',
        'meetings': 'Meetings',
        'contracts': 'Contracts',
        'training': 'Training',
        'sales_dashboard': 'Sales Dashboard',
        'leads': 'Leads',
        'deals': 'Deals',
        'offers': 'Offers',
        'rate_service': 'Rate Service',
        'main_dashboard': 'Main Dashboard',
        'total_employees': 'Total Employees',
        'total_departments': 'Departments',
        'avg_salary': 'Average Salary',
        'active_employees': 'Active Employees',
        'recent_records': 'Recent Records',
        'no_data': 'No data available',
        'view_all_data': 'View All Data',
        'search_placeholder': 'Search by name, department, position...',
        'filter_dept': 'Filter by Department:',
        'filter_status': 'Filter by Status:',
        'filter_role': 'Filter by Role:',
        'no_display': 'No data to display',
        'add_new_record': 'Add New Record',
        'account_settings_hdr': 'Account Settings',
        'create_login': 'Create login account for this employee',
        'account_role': 'Account Role',
        'auto_gen_pass': 'Auto-generate password',
        'personal_info': 'Personal Information',
        'employee_name': 'Employee Name *',
        'phone_field': 'Phone',
        'hire_date_field': 'Hire Date *',
        'reg_date': 'Registration Date *',
        'salary_field': 'Salary ($) *',
        'position_field': 'Position *',
        'dept_field': 'Department *',
        'all': 'All',
        'active': 'Active',
        'inactive': 'Inactive',
        'on_leave': 'On Leave',
        'add_record_btn': '➕ Add Record',
        'edit_existing': 'Edit Existing Record',
        'select_edit': 'Select record to edit:',
        'save_changes': '💾 Save Changes',
        'delete_record_hdr': 'Delete Record',
        'select_delete': 'Select record to delete:',
        'confirm_del': 'Type "DELETE" to confirm:',
        'delete_btn': '🗑️ Delete Record',
        'it_pages_label': 'IT Pages',
        'db_management': 'Database Management',
        'security_mgmt': 'Security Management',
        'performance_monitor': 'Performance Monitoring',
        'bug_tracking': 'Bug Tracking',
        'activity_logs': 'Activity Logs',
        'manage_users_hdr': 'Manage Users',
        'manage_leaves_hdr': 'Manage Leave Requests',
        'leave_requests_hdr': 'Leave Requests',
        'my_leave_requests': 'My Leave Requests',
        'no_leave_requests': 'No leave requests found.',
        'request_leave_hdr': 'Request Leave',
        'submit_leave': 'Submit Request',
        'update_status': 'Update Status',
        'response_note': 'Response note',
        'filter_priority': 'Filter Priority',
        'filter_status_lbl': 'Filter Status',
        'search_lbl': 'Search employee',
        'all_priorities': 'All Priorities',
        'all_statuses': 'All Statuses',
        'no_messages': 'No messages.',
        'compose': 'Compose',
        'inbox': 'Inbox',
        'sent_msgs': 'Sent',
        'subject': 'Subject',
        'message_body': 'Message Body',
        'send_to': 'Send To',
        'reply_btn': 'Reply',
        'delete_msg': 'Delete',
        'send': 'Send',
        'priority': 'Priority',
        'low': 'Low',
        'medium': 'Medium',
        'high': 'High',
        'urgent': 'Urgent',
        'open': 'Open',
        'in_progress': 'In Progress',
        'resolved': 'Resolved',
        'closed': 'Closed',
        'tab_distribution': 'Distribution',
        'tab_salaries': 'Salaries',
        'tab_trends': 'Trends',
        'tab_details': 'Details',
        'tab_cargo_items': 'Cargo Items',
        'tab_tracking': 'Tracking',
        'tab_documents': 'Documents',
        'tab_all_invoices': 'All Invoices',
        'tab_new_invoice': 'New Invoice',
        'tab_payment_history': 'Payment History',
        'tab_record_payment': 'Record Payment',
        'tab_all_expenses': 'All Expenses',
        'tab_add_expense': 'Add Expense',
        'tab_all_campaigns': 'All Campaigns',
        'tab_new_campaign': 'New Campaign',
        'tab_add_metrics': 'Add Metrics',
        'tab_performance': 'Performance',
        'tab_customer_behavior': 'Customer Behavior',
        'tab_competitors': 'Competitors',
        'tab_posts': 'Posts',
        'tab_new_post': 'New Post',
        'tab_stats': 'Stats',
        'tab_new_ticket': 'New Ticket',
        'tab_all_tickets': 'All Tickets',
        'tab_security_events': 'Security Events',
        'tab_user_activity': 'User Activity',
        'tab_all_shipments': 'All Shipments',
        'tab_new_shipment': 'New Shipment',
        'tab_tracking_timeline': 'Tracking Timeline',
        'tab_active_routes': 'Active Routes',
        'tab_add_route': 'Add Route',
        'tab_unassigned': 'Unassigned',
        'tab_active_assignments': 'Active Assignments',
        'tab_completed': 'Completed',
        'tab_all_docs': 'All Documents',
        'tab_add_doc': 'Add Document',
        'tab_upcoming': 'Upcoming',
        'tab_past_meetings': 'Past Meetings',
        'tab_schedule': 'Schedule Meeting',
        'tab_all_contracts': 'All Contracts',
        'tab_new_contract': 'New Contract',
        'tab_overview': 'Overview',
        'tab_all_programs': 'All Programs',
        'tab_add_program': 'Add Program',
        'tab_info': 'Info',
        'tab_security': 'Security',
        'tab_activity': 'My Activity',
        'tab_feedback': 'Feedback Received',
        'tab_add_feedback': 'Add Feedback',
        'department': 'Department',
        'dark_mode': 'Dark Mode',
        'light_mode': 'Light Mode',
        'search_data': 'Search Data',
        'found_results': 'results found',
        'no_records_role': 'No records found for role',
        'confirm_delete_type': 'Type DELETE to confirm',
        'record_deleted': 'Record deleted successfully!',
        'record_added': 'Record added successfully!',
        'record_updated': 'Record updated successfully!',
        'changes_saved': 'Changes saved successfully!',
        'fill_required': 'Please fill all required fields',
        'passwords_no_match': 'Passwords do not match',
        'invalid_email': 'Invalid email format',
        'login_required_msg': 'Please log in to access these pages.',
        'go_login': 'Go to Login',
        'manager': 'Manager',
        'record_info': 'Record Information',
        'name_field': 'Name',
        'position': 'Position',
        'salary_display': 'Salary',
        'hire_date_display': 'Hire Date',
        'reg_date_display': 'Registration Date',
        'warning_undone': '⚠️ Warning: This action cannot be undone!',
        'cancel': 'Cancel',
        'confirm_delete_btn': '🗑️ Confirm Delete',
        'delete_cancelled': 'Delete operation cancelled.',
        'no_delete': 'No records to delete',
        'no_edit': 'No records to edit',
        'save_client_info': '💾 Save Client Info',
        'avg_salary_label': 'Average Salary',
        'highest_salary': 'Highest Salary',
        'lowest_salary': 'Lowest Salary',
        'total_payroll': 'Total Payroll',
        'emp_distribution': 'Employee Distribution by Department',
        'emp_status': 'Employee Status',
        'common_positions': 'Common Positions',
        'salary_analysis': 'Salary Analysis',
        'hire_trends': 'Hiring Trends',
        'monthly_hires': 'Monthly Hires',
        'emp_trends': 'Employee Trends',
        'hiring_timeline': 'Hiring Timeline',
        'emp_added_time': 'Employees Added Over Time',
        'dept_sizes': 'Department Sizes',
        'status_breakdown': 'Status Breakdown',
        'avg_salary_by_dept': 'Average Salary by Department',
        'salary_dist': 'Salary Distribution',
        'export_employees_hdr': 'Export Employees',
        'export_shipments_hdr': 'Export Shipments',
        'select_columns': 'Select columns to include',
        'columns_label': 'Columns',
        'preview_columns': 'Preview selected columns',
        'show_all_rows': 'Show all rows',
        'export_format': 'Export format',
        'format_label': 'Format',
        'download_csv': '⬇️ Download CSV',
        'download_excel': '⬇️ Download Excel',
        'download_csv_fallback': '⬇️ Download CSV (fallback)',
        'select_one_col': 'Select at least one column to export.',
        'no_records_export': 'No employee records to export.',
        'no_shipments_export': 'No shipments to export.',
        'submit_leave_card': 'Submit Leave Request',
        'date_range_err': 'End date must be the same or after start date.',
        'leave_submitted': 'Leave request submitted successfully.',
        'date_label': 'Date',
        'employees_added': 'Employees Added',
        'count_label': 'Count',
        'forgot_pass_btn': '🔑 Forgot Password?',
        'back_to_login': '⬅️ Back to Login',
        'reset_your_pass': '🔑 Reset Your Password',
        'your_email': '📧 Your Registered Email:',
        'contact_email': '📬 Contact Email (where manager will send new password):',
        'reset_info': '💡 Enter a personal email (Gmail, Outlook, etc.) where the manager can send your new password.',
        'submit_request_btn': '📤 Submit Request',
        'full_name_signup': 'Full Name *',
        'phone_number': 'Phone Number',
        'emp_status_field': 'Employment Status *',
        'change_type': '⬅ Change',
        'server_status': 'Server Status',
        'running': 'Running',
        'restart_server': 'Restart Server',
        'check_status': 'Check System Status',
        'server_restarted': 'Server restart command sent!',
        'status_checked': 'System status checked.',
        'last_backup': 'Last Backup',
        'db_size': 'Database Size',
        'create_backup': 'Create Backup',
        'restore_backup': 'Restore Backup',
        'backup_created': 'Backup created!',
        'restore_started': 'Restore started (demo only)',
        'bug_title': 'Bug Title',
        'bug_desc': 'Bug Description',
        'log_bug': 'Log Bug',
        'bug_logged': 'Bug logged!',
        'cpu_usage': 'CPU Usage',
        'mem_usage': 'Memory Usage',
        'active_users': 'Active Users',
        'recent_activities': 'Recent Activities',
        'manage_leave_response': 'Response to user:',
        'account_info': 'Account Information',
        'employment_details': 'Employment Details',
        'email_address': 'Email Address *',
        'create_emp_account': 'Create Employee Account',
        'create_client_account': 'Create Client Account',
        'provide_email_pass': 'Please provide email and password',
        'enter_full_name': 'Please enter your full name',
        'support_ticket_sys': 'Support Ticket System',
        'create_ticket': 'Create a new support ticket:',
        'ticket_title_lbl': 'Title',
        'ticket_desc_lbl': 'Description',
        'submit_ticket_btn': 'Submit Ticket',
        'fill_all_fields': 'Please fill all fields.',
        'all_tickets_lbl': 'All Tickets:',
        'no_tickets': 'No tickets found.',
        'user_list': 'User List:',
        'change_user_role': 'Change user role/status (demo only)',
        'last_check': 'Last Check:',
        'log_new_bug': 'Log a new bug:',
        'all_bugs': 'All Bugs:',
        'recent_activities_lbl': 'Recent Activities:',
        'select_section': 'Select a section from the sidebar.',
        # Manage Users
        'change_user_role': 'Change User Role',
        'select_user': 'Select user:',
        'new_role': 'New role:',
        'update_role': 'Update Role',
        'reset_user_pass': '🔐 Reset User Password',
        'reset_pass_info': 'Generate a new random password for a user who forgot their credentials.',
        'select_reset_user': 'Select user to reset password:',
        'reset_pass_btn': '🔄 Reset Password',
        'pass_reset_success': '✅ Password reset successfully!',
        'pass_save_warning': '⚠️ Please save this password and provide it to the user. It will not be shown again.',
        # Finance Dashboard
        'total_revenue': 'Total Revenue',
        'total_expenses': 'Total Expenses',
        'net_profit': 'Net Profit',
        'outstanding': 'Outstanding',
        'overdue': 'Overdue',
        # Finance - Invoices
        'filter_by_status': 'Filter by Status',
        'create_new_invoice': 'Create New Invoice',
        'client_name_field': 'Client Name *',
        'shipment_ref': 'Shipment Reference',
        'amount_usd': 'Amount (USD) *',
        'tax_rate': 'Tax Rate (%)',
        'due_date': 'Due Date *',
        'desc_services': 'Description / Services',
        'create_invoice_btn': 'Create Invoice',
        'mark_sent': 'Mark Sent',
        'mark_paid': 'Mark Paid',
        'mark_overdue': 'Mark Overdue',
        'delete_btn': 'Delete',
        'no_invoices': 'No invoices found.',
        'invoice_created': 'Invoice created successfully.',
        'client_amount_required': 'Client name and amount are required.',
        'shipment_ref_lbl': 'Shipment Ref',
        'issue_date_lbl': 'Issue Date',
        'due_date_lbl': 'Due Date',
        'amount_lbl': 'Amount',
        'tax_lbl': 'Tax',
        'total_lbl': 'Total',
        'cost_breakdown': 'Cost Breakdown',
        'cargo_value_lbl': 'Cargo / Shipment Value',
        'freight_charge_lbl': 'Freight Charge',
        'handling_fee_lbl': 'Handling Fee',
        'insurance_lbl': 'Insurance (0.4%)',
        'services_subtotal_lbl': 'Services Subtotal',
        'vat_lbl': 'VAT',
        'grand_total_lbl': 'Grand Total',
        'amount_usd_col': 'Amount (USD)',
        'download_pdf': 'Download Invoice PDF',
        'total_invoices_lbl': 'Total Invoices',
        'paid_lbl': 'Paid',
        'outstanding_lbl': 'Outstanding',
        'col_invoice': 'Invoice #',
        'col_client': 'Client',
        'col_total_usd': 'Total (USD)',
        'col_due_date': 'Due Date',
        'col_date': 'Date',
        'col_amount_usd': 'Amount (USD)',
        'col_method': 'Method',
        'col_reference': 'Reference',
        'col_month': 'Month',
        'col_revenue': 'Revenue',
        'col_net_profit': 'Net Profit',
        'chart_rev_vs_exp': 'Revenue vs Expenses (Monthly)',
        'chart_invoice_status': 'Invoice Status Breakdown',
        'chart_payments_method': 'Payments by Method',
        'chart_monthly_payments': 'Monthly Payments Received',
        'chart_approved_expenses': 'Approved Expenses by Category',
        'chart_monthly_revenue': 'Monthly Revenue',
        'chart_monthly_trend': 'Monthly Revenue, Expenses & Profit',
        'chart_expense_breakdown': 'Expense Breakdown by Category',
        'margin_lbl': 'margin',
        'paid_lbl': 'paid',
        'no_financial_data_short': 'No financial data yet.',
        'no_invoices_yet': 'No invoices yet.',
        # Finance - Payments
        'no_payments': 'No payments recorded yet.',
        'record_payment': 'Record a Payment',
        'select_invoice': 'Select Invoice *',
        'amount_received': 'Amount Received (USD) *',
        'payment_date': 'Payment Date *',
        'payment_method': 'Payment Method',
        'ref_no': 'Reference / Transaction No.',
        'record_payment_btn': 'Record Payment',
        'amount_gt_zero': 'Amount must be greater than zero.',
        'payment_recorded': 'Payment recorded and invoice marked as Paid.',
        # Finance - Expenses
        'filter_by_cat': 'Filter by Category',
        'total_filtered': 'Total (filtered)',
        'approve_btn': 'Approve',
        'reject_btn': 'Reject',
        'add_new_expense': 'Add New Expense',
        'category_field': 'Category *',
        'expense_date': 'Expense Date *',
        'vendor_supplier': 'Vendor / Supplier',
        'receipt_ref': 'Receipt Reference',
        'add_expense_btn': 'Add Expense',
        'expense_added': 'Expense added and pending approval.',
        'vendor_lbl': 'Vendor',
        'receipt_ref_lbl': 'Receipt Ref',
        'cargo_cost_opt': 'Cargo Cost', 'customs_duty_opt': 'Customs Duty',
        'freight_fee_opt': 'Freight Fee', 'fuel_transport_opt': 'Fuel & Transport',
        'staff_salary_opt': 'Staff Salary', 'office_rent_opt': 'Office Rent',
        'insurance_opt': 'Insurance', 'port_handling_opt': 'Port Handling',
        'bank_transfer_opt': 'Bank Transfer', 'letter_credit_opt': 'Letter of Credit',
        'cash_opt': 'Cash', 'credit_card_opt': 'Credit Card', 'cheque_opt': 'Cheque',
        # Finance - Reports
        'total_invoices': 'Total Invoices',
        'profit_loss': 'Profit & Loss Summary',
        'no_financial_data': 'No financial data yet. Create invoices and record expenses to see reports.',
        # Marketing Dashboard
        'campaign_performance': 'Campaign Performance',
        'budget_vs_spent': 'Budget vs Spent',
        'recent_campaigns': 'Recent Campaigns',
        'no_campaign_metrics': 'No campaign metrics yet.',
        'no_budget_data': 'No budget data yet.',
        'no_campaigns': 'No campaigns yet.',
        # Campaigns
        'filter_by_status_lbl': 'Filter by Status',
        'change_status': 'Change Status',
        'update_btn': 'Update',
        'campaign_name': 'Campaign Name',
        'type_field': 'Type',
        'budget_field': 'Budget ($)',
        'start_date_field': 'Start Date',
        'end_date_field': 'End Date',
        'target_audience': 'Target Audience',
        'create_campaign_btn': 'Create Campaign',
        'campaign_name_required': 'Campaign name is required.',
        'create_campaign_first': 'Create a campaign first.',
        'date_field': 'Date',
        'impressions': 'Impressions',
        'clicks': 'Clicks',
        'conversions': 'Conversions',
        'users': 'Users',
        'revenue_field': 'Revenue ($)',
        'save_metrics': 'Save Metrics',
        'metrics_saved': 'Metrics saved!',
        # Marketing Analytics
        'filter_by_campaign': 'Filter by Campaign',
        'customer_behavior': 'Customer Behavior Analysis',
        'competitor_analysis': 'Competitor Analysis',
        'add_competitor': 'Add Competitor',
        'name_field': 'Name',
        'website_field': 'Website',
        'strengths_field': 'Strengths',
        'weaknesses_field': 'Weaknesses',
        'no_competitors': 'No competitors added yet.',
        'no_metrics_data': 'No metrics data yet. Add metrics from the Campaigns page.',
        # Social Media
        'platform_field': 'Platform',
        'content_field': 'Content',
        'schedule_date': 'Schedule Date',
        'link_to_campaign': 'Link to Campaign (optional)',
        'save_post': 'Save Post',
        'update_engagement': 'Update Engagement',
        'select_post': 'Select Post',
        'likes_field': 'Likes',
        'shares_field': 'Shares',
        'comments_field': 'Comments',
        # IT Dashboard
        'ram_usage': 'RAM Usage',
        'db_response_time': 'DB Response Time',
        'open_tickets': 'Open Tickets',
        'system_performance': 'System Performance',
        'total_users': 'Total Users',
        'new_this_month': 'New This Month',
        'security_last7': 'Security (last 7 days)',
        'failed_logins': 'Failed Login Attempts',
        'recent_activity_log': 'Recent Activity Log',
        'no_activity': 'No activity logged yet.',
        'live_metrics': 'Live System Metrics',
        'perf_history': 'Performance History (last 30 snapshots)',
        'network_io': 'Network I/O',
        'total_sent': 'Total Sent',
        'total_received': 'Total Received',
        # Support Tickets (employee)
        'ticket_submitted': 'Ticket submitted successfully!',
        'change_status_lbl': 'Change Status',
        'status_updated': 'Status updated!',
        'ticket_deleted': 'Ticket deleted.',
        'delete_ticket_btn': 'Delete Ticket',
        'edit_ticket_btn': 'Edit Ticket',
        'save_ticket_btn': 'Save Changes',
        'cancel_btn': 'Cancel',
        'ticket_edited_success': 'Ticket updated successfully.',
        'cs_ticket_deleted_ok': 'Ticket deleted.',
        'confirm_delete_ticket': 'Are you sure you want to delete this ticket?',
        # Security page
        'security_events_log': 'Security Events Log',
        'filter_event_type': 'Filter by Event Type',
        'registered_users': 'Registered Users',
        # Logistics
        'total_shipments': 'Total Shipments',
        'delivered_month': 'Delivered (Month)',
        'pending_customs': 'Pending / Customs',
        'overdue': 'Overdue',
        'recent_shipments': 'Recent Shipments',
        # Shipments
        'cargo_items_lbl': 'Cargo Items:',
        'tracking_history_lbl': 'Tracking History:',
        'change_status_lbl2': 'Change Status',
        'save_update': 'Save Update',
        'create_new_shipment': 'Create New Shipment',
        'freight_mode': 'Freight Mode *',
        'origin_country': 'Origin Country *',
        'dest_country': 'Destination Country *',
        'departure_date': 'Departure Date *',
        'expected_arrival_field': 'Expected Arrival *',
        'carrier_name_field': 'Carrier Name',
        'container_type': 'Container Type',
        'incoterms_field': 'Incoterms',
        'total_weight_field': 'Total Weight (kg)',
        'total_value_field': 'Total Value (USD)',
        'notes_desc': 'Notes / Description',
        'create_shipment_btn': 'Create Shipment',
        'shipment_tracking': 'Shipment Tracking Timeline',
        'select_shipment': 'Select Shipment',
        # Routes
        'filter_mode': 'Filter Mode',
        'add_new_route': 'Add New Route',
        'route_name': 'Route Name *  (e.g. Istanbul → Dubai)',
        'origin_port': 'Origin Port / Airport *',
        'dest_port': 'Destination Port / Airport *',
        'transit_days': 'Transit Days',
        'frequency': 'Frequency  (e.g. Weekly)',
        'add_route_btn': 'Add Route',
        'route_required': 'Route name, origin, and destination are required.',
        # Delivery Assignments
        'assign_carrier_hdr': '**Assign Carrier:**',
        'carrier_name_req': 'Carrier Name *',
        'assign_carrier_btn': 'Assign Carrier',
        'carrier_required': 'Carrier name required.',
        'all_assigned': 'All shipments are assigned to carriers.',
        'update_status_lbl': 'Update Status',
        'mark_delivered': '✅ Mark Delivered',
        'no_active_assignments': 'No active assignments.',
        'no_completed': 'No completed deliveries yet.',
        # Customer Service Dashboard
        'total_tickets': 'Total Tickets',
        'open_status': 'Open',
        'resolved_today': 'Resolved Today',
        'avg_rating': 'Avg Rating',
        'open_urgent_tickets': 'Open & Urgent Tickets',
        'no_open_tickets': 'No open tickets. All resolved!',
        'needs_attention': 'Needs attention',
        'chart_tickets_status': 'Tickets by Status',
        'chart_tickets_category': 'Tickets by Category',
        'unread_feedback_msg': '{n} unread customer feedback awaiting review.',
        'tickets_found': 'ticket(s) found',
        'created_lbl': 'Created',
        'resolved_at_lbl': 'Resolved at',
        'resolution_lbl': 'Resolution',
        'chart_rating_dist': 'Rating Distribution',
        'total_word': 'total',
        # Department names (sidebar)
        'dept_finance': 'Finance', 'dept_marketing': 'Marketing', 'dept_it': 'IT',
        'dept_logistics': 'Logistics', 'dept_customer_service': 'Customer Service',
        'dept_administration': 'Administration', 'dept_sales': 'Sales',
        # Client Tickets
        'create_new_ticket': 'Create New Ticket',
        'update_status_btn': 'Update Status',
        'resolution_response': 'Resolution / Response',
        'update_ticket_btn': 'Update Ticket',
        'client_subj_desc_required': 'Client, subject, and description are required.',
        'create_ticket_btn': 'Create Ticket',
        # Ticket Replies
        'ticket_conversation': 'Conversation',
        'reply_from_cs': 'Customer Service',
        'reply_from_client': 'Client',
        'type_reply': 'Write a reply...',
        'send_reply_btn': 'Send Reply',
        'reply_sent': 'Reply sent.',
        'reply_required': 'Reply cannot be empty.',
        'no_replies_yet': 'No messages yet. Start the conversation.',
        # Customer Feedback
        'show_unread_only': 'Show unread only',
        'total_feedback': 'Total Feedback',
        'average_rating': 'Average Rating',
        'unread_count': 'Unread',
        'mark_as_read': 'Mark as Read',
        'add_cust_feedback': 'Add Customer Feedback',
        'rating_field': 'Rating *',
        'comment_review': 'Comment / Review',
        'submit_feedback_btn': 'Submit Feedback',
        'select_client': 'Please select a client.',
        # Administration Dashboard
        'total_documents': 'Total Documents',
        'active_contracts': 'Active Contracts',
        'upcoming_meetings': 'Upcoming Meetings',
        'expiring_30d': 'Expiring in 30 days',
        'total_contracts_lbl': 'Total Contracts',
        'upcoming_meetings_hdr': 'Upcoming Meetings',
        'contracts_expiring': 'Contracts Expiring Soon',
        'recent_docs': 'Recent Documents',
        'no_upcoming_meetings': 'No upcoming meetings.',
        'no_contracts_expiring': 'No contracts expiring in the next 90 days.',
        # Documents
        'add_new_doc': 'Add New Document',
        'doc_title': 'Document Title *',
        'expiry_date_field': 'Expiry Date (optional)',
        'file_name_field': 'File Name (e.g. contract_2026.pdf)',
        'set_expiry': 'Set expiry date',
        'add_doc_btn': 'Add Document',
        'archive_btn': 'Archive',
        'restore_btn': 'Restore',
        # Meetings
        'update_meeting': 'Update',
        'meeting_minutes': 'Meeting Minutes',
        'save_btn': 'Save',
        'schedule_new_meeting': 'Schedule New Meeting',
        'meeting_title': 'Title *',
        'meeting_type': 'Meeting Type *',
        'date_field2': 'Date *',
        'time_field': 'Time *',
        'location_platform': 'Location / Platform',
        'duration_min': 'Duration (min)',
        'attendees_field': 'Attendees (names or roles)',
        'agenda_field': 'Agenda',
        'schedule_meeting_btn': 'Schedule Meeting',
        'meeting_scheduled': "Meeting scheduled.",
        'meeting_title_required': 'Title is required.',
        # Contracts
        'party_type_filter': 'Party Type',
        'active_contracts_value': 'Active Contracts Total Value',
        'add_new_contract': 'Add New Contract',
        'contract_title': 'Contract Title *',
        'party_name': 'Party Name *  (company/person)',
        'party_type_field': 'Party Type *',
        'contract_type_field': 'Contract Type *',
        'value_field': 'Value',
        'currency_field': 'Currency',
        'start_date_lbl': 'Start Date *',
        'end_date_lbl': 'End Date *',
        'add_contract_btn': 'Add Contract',
        'title_party_required': 'Title and party name are required.',
        'end_after_start': 'End date must be after start date.',
        'contract_added_success': 'Contract added successfully.',
        'party_type_lbl': 'Party Type', 'contract_type_lbl': 'Contract Type',
        'value_lbl': 'Value', 'start_lbl': 'Start', 'end_lbl': 'End',
        'update_btn_lbl': 'Update',
        'carrier_opt': 'Carrier', 'supplier_opt': 'Supplier', 'government_opt': 'Government',
        'terminated_opt': 'Terminated',
        'service_agreement_opt': 'Service Agreement', 'rate_agreement_opt': 'Rate Agreement',
        'agency_agreement_opt': 'Agency Agreement', 'insurance_policy_opt': 'Insurance Policy',
        'software_license_opt': 'Software License', 'lease_agreement_opt': 'Lease Agreement',
        'nda_opt': 'NDA',
        # Sales Dashboard
        'total_leads': 'Total Leads',
        'active_leads': 'Active Leads',
        'open_deals': 'Open Deals',
        'pipeline_value': 'Pipeline Value',
        'won_deals': 'Won Deals',
        'win_rate': 'Win Rate',
        'recent_leads': 'Recent Leads',
        # Leads
        'source_filter': 'Source',
        'freight_filter': 'Freight',
        'update_status_lead': 'Update Status',
        'add_new_lead': '➕ Add New Lead',
        'lead_name': 'Lead Name *',
        'company_name': 'Company Name',
        'email_field': 'Email',
        'phone_field': 'Phone',
        'country_field': 'Country',
        'source_field': 'Source *',
        'freight_interest': 'Freight Interest',
        'add_lead_btn': 'Add Lead',
        'lead_name_required': 'Lead name is required.',
        # Deals
        'stage_filter': 'Stage',
        'freight_type_filter': 'Freight Type',
        'add_new_deal': '➕ Add New Deal',
        'deal_title_field': 'Deal Title *',
        'value_usd_field': 'Value (USD)',
        'probability_pct': 'Probability %',
        'expected_close': 'Expected Close Date',
        'origin_lbl': 'Origin',
        'destination_lbl': 'Destination',
        'notes_lbl': 'Notes',
        'add_deal_btn': 'Add Deal',
        'deal_title_client_req': 'Title and Client Name are required.',
        'failed_add_deal': 'Failed to add deal.',
        'move_to_stage': 'Move to Stage',
        'update_stage_btn': 'Update Stage',
        'no_deals': 'No deals found matching filters.',
        'freight_lbl': '**Freight:**',
        'route_lbl': '**Route:**',
        'probability_lbl': '**Probability:**',
        'close_date_lbl': '**Close Date:**',
        # Offers
        'freight_filter_lbl': 'Freight',
        'total_offers': 'Total Offers',
        'pending_sent': 'Pending (Sent)',
        'accepted_value': 'Accepted Value',
        'commodity_lbl': '**Commodity:**',
        'valid_until_lbl': '**Valid Until:**',
        'weight_lbl': '**Weight:**',
        'volume_lbl': '**Volume:**',
        'update_status_offer': 'Update Status',
        'no_offers': 'No offers found matching filters.',
        'create_new_offer_exp': '➕ Create New Offer',
        'offer_num_field': 'Offer Number *',
        'client_email_field': 'Client Email',
        'commodity_desc_field': 'Commodity / Cargo Description',
        'weight_kg_field': 'Weight (kg)',
        'volume_cbm_field': 'Volume (CBM)',
        'total_value_usd': 'Total Value (USD)',
        'valid_until_field': 'Valid Until',
        'notes_incl': 'Notes / Inclusions',
        'create_offer_btn': 'Create Offer',
        'offer_num_client_req': 'Offer number and client name are required.',
        'failed_create_offer': 'Failed to create offer. Offer number may already exist.',
        'leads_by_status': 'Leads by Status',
        'deals_pipeline_chart': 'Deals Pipeline',
        'leads_found': 'leads found',
        'deals_found': 'deals',
        'pipeline_lbl': 'Pipeline',
        'col_name': 'Name', 'col_company': 'Company', 'col_country': 'Country',
        'col_source': 'Source', 'col_freight': 'Freight', 'col_created': 'Created',
        'phone_lbl': 'Phone',
        'lead_added_success': "Lead added successfully.",
        'deal_added_success': "Deal added successfully.",
        'offer_created_success': "Offer created successfully.",
        'stage_updated_success': 'Stage updated.',
        'offer_status_updated': 'Offer status updated.',
        'discovery_opt': 'Discovery', 'proposal_opt': 'Proposal',
        'negotiation_opt': 'Negotiation',
        # Training
        'total_programs': 'Total Programs',
        'upcoming_lbl': 'Upcoming',
        'total_enrollments': 'Total Enrollments',
        'completions_lbl': 'Completions',
        'upcoming_training': 'Upcoming Training',
        'programs_by_cat': 'Programs by Category',
        'cat_filter_lbl': 'Category',
        'enroll_employee': 'Enroll Employee',
        'enroll_btn': 'Enroll',
        'update_status_prog': 'Update Status',
        'update_prog_status': 'Update Program Status',
        'enrolled_participants': '**Enrolled Participants:**',
        'no_programs': "No programs yet. Add one from the 'Add Program' tab.",
        'add_program_hdr': 'Add Training Program',
        'program_title': 'Program Title *',
        'scheduled_date': 'Scheduled Date *',
        'duration_hours': 'Duration (hours)',
        'trainer_instructor': 'Trainer / Instructor',
        'max_participants': 'Max Participants',
        'description_area': 'Description',
        'add_program_btn': 'Add Program',
        'title_required_err': 'Title is required.',
        # Profile page
        'please_login': 'Please log in to view profiles.',
        'no_users_avail': 'No users available.',
        'select_user_view': 'Select user to view',
        'my_profile_btn': 'My Profile',
        'user_not_found': 'User not found.',
        'change_profile_pic': 'Change Profile Picture',
        'upload_photo': 'Upload photo (JPG / PNG, max 2 MB)',
        'file_too_large': 'File too large — max 2 MB.',
        'save_photo': 'Save Photo',
        'profile_pic_updated': 'Profile picture updated!',
        'remove_photo': 'Remove Photo',
        'photo_removed': 'Photo removed.',
        'leave_requests_lbl': 'Leave Requests',
        'employees_lbl': 'Employees',
        'clients_lbl': 'Clients',
        'leave_req_pending': 'Leave Requests Pending',
        'active_lbl': 'Active',
        'delivered_lbl': 'Delivered',
        'customs_cleared': 'Customs Cleared',
        'full_name_lbl': 'Full Name',
        'phone_lbl': 'Phone',
        'role_lbl': 'Role',
        'save_info': 'Save Info',
        'profile_updated': 'Profile updated successfully.',
        'new_pass_auto': 'New password (blank = auto-generate)',
        'reset_password_btn': 'Reset Password',
        'password_reset_done': 'Password reset.',
        'failed_short': 'Failed.',
        'current_password': 'Current Password',
        'new_password': 'New Password',
        'confirm_new_password': 'Confirm New Password',
        'update_password_btn': 'Update Password',
        'pass_no_match_new': 'New passwords do not match.',
        'current_pass_incorrect': 'Current password is incorrect.',
        'password_updated': 'Password updated successfully.',
        'failed_update_pass': 'Failed to update password.',
        'no_activity_yet': 'No activity recorded yet. Actions like login, adding data, submitting requests will appear here.',
        # Messages page
        'your_messages': 'Your Messages',
        'no_messages': '📭 No messages yet.',
        'from_lbl': 'From',
        'to_lbl': 'To',
        'subject_lbl': '**Subject:**',
        'from_lbl2': '**From:**',
        'to_lbl2': '**To:**',
        'shipment_lbl': '**Shipment:**',
        'date_lbl': '**Date:**',
        'message_lbl': '**Message:**',
        'mark_as_read_btn': '✓ Mark as Read',
        'reply_btn': '↩️ Reply',
        'reply_hdr': 'Reply',
        'to_field': 'To:',
        'subject_field': 'Subject:',
        'message_field': 'Message:',
        'send_reply_btn': '📤 Send Reply',
        'fill_subject_msg': 'Please fill in subject and message content.',
        'recipient_not_found': 'Recipient not found.',
        'reply_sent': '✅ Reply sent successfully!',
        'error_sending_reply': '❌ Error sending reply.',
        'send_new_msg': 'Send New Message',
        'no_users_msg': 'No users available to send messages to.',
        'related_shipment': 'Related Shipment (optional):',
        'send_msg_btn': '📤 Send Message',
        'msg_sent': '✅ Message sent successfully!',
        'error_sending_msg': '❌ Error sending message.',
        'showing_x_of_y': 'Showing {x} of {y} requests',
        'request_num': 'Request',
        'download_attach': 'Download attachment',
        'attachment_lbl': 'Attachment',
        'user_lbl': 'User',
        'last_updated': 'Last updated',
        'delete_action': 'Delete',
        'name_lbl': 'Name',
        'total_impressions_kpi': 'Total Impressions',
        'total_clicks_kpi': 'Total Clicks',
        'avg_ctr': 'Avg CTR',
        'total_revenue_kpi': 'Total Revenue',
        'impressions_over_time': 'Impressions Over Time',
        'clicks_conv_over_time': 'Clicks & Conversions Over Time',
        'daily_revenue': 'Daily Revenue',
        'campaigns_by_type': 'Campaigns by Type',
        'ctr_by_campaign': 'Click-Through Rate by Campaign',
        'cvr_by_campaign': 'Conversion Rate by Campaign',
        'date_lbl': 'Date',
        'count_lbl': 'Count',
        'metric_lbl': 'Metric',
        'impressions_lbl': 'Impressions',
        'rate_service_caption': 'Share your experience with EIMS Logistics — your feedback helps us improve.',
        'no_delivered_ships': 'You have no delivered shipments to rate yet. Ratings become available once a shipment is marked Delivered.',
        'ships_awaiting_rating': 'Shipments Awaiting Your Rating ({n})',
        'delivered_lbl': 'Delivered',
        'rating_lbl': 'Rating *',
        'what_rating': 'What are you rating?',
        'comment_optional': 'Your comment (optional)',
        'submit_rating': 'Submit Rating',
        'rating_submitted': 'Thank you! Your rating has been submitted.',
        'all_ships_rated': '✅ You have rated all your delivered shipments. Thank you!',
        'already_rated': 'Already Rated ({n})',
        'edit_rating_btn': 'Edit Rating',
        'delete_rating_btn': 'Delete Rating',
        'confirm_delete_rating': 'Are you sure you want to delete this rating?',
        'rating_edited_success': 'Rating updated successfully.',
        'rating_deleted_ok': 'Rating deleted.',
        'cat_overall': 'Overall',
        'cat_delivery_time': 'Delivery Time',
        'cat_service_quality': 'Service Quality',
        'cat_documentation': 'Documentation',
        'cat_communication': 'Communication',
        'cat_pricing': 'Pricing',
        'track_btn': '🔍 Track',
        'ship_not_found': 'Shipment not found.',
        'own_ships_only': 'You can only track your own shipments.',
        'ship_found': 'Shipment Found: {n}',
        'tracking_history': 'Tracking History',
        'no_tracking_updates': 'No tracking updates available yet.',
        'enter_ship_number': 'Please enter a shipment number.',
        'origin_lbl': 'Origin',
        'destination_lbl': 'Destination',
        'likes_lbl': 'Likes',
        'shares_lbl': 'Shares',
        'comments_lbl': 'Comments',
        'write_post_ph': 'Write your post...',
        'engagement_by_platform': 'Engagement by Platform',
        'posts_distribution': 'Posts Distribution',
        'posts_lbl': 'Posts',
        'platform_lbl': 'Platform',
        'update_engagement_btn': 'Update Engagement',
        'users_by_role': 'Users by Role',
        'events_by_type': 'Events by Type',
        'role_col': 'Role',
        'shipments_by_status': 'Shipments by Status',
        'monthly_ship_vol': 'Monthly Shipment Volume',
        'needs_action': 'Needs action',
        'shipments_lbl': 'Shipments',
        'month_lbl': 'Month',
        'shipment_hash': 'Shipment #',
        'mode_lbl': 'Mode',
        'carrier_lbl': 'Carrier',
        'departure_lbl': 'Departure',
        'eta_lbl': 'ETA',
        'footer_text': 'EIMS | Company Data Management System',
        'category_lbl': 'Category',
        'expiry_lbl': 'Expiry',
        'x_documents': '{n} document(s)',
        'active_opt': 'Active',
        'inactive_opt': 'Inactive',
        'on_leave_opt': 'On Leave',
        'draft_opt': 'Draft',
        'scheduled_opt': 'Scheduled',
        'published_opt': 'Published',
        'paused_opt': 'Paused',
        'completed_opt': 'Completed',
        'cancelled_opt': 'Cancelled',
        'archived_opt': 'Archived',
        'expired_opt': 'Expired',
        'open_opt': 'Open',
        'closed_opt': 'Closed',
        'in_progress_opt': 'In Progress',
        'low_opt': 'Low',
        'medium_opt': 'Medium',
        'high_opt': 'High',
        'urgent_opt': 'Urgent',
        'import_opt': 'Import',
        'export_opt': 'Export',
        'sea_opt': 'Sea',
        'air_opt': 'Air',
        'land_opt': 'Land',
        'paid_opt': 'Paid',
        'unpaid_opt': 'Unpaid',
        'sick_opt': 'Sick',
        'other_opt': 'Other',
        'employee_opt': 'Employee',
        'client_opt': 'Client',
        'manager_opt': 'Manager',
        'male_opt': 'Male',
        'female_opt': 'Female',
        'fulltime_opt': 'Full-time',
        'parttime_opt': 'Part-time',
        'select_placeholder': '— Select —',
        'overall_opt': 'Overall',
        'delivery_time_opt': 'Delivery Time',
        'service_quality_opt': 'Service Quality',
        'documentation_opt': 'Documentation',
        'communication_opt': 'Communication',
        'pricing_opt': 'Pricing',
        'sent_opt': 'Sent',
        'overdue_opt': 'Overdue',
        'email_type_opt': 'Email',
        'social_media_opt': 'Social Media',
        'content_type_opt': 'Content',
        'event_opt': 'Event',
        'none_opt': 'None',
        'category_label': 'Category',
        'doc_type_lbl': 'Document Type',
        'confirmed_opt': 'Confirmed',
        'in_transit_opt': 'In Transit',
        'customs_opt': 'Customs',
        'delivered_opt': 'Delivered',
        'resolved_opt': 'Resolved',
        'departed_opt': 'Departed',
        'arrived_port_opt': 'Arrived at Port',
        'out_for_delivery_opt': 'Out for Delivery',
        'shipment_inquiry_opt': 'Shipment Inquiry',
        'doc_request_opt': 'Document Request',
        'customs_issue_opt': 'Customs Issue',
        'complaint_opt': 'Complaint',
        'rate_request_opt': 'Rate Request',
        'general_inquiry_opt': 'General Inquiry',
        'new_opt': 'New',
        'contacted_opt': 'Contacted',
        'qualified_opt': 'Qualified',
        'proposal_sent_opt': 'Proposal Sent',
        'won_opt': 'Won',
        'lost_opt': 'Lost',
        'cold_call_opt': 'Cold Call',
        'referral_opt': 'Referral',
        'trade_show_opt': 'Trade Show',
        'website_opt': 'Website',
        'email_campaign_opt': 'Email Campaign',
        'sea_fcl_opt': 'Sea FCL',
        'sea_lcl_opt': 'Sea LCL',
        'air_cargo_opt': 'Air Cargo',
        'road_opt': 'Road',
        'accepted_opt': 'Accepted',
        'invoice_doc_opt': 'Invoice',
        'bill_of_lading_opt': 'Bill of Lading',
        'customs_declaration_opt': 'Customs Declaration',
        'certificate_origin_opt': 'Certificate of Origin',
        'packing_list_opt': 'Packing List',
        'currency_lbl': 'Currency',
        'filter_by_type': 'Filter by Type:',
        'select_your_shipment': 'Select Your Shipment:',
        'select_cargo_item': 'Select Cargo Item:',
        'select_shipment_edit': 'Select Shipment to Edit:',
        'select_shipment_delete': 'Select Shipment to Delete:',
        'action_lbl': 'Action:',
        'type_lbl': 'Type',
        'location_lbl': 'Location',
        'duration_lbl': 'Duration',
        'min_lbl': 'min',
        'attendees_lbl': 'Attendees',
        'agenda_lbl': 'Agenda',
        'minutes_lbl': 'Minutes',
        'no_upcoming_meetings': 'No upcoming meetings.',
        'no_past_meetings': 'No past meetings.',
        'active_campaigns_kpi': 'Active Campaigns',
        'total_budget_kpi': 'Total Budget',
        'scheduled_posts_kpi': 'Scheduled Posts',
        'x_total': '{n} total',
        'x_spent': '${n} spent',
        'x_clicks': '{n} clicks',
        'budget_lbl': 'Budget',
        'spent_lbl': 'Spent',
        'period_lbl': 'Period',
        'target_aud_lbl': 'Target Audience',
        'no_campaign_metrics': 'No campaign metrics yet.',
        'no_budget_data': 'No budget data yet.',
        'campaign_created_ok': "Campaign '{name}' created!",
        'metrics_saved_ok': 'Metrics saved!',
        'email_not_found': 'No account found with that email. Please sign up.',
        'incorrect_password': 'Incorrect password',
        'login_error': 'An error occurred during login. Please try again.',
        'forgot_pass_desc': 'Enter your information to request a password reset',
        'enter_reg_email': 'Please enter your registered email address',
        'enter_contact_email': 'Please enter a contact email address',
        'reset_email_sent_maybe': 'If this email is registered, a reset request will be sent to the manager.',
        'no_managers_available': 'No managers available. Please contact technical support.',
        'request_submitted_success': '✅ Request submitted successfully!',
        'error_sending_request': 'An error occurred while sending the request. Please try again later.',
        'error_try_again': 'An error occurred. Please try again later.',
        'create_new_account': 'Create New Account',
        'choose_account_type': 'Choose the type of account you want to register',
        'employee_account_lbl': 'Employee Account',
        'client_account_lbl': 'Client Account',
        'invalid_email_format': '❌ Please enter a valid email address',
        'email_already_exists': 'An account with this email already exists.',
        'failed_create_account': 'Failed to create account. Try again.',
        'emp_registration_hdr': 'Employee Registration',
        'client_registration_hdr': 'Client Registration',
        'register_emp_help': 'Register as an employee with complete access',
        'register_client_help': 'Register as a client with basic access',
        'origin_dest_same': 'Origin and destination cannot be the same country.',
        'failed_create_shipment_err': 'Failed to create shipment.',
        'failed_add_route_err': 'Failed to add route.',
        'no_psutil': 'The psutil package is not installed. Install it to use System Management.',
        'content_required': 'Content is required.',
        'please_enter_location': 'Please enter location',
        'please_select_file': 'Please select a file',
        'only_managers_export': 'Only managers can export employee data.',
        'only_emp_mgr_export': 'Only employees or managers can export shipments.',
        'only_mgr_emp_ships': 'You must be a manager or employee to manage shipments.',
        'no_clients_add_ship': 'Please create at least one client user before adding shipments.',
        'page_for_clients': 'This page is only for clients.',
        'must_be_manager_leave': 'You must be a manager to manage leave requests.',
        'must_be_manager_users_err': 'You must be a manager to manage users.',
        'failed_update_role_db': 'Failed to update role in database.',
        'failed_login_submit': 'Please log in to submit a leave request.',
        'failed_create_ticket_err': 'Failed to create ticket.',
        'failed_submit_feedback_err': 'Failed to submit feedback.',
        'failed_add_document_err': 'Failed to add document.',
        'failed_schedule_meeting_err': 'Failed to schedule meeting.',
        'failed_add_contract_err': 'Failed to add contract.',
        'failed_add_lead_err': 'Failed to add lead.',
        'failed_create_invoice_err': 'Failed to create invoice.',
        'failed_record_payment_err': 'Failed to record payment.',
        'failed_add_expense_err': 'Failed to add expense.',
        'status_updated_success': '✅ Status updated successfully!',
        'cargo_item_updated': '✅ Cargo item updated successfully!',
        'item_deleted_success': '✅ Item deleted successfully!',
        'cargo_item_added': '✅ Cargo item added successfully!',
        'tracking_added': '✅ Tracking update added successfully!',
        'document_uploaded': 'Document uploaded!',
        'request_updated': 'Request updated.',
        'user_role_updated': 'User role updated.',
        'ticket_updated_success': 'Ticket updated.',
        'feedback_submitted_success': 'Feedback submitted.',
        'document_added_success': 'Document added.',
        'updated_label': 'Updated!',
        'deleted_label': 'Deleted.',
        'added_label': 'Added!',
        'post_saved': 'Post saved!',
        'no_shipments_start': 'No shipments found. Add a new shipment to get started.',
        'no_cargo_items': 'No cargo items found.',
        'no_tracking_yet': 'No tracking updates yet.',
        'no_documents_yet': 'No documents uploaded yet.',
        'no_shipments_client': 'You have no shipments yet.',
        'no_cargo_listed': 'No cargo items listed.',
        'no_unpaid_invoices': 'No unpaid invoices. All invoices are settled.',
        'no_expenses_found': 'No expenses found.',
        'no_perf_data': 'No performance data yet.',
        'no_devices_registered': 'No devices registered yet. Run monitoring_agent.py on IT machines.',
        'no_server_data_yet': 'No server data yet. Open IT Dashboard first to collect a snapshot.',
        'no_security_events_yet': 'No security events recorded yet.',
        'no_shipments_found': 'No shipments found.',
        'no_tracking_ship': 'No tracking updates for this shipment yet.',
        'no_ship_available': 'No shipments available.',
        'no_routes_found': 'No routes found.',
        'no_documents_found': 'No documents found.',
        'no_contracts_found': 'No contracts found.',
        'no_leads_filters_msg': 'No leads found matching filters.',
        'no_leave_reqs_found': 'No leave requests found.',
        'no_reqs_filters': 'No requests match your filters.',
        'no_users_found_msg': 'No users found.',
        'no_tickets_found': 'No tickets found.',
        'no_campaigns_go': 'No campaigns yet. Go to Campaigns page to create one.',
        'no_metrics_campaigns': 'No metrics data yet. Add metrics from the Campaigns page.',
        'no_data_available_yet': 'No data available yet.',
        'no_posts_yet': 'No posts yet.',
        'no_published_posts': 'No published posts yet.',
        'psutil_missing': '⚠️ System metrics unavailable — psutil not found in this Python environment.',
        'no_users_send_msg': 'No users available to send messages to.',
        'no_feedback_yet': 'No feedback received yet.',
        'no_clients_warning': "⚠️ No clients found. Please create a user with 'client' role first.",
        'action_undone_warn': '⚠️ This action cannot be undone!',
        'no_users_email_warn': 'No users with email addresses found.',
        'ship_details_mgmt': 'Shipment Details & Management',
        'update_status_hdr': 'Update Status',
        'ship_details_hdr': 'Shipment Details',
        'rate_shipment': '⭐ Rate This Shipment',
        'filter_by_type_lbl': 'Filter by Type:',
        'filter_by_status_lbl2': 'Filter by Status:',
        'search_lbl2': '\U0001f50d Search',
        'search_ship_ph': 'Shipment number, client...',
        'filter_priority_lbl': 'Filter by Priority',
        'filter_status_leave_lbl': 'Filter by Status',
        'search_email_reason': 'Search by email or reason',
        'type_to_search': 'Type to search...',
        'item_name_lbl': 'Item Name',
        'quantity_lbl': 'Quantity',
        'weight_kg_lbl': 'Weight (kg)',
        'unit_lbl': 'Unit',
        'value_lbl': 'Value',
        'hs_code_lbl': 'HS Code',
        'hs_code_optional': 'HS Code (Optional)',
        'update_status_star': 'Status *',
        'update_date_star': 'Update Date *',
        'doc_type_lbl': 'Document Type',
        'choose_file_lbl': 'Choose file',
        'type_imp_exp_star': 'Type *',
        'client_star': 'Client *',
        'enter_shipment_num_lbl': 'Enter Shipment Number:',
        'select_shipment_lbl': 'Select Shipment:',
        'select_item_edit': 'Select item to edit:',
        'select_post_lbl': 'Select Post',
        'update_status_ticket_lbl': 'Update Status',

        'act_login_success': 'Login Success',
        'act_logout': 'Logout',
        'act_login_failed': 'Login Failed',
        'act_leave_request': 'Leave Request',
        'act_leave_updated': 'Leave Updated',
        'act_add_record': 'Add Record',
        'act_edit_record': 'Edit Record',
        'act_delete_record': 'Delete Record',
        'act_role_update': 'Role Update',
        'act_profile_update': 'Profile Update',
        'act_avatar_update': 'Avatar Update',
        'act_password_change': 'Password Change',
        'act_ticket_created': 'Ticket Created',
        'act_ticket_updated': 'Ticket Updated',
        'act_ticket_deleted': 'Ticket Deleted',
        'unread_messages_warn': 'You have {count} unread message(s)',
    },
    'tr': {
        'title': 'EIMS',
        'login': 'Giriş Yap',
        'signup': 'Kaydol',
        'forgot_password': 'Şifremi Unuttum',
        'dashboard': 'Gösterge Paneli',
        'view': 'Verileri Görüntüle',
        'add': 'Veri Ekle',
        'edit': 'Verileri Düzenle',
        'delete': 'Verileri Sil',
        'analytics': 'Analitikler ve Grafikler',
        'export': 'Verileri Dışa Aktar',
        'request_leave': 'İzin Talep Et',
        'manage_leaves': 'İzin Taleplerini Yönet',
        'manage_users': 'Kullanıcıları Yönet',
        'logout': 'Çıkış Yap',
        'select_page': 'Sayfa Seç:',
        'project_info': 'Proje Bilgileri',
        'graduation_project': 'Mezuniyet Projesi - Yazılım Mühendisliği\nŞirket Veri Yönetim Sistemi',
        'email': 'E-posta',
        'password': 'Şifre',
        'confirm_password': 'Şifreyi Onayla',
        'role': 'Rol',
        'create_account': 'Hesap Oluştur',
        'login_button': 'Giriş Yap',
        'start_date': 'Başlama Tarihi',
        'end_date': 'Bitiş Tarihi',
        'leave_type': 'İzin Türü',
        'reason': 'İzin Nedeni',
        'attachment': 'İsteğe Bağlı Ek (tıbbi not, vb.)',
        'submit': 'Talebi Gönder',
        'my_requests': 'Benim İzin Taleplerimin',
        'from': 'Başlama',
        'to': 'Bitiş',
        'type': 'Tür',
        'manager_response': 'Yönetici Yanıtı',
        'status': 'Durum',
        'pending': 'Beklemede',
        'approved': 'Onaylandı',
        'rejected': 'Reddedildi',
        'language': '🌐 Dil',
        'no_requests': 'İzin talebiniz yok.',
        'set_status': 'Durumu Ayarla:',
        'response_user': 'Kullanıcıya Yanıt:',
        'save': 'Kaydet',
        'user': 'Kullanıcı',
        'employee': 'Çalışan',
        'client': 'Müşteri',
        'admin': 'Yönetici',
        'shipments': 'Gönderiler',
        'add_shipment': 'Gönderi Ekle',
        'client_dashboard': 'Genel Bakış',
        'my_shipments': 'Gönderilerim',
        'my_tickets': 'Taleplerim',
        'my_invoices': 'Faturalarım',
        'my_offers': 'Tekliflerim',
        'accept_offer': 'Teklifi Kabul Et',
        'reject_offer': 'Teklifi Reddet',
        'offer_accepted': 'Teklif başarıyla kabul edildi.',
        'offer_rejected': 'Teklif reddedildi.',
        'no_offers_client': 'Henüz size gönderilmiş teklif bulunmamaktadır.',
        # Client Dashboard
        'welcome_back': 'Tekrar hoş geldiniz',
        'active_shipments': 'Aktif Gönderiler',
        'open_tickets': 'Açık Talepler',
        'unpaid_invoices': 'Ödenmemiş Faturalar',
        'new_offers': 'Yeni Teklifler',
        'unread_messages': 'Okunmamış Mesajlar',
        'recent_shipments': 'Son Gönderiler',
        'recent_tickets': 'Son Talepler',
        'recent_offers': 'Son Teklifler',
        'no_recent_activity': 'Son aktivite bulunmamaktadır.',
        'view_all': 'Tümünü Gör',
        'due_lbl': 'Son Ödeme',
        'dash_origin': 'Kaynak',
        'dash_dest': 'Hedef',
        'track_shipment': 'Gönderi Takibi',
        'manage_shipments': 'Gönderileri Yönet',
        'shipment_analytics': 'Gönderi Analizleri',
        'import': 'İthalat',
        'export': 'İhracat',
        'shipment_number': 'Gönderi Numarası',
        'origin': 'Çıkış Ülkesi',
        'destination': 'Varış Ülkesi',
        'departure_date': 'Kalkış Tarihi',
        'expected_arrival': 'Beklenen Varış',
        'actual_arrival': 'Gerçek Varış',
        'shipment_status': 'Gönderi Durumu',
        'cargo_items': 'Kargo Kalemleri',
        'add_cargo': 'Kargo Kalemi Ekle',
        'tracking': 'Takip',
        'documents': 'Belgeler',
        'total_weight': 'Toplam Ağırlık (kg)',
        'total_value': 'Toplam Değer',
        'customs_cleared': 'Gümrük Geçişi',
        'in_transit': 'Yolda',
        'delivered': 'Teslim Edildi',
        'cancelled': 'İptal Edildi',
        'item_name': 'Ürün Adı',
        'quantity': 'Miktar',
        'weight': 'Ağırlık (kg)',
        'value': 'Değer',
        'description': 'Açıklama',
        'location': 'Konum',
        'update_tracking': 'Takip Güncelle',
        'upload_document': 'Belge Yükle',
        'cargo_requests': 'Kargo Talepleri',
        'request_cargo_change': 'Kargo Değişikliği Talep Et',
        'manage_cargo_requests': 'Kargo Taleplerini Yönet',
        'request_type': 'Talep Türü',
        'modify': 'Düzenle',
        'remove': 'Kaldır',
        'request_reason': 'Talep Nedeni',
        'my_cargo_requests': 'Kargo Taleplerim',
        'approve': 'Onayla',
        'reject': 'Reddet',
        'request_status': 'Talep Durumu',
        'client_name': 'Müşteri Adı',
        'edit_shipment': '✏️ Gönderiyi Düzenle',
        'delete_shipment': '🗑️ Gönderiyi Sil',
        'profile': 'Profil',
        'messages': 'Mesajlar',
        'export_employees': 'Çalışan Verilerini Dışa Aktar',
        'export_shipments': 'Gönderi Verilerini Dışa Aktar',
        'finance_dashboard': 'Finans Gösterge Paneli',
        'invoices': 'Faturalar',
        'payments': 'Ödemeler',
        'expenses': 'Giderler',
        'financial_reports': 'Finansal Raporlar',
        'marketing_dashboard': 'Pazarlama Gösterge Paneli',
        'campaigns': 'Kampanyalar',
        'marketing_analytics': 'Analitik',
        'social_media': 'Sosyal Medya',
        'it_dashboard': 'BT Gösterge Paneli',
        'system_management': 'Sistem Yönetimi',
        'support_tickets': 'Destek Talepleri',
        'security': 'Güvenlik',
        'logistics_dashboard': 'Lojistik Gösterge Paneli',
        'routes': 'Güzergahlar',
        'delivery_assignments': 'Teslimat Görevleri',
        'customer_service_dashboard': 'Müşteri Hizmetleri Paneli',
        'client_tickets': 'Müşteri Talepleri',
        'customer_feedback': 'Müşteri Geri Bildirimleri',
        'administration_dashboard': 'Yönetim Gösterge Paneli',
        'meetings': 'Toplantılar',
        'contracts': 'Sözleşmeler',
        'training': 'Eğitim',
        'sales_dashboard': 'Satış Gösterge Paneli',
        'leads': 'Potansiyel Müşteriler',
        'deals': 'Anlaşmalar',
        'offers': 'Teklifler',
        'rate_service': 'Hizmeti Değerlendir',
        'main_dashboard': 'Ana Gösterge Paneli',
        'total_employees': 'Toplam Çalışanlar',
        'total_departments': 'Departmanlar',
        'avg_salary': 'Ortalama Maaş',
        'active_employees': 'Aktif Çalışanlar',
        'recent_records': 'Son Kayıtlar',
        'no_data': 'Veri bulunamadı',
        'view_all_data': 'Tüm Verileri Görüntüle',
        'search_placeholder': 'Ad, departman, pozisyona göre ara...',
        'filter_dept': 'Departmana Göre Filtrele:',
        'filter_status': 'Duruma Göre Filtrele:',
        'filter_role': 'Role Göre Filtrele:',
        'no_display': 'Görüntülenecek veri yok',
        'add_new_record': 'Yeni Kayıt Ekle',
        'account_settings_hdr': 'Hesap Ayarları',
        'create_login': 'Bu çalışan için giriş hesabı oluştur',
        'account_role': 'Hesap Rolü',
        'auto_gen_pass': 'Şifre otomatik oluştur',
        'personal_info': 'Kişisel Bilgiler',
        'employee_name': 'Çalışan Adı *',
        'phone_field': 'Telefon',
        'hire_date_field': 'İşe Başlama Tarihi *',
        'reg_date': 'Kayıt Tarihi *',
        'salary_field': 'Maaş ($) *',
        'position_field': 'Pozisyon *',
        'dept_field': 'Departman *',
        'all': 'Tümü',
        'active': 'Aktif',
        'inactive': 'Pasif',
        'on_leave': 'İzinde',
        'add_record_btn': '➕ Kayıt Ekle',
        'edit_existing': 'Mevcut Kaydı Düzenle',
        'select_edit': 'Düzenlemek için kayıt seçin:',
        'save_changes': '💾 Değişiklikleri Kaydet',
        'delete_record_hdr': 'Kayıt Sil',
        'select_delete': 'Silmek için kayıt seçin:',
        'confirm_del': 'Onaylamak için "DELETE" yazın:',
        'delete_btn': '🗑️ Kayıt Sil',
        'it_pages_label': 'BT Sayfaları',
        'db_management': 'Veritabanı Yönetimi',
        'security_mgmt': 'Güvenlik Yönetimi',
        'performance_monitor': 'Performans İzleme',
        'bug_tracking': 'Hata Takibi',
        'activity_logs': 'Aktivite Kayıtları',
        'manage_users_hdr': 'Kullanıcıları Yönet',
        'manage_leaves_hdr': 'İzin Taleplerini Yönet',
        'leave_requests_hdr': 'İzin Talepleri',
        'my_leave_requests': 'İzin Taleplerim',
        'no_leave_requests': 'İzin talebi bulunamadı.',
        'request_leave_hdr': 'İzin Talep Et',
        'submit_leave': 'Talebi Gönder',
        'update_status': 'Durumu Güncelle',
        'response_note': 'Yanıt notu',
        'filter_priority': 'Öncelik Filtrele',
        'filter_status_lbl': 'Durum Filtrele',
        'search_lbl': 'Çalışan ara',
        'all_priorities': 'Tüm Öncelikler',
        'all_statuses': 'Tüm Durumlar',
        'no_messages': 'Mesaj bulunamadı.',
        'compose': 'Oluştur',
        'inbox': 'Gelen Kutusu',
        'sent_msgs': 'Gönderilenler',
        'subject': 'Konu',
        'message_body': 'Mesaj İçeriği',
        'send_to': 'Alıcı',
        'reply_btn': 'Yanıtla',
        'delete_msg': 'Sil',
        'send': 'Gönder',
        'priority': 'Öncelik',
        'low': 'Düşük',
        'medium': 'Orta',
        'high': 'Yüksek',
        'urgent': 'Acil',
        'open': 'Açık',
        'in_progress': 'Devam Ediyor',
        'resolved': 'Çözüldü',
        'closed': 'Kapatıldı',
        'tab_distribution': 'Dağılım',
        'tab_salaries': 'Maaşlar',
        'tab_trends': 'Trendler',
        'tab_details': 'Detaylar',
        'tab_cargo_items': 'Kargo Kalemleri',
        'tab_tracking': 'Takip',
        'tab_documents': 'Belgeler',
        'tab_all_invoices': 'Tüm Faturalar',
        'tab_new_invoice': 'Yeni Fatura',
        'tab_payment_history': 'Ödeme Geçmişi',
        'tab_record_payment': 'Ödeme Kaydet',
        'tab_all_expenses': 'Tüm Giderler',
        'tab_add_expense': 'Gider Ekle',
        'tab_all_campaigns': 'Tüm Kampanyalar',
        'tab_new_campaign': 'Yeni Kampanya',
        'tab_add_metrics': 'Metrik Ekle',
        'tab_performance': 'Performans',
        'tab_customer_behavior': 'Müşteri Davranışı',
        'tab_competitors': 'Rakipler',
        'tab_posts': 'Gönderiler',
        'tab_new_post': 'Yeni Gönderi',
        'tab_stats': 'İstatistikler',
        'tab_new_ticket': 'Yeni Talep',
        'tab_all_tickets': 'Tüm Talepler',
        'tab_security_events': 'Güvenlik Olayları',
        'tab_user_activity': 'Kullanıcı Aktivitesi',
        'tab_all_shipments': 'Tüm Gönderiler',
        'tab_new_shipment': 'Yeni Gönderi',
        'tab_tracking_timeline': 'Takip Zaman Çizelgesi',
        'tab_active_routes': 'Aktif Güzergahlar',
        'tab_add_route': 'Güzergah Ekle',
        'tab_unassigned': 'Atanmamış',
        'tab_active_assignments': 'Aktif Görevler',
        'tab_completed': 'Tamamlandı',
        'tab_all_docs': 'Tüm Belgeler',
        'tab_add_doc': 'Belge Ekle',
        'tab_upcoming': 'Yaklaşan',
        'tab_past_meetings': 'Geçmiş Toplantılar',
        'tab_schedule': 'Toplantı Planla',
        'tab_all_contracts': 'Tüm Sözleşmeler',
        'tab_new_contract': 'Yeni Sözleşme',
        'tab_overview': 'Genel Bakış',
        'tab_all_programs': 'Tüm Programlar',
        'tab_add_program': 'Program Ekle',
        'tab_info': 'Bilgi',
        'tab_security': 'Güvenlik',
        'tab_activity': 'Aktivitelerim',
        'tab_feedback': 'Alınan Geri Bildirimler',
        'tab_add_feedback': 'Geri Bildirim Ekle',
        'department': 'Departman',
        'dark_mode': 'Karanlık Mod',
        'light_mode': 'Aydınlık Mod',
        'search_data': 'Veri Ara',
        'found_results': 'sonuç bulundu',
        'no_records_role': 'Bu rol için kayıt bulunamadı',
        'confirm_delete_type': 'Onaylamak için DELETE yazın',
        'record_deleted': 'Kayıt başarıyla silindi!',
        'record_added': 'Kayıt başarıyla eklendi!',
        'record_updated': 'Kayıt başarıyla güncellendi!',
        'changes_saved': 'Değişiklikler kaydedildi!',
        'fill_required': 'Lütfen tüm zorunlu alanları doldurun',
        'passwords_no_match': 'Şifreler eşleşmiyor',
        'invalid_email': 'Geçersiz e-posta formatı',
        'login_required_msg': 'Bu sayfalara erişmek için giriş yapın.',
        'go_login': 'Girişe Git',
        'manager': 'Yönetici',
        'record_info': 'Kayıt Bilgisi',
        'name_field': 'Ad',
        'position': 'Pozisyon',
        'salary_display': 'Maaş',
        'hire_date_display': 'İşe Başlama Tarihi',
        'reg_date_display': 'Kayıt Tarihi',
        'warning_undone': '⚠️ Uyarı: Bu işlem geri alınamaz!',
        'cancel': 'İptal',
        'confirm_delete_btn': '🗑️ Silmeyi Onayla',
        'delete_cancelled': 'Silme işlemi iptal edildi.',
        'no_delete': 'Silinecek kayıt yok',
        'no_edit': 'Düzenlenecek kayıt yok',
        'save_client_info': '💾 Müşteri Bilgilerini Kaydet',
        'avg_salary_label': 'Ortalama Maaş',
        'highest_salary': 'En Yüksek Maaş',
        'lowest_salary': 'En Düşük Maaş',
        'total_payroll': 'Toplam Maaş Bordrosu',
        'emp_distribution': 'Departmana Göre Çalışan Dağılımı',
        'emp_status': 'Çalışan Durumu',
        'common_positions': 'Yaygın Pozisyonlar',
        'salary_analysis': 'Maaş Analizi',
        'hire_trends': 'İşe Alım Trendleri',
        'monthly_hires': 'Aylık İşe Alımlar',
        'emp_trends': 'Çalışan Trendleri',
        'hiring_timeline': 'İşe Alım Zaman Çizelgesi',
        'emp_added_time': 'Zaman İçinde Eklenen Çalışanlar',
        'dept_sizes': 'Departman Büyüklükleri',
        'status_breakdown': 'Durum Dağılımı',
        'avg_salary_by_dept': 'Departmana Göre Ortalama Maaş',
        'salary_dist': 'Maaş Dağılımı',
        'export_employees_hdr': 'Çalışanları Dışa Aktar',
        'export_shipments_hdr': 'Gönderileri Dışa Aktar',
        'select_columns': 'Dahil edilecek sütunları seçin',
        'columns_label': 'Sütunlar',
        'preview_columns': 'Seçili sütunları önizle',
        'show_all_rows': 'Tüm satırları göster',
        'export_format': 'Dışa aktarma formatı',
        'format_label': 'Format',
        'download_csv': '⬇️ CSV İndir',
        'download_excel': '⬇️ Excel İndir',
        'download_csv_fallback': '⬇️ CSV İndir (yedek)',
        'select_one_col': 'Dışa aktarmak için en az bir sütun seçin.',
        'no_records_export': 'Dışa aktarılacak kayıt yok.',
        'no_shipments_export': 'Dışa aktarılacak gönderi yok.',
        'submit_leave_card': 'İzin Talebi Gönder',
        'date_range_err': 'Bitiş tarihi başlangıç tarihinden önce olamaz.',
        'leave_submitted': 'İzin talebi başarıyla gönderildi.',
        'date_label': 'Tarih',
        'employees_added': 'Eklenen Çalışanlar',
        'count_label': 'Sayı',
        'forgot_pass_btn': '🔑 Şifremi Unuttum?',
        'back_to_login': '⬅️ Girişe Dön',
        'reset_your_pass': '🔑 Şifrenizi Sıfırlayın',
        'your_email': '📧 Kayıtlı E-postanız:',
        'contact_email': '📬 İletişim E-postası (yönetici yeni şifreyi buraya gönderecek):',
        'reset_info': '💡 Yöneticinin yeni şifrenizi gönderebileceği kişisel bir e-posta girin.',
        'submit_request_btn': '📤 Talep Gönder',
        'full_name_signup': 'Tam Ad *',
        'phone_number': 'Telefon Numarası',
        'emp_status_field': 'İstihdam Durumu *',
        'change_type': '⬅ Değiştir',
        'server_status': 'Sunucu Durumu',
        'running': 'Çalışıyor',
        'restart_server': 'Sunucuyu Yeniden Başlat',
        'check_status': 'Sistem Durumunu Kontrol Et',
        'server_restarted': 'Sunucu yeniden başlatma komutu gönderildi!',
        'status_checked': 'Sistem durumu kontrol edildi.',
        'last_backup': 'Son Yedekleme',
        'db_size': 'Veritabanı Boyutu',
        'create_backup': 'Yedek Oluştur',
        'restore_backup': 'Yedeği Geri Yükle',
        'backup_created': 'Yedek oluşturuldu!',
        'restore_started': 'Geri yükleme başladı (demo)',
        'bug_title': 'Hata Başlığı',
        'bug_desc': 'Hata Açıklaması',
        'log_bug': 'Hata Kaydet',
        'bug_logged': 'Hata kaydedildi!',
        'cpu_usage': 'CPU Kullanımı',
        'mem_usage': 'Bellek Kullanımı',
        'active_users': 'Aktif Kullanıcılar',
        'recent_activities': 'Son Aktiviteler',
        'manage_leave_response': 'Kullanıcıya yanıt:',
        'account_info': 'Hesap Bilgileri',
        'employment_details': 'İstihdam Detayları',
        'email_address': 'E-posta Adresi *',
        'create_emp_account': 'Çalışan Hesabı Oluştur',
        'create_client_account': 'Müşteri Hesabı Oluştur',
        'provide_email_pass': 'Lütfen e-posta ve şifre girin',
        'enter_full_name': 'Lütfen tam adınızı girin',
        'support_ticket_sys': 'Destek Talep Sistemi',
        'create_ticket': 'Yeni destek talebi oluştur:',
        'ticket_title_lbl': 'Başlık',
        'ticket_desc_lbl': 'Açıklama',
        'submit_ticket_btn': 'Talep Gönder',
        'fill_all_fields': 'Lütfen tüm alanları doldurun.',
        'all_tickets_lbl': 'Tüm Talepler:',
        'no_tickets': 'Talep bulunamadı.',
        'user_list': 'Kullanıcı Listesi:',
        'change_user_role': 'Kullanıcı rolü/durumu değiştir (demo)',
        'last_check': 'Son Kontrol:',
        'log_new_bug': 'Yeni hata kaydet:',
        'all_bugs': 'Tüm Hatalar:',
        'recent_activities_lbl': 'Son Aktiviteler:',
        'select_section': 'Kenar çubuğundan bir bölüm seçin.',
        # Manage Users
        'change_user_role': 'Kullanıcı Rolünü Değiştir',
        'select_user': 'Kullanıcı seçin:',
        'new_role': 'Yeni rol:',
        'update_role': 'Rolü Güncelle',
        'reset_user_pass': '🔐 Kullanıcı Şifresini Sıfırla',
        'reset_pass_info': 'Kimlik bilgilerini unutan kullanıcı için yeni bir rastgele şifre oluşturun.',
        'select_reset_user': 'Şifresini sıfırlamak için kullanıcı seçin:',
        'reset_pass_btn': '🔄 Şifreyi Sıfırla',
        'pass_reset_success': '✅ Şifre başarıyla sıfırlandı!',
        'pass_save_warning': '⚠️ Bu şifreyi kaydedin ve kullanıcıya verin. Bir daha gösterilmeyecektir.',
        # Finance Dashboard
        'total_revenue': 'Toplam Gelir',
        'total_expenses': 'Toplam Gider',
        'net_profit': 'Net Kâr',
        'outstanding': 'Bekleyen',
        'overdue': 'Vadesi Geçmiş',
        # Finance - Invoices
        'filter_by_status': 'Duruma Göre Filtrele',
        'create_new_invoice': 'Yeni Fatura Oluştur',
        'client_name_field': 'Müşteri Adı *',
        'shipment_ref': 'Gönderi Referansı',
        'amount_usd': 'Tutar (USD) *',
        'tax_rate': 'Vergi Oranı (%)',
        'due_date': 'Vade Tarihi *',
        'desc_services': 'Açıklama / Hizmetler',
        'create_invoice_btn': 'Fatura Oluştur',
        'mark_sent': 'Gönderildi',
        'mark_paid': 'Ödendi',
        'mark_overdue': 'Vadesi Geçti',
        'delete_btn': 'Sil',
        'no_invoices': 'Fatura bulunamadı.',
        'invoice_created': 'Fatura başarıyla oluşturuldu.',
        'client_amount_required': 'Müşteri adı ve tutar gereklidir.',
        'shipment_ref_lbl': 'Sevkiyat Referansı',
        'issue_date_lbl': 'Düzenleme Tarihi',
        'due_date_lbl': 'Vade Tarihi',
        'amount_lbl': 'Tutar',
        'tax_lbl': 'Vergi',
        'total_lbl': 'Toplam',
        'cost_breakdown': 'Maliyet Dökümü',
        'cargo_value_lbl': 'Kargo / Sevkiyat Değeri',
        'freight_charge_lbl': 'Navlun Ücreti',
        'handling_fee_lbl': 'Elleçleme Ücreti',
        'insurance_lbl': 'Sigorta (%0.4)',
        'services_subtotal_lbl': 'Hizmetler Ara Toplamı',
        'vat_lbl': 'KDV',
        'grand_total_lbl': 'Genel Toplam',
        'amount_usd_col': 'Tutar (USD)',
        'download_pdf': 'Fatura PDF İndir',
        'total_invoices_lbl': 'Toplam Fatura',
        'paid_lbl': 'Ödenen',
        'outstanding_lbl': 'Bekleyen',
        'col_invoice': 'Fatura #',
        'col_client': 'Müşteri',
        'col_total_usd': 'Toplam (USD)',
        'col_due_date': 'Vade Tarihi',
        'col_date': 'Tarih',
        'col_amount_usd': 'Tutar (USD)',
        'col_method': 'Yöntem',
        'col_reference': 'Referans',
        'col_month': 'Ay',
        'col_revenue': 'Gelir',
        'col_net_profit': 'Net Kâr',
        'chart_rev_vs_exp': 'Gelir ve Gider Karşılaştırması (Aylık)',
        'chart_invoice_status': 'Fatura Durumu Dağılımı',
        'chart_payments_method': 'Ödeme Yöntemine Göre',
        'chart_monthly_payments': 'Aylık Alınan Ödemeler',
        'chart_approved_expenses': 'Kategoriye Göre Onaylı Giderler',
        'chart_monthly_revenue': 'Aylık Gelir',
        'chart_monthly_trend': 'Aylık Gelir, Gider ve Kâr',
        'chart_expense_breakdown': 'Kategoriye Göre Gider Dağılımı',
        'margin_lbl': 'marj',
        'paid_lbl': 'ödendi',
        'no_financial_data_short': 'Henüz finansal veri yok.',
        'no_invoices_yet': 'Henüz fatura yok.',
        # Finance - Payments
        'no_payments': 'Henüz ödeme kaydedilmedi.',
        'record_payment': 'Ödeme Kaydet',
        'select_invoice': 'Fatura Seçin *',
        'amount_received': 'Alınan Tutar (USD) *',
        'payment_date': 'Ödeme Tarihi *',
        'payment_method': 'Ödeme Yöntemi',
        'ref_no': 'Referans / İşlem No.',
        'record_payment_btn': 'Ödeme Kaydet',
        'amount_gt_zero': 'Tutar sıfırdan büyük olmalıdır.',
        'payment_recorded': 'Ödeme kaydedildi ve fatura Ödendi olarak işaretlendi.',
        # Finance - Expenses
        'filter_by_cat': 'Kategoriye Göre Filtrele',
        'total_filtered': 'Toplam (Filtreli)',
        'approve_btn': 'Onayla',
        'reject_btn': 'Reddet',
        'add_new_expense': 'Yeni Gider Ekle',
        'category_field': 'Kategori *',
        'expense_date': 'Gider Tarihi *',
        'vendor_supplier': 'Tedarikçi',
        'receipt_ref': 'Makbuz Referansı',
        'add_expense_btn': 'Gider Ekle',
        'expense_added': 'Gider eklendi ve onay bekliyor.',
        'vendor_lbl': 'Tedarikçi',
        'receipt_ref_lbl': 'Makbuz Ref.',
        'cargo_cost_opt': 'Kargo Maliyeti', 'customs_duty_opt': 'Gümrük Vergisi',
        'freight_fee_opt': 'Navlun Ücreti', 'fuel_transport_opt': 'Yakıt ve Taşıma',
        'staff_salary_opt': 'Personel Maaşı', 'office_rent_opt': 'Ofis Kirası',
        'insurance_opt': 'Sigorta', 'port_handling_opt': 'Liman İşleme',
        'bank_transfer_opt': 'Banka Transferi', 'letter_credit_opt': 'Akreditif',
        'cash_opt': 'Nakit', 'credit_card_opt': 'Kredi Kartı', 'cheque_opt': 'Çek',
        # Finance - Reports
        'total_invoices': 'Toplam Faturalar',
        'profit_loss': 'Kâr & Zarar Özeti',
        'no_financial_data': 'Henüz finansal veri yok. Rapor görmek için fatura oluşturun ve gider kaydedin.',
        # Marketing Dashboard
        'campaign_performance': 'Kampanya Performansı',
        'budget_vs_spent': 'Bütçe vs Harcama',
        'recent_campaigns': 'Son Kampanyalar',
        'no_campaign_metrics': 'Henüz kampanya metriği yok.',
        'no_budget_data': 'Henüz bütçe verisi yok.',
        'no_campaigns': 'Henüz kampanya yok.',
        # Campaigns
        'filter_by_status_lbl': 'Duruma Göre Filtrele',
        'change_status': 'Durumu Değiştir',
        'update_btn': 'Güncelle',
        'campaign_name': 'Kampanya Adı',
        'type_field': 'Tür',
        'budget_field': 'Bütçe ($)',
        'start_date_field': 'Başlangıç Tarihi',
        'end_date_field': 'Bitiş Tarihi',
        'target_audience': 'Hedef Kitle',
        'create_campaign_btn': 'Kampanya Oluştur',
        'campaign_name_required': 'Kampanya adı gereklidir.',
        'create_campaign_first': 'Önce bir kampanya oluşturun.',
        'date_field': 'Tarih',
        'impressions': 'Gösterimler',
        'clicks': 'Tıklamalar',
        'conversions': 'Dönüşümler',
        'users': 'Kullanıcılar',
        'revenue_field': 'Gelir ($)',
        'save_metrics': 'Metrikleri Kaydet',
        'metrics_saved': 'Metrikler kaydedildi!',
        # Marketing Analytics
        'filter_by_campaign': 'Kampanyaya Göre Filtrele',
        'customer_behavior': 'Müşteri Davranış Analizi',
        'competitor_analysis': 'Rakip Analizi',
        'add_competitor': 'Rakip Ekle',
        'name_field': 'İsim',
        'website_field': 'Web Sitesi',
        'strengths_field': 'Güçlü Yönler',
        'weaknesses_field': 'Zayıf Yönler',
        'no_competitors': 'Henüz rakip eklenmedi.',
        'no_metrics_data': 'Henüz metrik verisi yok. Kampanyalar sayfasından metrik ekleyin.',
        # Social Media
        'platform_field': 'Platform',
        'content_field': 'İçerik',
        'schedule_date': 'Planlama Tarihi',
        'link_to_campaign': 'Kampanyaya Bağla (isteğe bağlı)',
        'save_post': 'Gönderiyi Kaydet',
        'update_engagement': 'Etkileşim Güncelle',
        'select_post': 'Gönderi Seçin',
        'likes_field': 'Beğeniler',
        'shares_field': 'Paylaşımlar',
        'comments_field': 'Yorumlar',
        # IT Dashboard
        'ram_usage': 'RAM Kullanımı',
        'db_response_time': 'DB Yanıt Süresi',
        'open_tickets': 'Açık Talepler',
        'system_performance': 'Sistem Performansı',
        'total_users': 'Toplam Kullanıcılar',
        'new_this_month': 'Bu Ay Yeni',
        'security_last7': 'Güvenlik (son 7 gün)',
        'failed_logins': 'Başarısız Giriş Denemeleri',
        'recent_activity_log': 'Son Aktivite Kaydı',
        'no_activity': 'Henüz aktivite kaydedilmedi.',
        'live_metrics': 'Canlı Sistem Metrikleri',
        'perf_history': 'Performans Geçmişi (son 30 anlık görüntü)',
        'network_io': 'Ağ G/Ç',
        'total_sent': 'Toplam Gönderildi',
        'total_received': 'Toplam Alındı',
        # Support Tickets (employee)
        'ticket_submitted': 'Talep başarıyla gönderildi!',
        'change_status_lbl': 'Durumu Değiştir',
        'status_updated': 'Durum güncellendi!',
        'ticket_deleted': 'Talep silindi.',
        'delete_ticket_btn': 'Talebi Sil',
        'edit_ticket_btn': 'Talebi Düzenle',
        'save_ticket_btn': 'Değişiklikleri Kaydet',
        'cancel_btn': 'İptal',
        'ticket_edited_success': 'Talep başarıyla güncellendi.',
        'cs_ticket_deleted_ok': 'Talep silindi.',
        'confirm_delete_ticket': 'Bu talebi silmek istediğinizden emin misiniz?',
        # Security page
        'security_events_log': 'Güvenlik Olayları Kaydı',
        'filter_event_type': 'Olay Türüne Göre Filtrele',
        'registered_users': 'Kayıtlı Kullanıcılar',
        # Logistics
        'total_shipments': 'Toplam Gönderiler',
        'delivered_month': 'Teslim Edilen (Ay)',
        'pending_customs': 'Bekliyor / Gümrük',
        'overdue': 'Vadesi Geçmiş',
        'recent_shipments': 'Son Gönderiler',
        # Shipments
        'cargo_items_lbl': 'Kargo Kalemleri:',
        'tracking_history_lbl': 'Takip Geçmişi:',
        'change_status_lbl2': 'Durumu Değiştir',
        'save_update': 'Güncellemeyi Kaydet',
        'create_new_shipment': 'Yeni Gönderi Oluştur',
        'freight_mode': 'Nakliye Modu *',
        'origin_country': 'Çıkış Ülkesi *',
        'dest_country': 'Varış Ülkesi *',
        'departure_date': 'Kalkış Tarihi *',
        'expected_arrival_field': 'Beklenen Varış *',
        'carrier_name_field': 'Taşıyıcı Adı',
        'container_type': 'Konteyner Tipi',
        'incoterms_field': 'İncoterms',
        'total_weight_field': 'Toplam Ağırlık (kg)',
        'total_value_field': 'Toplam Değer (USD)',
        'notes_desc': 'Notlar / Açıklama',
        'create_shipment_btn': 'Gönderi Oluştur',
        'shipment_tracking': 'Gönderi Takip Zaman Çizelgesi',
        'select_shipment': 'Gönderi Seçin',
        # Routes
        'filter_mode': 'Mod Filtrele',
        'add_new_route': 'Yeni Güzergah Ekle',
        'route_name': 'Güzergah Adı *  (örn. İstanbul → Dubai)',
        'origin_port': 'Çıkış Limanı / Havalimanı *',
        'dest_port': 'Varış Limanı / Havalimanı *',
        'transit_days': 'Geçiş Günleri',
        'frequency': 'Sıklık  (örn. Haftalık)',
        'add_route_btn': 'Güzergah Ekle',
        'route_required': 'Güzergah adı, çıkış ve varış gereklidir.',
        # Delivery Assignments
        'assign_carrier_hdr': '**Taşıyıcı Ata:**',
        'carrier_name_req': 'Taşıyıcı Adı *',
        'assign_carrier_btn': 'Taşıyıcı Ata',
        'carrier_required': 'Taşıyıcı adı gereklidir.',
        'all_assigned': 'Tüm gönderiler taşıyıcılara atanmıştır.',
        'update_status_lbl': 'Durumu Güncelle',
        'mark_delivered': '✅ Teslim Edildi',
        'no_active_assignments': 'Aktif görev yok.',
        'no_completed': 'Henüz tamamlanmış teslimat yok.',
        # Customer Service Dashboard
        'total_tickets': 'Toplam Talepler',
        'open_status': 'Açık',
        'resolved_today': 'Bugün Çözüldü',
        'avg_rating': 'Ort. Puan',
        'open_urgent_tickets': 'Açık ve Acil Talepler',
        'no_open_tickets': 'Açık talep yok. Hepsi çözüldü!',
        'needs_attention': 'Dikkat gerekiyor',
        'chart_tickets_status': 'Duruma Göre Talepler',
        'chart_tickets_category': 'Kategoriye Göre Talepler',
        'unread_feedback_msg': '{n} okunmamış müşteri geri bildirimi inceleme bekliyor.',
        'tickets_found': 'talep bulundu',
        'created_lbl': 'Oluşturuldu',
        'resolved_at_lbl': 'Çözüm tarihi',
        'resolution_lbl': 'Çözüm',
        'chart_rating_dist': 'Puan Dağılımı',
        'total_word': 'toplam',
        # Department names (sidebar)
        'dept_finance': 'Finans', 'dept_marketing': 'Pazarlama', 'dept_it': 'BT',
        'dept_logistics': 'Lojistik', 'dept_customer_service': 'Müşteri Hizmetleri',
        'dept_administration': 'Yönetim', 'dept_sales': 'Satış',
        # Client Tickets
        'create_new_ticket': 'Yeni Talep Oluştur',
        'update_status_btn': 'Durumu Güncelle',
        'resolution_response': 'Çözüm / Yanıt',
        'update_ticket_btn': 'Talebi Güncelle',
        'client_subj_desc_required': 'Müşteri, konu ve açıklama gereklidir.',
        'create_ticket_btn': 'Talep Oluştur',
        # Ticket Replies
        'ticket_conversation': 'Konuşma',
        'reply_from_cs': 'Müşteri Hizmetleri',
        'reply_from_client': 'Müşteri',
        'type_reply': 'Yanıt yazın...',
        'send_reply_btn': 'Yanıt Gönder',
        'reply_sent': 'Yanıt gönderildi.',
        'reply_required': 'Yanıt boş olamaz.',
        'no_replies_yet': 'Henüz mesaj yok. Konuşmayı başlatın.',
        # Customer Feedback
        'show_unread_only': 'Sadece okunmamışları göster',
        'total_feedback': 'Toplam Geri Bildirim',
        'average_rating': 'Ortalama Puan',
        'unread_count': 'Okunmamış',
        'mark_as_read': 'Okundu İşaretle',
        'add_cust_feedback': 'Müşteri Geri Bildirimi Ekle',
        'rating_field': 'Puan *',
        'comment_review': 'Yorum / Değerlendirme',
        'submit_feedback_btn': 'Geri Bildirimi Gönder',
        'select_client': 'Lütfen bir müşteri seçin.',
        # Administration Dashboard
        'total_documents': 'Toplam Belgeler',
        'active_contracts': 'Aktif Sözleşmeler',
        'upcoming_meetings': 'Yaklaşan Toplantılar',
        'expiring_30d': '30 Günde Süresi Dolacak',
        'total_contracts_lbl': 'Toplam Sözleşmeler',
        'upcoming_meetings_hdr': 'Yaklaşan Toplantılar',
        'contracts_expiring': 'Yakında Süresi Dolacak Sözleşmeler',
        'recent_docs': 'Son Belgeler',
        'no_upcoming_meetings': 'Yaklaşan toplantı yok.',
        'no_contracts_expiring': 'Önümüzdeki 90 gün içinde süresi dolacak sözleşme yok.',
        # Documents
        'add_new_doc': 'Yeni Belge Ekle',
        'doc_title': 'Belge Başlığı *',
        'expiry_date_field': 'Son Kullanma Tarihi (isteğe bağlı)',
        'file_name_field': 'Dosya Adı (örn. contract_2026.pdf)',
        'set_expiry': 'Son kullanma tarihi belirle',
        'add_doc_btn': 'Belge Ekle',
        'archive_btn': 'Arşivle',
        'restore_btn': 'Geri Yükle',
        # Meetings
        'update_meeting': 'Güncelle',
        'meeting_minutes': 'Toplantı Tutanağı',
        'save_btn': 'Kaydet',
        'schedule_new_meeting': 'Yeni Toplantı Planla',
        'meeting_title': 'Başlık *',
        'meeting_type': 'Toplantı Türü *',
        'date_field2': 'Tarih *',
        'time_field': 'Saat *',
        'location_platform': 'Konum / Platform',
        'duration_min': 'Süre (dak)',
        'attendees_field': 'Katılımcılar (isimler veya roller)',
        'agenda_field': 'Gündem',
        'schedule_meeting_btn': 'Toplantı Planla',
        'meeting_scheduled': 'Toplantı planlandı.',
        'meeting_title_required': 'Başlık gereklidir.',
        # Contracts
        'party_type_filter': 'Taraf Türü',
        'active_contracts_value': 'Aktif Sözleşmeler Toplam Değeri',
        'add_new_contract': 'Yeni Sözleşme Ekle',
        'contract_title': 'Sözleşme Başlığı *',
        'party_name': 'Taraf Adı *  (şirket/kişi)',
        'party_type_field': 'Taraf Türü *',
        'contract_type_field': 'Sözleşme Türü *',
        'value_field': 'Değer',
        'currency_field': 'Para Birimi',
        'start_date_lbl': 'Başlangıç Tarihi *',
        'end_date_lbl': 'Bitiş Tarihi *',
        'add_contract_btn': 'Sözleşme Ekle',
        'title_party_required': 'Başlık ve taraf adı gereklidir.',
        'end_after_start': 'Bitiş tarihi başlangıç tarihinden sonra olmalıdır.',
        'contract_added_success': 'Sözleşme başarıyla eklendi.',
        'party_type_lbl': 'Taraf Türü', 'contract_type_lbl': 'Sözleşme Türü',
        'value_lbl': 'Değer', 'start_lbl': 'Başlangıç', 'end_lbl': 'Bitiş',
        'update_btn_lbl': 'Güncelle',
        'carrier_opt': 'Taşıyıcı', 'supplier_opt': 'Tedarikçi', 'government_opt': 'Devlet',
        'terminated_opt': 'Sonlandırıldı',
        'service_agreement_opt': 'Hizmet Anlaşması', 'rate_agreement_opt': 'Ücret Anlaşması',
        'agency_agreement_opt': 'Acenta Anlaşması', 'insurance_policy_opt': 'Sigorta Poliçesi',
        'software_license_opt': 'Yazılım Lisansı', 'lease_agreement_opt': 'Kira Anlaşması',
        'nda_opt': 'GGS',
        # Sales Dashboard
        'total_leads': 'Toplam Adaylar',
        'active_leads': 'Aktif Adaylar',
        'open_deals': 'Açık Anlaşmalar',
        'pipeline_value': 'Boru Hattı Değeri',
        'won_deals': 'Kazanılan Anlaşmalar',
        'win_rate': 'Kazanma Oranı',
        'recent_leads': 'Son Adaylar',
        # Leads
        'source_filter': 'Kaynak',
        'freight_filter': 'Yük',
        'update_status_lead': 'Durumu Güncelle',
        'add_new_lead': '➕ Yeni Aday Ekle',
        'lead_name': 'Aday Adı *',
        'company_name': 'Şirket Adı',
        'email_field': 'E-posta',
        'phone_field': 'Telefon',
        'country_field': 'Ülke',
        'source_field': 'Kaynak *',
        'freight_interest': 'Yük İlgi Alanı',
        'add_lead_btn': 'Aday Ekle',
        'lead_name_required': 'Aday adı gereklidir.',
        # Deals
        'stage_filter': 'Aşama',
        'freight_type_filter': 'Yük Türü',
        'add_new_deal': '➕ Yeni Anlaşma Ekle',
        'deal_title_field': 'Anlaşma Başlığı *',
        'value_usd_field': 'Değer (USD)',
        'probability_pct': 'Olasılık %',
        'expected_close': 'Beklenen Kapanış Tarihi',
        'origin_lbl': 'Kaynak',
        'destination_lbl': 'Varış Yeri',
        'notes_lbl': 'Notlar',
        'add_deal_btn': 'Anlaşma Ekle',
        'deal_title_client_req': 'Başlık ve Müşteri Adı gereklidir.',
        'failed_add_deal': 'Anlaşma eklenemedi.',
        'move_to_stage': 'Aşamaya Taşı',
        'update_stage_btn': 'Aşamayı Güncelle',
        'no_deals': 'Filtreyle eşleşen anlaşma bulunamadı.',
        'freight_lbl': '**Yük:**',
        'route_lbl': '**Rota:**',
        'probability_lbl': '**Olasılık:**',
        'close_date_lbl': '**Kapanış Tarihi:**',
        # Offers
        'freight_filter_lbl': 'Yük',
        'total_offers': 'Toplam Teklifler',
        'pending_sent': 'Bekleyen (Gönderildi)',
        'accepted_value': 'Kabul Edilen Değer',
        'commodity_lbl': '**Emtia:**',
        'valid_until_lbl': '**Geçerlilik:**',
        'weight_lbl': '**Ağırlık:**',
        'volume_lbl': '**Hacim:**',
        'update_status_offer': 'Durumu Güncelle',
        'no_offers': 'Filtreyle eşleşen teklif bulunamadı.',
        'create_new_offer_exp': '➕ Yeni Teklif Oluştur',
        'offer_num_field': 'Teklif Numarası *',
        'client_email_field': 'Müşteri E-postası',
        'commodity_desc_field': 'Emtia / Kargo Açıklaması',
        'weight_kg_field': 'Ağırlık (kg)',
        'volume_cbm_field': 'Hacim (CBM)',
        'total_value_usd': 'Toplam Değer (USD)',
        'valid_until_field': 'Geçerlilik Tarihi',
        'notes_incl': 'Notlar / Dahil Olanlar',
        'create_offer_btn': 'Teklif Oluştur',
        'offer_num_client_req': 'Teklif numarası ve müşteri adı gereklidir.',
        'failed_create_offer': 'Teklif oluşturulamadı. Teklif numarası zaten mevcut olabilir.',
        'leads_by_status': 'Duruma Göre Potansiyel Müşteriler',
        'deals_pipeline_chart': 'Anlaşma Hattı',
        'leads_found': 'potansiyel müşteri bulundu',
        'deals_found': 'anlaşma',
        'pipeline_lbl': 'Hat',
        'col_name': 'Ad', 'col_company': 'Şirket', 'col_country': 'Ülke',
        'col_source': 'Kaynak', 'col_freight': 'Yük', 'col_created': 'Oluşturuldu',
        'phone_lbl': 'Telefon',
        'lead_added_success': "Potansiyel müşteri başarıyla eklendi.",
        'deal_added_success': "Anlaşma başarıyla eklendi.",
        'offer_created_success': "Teklif başarıyla oluşturuldu.",
        'stage_updated_success': 'Aşama güncellendi.',
        'offer_status_updated': 'Teklif durumu güncellendi.',
        'discovery_opt': 'Keşif', 'proposal_opt': 'Teklif',
        'negotiation_opt': 'Müzakere',
        # Training
        'total_programs': 'Toplam Programlar',
        'upcoming_lbl': 'Yaklaşan',
        'total_enrollments': 'Toplam Kayıtlar',
        'completions_lbl': 'Tamamlamalar',
        'upcoming_training': 'Yaklaşan Eğitim',
        'programs_by_cat': 'Kategoriye Göre Programlar',
        'cat_filter_lbl': 'Kategori',
        'enroll_employee': 'Çalışanı Kaydet',
        'enroll_btn': 'Kaydet',
        'update_status_prog': 'Durumu Güncelle',
        'update_prog_status': 'Program Durumunu Güncelle',
        'enrolled_participants': '**Kayıtlı Katılımcılar:**',
        'no_programs': "Henüz program yok. 'Program Ekle' sekmesinden bir tane ekleyin.",
        'add_program_hdr': 'Eğitim Programı Ekle',
        'program_title': 'Program Başlığı *',
        'scheduled_date': 'Planlanan Tarih *',
        'duration_hours': 'Süre (saat)',
        'trainer_instructor': 'Eğitmen',
        'max_participants': 'Maksimum Katılımcı',
        'description_area': 'Açıklama',
        'add_program_btn': 'Program Ekle',
        'title_required_err': 'Başlık gereklidir.',
        # Profile page
        'please_login': 'Profilleri görüntülemek için giriş yapın.',
        'no_users_avail': 'Kullanıcı yok.',
        'select_user_view': 'Görüntülenecek kullanıcı seçin',
        'my_profile_btn': 'Profilim',
        'user_not_found': 'Kullanıcı bulunamadı.',
        'change_profile_pic': 'Profil Fotoğrafını Değiştir',
        'upload_photo': 'Fotoğraf yükle (JPG / PNG, max 2 MB)',
        'file_too_large': 'Dosya çok büyük — max 2 MB.',
        'save_photo': 'Fotoğrafı Kaydet',
        'profile_pic_updated': 'Profil fotoğrafı güncellendi!',
        'remove_photo': 'Fotoğrafı Kaldır',
        'photo_removed': 'Fotoğraf kaldırıldı.',
        'leave_requests_lbl': 'İzin Talepleri',
        'employees_lbl': 'Çalışanlar',
        'clients_lbl': 'Müşteriler',
        'leave_req_pending': 'Bekleyen İzin Talepleri',
        'active_lbl': 'Aktif',
        'delivered_lbl': 'Teslim Edildi',
        'customs_cleared': 'Gümrük Temizlendi',
        'full_name_lbl': 'Ad Soyad',
        'phone_lbl': 'Telefon',
        'role_lbl': 'Rol',
        'save_info': 'Bilgileri Kaydet',
        'profile_updated': 'Profil başarıyla güncellendi.',
        'new_pass_auto': 'Yeni şifre (boş = otomatik oluştur)',
        'reset_password_btn': 'Şifreyi Sıfırla',
        'password_reset_done': 'Şifre sıfırlandı.',
        'failed_short': 'Başarısız.',
        'current_password': 'Mevcut Şifre',
        'new_password': 'Yeni Şifre',
        'confirm_new_password': 'Yeni Şifreyi Onayla',
        'update_password_btn': 'Şifreyi Güncelle',
        'pass_no_match_new': 'Yeni şifreler eşleşmiyor.',
        'current_pass_incorrect': 'Mevcut şifre yanlış.',
        'password_updated': 'Şifre başarıyla güncellendi.',
        'failed_update_pass': 'Şifre güncellenemedi.',
        'no_activity_yet': 'Henüz aktivite kaydedilmedi. Giriş, veri ekleme, talep gönderme gibi işlemler burada görünecek.',
        # Messages page
        'your_messages': 'Mesajlarınız',
        'no_messages': '📭 Henüz mesaj yok.',
        'from_lbl': 'Gönderen',
        'to_lbl': 'Alıcı',
        'subject_lbl': '**Konu:**',
        'from_lbl2': '**Gönderen:**',
        'to_lbl2': '**Alıcı:**',
        'shipment_lbl': '**Gönderi:**',
        'date_lbl': '**Tarih:**',
        'message_lbl': '**Mesaj:**',
        'mark_as_read_btn': '✓ Okundu İşaretle',
        'reply_btn': '↩️ Yanıtla',
        'reply_hdr': 'Yanıtla',
        'to_field': 'Alıcı:',
        'subject_field': 'Konu:',
        'message_field': 'Mesaj:',
        'send_reply_btn': '📤 Yanıt Gönder',
        'fill_subject_msg': 'Lütfen konu ve mesaj içeriğini doldurun.',
        'recipient_not_found': 'Alıcı bulunamadı.',
        'reply_sent': '✅ Yanıt başarıyla gönderildi!',
        'error_sending_reply': '❌ Yanıt gönderilirken hata oluştu.',
        'send_new_msg': 'Yeni Mesaj Gönder',
        'no_users_msg': 'Mesaj göndermek için kullanıcı yok.',
        'related_shipment': 'İlgili Gönderi (isteğe bağlı):',
        'send_msg_btn': '📤 Mesaj Gönder',
        'msg_sent': '✅ Mesaj başarıyla gönderildi!',
        'error_sending_msg': '❌ Mesaj gönderilirken hata oluştu.',
        'showing_x_of_y': '{x} / {y} talep gösteriliyor',
        'request_num': 'Talep',
        'download_attach': 'Eki indir',
        'attachment_lbl': 'Ek',
        'user_lbl': 'Kullanıcı',
        'last_updated': 'Son güncelleme',
        'delete_action': 'Sil',
        'name_lbl': 'İsim',
        'total_impressions_kpi': 'Toplam Gösterim',
        'total_clicks_kpi': 'Toplam Tıklama',
        'avg_ctr': 'Ort. TTO',
        'total_revenue_kpi': 'Toplam Gelir',
        'impressions_over_time': 'Zaman İçinde Gösterimler',
        'clicks_conv_over_time': 'Zaman İçinde Tıklama ve Dönüşümler',
        'daily_revenue': 'Günlük Gelir',
        'campaigns_by_type': 'Türe Göre Kampanyalar',
        'ctr_by_campaign': 'Kampanyaya Göre TTO',
        'cvr_by_campaign': 'Kampanyaya Göre Dönüşüm Oranı',
        'date_lbl': 'Tarih',
        'count_lbl': 'Sayı',
        'metric_lbl': 'Metrik',
        'impressions_lbl': 'Gösterimler',
        'rate_service_caption': 'EIMS Lojistik ile deneyiminizi paylaşın — geri bildiriminiz iyileştirmemize yardımcı olur.',
        'no_delivered_ships': 'Henüz teslim edilmiş ve değerlendirilecek gönderiniz yok. Değerlendirme, gönderi Teslim Edildi olarak işaretlendikten sonra kullanılabilir.',
        'ships_awaiting_rating': 'Değerlendirme Bekleyen Gönderiler ({n})',
        'delivered_lbl': 'Teslim Edildi',
        'rating_lbl': 'Puan *',
        'what_rating': 'Neyi değerlendiriyorsunuz?',
        'comment_optional': 'Yorumunuz (isteğe bağlı)',
        'submit_rating': 'Puanı Gönder',
        'rating_submitted': 'Teşekkürler! Puanınız gönderildi.',
        'all_ships_rated': '✅ Teslim edilen tüm gönderilerinizi değerlendirdiniz. Teşekkürler!',
        'already_rated': 'Değerlendirilmiş ({n})',
        'edit_rating_btn': 'Değerlendirmeyi Düzenle',
        'delete_rating_btn': 'Değerlendirmeyi Sil',
        'confirm_delete_rating': 'Bu değerlendirmeyi silmek istediğinizden emin misiniz?',
        'rating_edited_success': 'Değerlendirme başarıyla güncellendi.',
        'rating_deleted_ok': 'Değerlendirme silindi.',
        'cat_overall': 'Genel',
        'cat_delivery_time': 'Teslimat Süresi',
        'cat_service_quality': 'Hizmet Kalitesi',
        'cat_documentation': 'Belgeler',
        'cat_communication': 'İletişim',
        'cat_pricing': 'Fiyatlandırma',
        'track_btn': '🔍 Takip Et',
        'ship_not_found': 'Gönderi bulunamadı.',
        'own_ships_only': 'Yalnızca kendi gönderilerinizi takip edebilirsiniz.',
        'ship_found': 'Gönderi Bulundu: {n}',
        'tracking_history': 'Takip Geçmişi',
        'no_tracking_updates': 'Henüz takip güncellemesi yok.',
        'enter_ship_number': 'Lütfen bir gönderi numarası girin.',
        'origin_lbl': 'Köken',
        'destination_lbl': 'Varış Noktası',
        'likes_lbl': 'Beğeniler',
        'shares_lbl': 'Paylaşımlar',
        'comments_lbl': 'Yorumlar',
        'write_post_ph': 'Gönderi yazın...',
        'engagement_by_platform': 'Platforma Göre Etkileşim',
        'posts_distribution': 'Gönderi Dağılımı',
        'posts_lbl': 'Gönderiler',
        'platform_lbl': 'Platform',
        'update_engagement_btn': 'Etkileşimi Güncelle',
        'users_by_role': 'Role Göre Kullanıcılar',
        'events_by_type': 'Türe Göre Olaylar',
        'role_col': 'Rol',
        'shipments_by_status': 'Duruma Göre Gönderiler',
        'monthly_ship_vol': 'Aylık Gönderi Hacmi',
        'needs_action': 'İşlem gerekiyor',
        'shipments_lbl': 'Gönderiler',
        'month_lbl': 'Ay',
        'shipment_hash': 'Gönderi #',
        'mode_lbl': 'Mod',
        'carrier_lbl': 'Taşıyıcı',
        'departure_lbl': 'Kalkış',
        'eta_lbl': 'TVT',
        'footer_text': 'EIMS | Şirket Veri Yönetim Sistemi',
        'category_lbl': 'Kategori',
        'expiry_lbl': 'Son Kullanma',
        'x_documents': '{n} belge',
        'active_opt': 'Aktif',
        'inactive_opt': 'Pasif',
        'on_leave_opt': 'İzinde',
        'draft_opt': 'Taslak',
        'scheduled_opt': 'Planlandı',
        'published_opt': 'Yayınlandı',
        'paused_opt': 'Duraklatıldı',
        'completed_opt': 'Tamamlandı',
        'cancelled_opt': 'İptal Edildi',
        'archived_opt': 'Arşivlendi',
        'expired_opt': 'Süresi Doldu',
        'open_opt': 'Açık',
        'closed_opt': 'Kapalı',
        'in_progress_opt': 'Devam Ediyor',
        'low_opt': 'Düşük',
        'medium_opt': 'Orta',
        'high_opt': 'Yüksek',
        'urgent_opt': 'Acil',
        'import_opt': 'İthalat',
        'export_opt': 'İhracat',
        'sea_opt': 'Deniz',
        'air_opt': 'Hava',
        'land_opt': 'Kara',
        'paid_opt': 'Ücretli',
        'unpaid_opt': 'Ücretsiz',
        'sick_opt': 'Hastalık',
        'other_opt': 'Diğer',
        'employee_opt': 'Çalışan',
        'client_opt': 'Müşteri',
        'manager_opt': 'Müdür',
        'male_opt': 'Erkek',
        'female_opt': 'Kadın',
        'fulltime_opt': 'Tam zamanlı',
        'parttime_opt': 'Yarı zamanlı',
        'select_placeholder': '— Seçin —',
        'overall_opt': 'Genel',
        'delivery_time_opt': 'Teslimat Süresi',
        'service_quality_opt': 'Hizmet Kalitesi',
        'documentation_opt': 'Dokümantasyon',
        'communication_opt': 'İletişim',
        'pricing_opt': 'Fiyatlandırma',
        'sent_opt': 'Gönderildi',
        'overdue_opt': 'Gecikmiş',
        'email_type_opt': 'E-posta',
        'social_media_opt': 'Sosyal Medya',
        'content_type_opt': 'İçerik',
        'event_opt': 'Etkinlik',
        'none_opt': 'Yok',
        'category_label': 'Kategori',
        'doc_type_lbl': 'Belge Türü',
        'confirmed_opt': 'Onaylandı',
        'in_transit_opt': 'Taşımada',
        'customs_opt': 'Gümrük',
        'delivered_opt': 'Teslim Edildi',
        'resolved_opt': 'Çözüldü',
        'departed_opt': 'Hareket Etti',
        'arrived_port_opt': 'Limana Ulaştı',
        'out_for_delivery_opt': 'Dağıtımda',
        'shipment_inquiry_opt': 'Sevkiyat Sorgusu',
        'doc_request_opt': 'Belge Talebi',
        'customs_issue_opt': 'Gümrük Sorunu',
        'complaint_opt': 'Şikayet',
        'rate_request_opt': 'Ücret Talebi',
        'general_inquiry_opt': 'Genel Sorgu',
        'new_opt': 'Yeni',
        'contacted_opt': 'İletişime Geçildi',
        'qualified_opt': 'Nitelikli',
        'proposal_sent_opt': 'Teklif Gönderildi',
        'won_opt': 'Kazanıldı',
        'lost_opt': 'Kaybedildi',
        'cold_call_opt': 'Soğuk Arama',
        'referral_opt': 'Tavsiye',
        'trade_show_opt': 'Fuar',
        'website_opt': 'Web Sitesi',
        'email_campaign_opt': 'E-posta Kampanyası',
        'sea_fcl_opt': 'Deniz FCL',
        'sea_lcl_opt': 'Deniz LCL',
        'air_cargo_opt': 'Hava Kargo',
        'road_opt': 'Kara',
        'accepted_opt': 'Kabul Edildi',
        'invoice_doc_opt': 'Fatura',
        'bill_of_lading_opt': 'Konşimento',
        'customs_declaration_opt': 'Gümrük Beyannamesi',
        'certificate_origin_opt': 'Menşei Belgesi',
        'packing_list_opt': 'Ambalaj Listesi',
        'currency_lbl': 'Para Birimi',
        'filter_by_type': 'Türe Göre Filtrele:',
        'select_your_shipment': 'Sevkiyatınızı Seçin:',
        'select_cargo_item': 'Kargo Öğesi Seçin:',
        'select_shipment_edit': 'Düzenlenecek Sevkiyatı Seçin:',
        'select_shipment_delete': 'Silinecek Sevkiyatı Seçin:',
        'action_lbl': 'İşlem:',
        'type_lbl': 'Tür',
        'location_lbl': 'Konum',
        'duration_lbl': 'Süre',
        'min_lbl': 'dak',
        'attendees_lbl': 'Katılımcılar',
        'agenda_lbl': 'Gündem',
        'minutes_lbl': 'Tutanak',
        'no_upcoming_meetings': 'Yaklaşan toplantı yok.',
        'no_past_meetings': 'Geçmiş toplantı yok.',
        'active_campaigns_kpi': 'Aktif Kampanyalar',
        'total_budget_kpi': 'Toplam Bütçe',
        'scheduled_posts_kpi': 'Planlanmış Gönderiler',
        'x_total': 'Toplam {n}',
        'x_spent': '${n} harcandı',
        'x_clicks': '{n} tıklama',
        'budget_lbl': 'Bütçe',
        'spent_lbl': 'Harcanan',
        'period_lbl': 'Dönem',
        'target_aud_lbl': 'Hedef Kitle',
        'no_campaign_metrics': 'Henüz kampanya metrikleri yok.',
        'no_budget_data': 'Henüz bütçe verisi yok.',
        'campaign_created_ok': "'{name}' kampanyası oluşturuldu!",
        'metrics_saved_ok': 'Metrikler kaydedildi!',
        'email_not_found': 'Bu e-posta ile kayıtlı hesap bulunamadı. Lütfen kaydolun.',
        'incorrect_password': 'Yanlış şifre',
        'login_error': 'Giriş sırasında bir hata oluştu. Lütfen tekrar deneyin.',
        'forgot_pass_desc': 'Şifre sıfırlama talebinde bulunmak için bilgilerinizi girin',
        'enter_reg_email': 'Lütfen kayıtlı e-posta adresinizi girin',
        'enter_contact_email': 'Lütfen iletişim e-posta adresinizi girin',
        'reset_email_sent_maybe': 'Bu e-posta kayıtlıysa, yöneticiye sıfırlama talebi gönderilecektir.',
        'no_managers_available': 'Sistemde yönetici bulunamadı. Lütfen teknik destek ile iletişime geçin.',
        'request_submitted_success': '✅ Talep başarıyla gönderildi!',
        'error_sending_request': 'İstek gönderilirken bir hata oluştu. Lütfen daha sonra tekrar deneyin.',
        'error_try_again': 'Bir hata oluştu. Lütfen daha sonra tekrar deneyin.',
        'create_new_account': 'Yeni Hesap Oluştur',
        'choose_account_type': 'Kayıt olmak istediğiniz hesap türünü seçin',
        'employee_account_lbl': 'Çalışan Hesabı',
        'client_account_lbl': 'Müşteri Hesabı',
        'invalid_email_format': '❌ Lütfen geçerli bir e-posta adresi girin',
        'email_already_exists': 'Bu e-posta ile bir hesap zaten mevcut.',
        'failed_create_account': 'Hesap oluşturulamadı. Tekrar deneyin.',
        'emp_registration_hdr': 'Çalışan Kayıdı',
        'client_registration_hdr': 'Müşteri Kayıdı',
        'register_emp_help': 'Tam erişimle çalışan olarak kaydolun',
        'register_client_help': 'Temel erişimle müşteri olarak kaydolun',
        'origin_dest_same': 'Çıkış ve varış ülkesi aynı olamaz.',
        'failed_create_shipment_err': 'Gönderi oluşturulamadı.',
        'failed_add_route_err': 'Güzergh eklenemedi.',
        'no_psutil': 'psutil paketi yüklü değil. Sistem Yönetimini kullanmak için yükleyin.',
        'content_required': 'İçerik zorunludur.',
        'please_enter_location': 'Lütfen konum girin',
        'please_select_file': 'Lütfen bir dosya seçin',
        'only_managers_export': 'Yalnızca yöneticiler çalışan verilerini dışa aktarabilir.',
        'only_emp_mgr_export': 'Yalnızca çalışanlar veya yöneticiler gönderileri dışa aktarabilir.',
        'only_mgr_emp_ships': 'Gönderileri yönetmek için yönetici veya çalışan olmanız gerekir.',
        'no_clients_add_ship': 'Gönderi eklemeden önce lütfen en az bir müşteri kullanıcısı oluşturun.',
        'page_for_clients': 'Bu sayfa yalnızca müşteriler içindir.',
        'must_be_manager_leave': 'İzin taleplerini yönetmek için yönetici olmanız gerekir.',
        'must_be_manager_users_err': 'Kullanıcıları yönetmek için yönetici olmanız gerekir.',
        'failed_update_role_db': 'Veritabanında rol güncellenemedi.',
        'failed_login_submit': 'İzin talebi göndermek için lütfen giriş yapın.',
        'failed_create_ticket_err': 'Bilet oluşturulamadı.',
        'failed_submit_feedback_err': 'Geri bildirim gönderilemedi.',
        'failed_add_document_err': 'Belge eklenemedi.',
        'failed_schedule_meeting_err': 'Toplantı planlanamadı.',
        'failed_add_contract_err': 'Sözleşme eklenemedi.',
        'failed_add_lead_err': 'Potansiyel müşteri eklenemedi.',
        'failed_create_invoice_err': 'Fatura oluşturulamadı.',
        'failed_record_payment_err': 'Ödeme kaydedilemedi.',
        'failed_add_expense_err': 'Masraf eklenemedi.',
        'status_updated_success': '✅ Durum başarıyla güncellendi!',
        'cargo_item_updated': '✅ Kargo kalemi başarıyla güncellendi!',
        'item_deleted_success': '✅ Öğe başarıyla silindi!',
        'cargo_item_added': '✅ Kargo kalemi başarıyla eklendi!',
        'tracking_added': '✅ Takip güncellemesi başarıyla eklendi!',
        'document_uploaded': 'Belge yüklendi!',
        'request_updated': 'Talep güncellendi.',
        'user_role_updated': 'Kullanıcı rolü güncellendi.',
        'ticket_updated_success': 'Bilet güncellendi.',
        'feedback_submitted_success': 'Geri bildirim gönderildi.',
        'document_added_success': 'Belge eklendi.',
        'updated_label': 'Güncellendi!',
        'deleted_label': 'Silindi.',
        'added_label': 'Eklendi!',
        'post_saved': 'Gönderi kaydedildi!',
        'no_shipments_start': 'Gönderi bulunamadı. Başlamak için yeni bir gönderi ekleyin.',
        'no_cargo_items': 'Kargo kalemi bulunamadı.',
        'no_tracking_yet': 'Henüz takip güncellemesi yok.',
        'no_documents_yet': 'Henüz belge yüklenmedi.',
        'no_shipments_client': 'Henüz gönderiniz yok.',
        'no_cargo_listed': 'Listelenen kargo kalemi yok.',
        'no_unpaid_invoices': 'Ödenmemiş fatura yok. Tüm faturalar ödendi.',
        'no_expenses_found': 'Masraf bulunamadı.',
        'no_perf_data': 'Henüz performans verisi yok.',
        'no_devices_registered': 'Henüz kayıtlı cihaz yok. BT makinelerine monitoring_agent.py çalıştırın.',
        'no_server_data_yet': 'Henüz sunucu verisi yok. Önce BT Panosunu açın.',
        'no_security_events_yet': 'Henüz güvenlik olayı kaydedilmedi.',
        'no_shipments_found': 'Gönderi bulunamadı.',
        'no_tracking_ship': 'Bu gönderi için henüz takip güncellemesi yok.',
        'no_ship_available': 'Mevcut gönderi yok.',
        'no_routes_found': 'Güzergh bulunamadı.',
        'no_documents_found': 'Belge bulunamadı.',
        'no_contracts_found': 'Sözleşme bulunamadı.',
        'no_leads_filters_msg': 'Filtrelere uyan potansiyel müşteri bulunamadı.',
        'no_leave_reqs_found': 'İzin talebi bulunamadı.',
        'no_reqs_filters': 'Filtrelerinize uyan talep yok.',
        'no_users_found_msg': 'Kullanıcı bulunamadı.',
        'no_tickets_found': 'Bilet bulunamadı.',
        'no_campaigns_go': 'Henüz kampanya yok. Kampanya oluşturmak için Kampanyalar sayfasına gidin.',
        'no_metrics_campaigns': 'Henüz metrik verisi yok. Kampanyalar sayfasından metrik ekleyin.',
        'no_data_available_yet': 'Henüz veri yok.',
        'no_posts_yet': 'Henüz gönderi yok.',
        'no_published_posts': 'Henüz yayınlanmış gönderi yok.',
        'psutil_missing': '⚠️ Sistem metrikleri kullanılamıyor — psutil bu Python ortamında bulunamadı.',
        'no_users_send_msg': 'Mesaj gönderecek kullanıcı yok.',
        'no_feedback_yet': 'Henüz geri bildirim alınmadı.',
        'no_clients_warning': "⚠️ Müşteri bulunamadı. Lütfen önce 'müşteri' rolünde bir kullanıcı oluşturun.",
        'action_undone_warn': '⚠️ Bu işlem geri alınamaz!',
        'no_users_email_warn': 'E-posta adresi olan kullanıcı bulunamadı.',
        'ship_details_mgmt': 'Gönderi Detayları ve Yönetimi',
        'update_status_hdr': 'Durumu Güncelle',
        'ship_details_hdr': 'Gönderi Detayları',
        'rate_shipment': '⭐ Bu Gönderiye Puan Ver',
        'filter_by_type_lbl': 'Türe Göre Filtrele:',
        'filter_by_status_lbl2': 'Duruma Göre Filtrele:',
        'search_lbl2': '\U0001f50d Ara',
        'search_ship_ph': 'Gönderi numarası, müşteri...',
        'filter_priority_lbl': 'Önceliğe Göre Filtrele',
        'filter_status_leave_lbl': 'Duruma Göre Filtrele',
        'search_email_reason': 'E-posta veya nedene göre ara',
        'type_to_search': 'Aramak için yazın...',
        'item_name_lbl': 'Ürün Adı',
        'quantity_lbl': 'Miktar',
        'weight_kg_lbl': 'Ağırlık (kg)',
        'unit_lbl': 'Birim',
        'value_lbl': 'Değer',
        'hs_code_lbl': 'HS Kodu',
        'hs_code_optional': 'HS Kodu (İsteğe Bağlı)',
        'update_status_star': 'Durum *',
        'update_date_star': 'Güncelleme Tarihi *',
        'doc_type_lbl': 'Belge Türü',
        'choose_file_lbl': 'Dosya Seç',
        'type_imp_exp_star': 'Tür *',
        'client_star': 'Müşteri *',
        'enter_shipment_num_lbl': 'Gönderi Numarasını Girin:',
        'select_shipment_lbl': 'Gönderi Seç:',
        'select_item_edit': 'Düzenlenmecek öğeyi seçin:',
        'select_post_lbl': 'Gönderi Seç',
        'update_status_ticket_lbl': 'Durumu Güncelle',

        'act_login_success': 'Giriş Başarılı',
        'act_logout': 'Çıkış',
        'act_login_failed': 'Giriş Başarısız',
        'act_leave_request': 'İzin Talebi',
        'act_leave_updated': 'İzin Güncellendi',
        'act_add_record': 'Kayıt Eklendi',
        'act_edit_record': 'Kayıt Düzenlendi',
        'act_delete_record': 'Kayıt Silindi',
        'act_role_update': 'Rol Güncelleme',
        'act_profile_update': 'Profil Güncelleme',
        'act_avatar_update': 'Fotoğraf Güncelleme',
        'act_password_change': 'Şifre Değişikliği',
        'act_ticket_created': 'Bilet Oluşturuldu',
        'act_ticket_updated': 'Bilet Güncellendi',
        'act_ticket_deleted': 'Bilet Silindi',
        'unread_messages_warn': '{count} okunmamış mesajınız var',
    }
}


@st.cache_resource(show_spinner=False)
def init_database():
    return Database()

@st.cache_resource(show_spinner=False)
def init_data_manager():
    return DataManager()

db = init_database()
data_manager = init_data_manager()

# Initialize language in session state (only if not already set)
if 'language' not in st.session_state:
    st.session_state['language'] = 'en'
# Initialize theme in session state ('light' or 'dark')
if 'theme' not in st.session_state:
    st.session_state['theme'] = 'light'

def t(key: str) -> str:
    """Get translation for a key in the current language."""
    lang = st.session_state.get('language', 'en')
    return TRANSLATIONS.get(lang, TRANSLATIONS['en']).get(key, key)

# Maps English option values → translation keys for use with format_func=_topt
_OPT_KEYS = {
    'All': 'all',
    'Active': 'active_opt', 'Inactive': 'inactive_opt', 'On Leave': 'on_leave_opt',
    'Draft': 'draft_opt', 'Scheduled': 'scheduled_opt', 'Published': 'published_opt',
    'Paused': 'paused_opt', 'Completed': 'completed_opt', 'Cancelled': 'cancelled_opt',
    'Archived': 'archived_opt', 'Expired': 'expired_opt',
    'Open': 'open_opt', 'Closed': 'closed_opt', 'In Progress': 'in_progress_opt',
    'Pending': 'pending', 'Approved': 'approved', 'Rejected': 'rejected',
    'Low': 'low_opt', 'Medium': 'medium_opt', 'High': 'high_opt', 'Urgent': 'urgent_opt',
    'Import': 'import_opt', 'Export': 'export_opt',
    'Sea': 'sea_opt', 'Air': 'air_opt', 'Land': 'land_opt',
    'Paid': 'paid_opt', 'Unpaid': 'unpaid_opt', 'Sick': 'sick_opt', 'Other': 'other_opt',
    'employee': 'employee_opt', 'client': 'client_opt', 'manager': 'manager_opt',
    'Male': 'male_opt', 'Female': 'female_opt',
    'Full-time': 'fulltime_opt', 'Part-time': 'parttime_opt',
    '— Select —': 'select_placeholder',
    'Overall': 'overall_opt', 'Delivery Time': 'delivery_time_opt',
    'Service Quality': 'service_quality_opt', 'Documentation': 'documentation_opt',
    'Communication': 'communication_opt', 'Pricing': 'pricing_opt',
    'Sent': 'sent_opt', 'Overdue': 'overdue_opt',
    'Email': 'email_type_opt', 'Social Media': 'social_media_opt',
    'Content': 'content_type_opt', 'Event': 'event_opt', 'None': 'none_opt',
    'Confirmed': 'confirmed_opt', 'In Transit': 'in_transit_opt',
    'Customs': 'customs_opt', 'Delivered': 'delivered_opt', 'Resolved': 'resolved_opt',
    'Departed': 'departed_opt', 'Arrived at Port': 'arrived_port_opt',
    'Out for Delivery': 'out_for_delivery_opt',
    'Shipment Inquiry': 'shipment_inquiry_opt', 'Document Request': 'doc_request_opt',
    'Customs Issue': 'customs_issue_opt', 'Complaint': 'complaint_opt',
    'Rate Request': 'rate_request_opt', 'General Inquiry': 'general_inquiry_opt',
    'New': 'new_opt', 'Contacted': 'contacted_opt', 'Qualified': 'qualified_opt',
    'Proposal Sent': 'proposal_sent_opt', 'Won': 'won_opt', 'Lost': 'lost_opt',
    'Cold Call': 'cold_call_opt', 'Referral': 'referral_opt',
    'Trade Show': 'trade_show_opt', 'Website': 'website_opt',
    'Email Campaign': 'email_campaign_opt',
    'Sea FCL': 'sea_fcl_opt', 'Sea LCL': 'sea_lcl_opt',
    'Air Cargo': 'air_cargo_opt', 'Road': 'road_opt', 'Accepted': 'accepted_opt',
    'Invoice': 'invoice_doc_opt', 'Bill of Lading': 'bill_of_lading_opt',
    'Customs Declaration': 'customs_declaration_opt',
    'Certificate of Origin': 'certificate_origin_opt',
    'Packing List': 'packing_list_opt',
    'Discovery': 'discovery_opt', 'Proposal': 'proposal_opt', 'Negotiation': 'negotiation_opt',
    'Carrier': 'carrier_opt', 'Supplier': 'supplier_opt', 'Government': 'government_opt',
    'Terminated': 'terminated_opt',
    'Service Agreement': 'service_agreement_opt', 'Rate Agreement': 'rate_agreement_opt',
    'Agency Agreement': 'agency_agreement_opt', 'Insurance Policy': 'insurance_policy_opt',
    'Software License': 'software_license_opt', 'Lease Agreement': 'lease_agreement_opt',
    'NDA': 'nda_opt',
    'Cargo Cost': 'cargo_cost_opt', 'Customs Duty': 'customs_duty_opt',
    'Freight Fee': 'freight_fee_opt', 'Fuel & Transport': 'fuel_transport_opt',
    'Staff Salary': 'staff_salary_opt', 'Office Rent': 'office_rent_opt',
    'Insurance': 'insurance_opt', 'Port Handling': 'port_handling_opt',
    'Bank Transfer': 'bank_transfer_opt', 'Letter of Credit': 'letter_credit_opt',
    'Cash': 'cash_opt', 'Credit Card': 'credit_card_opt', 'Cheque': 'cheque_opt',
}

def _topt(val: str) -> str:
    """Translate a dropdown option value. Falls back to the original if no key found."""
    key = _OPT_KEYS.get(val)
    return t(key) if key else val

_DEPT_KEYS = {
    'Finance': 'dept_finance', 'Marketing': 'dept_marketing', 'IT': 'dept_it',
    'Logistics': 'dept_logistics', 'Customer Service': 'dept_customer_service',
    'Administration': 'dept_administration', 'Sales': 'dept_sales',
}

def _tdept(dept: str) -> str:
    key = _DEPT_KEYS.get(dept)
    return t(key) if key else dept

def apply_theme():
    """Inject CSS for light or dark theme based on session state."""
    theme = st.session_state.get('theme', 'light')
    if theme == 'dark':
        css = """
        <style>
        /* ── BACKGROUNDS — only top-level containers ── */
        .stApp, body, [data-testid="stAppViewContainer"],
        [data-testid="stAppViewBlockContainer"],
        .main, .main > div, .block-container,
        [data-testid="block-container"] {
            background-color: #0e1117 !important;
        }
        /* inner layout containers → transparent so they don't paint black over cards/forms */
        .stVerticalBlock, .stVerticalBlockBorderWrapper,
        .element-container, [data-testid="column"],
        [data-testid="stColumns"], [data-testid="stHorizontalBlock"] {
            background-color: transparent !important;
        }

        /* ── SIDEBAR ── */
        [data-testid="stSidebar"],
        [data-testid="stSidebar"] > div,
        [data-testid="stSidebarContent"] {
            background-color: #262730 !important;
        }
        [data-testid="stSidebar"] .stVerticalBlock,
        [data-testid="stSidebar"] .element-container,
        [data-testid="stSidebar"] [data-testid="column"],
        [data-testid="stSidebar"] .stVerticalBlockBorderWrapper {
            background-color: #262730 !important;
        }
        [data-testid="stSidebar"] hr { border-color: #4a4a5a !important; }

        /* ── TEXT ── */
        html, body, .stApp,
        h1, h2, h3, h4, h5, h6,
        p, span, label, li, a,
        [data-testid="stMarkdownContainer"],
        [data-testid="stMarkdownContainer"] *,
        [data-testid="stText"], .stMarkdown { color: #fafafa !important; }

        /* ── INPUTS ── */
        input, textarea,
        .stTextInput input, .stTextArea textarea, .stNumberInput input,
        .stDateInput input { background-color: #1e2030 !important; color: #e2e8f0 !important; border-color: #4a4a5a !important; }
        .stTextInput > div > div, .stTextArea > div > div,
        .stNumberInput > div > div, .stDateInput > div > div { background-color: #1e2030 !important; border-color: #4a4a5a !important; }

        /* Form field labels */
        .stTextInput label, .stTextArea label, .stNumberInput label,
        .stSelectbox label, .stDateInput label, .stMultiSelect label,
        /* ── LABELS (all form field labels) ── */
        [data-testid="stWidgetLabel"],
        [data-testid="stWidgetLabel"] p,
        [data-testid="stWidgetLabel"] label,
        .stTextInput label, .stTextInput label p,
        .stTextArea label, .stTextArea label p,
        .stNumberInput label, .stNumberInput label p,
        .stSelectbox label, .stSelectbox label p,
        .stDateInput label, .stDateInput label p,
        .stMultiSelect label, .stMultiSelect label p,
        .stCheckbox label p, .stRadio label p { color: #94a3b8 !important; font-size: 13px !important; }

        /* Password show/hide button */
        .stTextInput button, .stTextInput [data-testid="InputInstructions"],
        input[type="password"] + button,
        [data-testid="stTextInput"] button {
            background-color: #1e2030 !important;
            border: none !important;
            color: #94a3b8 !important;
        }
        [data-testid="stTextInput"] > div { background-color: #1e2030 !important; border-color: #4a4a5a !important; }

        /* ── SELECT / MULTISELECT — input field ── */
        .stSelectbox > div > div, .stMultiSelect > div > div,
        [data-baseweb="select"] > div { background-color: #1e2030 !important; border-color: #4a4a5a !important; color: #fafafa !important; }

        /* ── DROPDOWN POPUP (popover + menu) ── */
        [data-baseweb="popover"],
        [data-baseweb="popover"] > div,
        [data-baseweb="popover"] > div > div,
        [data-baseweb="menu"],
        [data-baseweb="menu"] > div,
        [data-baseweb="menu"] ul,
        [data-baseweb="menu"] li {
            background-color: #1e2030 !important;
            border-color: #4a4a5a !important;
            color: #e2e8f0 !important;
        }
        [data-baseweb="popover"] {
            border: 1px solid #4a4a5a !important;
            border-radius: 6px !important;
            box-shadow: 0 4px 16px rgba(0,0,0,0.5) !important;
        }
        [data-baseweb="option"] { background-color: #1e2030 !important; color: #e2e8f0 !important; }
        [data-baseweb="option"]:hover { background-color: #3a3a4a !important; color: #ffffff !important; }
        [data-baseweb="option"][aria-selected="true"] { background-color: #2a3550 !important; color: #93c5fd !important; }
        [data-baseweb="tag"] { background-color: #3a3a4a !important; color: #fafafa !important; }

        /* ── DROPDOWN PORTAL (renders in body, outside normal DOM) ── */
        body > div[data-baseweb="layer"],
        body > div[data-baseweb="layer"] *,
        body > div > [data-baseweb="popover"],
        body > div > [data-baseweb="popover"] *,
        [role="listbox"],
        [role="listbox"] *,
        [role="option"],
        ul[data-baseweb="menu"],
        ul[data-baseweb="menu"] li,
        li[role="option"] {
            background-color: #1e2030 !important;
            color: #e2e8f0 !important;
            border-color: #4a4a5a !important;
        }
        [role="option"]:hover,
        li[role="option"]:hover { background-color: #3a3a4a !important; color: #ffffff !important; }
        [role="option"][aria-selected="true"] { background-color: #2a3550 !important; color: #93c5fd !important; }
        [data-baseweb="select"] [data-baseweb="input"],
        [data-baseweb="select"] input { background-color: #1e2030 !important; color: #e2e8f0 !important; }

        /* ── BUTTONS ── */
        .stButton > button, button[kind="secondary"], button[kind="primary"] {
            background: #262730 !important; color: #fafafa !important;
            border: 1px solid #4a4a5a !important;
        }
        .stButton > button:hover { background: #3a3a4a !important; border-color: #6a6a7a !important; }
        .stFormSubmitButton > button {
            background: #5e81ac !important;
            color: #ffffff !important;
            border: none !important;
            border-radius: 8px !important;
            font-weight: 600 !important;
            transition: background 0.2s ease !important;
        }
        .stFormSubmitButton > button:hover { background: #4a6d96 !important; }

        /* ── FORM ── */
        [data-testid="stForm"] { background-color: #1a1d27 !important; border: 1px solid #3a3a4a !important; border-radius: 10px !important; padding: 16px !important; }
        [data-testid="stForm"] > div,
        [data-testid="stForm"] .stVerticalBlock,
        [data-testid="stForm"] .stVerticalBlockBorderWrapper,
        [data-testid="stForm"] .element-container,
        [data-testid="stForm"] [data-testid="stWidgetLabel"],
        [data-testid="stForm"] [data-testid="stWidgetLabel"] * { background-color: transparent !important; border: none !important; }

        /* ── REGULAR BUTTONS (non-submit) ── */
        .stButton > button {
            background: #262730 !important;
            color: #e2e8f0 !important;
            border: 1px solid #3a3a4a !important;
            border-radius: 8px !important;
        }
        .stButton > button:hover { background: #3a3a4a !important; }

        /* ── EXPANDER ── */
        [data-testid="stExpander"], [data-testid="stExpander"] > div,
        [data-testid="stExpander"] summary { background-color: #262730 !important; color: #fafafa !important; border-color: #4a4a5a !important; }

        /* ── TABS ── */
        [data-baseweb="tab-list"] { background-color: #262730 !important; border-color: #4a4a5a !important; }
        [data-baseweb="tab"] { color: #aaaaaa !important; background-color: transparent !important; }
        [aria-selected="true"][data-baseweb="tab"] { color: #ff4b4b !important; border-bottom: 2px solid #ff4b4b !important; }
        [data-baseweb="tab-panel"] { background-color: #0e1117 !important; }

        /* ── METRICS ── */
        [data-testid="metric-container"] { background-color: #1e2030 !important; border: 1px solid #4a4a5a !important; border-radius: 10px !important; padding: 12px !important; }
        [data-testid="metric-container"] label,
        [data-testid="stMetricLabel"] { color: #94a3b8 !important; font-size: 13px !important; }
        [data-testid="stMetricValue"],
        [data-testid="stMetricValue"] > div { color: #e2f0ff !important; font-size: 2rem !important; font-weight: 700 !important; }
        [data-testid="stMetricDelta"] { color: #4ade80 !important; }
        [data-testid="stMetricDelta"][data-direction="down"] { color: #f87171 !important; }

        /* ── ALERTS ── */
        [data-testid="stAlert"], [data-testid="stAlert"] > div { background-color: #262730 !important; color: #fafafa !important; }
        [data-testid="stAlert"] p { color: #fafafa !important; }

        /* ── DATAFRAME (canvas-based) ── */
        [data-testid="stDataFrame"],
        [data-testid="stDataFrame"] > div,
        [data-testid="stDataFrame"] > div > div {
            filter: invert(1) hue-rotate(180deg) brightness(0.88) contrast(0.95) !important;
            border-radius: 8px !important;
            overflow: hidden !important;
        }
        /* st.table HTML table */
        [data-testid="stTable"] table { background-color: #0e1117 !important; border-collapse: collapse !important; width: 100% !important; }
        [data-testid="stTable"] th { background-color: #1e2030 !important; color: #94a3b8 !important; border: 1px solid #4a4a5a !important; padding: 10px 14px !important; font-weight: 600 !important; }
        [data-testid="stTable"] td { background-color: #0e1117 !important; color: #e2e8f0 !important; border: 1px solid #3a3a4a !important; padding: 10px 14px !important; }
        [data-testid="stTable"] tr:hover td { background-color: #1e2030 !important; }

        /* ── FILE UPLOADER ── */
        [data-testid="stFileUploader"],
        [data-testid="stFileUploadDropzone"],
        [data-testid="stFileUploader"] > div,
        [data-testid="stFileUploader"] section,
        [data-testid="stFileUploader"] section > div,
        [data-testid="stFileUploader"] section > input + div,
        [data-testid="stFileUploaderDropzone"],
        [data-testid="stFileUploaderDropzoneInstructions"],
        [data-testid="stFileUploaderDropzone"] > div {
            background-color: #1e2030 !important;
            border-color: #4a4a5a !important;
            border-radius: 8px !important;
            color: #e2e8f0 !important;
        }
        [data-testid="stFileUploader"] section { border: 1px dashed #4a4a5a !important; }
        [data-testid="stFileUploader"] small,
        [data-testid="stFileUploader"] span,
        [data-testid="stFileUploader"] p { color: #e2e8f0 !important; }

        /* ── RADIO / CHECKBOX ── */
        .stRadio, .stCheckbox,
        .stRadio > div, .stCheckbox > div,
        .stRadio > label, .stCheckbox > label { background-color: transparent !important; }
        .stRadio label, .stCheckbox label,
        .stRadio p, .stCheckbox p { color: #e2e8f0 !important; }

        /* ── INLINE CODE (backticks) ── */
        code, pre, .stMarkdown code, .stMarkdown pre,
        [data-testid="stMarkdownContainer"] code,
        [data-testid="stMarkdownContainer"] pre {
            background-color: #1e2030 !important;
            color: #93c5fd !important;
            border: 1px solid #3a3a4a !important;
            border-radius: 4px !important;
            padding: 1px 6px !important;
        }

        /* ── MISC ── */
        hr { border-color: #4a4a5a !important; }
        .card, .metric-card { background-color: #262730 !important; color: #fafafa !important; border-color: #4a4a5a !important; }
        [data-testid="stToolbar"], header { background-color: #0e1117 !important; }
        [data-testid="stDecoration"] { background-color: #0e1117 !important; }

        /* ── column divider fix (sidebar) ── */
        [data-testid="stSidebar"] [data-testid="stHorizontalBlock"] {
            gap: 6px !important;
            background: transparent !important;
        }
        [data-testid="stSidebar"] [data-testid="stHorizontalBlock"]::before,
        [data-testid="stSidebar"] [data-testid="stHorizontalBlock"]::after {
            display: none !important;
        }
        [data-testid="stSidebar"] [data-testid="column"] {
            background: transparent !important;
            border: none !important;
            padding: 0 !important;
        }

        /* ── MAIN CONTENT: make all columns & containers transparent ── */
        [data-testid="stHorizontalBlock"],
        [data-testid="column"],
        [data-testid="stVerticalBlock"],
        [data-testid="stVerticalBlockBorderWrapper"],
        [data-testid="stMarkdownContainer"] {
            background-color: transparent !important;
        }
        </style>
        """
    else:
        css = """
        <style>
        </style>
        """
    st.markdown(css, unsafe_allow_html=True)

# Apply theme on every rerun so colors stay consistent
apply_theme()


# ── Plotly dark/light theme wrapper ──────────────────────────────────────────
_orig_plotly_chart = st.plotly_chart

def _themed_plotly_chart(fig, *args, **kwargs):
    if st.session_state.get('theme') == 'dark':
        fig.update_layout(
            paper_bgcolor='#0e1117',
            plot_bgcolor='#0e1117',
            font=dict(color='#fafafa'),
            xaxis=dict(gridcolor='#3a3a4a', linecolor='#4a4a5a', zerolinecolor='#4a4a5a'),
            yaxis=dict(gridcolor='#3a3a4a', linecolor='#4a4a5a', zerolinecolor='#4a4a5a'),
            legend=dict(bgcolor='#262730', bordercolor='#4a4a5a'),
        )
    else:
        fig.update_layout(
            paper_bgcolor='#ffffff',
            plot_bgcolor='#ffffff',
            font=dict(color='#31333F'),
        )
    return _orig_plotly_chart(fig, *args, **kwargs)

st.plotly_chart = _themed_plotly_chart
# ─────────────────────────────────────────────────────────────────────────────

def _strip_nav_badge(label: str) -> str:
    """Remove unread badge suffix added to navigation labels, e.g. 'My Tickets  🔴 2' → 'My Tickets'."""
    for sep in ('  🔴', ' 🔴'):
        if sep in label:
            return label[:label.index(sep)]
    return label

def generate_invoice_pdf(row, lang='en'):
    try:
        from fpdf import FPDF
    except ImportError:
        return None
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # Header
    pdf.set_fill_color(15, 52, 96)
    pdf.rect(0, 0, 210, 32, 'F')
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 20)
    pdf.set_xy(10, 8)
    pdf.cell(130, 10, "EIMS - Logistics Management", ln=False)
    pdf.set_font("Helvetica", "", 10)
    pdf.set_xy(10, 20)
    pdf.cell(130, 6, "Enterprise Information Management System")
    pdf.set_font("Helvetica", "B", 14)
    pdf.set_xy(140, 12)
    pdf.cell(60, 10, "INVOICE", align="R")
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(0, 38)

    # Invoice details
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(95, 7, f"Invoice #: {row.get('invoice_number','—')}", ln=False)
    pdf.cell(95, 7, f"Issue Date: {str(row.get('issue_date','—'))[:10]}", align="R", ln=True)
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(95, 6, f"Client: {row.get('client_name','—')}", ln=False)
    pdf.cell(95, 6, f"Due Date: {str(row.get('due_date','—'))[:10]}", align="R", ln=True)
    pdf.cell(95, 6, f"Shipment Ref: {row.get('shipment_ref','—') or '—'}", ln=False)
    pdf.set_font("Helvetica", "B", 10)
    status = str(row.get('status','—'))
    pdf.cell(95, 6, f"Status: {status}", align="R", ln=True)
    pdf.ln(4)

    # Divider
    pdf.set_draw_color(15, 52, 96)
    pdf.set_line_width(0.5)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(4)

    # Cost breakdown table header
    pdf.set_fill_color(230, 236, 245)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(130, 8, "Description", fill=True, border=1)
    pdf.cell(60, 8, "Amount (USD)", fill=True, border=1, align="R", ln=True)

    cargo_val  = float(row.get("cargo_value", 0) or 0)
    freight    = float(row.get("freight_charge", 0) or 0)
    handling   = float(row.get("handling_fee", 0) or 0)
    insurance  = float(row.get("insurance_fee", 0) or 0)
    tax_amt    = float(row.get("tax_amount", 0) or 0)
    tax_rate   = float(row.get("tax_rate", 0) or 0)
    total      = float(row.get("total", 0) or 0)
    services   = freight + handling + insurance

    rows_data = [
        ("Cargo Value",           cargo_val),
        ("Freight Charge",        freight),
        ("Handling Fee",          handling),
        ("Insurance (0.4%)",      insurance),
        ("Services Subtotal",     services),
        (f"VAT ({tax_rate:.0f}%)", tax_amt),
    ]
    pdf.set_font("Helvetica", "", 10)
    for i, (desc, amt) in enumerate(rows_data):
        fill = i % 2 == 0
        if fill: pdf.set_fill_color(248, 250, 252)
        pdf.cell(130, 7, desc, border=1, fill=fill)
        pdf.cell(60, 7, f"${amt:,.2f}", border=1, align="R", fill=fill, ln=True)

    # Grand total row
    pdf.set_fill_color(15, 52, 96)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(130, 9, "GRAND TOTAL", border=1, fill=True)
    pdf.cell(60, 9, f"${total:,.2f}", border=1, align="R", fill=True, ln=True)
    pdf.set_text_color(0, 0, 0)

    if row.get("description"):
        pdf.ln(4)
        pdf.set_font("Helvetica", "I", 9)
        pdf.multi_cell(0, 5, f"Notes: {row['description']}")

    # Footer
    pdf.ln(6)
    pdf.set_font("Helvetica", "I", 8)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(0, 5, "Thank you for your business. For inquiries, contact Customer Service through the EIMS portal.", align="C")

    return bytes(pdf.output())

def page_matches(page_label: str, page_key: str) -> bool:
    """Check if a page label matches a page key in any language."""
    clean = _strip_nav_badge(page_label)
    return clean in [TRANSLATIONS['en'].get(page_key, ''), TRANSLATIONS['tr'].get(page_key, '')]

def get_page_key(page_label: str) -> str:
    """Get the key for a page label by checking all translations."""
    clean = _strip_nav_badge(page_label)
    for key in TRANSLATIONS['en']:
        if clean in [TRANSLATIONS['en'].get(key, ''), TRANSLATIONS['tr'].get(key, '')]:
            return key
    return ''

# ensure users table exists (safe to call repeatedly)
try:
    db.init_users_table()
except Exception:
    pass


# ensure leave_requests and support_tickets tables exist
try:
    db.init_leave_table()
    db.init_support_tickets_table()
    db.init_cs_tables()
except Exception:
    pass

# Initialize shipment-related tables
try:
    db.init_shipments_table()
    db.init_cargo_items_table()
    db.init_tracking_updates_table()
    db.init_documents_table()
    db.init_cargo_requests_table()
    db.init_messages_table()
except Exception as e:
    print(f"Error initializing shipment tables: {e}")

def _generate_salt():
    return secrets.token_hex(16)

def _hash_password(password: str, salt: str) -> str:
    return hashlib.sha256((password + salt).encode('utf-8')).hexdigest()

# Cache database queries for better performance with longer TTL
@st.cache_data(ttl=300, show_spinner=False, max_entries=20)
def get_cached_records():
    """Cache employee records for 5 minutes"""
    return db.get_all_records()

@st.cache_data(ttl=300, show_spinner=False, max_entries=20)
def get_cached_shipments():
    """Cache shipments for 5 minutes"""
    try:
        return db.get_all_shipments()
    except:
        return pd.DataFrame()

@st.cache_data(ttl=600, show_spinner=False, max_entries=10)
def get_cached_users():
    """Cache users for 10 minutes"""
    try:
        return db.get_all_users()
    except:
        return []

@st.cache_resource(show_spinner=False)
def get_db_connection():
    """Cache database connection"""
    return Database()

@st.cache_data(ttl=30, show_spinner=False, max_entries=50)
def get_cached_unread_messages(user_id):
    try:
        return db.count_unread_messages(user_id)
    except Exception:
        return 0

@st.cache_data(ttl=30, show_spinner=False, max_entries=50)
def get_cached_unread_tickets(role, user_id=None):
    try:
        return sum(db.get_all_unread_counts(role, user_id).values())
    except Exception:
        return 0

@st.cache_resource(show_spinner=False)
def _ensure_manager_once():
    try:
        _admin_email = 'aya@manager.com'
        existing = db.get_user_by_email(_admin_email)
        if existing and existing.get('role') != 'manager':
            db.update_user_role(existing['id'], 'manager')
    except Exception as _e:
        print(f"Manager check error: {_e}")

_ensure_manager_once()

import json as _json

@st.cache_data(ttl=45, show_spinner=False, max_entries=300)
def _qcache(query: str, params_json: str = "{}") -> "pd.DataFrame":
    _p = _json.loads(params_json) if params_json != "{}" else None
    return db.fetch_dataframe(query, _p)

def _q(query: str, params: dict = None) -> "pd.DataFrame":
    return _qcache(query, _json.dumps(params or {}, default=str))

@st.cache_data(ttl=45, show_spinner=False, max_entries=100)
def _get_shipments_by_client(client_id):
    return db.get_shipments_by_client(client_id)

@st.cache_data(ttl=45, show_spinner=False, max_entries=100)
def _get_cargo_by_shipment(shipment_id):
    return db.get_cargo_items_by_shipment(shipment_id)

@st.cache_data(ttl=45, show_spinner=False, max_entries=100)
def _get_cargo_requests_by_client(client_id):
    return db.get_cargo_requests_by_client(client_id)

@st.cache_data(ttl=30, show_spinner=False, max_entries=50)
def _get_user_messages(user_id):
    return db.get_user_messages(user_id)

@st.cache_data(ttl=60, show_spinner=False, max_entries=50)
def _get_invoices(status_filter=None):
    return db.get_invoices(status_filter)

@st.cache_data(ttl=60, show_spinner=False, max_entries=50)
def _get_leave_requests(user_id):
    return db.get_leave_requests_by_user(user_id)

def _safe_rerun():
    """Rerun the Streamlit script."""
    st.rerun()


# Show one-time success notifications stored in session state
try:
    if 'last_shipment_created' in st.session_state:
        info = st.session_state.pop('last_shipment_created', None)
        if info:
            st.success(f"✅ Shipment {info.get('number','')} created successfully — ID: {info.get('id')}")
            st.info(f"Shipment Number: {info.get('number','')}")
except Exception:
    pass

# Show one-time success notification for record updates
try:
    if 'last_record_updated' in st.session_state:
        info = st.session_state.pop('last_record_updated', None)
        if info:
            st.success(f"✅ Record '{info.get('name','')}' (ID: {info.get('id')}) updated successfully")
except Exception:
    pass


PAGES = {
    'login': 'Login',
    'signup': 'Sign Up',
    'dashboard': 'Dashboard',
    'view': 'View Data',
    'add': 'Add Data',
    'edit': 'Edit Data',
    'delete': 'Delete Data',
    'analytics': 'Analytics & Charts',
    'export': 'Export Data',
    'export_employees': 'Export Employees Data',
    'export_shipments': 'Export Shipments Data',
    'profile': 'Profile',
    'request_leave': 'Request Leave',
    'manage_leaves': 'Manage Requests',
    'manage_users': 'Manage employee'
}

st.markdown("""
    <style>
    .stButton>button {
        width: 100%;
    }
    .metric-card {
        background-color: #f0f2f6;
        padding: 20px;
        border-radius: 10px;
        text-align: center;
    }
    .card {
        background: #f8f9fa;
        border: 1px solid #e0e0e0;
        padding: 16px;
        border-radius: 8px;
        box-shadow: 0 2px 6px rgba(16,24,40,0.04);
        margin-bottom: 12px;
        color: #31333F;
    }
    .status-badge { display: inline-block; padding: 4px 10px; border-radius: 12px; font-weight: 600; color: #fff; font-size: 12px; }
    .status-pending { background: #f39c12; }
    .status-approved { background: #28a745; }
    .status-rejected { background: #dc3545; }
    .request-row { padding: 10px 0; border-bottom: 1px solid #f0f0f0; }
    </style>
""", unsafe_allow_html=True)

# Display title
st.title("📊 EIMS")
st.markdown("---")

with st.sidebar:
    # Display logo in sidebar
    logo_png = "assets/logo.png"
    logo_svg = "assets/logo.svg"
    if os.path.exists(logo_png):
        st.image(logo_png, width=80, output_format="PNG")
    elif os.path.exists(logo_svg):
        with open(logo_svg, "r", encoding="utf-8") as f:
            st.markdown(f'''<div style="text-align: center; margin: 10px 0; transform: scale(0.8);">{f.read()}</div>''', unsafe_allow_html=True)
    else:
        st.header("EIMS")

    # عرض عنوان القسم الحالي تحت اللوجو ومحاذاته لليسار
    if 'user' in st.session_state and st.session_state['user']:
        _u = st.session_state['user']
        department = _u.get('department')
        if department and _u.get('role') == 'employee':
            st.markdown(f"<div style='text-align:left; color:#1E88E5; margin-bottom:10px; margin-top:5px; font-size:18px; font-weight:bold;'>{t('department')}: {_tdept(department)}</div>", unsafe_allow_html=True)
    # build options depending on auth/role (exact mapping requested)
    # Generate page labels dynamically based on current language
    guest_pages = [t('login'), t('signup')]
    # employees can manage shipments and cargo
    # Department-specific pages mapping
    department_pages = {
        'Finance': [
            t('finance_dashboard'), t('invoices'), t('payments'), t('expenses'), t('financial_reports'), t('request_leave'), t('profile'), t('messages')
        ],
        'Marketing': [
            t('marketing_dashboard'), t('campaigns'), t('marketing_analytics'), t('social_media'), t('request_leave'), t('profile'), t('messages')
        ],
        'IT': [
            t('it_dashboard'), t('system_management'), t('support_tickets'), t('security'), t('request_leave'), t('profile'), t('messages')
        ],
        'Logistics': [
            t('logistics_dashboard'), t('shipments'), t('routes'), t('delivery_assignments'), t('request_leave'), t('profile'), t('messages')
        ],
        'Customer Service': [
            t('customer_service_dashboard'), t('client_tickets'), t('customer_feedback'), t('request_leave'), t('profile'), t('messages')
        ],
        'Administration': [
            t('administration_dashboard'), t('documents'), t('meetings'), t('contracts'), t('training'), t('request_leave'), t('profile'), t('messages')
        ],
        'Sales': [
            t('sales_dashboard'), t('leads'), t('deals'), t('offers'), t('request_leave'), t('profile'), t('messages')
        ],
    }
    # Default for employees with no department
    default_employee_pages = [t('request_leave'), t('profile'), t('messages')]

    # تحديد صفحات الشريط الجانبي بناءً على الدور أولاً، ثم القسم فقط إذا كان موظف
    page_options = guest_pages
    try:
        if 'user' in st.session_state and st.session_state['user']:
            role = st.session_state['user'].get('role', 'employee')
            user_department = st.session_state['user'].get('department', None)
            if role == 'manager':
                # صفحات المدير فقط (بدون تسجيل الخروج)
                page_options = [
                    t('dashboard'), t('view'), t('add'), t('edit'), t('delete'),
                    t('analytics'), t('export_employees'), t('export_shipments'),
                    t('financial_reports'), t('manage_leaves'), t('manage_users'), t('profile'), t('messages')
                ]
            elif role == 'client':
                page_options = [
                    t('client_dashboard'), t('my_shipments'), t('track_shipment'), t('my_tickets'), t('my_invoices'), t('my_offers'), t('rate_service'), t('profile'), t('messages')
                ]
            elif role == 'employee':
                # الموظف: استخدم القسم
                if user_department and user_department.lower() == 'it':
                    page_options = [t('it_dashboard')]
                    if user_department in department_pages:
                        page_options += [p for p in department_pages[user_department] if p != t('it_dashboard')]
                elif user_department and user_department in department_pages:
                    page_options = department_pages[user_department]
                else:
                    page_options = default_employee_pages
            else:
                page_options = default_employee_pages
    except Exception:
        page_options = guest_pages

    # ── Unread badges on navigation labels ──────────────────────────────────
    try:
        _nav_user = st.session_state.get('user') or {}
        _nav_uid  = _nav_user.get('id')
        _nav_role = _nav_user.get('role')
        _nav_dept = _nav_user.get('department')
        if _nav_uid:
            _unread_msgs = get_cached_unread_messages(_nav_uid)
            _msg_lbl     = t('messages')
            _unread_tkts = 0

            if _nav_role == 'client':
                _unread_tkts = get_cached_unread_tickets('client', _nav_uid)
                _tkt_lbl     = t('my_tickets')
            elif _nav_dept == 'Customer Service':
                _unread_tkts = get_cached_unread_tickets('cs')
                _tkt_lbl     = t('client_tickets')
            else:
                _tkt_lbl = None

            def _badge(base, count):
                return f"{base}  🔴 {count}" if count > 0 else base

            page_options = [
                _badge(p, _unread_tkts) if (_tkt_lbl and p == _tkt_lbl) else
                _badge(p, _unread_msgs) if p == _msg_lbl else p
                for p in page_options
            ]
    except Exception:
        pass
    # ────────────────────────────────────────────────────────────────────────

    # Radio widget key — constant so Streamlit preserves its state normally
    _radio_key = '_page_nav_radio'

    # default to first option unless query param 'page' requests a different one
    default_index = 0
    # Helper: find index of a clean label inside page_options that may have badges
    def _find_page_index(clean_label):
        for i, opt in enumerate(page_options):
            if _strip_nav_badge(opt) == clean_label:
                return i
        return None

    try:
        _cur_lang = st.session_state.get('language', 'en')
        _saved_key = st.session_state.pop('_nav_page_key', None)
        if _saved_key:
            target_label = TRANSLATIONS.get(_cur_lang, TRANSLATIONS['en']).get(_saved_key)
            if target_label:
                idx = _find_page_index(target_label)
                if idx is not None:
                    default_index = idx
                    st.session_state[_radio_key] = page_options[idx]
        else:
            qp = st.query_params
            if qp and 'page' in qp:
                requested = qp['page'][0] if isinstance(qp['page'], (list, tuple)) and qp['page'] else qp['page']
                target_label = TRANSLATIONS.get(_cur_lang, TRANSLATIONS['en']).get(requested)
                if target_label is None:
                    target_label = requested
                if target_label:
                    idx = _find_page_index(target_label)
                    if idx is not None:
                        default_index = idx
    except Exception:
        default_index = 0

    # Check if forgot_password is requested (special page not in sidebar)
    try:
        qp = st.query_params
        if qp and 'page' in qp:
            requested = qp['page'][0] if isinstance(qp['page'], (list, tuple)) and qp['page'] else qp['page']
            if requested == 'forgot_password':
                page = t('forgot_password')
            else:
                page = st.radio(
                    t('select_page'),
                    page_options,
                    index=default_index,
                    label_visibility="collapsed",
                    key=_radio_key
                )
        else:
            page = st.radio(
                t('select_page'),
                page_options,
                index=default_index,
                label_visibility="collapsed",
                key=_radio_key
            )
    except Exception:
        page = st.radio(
            t('select_page'),
            page_options,
            index=default_index,
            label_visibility="collapsed",
            key=_radio_key
        )
    
    # Fix page to match current language if it was from a previous language
    page_key = get_page_key(page)
    if page_key:
        page = t(page_key)

    st.markdown("---")

    # Animated Dark / Light mode toggle
    _is_dark = st.session_state.get('theme', 'light') == 'dark'
    _icon    = "🌙" if _is_dark else "☀️"
    _label   = t('dark_mode') if _is_dark else t('light_mode')
    _icon_pos = "right:3px" if _is_dark else "left:3px"

    st.markdown(f"""
    <style>
    [data-testid="stSidebar"] [data-testid="stToggle"] {{
        position: relative;
    }}
    [data-testid="stSidebar"] [data-testid="stToggle"] label {{
        font-size: 13px !important;
        font-weight: 600 !important;
        gap: 10px !important;
    }}
    /* emoji overlay on the circle */
    [data-testid="stSidebar"] [data-testid="stToggle"]::after {{
        content: '{_icon}';
        position: absolute;
        {_icon_pos};
        top: 50%;
        transform: translateY(-50%);
        font-size: 12px;
        z-index: 99;
        pointer-events: none;
        line-height: 1;
    }}
    /* track color when OFF (light mode) */
    [data-testid="stSidebar"] [data-testid="stToggle"] input:not(:checked) + div {{
        background-color: #dde6f5 !important;
    }}
    </style>
    """, unsafe_allow_html=True)

    tog = st.toggle(_label, value=_is_dark, key="theme_toggle_real")
    if tog != _is_dark:
        st.session_state['theme'] = 'dark' if tog else 'light'
        st.rerun()

    st.markdown("---")

    # Language toggle button
    col_lang1, col_lang2 = st.columns(2)
    with col_lang1:
        _en_active = st.session_state['language'] == 'en'
        if st.button("🇺🇸 English" + (" ✓" if _en_active else ""), width='stretch', disabled=_en_active):
            if page_key:
                st.session_state['_nav_page_key'] = page_key
            st.session_state['language'] = 'en'
            st.rerun()
    with col_lang2:
        _tr_active = st.session_state['language'] == 'tr'
        if st.button("🇹🇷 Türkçe" + (" ✓" if _tr_active else ""), width='stretch', disabled=_tr_active):
            if page_key:
                st.session_state['_nav_page_key'] = page_key
            st.session_state['language'] = 'tr'
            st.rerun()

    st.markdown("---")
    
    # Logout button for logged-in users
    if 'user' in st.session_state and st.session_state['user']:
        if st.button(t('logout'), width='stretch'):
            logged_out_email = st.session_state['user'].get('email', '')
            try:
                db.insert_activity_log("logout", "User logged out", logged_out_email)
            except Exception:
                pass
            st.session_state['user'] = None
            st.query_params["page"] = "login"
            _safe_rerun()
    
    st.markdown("---")
    st.markdown(f"### {t('project_info')}")
    st.info(t('graduation_project'))

### Authentication handling: Login / Sign Up pages and access control ###
if page_matches(page, 'login'):
    st.header(t('login'))
    with st.form("login_form"):
        email = st.text_input(t('email'), placeholder="name@role.com")
        password = st.text_input(t('password'), type="password")
        submitted = st.form_submit_button(t('login_button'))
        if submitted:
            if not email or not password:
                st.error(t('provide_email_pass'))
            else:
                _do_rerun = False
                try:
                    user = db.get_user_by_email(email.strip() if isinstance(email, str) else email)
                    if not user:
                        try:
                            db.add_security_event(email, 'failed_login', description='No account found with that email')
                            db.insert_activity_log("login_failed", "Login attempt failed: email not found", email)
                        except Exception:
                            pass
                        st.error(t('email_not_found'))
                        logger.warning(f"Failed login attempt for non-existent email: {email}")
                    else:
                        hash_val = _hash_password(password, user['salt'])
                        if hash_val == user['password_hash']:
                            user_role = user.get('role', 'employee')
                            if user_role != 'client' and not IS_DESKTOP:
                                st.error("⛔ This portal is for clients only. Employees and managers must use the desktop application.")
                                st.stop()
                            department = None
                            try:
                                db_records = db.fetch_dataframe("SELECT department FROM company_records WHERE email = :email ORDER BY id DESC LIMIT 1", {"email": user['email']})
                                if not db_records.empty:
                                    department = db_records.iloc[0]['department']
                            except Exception:
                                department = None
                            st.session_state['user'] = {'id': user['id'], 'email': user['email'], 'role': user_role, 'department': department}
                            logger.info(f"Successful login: {email} (role: {user_role}, department: {department})")
                            try:
                                db.insert_activity_log("login_success", f"User logged in | role: {user_role} | department: {department or 'N/A'}", email)
                            except Exception:
                                pass
                            if user_role == 'manager':
                                st.query_params["page"] = "dashboard"
                            elif user_role == 'employee':
                                st.query_params["page"] = "manage_shipments"
                            elif user_role == 'client':
                                st.query_params["page"] = "client_dashboard"
                            else:
                                st.query_params["page"] = "request_leave"
                            _do_rerun = True
                        else:
                            st.error(t('incorrect_password'))
                            logger.warning(f"Failed login attempt for {email}: incorrect password")
                            try:
                                db.add_security_event(email, 'failed_login', description='Incorrect password')
                                db.insert_activity_log("login_failed", "Login attempt failed: incorrect password", email)
                            except Exception:
                                pass
                except Exception as e:
                    st.error(t('login_error'))
                    logger.error(f"Login error for {email}: {str(e)}")
                if _do_rerun:
                    _safe_rerun()
    
    # Forgot password link
    st.markdown("<div style='height: 10px;'></div>", unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 1, 1])
    with col2:
        if st.button(t('forgot_pass_btn'), use_container_width=True):
            st.query_params["page"] = "forgot_password"
            _safe_rerun()

elif page_matches(page, 'forgot_password'):
    st.header(t('reset_your_pass'))
    st.markdown(f"### {t('forgot_pass_desc')}")
    
    with st.form("forgot_password_form"):
        email = st.text_input(t('your_email'), placeholder="name@role.com")
        contact_email = st.text_input(t('contact_email'),
                                      placeholder="name@gmail.com")
        st.info(t('reset_info'))
        
        submit = st.form_submit_button(t('submit_request_btn'))
        
        if submit:
            if not email or '@' not in email:
                st.error(t('enter_reg_email'))
            elif not contact_email or '@' not in contact_email:
                st.error(t('enter_contact_email'))
            else:
                try:
                    # Check if user exists
                    user = db.get_user_by_email(email)
                    if not user:
                        st.warning(t('reset_email_sent_maybe'))
                    else:
                        # Get all managers
                        with db.get_connection() as conn:
                            res = conn.execute(text("SELECT id, email FROM users WHERE role='manager'"))
                            managers = [(row[0], row[1]) for row in res.fetchall()]
                        
                        if not managers:
                            st.error(t('no_managers_available'))
                        else:
                            # Send message to all managers with contact email
                            subject = f"🔑 Password Reset Request - {email}"
                            content = f"""New password reset request:

Registered Email: {email}
Contact Email: {contact_email}

Please reset this user's password from the Manage Users page and send the new password to: {contact_email}

This request was automatically sent from the password recovery page."""
                            
                            messages_sent = 0
                            for manager_id, manager_email in managers:
                                if db.send_message(user['id'], manager_id, subject, content, None):
                                    messages_sent += 1
                            
                            if messages_sent > 0:
                                st.success(t('request_submitted_success'))
                                st.info(
                                    f"""Your password reset request has been sent to {messages_sent} manager(s).

The manager will:
1. Reset your password from Manage Users page
2. Send the new password to: **{contact_email}**

⏰ Please check your email ({contact_email}) within 24 hours.

📧 Manager contacts:"""
                                )
                                for _, manager_email in managers:
                                    st.write(f"• {manager_email}")
                                
                                logger.info(f"Password reset request sent for: {email}, contact: {contact_email}")
                            else:
                                st.error(t('error_sending_request'))
                except Exception as e:
                    st.error(t('error_try_again'))
                    logger.error(f"Forgot password error for {email}: {str(e)}")
    
    st.markdown("---")
    if st.button(t('back_to_login')):
        st.query_params["page"] = "login"
        _safe_rerun()

elif page_matches(page, 'signup'):
    # Web portal: skip account type selection, go straight to client form
    if 'signup_type' not in st.session_state and not IS_DESKTOP:
        st.session_state['signup_type'] = 'client'
        _safe_rerun()
    if False:  # disabled: account type selection
        # Show account type selection page
        st.markdown("<h1 style='text-align: center; color: #1E88E5; margin-bottom: 10px;'> Create New Account</h1>", unsafe_allow_html=True)
        st.markdown("<p style='text-align: center; color: #888; font-size: 18px; margin-bottom: 40px;'>Choose the type of account you want to register</p>", unsafe_allow_html=True)
        
        # Add spacing
        st.markdown("<br>", unsafe_allow_html=True)
        
        col_space1, col_main, col_space2 = st.columns([0.8, 2.4, 0.8])
        
        with col_main:
            col_btn1, col_space, col_btn2 = st.columns([1, 0.15, 1])
            
            with col_btn1:
                # Employee button as large card
                employee_clicked = st.button(
                    label="👨‍💼\n\nEmployee Account",
                    key="btn_employee",
                    use_container_width=True,
                    help="Register as an employee with complete access"
                )
                if employee_clicked:
                    st.session_state['signup_type'] = 'employee'
                    _safe_rerun()
                
                # Custom CSS for employee button - square and larger
                st.markdown("""
                    <style>
                    button[kind="secondary"]:has(div:first-child:contains("👨‍💼")) {
                        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%) !important;
                        border: none !important;
                        padding: 80px 50px !important;
                        min-height: 420px !important;
                        aspect-ratio: 1 / 1 !important;
                        border-radius: 25px !important;
                        box-shadow: 0 12px 30px rgba(102, 126, 234, 0.5) !important;
                        transition: all 0.3s ease !important;
                        color: white !important;
                        font-size: 36px !important;
                        font-weight: bold !important;
                        line-height: 1.8 !important;
                    }
                    button[kind="secondary"]:has(div:first-child:contains("👨‍💼")):hover {
                        transform: translateY(-10px) scale(1.02) !important;
                        box-shadow: 0 18px 50px rgba(102, 126, 234, 0.8) !important;
                    }
                    </style>
                """, unsafe_allow_html=True)
            
            with col_btn2:
                # Client button as large card
                client_clicked = st.button(
                    label="👤\n\nClient Account",
                    key="btn_client",
                    use_container_width=True,
                    help="Register as a client with basic access"
                )
                if client_clicked:
                    st.session_state['signup_type'] = 'client'
                    _safe_rerun()
                
                # Custom CSS for client button - square and larger
                st.markdown("""
                    <style>
                    button[kind="secondary"]:has(div:first-child:contains("👤")) {
                        background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%) !important;
                        border: none !important;
                        padding: 80px 50px !important;
                        min-height: 420px !important;
                        aspect-ratio: 1 / 1 !important;
                        border-radius: 25px !important;
                        box-shadow: 0 12px 30px rgba(240, 147, 251, 0.5) !important;
                        transition: all 0.3s ease !important;
                        color: white !important;
                        font-size: 36px !important;
                        font-weight: bold !important;
                        line-height: 1.8 !important;
                    }
                    button[kind="secondary"]:has(div:first-child:contains("👤")):hover {
                        transform: translateY(-10px) scale(1.02) !important;
                        box-shadow: 0 18px 50px rgba(240, 147, 251, 0.8) !important;
                    }
                        background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%) !important;
                        border: none !important;
                        padding: 60px 40px !important;
                        min-height: 350px !important;
                        border-radius: 20px !important;
                        box-shadow: 0 10px 25px rgba(240, 147, 251, 0.5) !important;
                        transition: all 0.3s ease !important;
                        color: white !important;
                        font-size: 32px !important;
                        font-weight: bold !important;
                        line-height: 1.8 !important;
                    }
                    button[kind="secondary"]:has(div:first-child:contains("👤")):hover {
                        transform: translateY(-8px) !important;
                        box-shadow: 0 15px 40px rgba(240, 147, 251, 0.7) !important;
                    }
                    </style>
                """, unsafe_allow_html=True)
        
        st.markdown("<br><br>", unsafe_allow_html=True)
        st.markdown("---")
        
        col_back1, col_back2, col_back3 = st.columns([1, 1, 1])
        with col_back2:
            if st.button(t('back_to_login'), use_container_width=True, key="back_to_login_btn"):
                st.query_params["page"] = "login"
                _safe_rerun()
    
    else:
        # Show registration form based on selected type
        role_choice = st.session_state['signup_type']
        
        # Back button on top left
        if st.button(t('change_type'), key="change_type", help="Change account type"):
            del st.session_state['signup_type']
            _safe_rerun()
        
        # Centered title
        icon = "👨‍💼" if role_choice == "employee" else "👤"
        account_type = "Employee" if role_choice == "employee" else "Client"
        st.markdown(f"<h2 style='text-align: center; color: #1E88E5; margin-top: 20px;'>{icon} {account_type} Registration</h2>", unsafe_allow_html=True)
        
        st.markdown("---")
        
        with st.form("signup_form"):
            st.markdown(f"#### {t('account_info')}")
            col1, col2 = st.columns(2)

            with col1:
                email = st.text_input(t('email_address'), placeholder=f"name@{role_choice}.com", help=f"Must end with @{role_choice}.com")
                password = st.text_input(t('password') + " *", type="password", help="Create a strong password")

            with col2:
                confirm = st.text_input(t('confirm_password') + " *", type="password", help="Re-enter your password")

            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown(f"#### {t('personal_info')}")

            col3, col4 = st.columns(2)

            with col3:
                employee_name = st.text_input(t('full_name_signup'), placeholder="Enter your full name")
                phone = st.text_input(t('phone_number'), placeholder="05xxxxxxxx")

            with col4:
                # Simplified form for clients
                if role_choice == "client":
                    hire_date = st.date_input(t('reg_date'), value=date.today())
                    status = st.selectbox(t('status') + " *", ["Active", "Inactive", "On Leave"], index=0, format_func=_topt)
                    # Set employee-specific fields to None
                    department = None
                    position = None
                    salary = None
                else:
                    # Full form for employees
                    department = st.selectbox(t('dept_field'), ["IT", "Sales", "Marketing", "Finance", "Administration", "Customer Service", "Logistics"])
                    position = st.text_input(t('position_field'), placeholder="e.g., Software Engineer")

            # Show additional employee fields only if not client
            if role_choice == "employee":
                st.markdown("<br>", unsafe_allow_html=True)
                st.markdown(f"#### {t('employment_details')}")
                col5, col6 = st.columns(2)
                with col5:
                    hire_date = st.date_input(t('hire_date_field'), value=date.today())
                with col6:
                    status = st.selectbox(t('emp_status_field'), ["Active", "Inactive", "On Leave"], index=0, format_func=_topt)
                
                # Salary is auto-set to 0 for new employees - manager can edit later
                salary = 0.0
            
            st.markdown("<br>", unsafe_allow_html=True)
            _create_btn_lbl = t('create_emp_account') if role_choice == "employee" else t('create_client_account')
            submitted = st.form_submit_button(f"✅ {_create_btn_lbl}", use_container_width=True, type="primary")

            if submitted:
                if not email or not password:
                    st.error(f"⚠️ {t('provide_email_pass')}")
                elif password != confirm:
                    st.error(f"⚠️ {t('passwords_no_match')}")
                elif role_choice == "employee" and (not employee_name or not department or not position):
                    st.error(f"⚠️ {t('fill_required')}")
                elif role_choice == "client" and not employee_name:
                    st.error(f"⚠️ {t('enter_full_name')}")
                elif '@' not in email or '.' not in email.split('@')[1] if '@' in email else False:
                    st.error(t('invalid_email_format'))
                else:
                    existing = db.get_user_by_email(email)
                    if existing:
                        st.error(t('email_already_exists'))
                    else:
                        _signup_ok = False
                        try:
                            salt = _generate_salt()
                            password_hash = _hash_password(password, salt)
                            ok = db.create_user(email, password_hash, salt, role=role_choice)
                            if ok:
                                db.add_record(
                                    employee_name=employee_name,
                                    department=department,
                                    position=position,
                                    salary=salary,
                                    hire_date=str(hire_date),
                                    email=email,
                                    phone=phone if phone else "",
                                    status=status,
                                    password=password
                                )
                                user = db.get_user_by_email(email)
                                st.session_state['user'] = {'id': user['id'], 'email': user['email'], 'role': user.get('role', 'employee')}
                                if 'signup_type' in st.session_state:
                                    del st.session_state['signup_type']
                                if role_choice == 'client':
                                    st.query_params["page"] = "client_dashboard"
                                else:
                                    st.query_params["page"] = "manage_shipments"
                                _signup_ok = True
                            else:
                                st.error(t('failed_create_account'))
                        except Exception as e:
                            st.error(f"❌ Error: {str(e)}")
                        if _signup_ok:
                            _safe_rerun()

else:
    # For all other pages require login
    if 'user' not in st.session_state or not st.session_state['user']:
        st.warning(t('login_required_msg'))
        if st.button(t('go_login')):
            st.query_params["page"] = "login"
            _safe_rerun()
        st.stop()

    # IT Dashboard: Only for IT employees and only if main page is 'it_dashboard'
    user = st.session_state['user']
    department = user.get('department')
    if isinstance(department, str) and department.lower() == 'it' and page_matches(page, 'it_dashboard'):
        st.header(t('it_dashboard'))
        st.markdown("---")
        # Sidebar navigation for IT features
        it_pages = [
            t('it_dashboard'),
            t('db_management'),
            t('support_tickets'),
            t('security_mgmt'),
            t('performance_monitor'),
            t('bug_tracking'),
            t('activity_logs')
        ]
        selected_it_page = st.sidebar.radio(t('it_pages_label'), it_pages, key="it_page")

        if page_matches(selected_it_page, 'it_dashboard'):
            st.info(t('select_section'))
        elif page_matches(selected_it_page, 'system_management'):
            st.subheader(f"1. {t('system_management')}")
            col1, col2 = st.columns(2)
            with col1:
                st.metric(t('server_status'), t('running'), delta="+0")
                st.write(f"{t('last_check')} 2026-03-26 10:00")
            with col2:
                if st.button(t('restart_server')):
                    st.success(t('server_restarted'))
                if st.button(t('check_status')):
                    st.info(t('status_checked'))

        elif page_matches(selected_it_page, 'db_management'):
            st.subheader(f"2. {t('db_management')}")
            st.write(f"{t('last_backup')}: 2026-03-25 23:00")
            st.write(f"{t('db_size')}: 120 MB")
            colb1, colb2 = st.columns(2)
            with colb1:
                if st.button(t('create_backup')):
                    st.success(t('backup_created'))
            with colb2:
                if st.button(t('restore_backup')):
                    st.warning(t('restore_started'))

        elif page_matches(selected_it_page, 'support_tickets'):
            st.subheader(f"3. {t('support_ticket_sys')}")
            st.write(t('create_ticket'))
            with st.form("ticket_form", clear_on_submit=True):
                ticket_title = st.text_input(t('ticket_title_lbl'))
                ticket_desc = st.text_area(t('ticket_desc_lbl'))
                submit_ticket = st.form_submit_button(t('submit_ticket_btn'))
                if submit_ticket:
                    if not ticket_title or not ticket_desc:
                        st.error(t('fill_all_fields'))
                    else:
                        try:
                            db.get_connection().execute(
                                text("""
                                    INSERT INTO support_tickets (user_id, title, description, status, created_at)
                                    VALUES (:user_id, :title, :desc, 'Open', CURRENT_TIMESTAMP)
                                """),
                                {"user_id": user['id'], "title": ticket_title, "desc": ticket_desc}
                            )
                            st.success(f"Ticket '{ticket_title}' submitted!")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error while adding ticket: {str(e)}")
                            import traceback
                            st.exception(e)

            st.write(t('all_tickets_lbl'))
            try:
                tickets_df = db.fetch_dataframe("SELECT t.id, t.title, t.description, t.status, t.created_at FROM support_tickets t ORDER BY t.created_at DESC")
                if not tickets_df.empty:
                    tickets_df = tickets_df.rename(columns={"title": "Title", "description": "Description", "status": "Status", "created_at": "Created"})
                    st.dataframe(tickets_df[["id", "Title", "Description", "Status", "Created"]].set_index("id"), use_container_width=True)
                else:
                    st.info(t('no_tickets'))
            except Exception as e:
                st.error(f"Error loading tickets: {str(e)}")
                import traceback
                st.exception(e)

        elif page_matches(selected_it_page, 'security_mgmt'):
            st.subheader(f"4. {t('security_mgmt')}")
            st.write(t('user_list'))
            st.table([
                {"Name": "Ali", "Role": "employee", "Status": "Active"},
                {"Name": "Sara", "Role": "manager", "Status": "Inactive"}
            ])
            st.write(t('change_user_role'))

        elif page_matches(selected_it_page, 'performance_monitor'):
            st.subheader(f"5. {t('performance_monitor')}")
            colp1, colp2, colp3 = st.columns(3)
            with colp1:
                st.metric(t('cpu_usage'), "23%", delta="+2%")
            with colp2:
                st.metric(t('mem_usage'), "1.2 GB", delta="-0.1 GB")
            with colp3:
                st.metric(t('active_users'), "8", delta="+1")

        elif page_matches(selected_it_page, 'bug_tracking'):
            st.subheader(f"6. {t('bug_tracking')}")
            st.write(t('log_new_bug'))
            with st.form("bug_form"):
                bug_title = st.text_input(t('bug_title'))
                bug_desc = st.text_area(t('bug_desc'))
                submit_bug = st.form_submit_button(t('log_bug'))
                if submit_bug:
                    st.success(f"{t('bug_logged').replace('!', '')} '{bug_title}'!")
            st.write(t('all_bugs'))
            st.table([
                {"Title": "Export Error", "Status": "Open"},
                {"Title": "UI Crash", "Status": "Fixed"}
            ])

        elif page_matches(selected_it_page, 'activity_logs'):
            st.subheader(f"7. {t('activity_logs')}")
            st.write(t('recent_activities_lbl'))
            st.table([
                {"User": "Ali", "Action": "Login", "Timestamp": "2026-03-26 09:00"},
                {"User": "Sara", "Action": "Data Export", "Timestamp": "2026-03-25 18:00"}
            ])

if page_matches(page, 'dashboard'):
    st.header(t('main_dashboard'))
    
    # Use cached data for statistics
    df = get_cached_records()
    if not df.empty:
        stats = {
            'total_employees': len(df),
            'total_departments': df['department'].nunique(),
            'avg_salary': df['salary'].mean(),
            'active_employees': len(df[df['status'] == 'Active'])
        }
    else:
        stats = {'total_employees': 0, 'total_departments': 0, 'avg_salary': 0, 'active_employees': 0}
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric(label=t('total_employees'), value=stats['total_employees'])
    with col2:
        st.metric(label=t('total_departments'), value=stats['total_departments'])
    with col3:
        st.metric(label=t('avg_salary'), value=f"{stats['avg_salary']:,.0f} $")
    with col4:
        st.metric(label=t('active_employees'), value=stats['active_employees'])

    st.markdown("---")

    st.subheader(t('recent_records'))
    df = get_cached_records()
    if not df.empty:
        df_display = df.drop(columns=['password'], errors='ignore')
        st.dataframe(df_display.head(5), width='stretch')
    else:
        st.info(t('no_data'))

elif page_matches(page, 'view'):
    st.header(t('view_all_data'))

    search_term = st.text_input("🔍 " + t('view_all_data'), placeholder=t('search_placeholder'))
    
    if search_term:
        df = db.search_records(search_term)
        st.info(f"Found {len(df)} results")
    else:
        df = get_cached_records()
    
    if not df.empty:
        col_f1, col_f2, col_f3 = st.columns(3)
        with col_f1:
            departments = [t('all')] + list(df['department'].dropna().unique())
            selected_dept = st.selectbox(t('filter_dept'), departments)

        with col_f2:
            statuses = [t('all')] + list(df['status'].unique())
            selected_status = st.selectbox(t('filter_status'), statuses)

        with col_f3:
            # Add role filter
            if 'role' in df.columns:
                roles = [t('all')] + [r for r in df['role'].unique() if pd.notna(r)]
                selected_role = st.selectbox(t('filter_role'), roles)
            else:
                selected_role = t('all')

        _all_val = t('all')
        if selected_dept != _all_val:
            df = df[df['department'] == selected_dept]

        if selected_status != _all_val:
            df = df[df['status'] == selected_status]

        if selected_role != _all_val and 'role' in df.columns:
            df = df[df['role'] == selected_role]

        # Display different columns based on role filter
        if selected_role == 'client':
            # For clients, show only: id, name, email, phone, status, hire_date, created_at
            client_cols = ['id', 'employee_name', 'email', 'phone', 'status', 'hire_date']
            if 'role' in df.columns:
                client_cols.insert(3, 'role')
            if 'created_at' in df.columns:
                client_cols.append('created_at')
            
            # Filter only existing columns
            display_cols = [col for col in client_cols if col in df.columns]
            df_display = df[display_cols]
        else:
            # For employees/managers, hide password column
            df_display = df.drop(columns=['password'], errors='ignore')
        
        # Use column_config for better performance
        st.dataframe(
            df_display, 
            use_container_width=True, 
            height=400,
            hide_index=True
        )
    else:
        st.warning(t('no_display'))

elif page_matches(page, 'add'):
    st.header(t('add_new_record'))

    # User account settings OUTSIDE form to allow dynamic updates
    st.subheader(t('account_settings_hdr'))
    col_acc1, col_acc2 = st.columns(2)
    with col_acc1:
        create_account = st.checkbox(t('create_login'), value=True)
    with col_acc2:
        if create_account:
            account_role = st.selectbox(t('account_role'), ["employee", "client", "manager"], index=0, format_func=_topt)
            auto_password = st.checkbox(t('auto_gen_pass'), value=True)
            user_password = None  # Initialize
            if not auto_password:
                user_password = st.text_input("Password", type="password", placeholder="Enter password")
        else:
            account_role = "employee"
            auto_password = True
            user_password = None
    
    st.markdown("---")
    
    # Check if adding a client
    is_client = create_account and account_role == "client"
    
    # Now the form with dynamic content
    with st.form("add_form"):
        st.subheader(t('personal_info'))

        if is_client:
            # Simplified form for clients (no department, position, salary)
            col1, col2 = st.columns(2)

            with col1:
                employee_name = st.text_input(t('client_name') + " *", placeholder="Enter full name")
                email = st.text_input(t('email') + " *", placeholder="client@client.com")
                hire_date = st.date_input(t('reg_date'), value=date.today())

            with col2:
                phone = st.text_input(t('phone_field'), placeholder="05xxxxxxxx")
                status = st.selectbox(t('status') + " *", ["Active", "Inactive", "On Leave"], format_func=_topt)

            # Set employee fields to None for clients
            department = None
            position = None
            salary = None
        else:
            # Full form for employees
            col1, col2 = st.columns(2)

            with col1:
                employee_name = st.text_input(t('employee_name'), placeholder="Enter full name")
                department = st.selectbox(t('dept_field'), ["IT", "Sales", "Marketing", "Finance", "Administration", "Customer Service", "Logistics"])
                position = st.text_input(t('position_field'), placeholder="Enter job title")
                salary = st.number_input(t('salary_field'), min_value=0.0, step=100.0, format="%.2f")

            with col2:
                hire_date = st.date_input(t('hire_date_field'), value=date.today())
                email = st.text_input(t('email') + " *", placeholder="example@company.com")
                phone = st.text_input(t('phone_field'), placeholder="05xxxxxxxx")
                status = st.selectbox(t('status') + " *", ["Active", "Inactive", "On Leave"], index=0, format_func=_topt)

        submitted = st.form_submit_button(t('add_record_btn'), width='stretch')
        
        if submitted:
            # Validate based on role
            if is_client:
                # For clients, only name and email are required
                if not employee_name or not email:
                    st.error(f"⚠️ {t('fill_required')} (*)")
                elif not email.endswith("@client.com"):
                    st.error("❌ Email must be in format: name@client.com")
                else:
                    proceed_with_add = True
            else:
                # For employees, all fields are required
                if not employee_name or not department or not position or salary <= 0 or not email:
                    st.error(f"⚠️ {t('fill_required')} (*)")
                elif create_account and not email.endswith(f"@{account_role}.com"):
                    st.error(f"❌ Email must be in format: name@{account_role}.com")
                else:
                    proceed_with_add = True
            
            if 'proceed_with_add' in locals() and proceed_with_add:
                    try:
                        # Add record to company_records (no password field)
                        db.add_record(
                            employee_name=employee_name,
                            department=department,  # None for clients
                            position=position,      # None for clients
                            salary=salary,          # None for clients
                            hire_date=str(hire_date),
                            email=email,
                            phone=phone,
                            status=status,
                            password=''  # Empty - authentication via users table only
                        )
                        db.insert_activity_log("add_record", f"Added record: {employee_name} | Dept: {department or 'N/A'}", st.session_state.get('user',{}).get('email',''))
                        
                        # Create user account if requested
                        if create_account:
                            # Check if user already exists
                            existing_user = db.get_user_by_email(email)
                            if not existing_user:
                                # Generate or use provided password
                                if auto_password:
                                    generated_password = secrets.token_urlsafe(10)
                                    password_to_use = generated_password
                                else:
                                    password_to_use = user_password if user_password else secrets.token_urlsafe(10)
                                
                                # Create user account
                                salt = _generate_salt()
                                password_hash = _hash_password(password_to_use, salt)
                                user_created = db.create_user(email, password_hash, salt, role=account_role)
                                
                                if user_created:
                                    if auto_password:
                                        st.info(f"📋 Login credentials — Email: `{email}` | Password: `{password_to_use}`")
                                    st.success("✅ Record added and user account created successfully!")
                                    _safe_rerun()
                                else:
                                    st.warning("✅ Record added but failed to create user account (email might already exist)")
                            else:
                                st.success("✅ Record added! User account already exists for this email.")
                        else:
                            st.success("✅ Record added successfully!")
                            _safe_rerun()
                    except Exception as e:
                        st.error(f"❌ Error: {str(e)}")
            else:
                st.error(f"⚠️ {t('fill_required')} (*)")

elif page_matches(page, 'edit'):
    st.header(t('edit_existing'))

    df = get_cached_records()

    if not df.empty:
        # Add role filter
        col_filter1, col_filter2 = st.columns([1, 3])
        with col_filter1:
            _re_en = ["All", "manager", "employee", "client"]
            _re_tr = [t('all'), t('manager'), t('employee'), t('client')]
            _re_idx = st.selectbox(t('filter_role'), range(len(_re_tr)), format_func=lambda i: _re_tr[i], index=0)
            role_filter = _re_en[_re_idx]

        if role_filter != "All":
            df_filtered = df[df['role'] == role_filter]
        else:
            df_filtered = df

        if df_filtered.empty:
            st.warning(f"{t('no_records_role')}: {_re_tr[_re_idx]}")
        else:
            if role_filter == 'client':
                record_options = [f"{row['id']} - {row['employee_name']} ({row.get('email', 'N/A')})" for _, row in df_filtered.iterrows()]
            else:
                record_options = [f"{row['id']} - {row['employee_name']} ({row['department']})" for _, row in df_filtered.iterrows()]

            selected_record = st.selectbox(t('select_edit'), record_options)

            if selected_record:
                record_id = int(selected_record.split(' - ')[0])
                record = df_filtered[df_filtered['id'] == record_id].iloc[0]

                record_role = None
                if pd.notna(record.get('role')):
                    record_role = record.get('role')
                else:
                    try:
                        email_for_role = record.get('email')
                        if email_for_role and pd.notna(email_for_role):
                            user_info = db.get_user_by_email(email_for_role)
                            if user_info:
                                record_role = user_info.get('role')
                    except Exception:
                        record_role = None

                is_client = (role_filter == 'client') or (record_role == 'client')

                st.markdown("---")

                if is_client:
                    with st.form("edit_form"):
                        col1, col2 = st.columns(2)
                        with col1:
                            employee_name = st.text_input(t('client_name') + " *", value=record['employee_name'])
                            email = st.text_input(t('email') + " *", value=record['email'] if pd.notna(record['email']) else "")
                            hire_date = st.date_input(t('reg_date'), value=pd.to_datetime(record['hire_date']).date())
                        with col2:
                            phone = st.text_input(t('phone_field'), value=record['phone'] if pd.notna(record['phone']) else "")
                            status = st.selectbox(t('status') + " *", ["Active", "Inactive"], index=["Active", "Inactive"].index(record['status']) if record['status'] in ["Active", "Inactive"] else 0, format_func=_topt)

                        # Set client-specific values
                        department = None
                        position = None
                        salary = None
                        password = None  # Not used - authentication via users table

                        submitted = st.form_submit_button(t('save_client_info'), width='stretch')
                else:
                    # Full form for employees/managers
                    with st.form("edit_form"):
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            employee_name = st.text_input(t('employee_name'), value=record['employee_name'])
                            _dept_list = ["IT", "Sales", "Marketing", "Finance", "Administration", "Customer Service", "Logistics"]
                            department = st.selectbox(t('dept_field'), _dept_list, index=_dept_list.index(record['department']) if record['department'] in _dept_list else 0)
                            position = st.text_input(t('position_field'), value=record['position'])
                            salary = st.number_input(t('salary_field'), value=float(record['salary']) if pd.notna(record.get('salary')) else 0.0, min_value=0.0, step=100.0, format="%.2f")
                        with col2:
                            hire_date = st.date_input(t('hire_date_field'), value=pd.to_datetime(record['hire_date']).date())
                            email = st.text_input(t('email'), value=record['email'] if pd.notna(record['email']) else "")
                            phone = st.text_input(t('phone_field'), value=record['phone'] if pd.notna(record['phone']) else "")
                            status = st.selectbox(t('status') + " *", ["Active", "Inactive", "On Leave"], index=["Active", "Inactive", "On Leave"].index(record['status']) if record['status'] in ["Active", "Inactive", "On Leave"] else 0, format_func=_topt)
                        
                        password = None  # Not used - authentication via users table
                        
                        submitted = st.form_submit_button(t('save_changes'), width='stretch')
                
                if submitted:
                    try:
                                ok = db.update_record(
                                    record_id=record_id,
                                    employee_name=employee_name,
                                    department=department,
                                    position=position,
                                    salary=salary,
                                    hire_date=str(hire_date),
                                    email=email,
                                    phone=phone,
                                    status=status,
                                    password=password
                                )
                                if ok:
                                    db.insert_activity_log("edit_record", f"Updated record: {employee_name} (ID:{record_id})", st.session_state.get('user',{}).get('email',''))
                                    # Immediate feedback before rerun
                                    st.success(f"✅ Record '{employee_name}' updated successfully!")
                                    try:
                                        st.session_state['last_record_updated'] = {'id': record_id, 'name': employee_name}
                                    except Exception:
                                        pass
                                    time.sleep(0.6)
                                    _safe_rerun()
                                else:
                                    st.error("❌ Failed to update record. Please try again.")
                    except Exception as e:
                        st.error(f"❌ Error: {str(e)}")
    else:
        st.warning(t('no_edit'))

elif page_matches(page, 'delete'):
    st.header(t('delete_record_hdr'))

    df = get_cached_records()

    if not df.empty:
        _role_en = ["All", "manager", "employee", "client"]
        _role_tr = [t('all'), t('manager'), t('employee'), t('client')]
        col_filter1, col_filter2 = st.columns([1, 3])
        with col_filter1:
            _rf_idx = st.selectbox(t('filter_role'), range(len(_role_tr)), format_func=lambda i: _role_tr[i], index=0)
            role_filter = _role_en[_rf_idx]

        if role_filter != "All":
            df_filtered = df[df['role'] == role_filter]
        else:
            df_filtered = df

        if df_filtered.empty:
            st.warning(f"{t('no_records_role')}: {_role_tr[_rf_idx]}")
        else:
            if role_filter == 'client':
                record_options = [f"{row['id']} - {row['employee_name']} ({row.get('email', 'N/A')})" for _, row in df_filtered.iterrows()]
            else:
                record_options = [f"{row['id']} - {row['employee_name']} ({row['department']})" for _, row in df_filtered.iterrows()]

            selected_record = st.selectbox(t('select_delete'), record_options)

            if selected_record:
                record_id = int(selected_record.split(' - ')[0])
                record = df_filtered[df_filtered['id'] == record_id].iloc[0]
                is_client = (role_filter == 'client') or (pd.notna(record.get('role')) and record.get('role') == 'client')

                st.markdown("---")
                st.subheader(t('record_info') + ":")

                if is_client:
                    col1, col2 = st.columns(2)
                    with col1:
                        st.write(f"**{t('client_name')}:** {record['employee_name']}")
                        st.write(f"**{t('email')}:** {record.get('email', 'N/A')}")
                        st.write(f"**{t('phone_field')}:** {record.get('phone', 'N/A')}")
                    with col2:
                        st.write(f"**{t('reg_date_display')}:** {record['hire_date']}")
                        st.write(f"**{t('status')}:** {record['status']}")
                        st.write(f"**{t('role')}:** {t('client')}")
                else:
                    col1, col2 = st.columns(2)
                    with col1:
                        st.write(f"**{t('name_field')}:** {record['employee_name']}")
                        st.write(f"**{t('department')}:** {record.get('department', 'N/A')}")
                        st.write(f"**{t('position')}:** {record.get('position', 'N/A')}")
                        st.write(f"**{t('email')}:** {record.get('email', 'N/A')}")
                    with col2:
                        st.write(f"**{t('salary_display')}:** {record.get('salary', 0):,.0f} $")
                        st.write(f"**{t('hire_date_display')}:** {record['hire_date']}")
                        st.write(f"**{t('status')}:** {record['status']}")

                st.markdown("---")
                st.warning(t('warning_undone'))

                col1, col2, col3 = st.columns([1, 1, 2])
                with col1:
                    if st.button(t('confirm_delete_btn'), width='stretch', type="primary"):
                        try:
                            db.delete_record(record_id)
                            db.insert_activity_log("delete_record", f"Deleted record ID:{record_id}", st.session_state.get('user',{}).get('email',''))
                            st.success(f"✅ {t('record_deleted')}")
                            _safe_rerun()
                        except Exception as e:
                            st.error(f"❌ {str(e)}")
                with col2:
                    if st.button(f"❌ {t('cancel')}", width='stretch'):
                        st.info(t('delete_cancelled'))
    else:
        st.warning(t('no_delete'))

elif page_matches(page, 'analytics'):
    df = get_cached_records()
    
    if not df.empty:
        # Filter out clients from analytics (only show employees)
        df_employees = df[df['role'].isin(['employee', 'manager'])] if 'role' in df.columns else df
        
        tab1, tab2, tab3 = st.tabs([t('tab_distribution'), t('tab_salaries'), t('tab_trends')])
        
        with tab1:
            st.subheader(t('emp_distribution'))
            col1, col2 = st.columns(2)
            with col1:
                st.subheader(t('emp_status'))
                fig3 = data_manager.create_status_pie_chart(df_employees)
                st.plotly_chart(fig3, use_container_width=True)
            with col2:
                st.subheader(t('common_positions'))
                fig4 = data_manager.create_position_chart(df_employees)
                st.plotly_chart(fig4, use_container_width=True)

        with tab2:
            st.subheader(t('salary_analysis'))
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                avg_salary = df_employees['salary'].mean()
                st.metric(t('avg_salary_label'), f"${avg_salary:,.0f}")
            with col2:
                max_salary = df_employees['salary'].max()
                st.metric(t('highest_salary'), f"${max_salary:,.0f}")
            with col3:
                min_salary = df_employees['salary'].min()
                st.metric(t('lowest_salary'), f"${min_salary:,.0f}")
            with col4:
                total_payroll = df_employees['salary'].sum()
                st.metric(t('total_payroll'), f"${total_payroll:,.0f}")
            
            st.markdown("---")
            
            # Salary charts
            col1, col2 = st.columns(2)
            with col1:
                st.subheader(t('avg_salary_by_dept'))
                fig_salary = data_manager.create_salary_chart(df_employees)
                st.plotly_chart(fig_salary, use_container_width=True)
            with col2:
                st.subheader(t('salary_dist'))
                fig_dist = px.histogram(
                    df_employees, x='salary', nbins=20,
                    title=t('salary_dist'),
                    labels={'salary': 'Salary ($)', 'count': t('count_label')},
                    color_discrete_sequence=['#2ecc71']
                )
                fig_dist.update_layout(xaxis_title="Salary ($)", yaxis_title=t('count_label'), showlegend=False)
                st.plotly_chart(fig_dist, use_container_width=True)

        with tab3:
            st.subheader(t('emp_trends'))
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric(t('total_employees'), len(df_employees))
            with col2:
                st.metric(t('active_employees'), len(df_employees[df_employees['status'] == 'Active']))
            with col3:
                st.metric(t('total_departments'), df_employees['department'].nunique())

            st.markdown("---")

            if 'created_at' in df_employees.columns:
                st.subheader(t('hiring_timeline'))
                df_employees['created_date'] = pd.to_datetime(df_employees['created_at']).dt.date
                hiring_trend = df_employees.groupby('created_date').size().reset_index(name='count')
                hiring_trend = hiring_trend.sort_values('created_date')
                fig_trend = px.line(
                    hiring_trend, x='created_date', y='count',
                    title=t('emp_added_time'),
                    labels={'created_date': t('date_label'), 'count': t('employees_added')},
                    markers=True
                )
                fig_trend.update_layout(xaxis_title=t('date_label'), yaxis_title=t('employees_added'), showlegend=False)
                st.plotly_chart(fig_trend, use_container_width=True)

            col1, col2 = st.columns(2)
            with col1:
                st.subheader(t('dept_sizes'))
                dept_counts = df_employees['department'].value_counts().reset_index()
                dept_counts.columns = [t('department'), t('count_label')]
                st.dataframe(dept_counts, use_container_width=True, hide_index=True)
            with col2:
                st.subheader(t('status_breakdown'))
                status_counts = df_employees['status'].value_counts().reset_index()
                status_counts.columns = [t('status'), t('count_label')]
                st.dataframe(status_counts, use_container_width=True, hide_index=True)
    else:
        st.warning(t('no_data'))

elif page_matches(page, 'request_leave'):
    st.header(t('request_leave'))
    user = st.session_state.get('user')
    if not user:
        st.error(t('failed_login_submit'))
        st.stop()

    # Request card
    _is_dark_lr = st.session_state.get('theme', 'light') == 'dark'
    _card_bg = "#262730" if _is_dark_lr else "#f0f4f8"
    _card_border = "#3a3a4a" if _is_dark_lr else "#d0d7e0"
    _card_text = "#e6eef8" if _is_dark_lr else "#1e2a3a"
    with st.container():
        st.markdown(f"<div style='background:{_card_bg};border:1px solid {_card_border};padding:16px;border-radius:8px;margin-bottom:16px;'><h3 style='margin:0 0 12px 0;color:{_card_text};'>{t('submit_leave_card')}</h3></div>", unsafe_allow_html=True)
        with st.form("request_leave_form"):
            c1, c2 = st.columns([1, 2])
            with c1:
                start_date = st.date_input(t('start_date'), value=date.today())
                end_date = st.date_input(t('end_date'), value=date.today())
                leave_type = st.selectbox(t('leave_type'), ["Other", "Paid", "Unpaid", "Sick"], index=0, format_func=_topt)
                priority = st.selectbox(t('priority'), ["Low", "Medium", "High", "Urgent"], index=1, format_func=_topt)
            with c2:
                reason = st.text_area(t('reason'), height=130, placeholder="Provide a short reason for your leave...")
                attachment_file = st.file_uploader(t('attachment'))
            submitted = st.form_submit_button(t('submit'), width='stretch')
            if submitted:
                if start_date > end_date:
                    st.error(t('date_range_err'))
                else:
                    try:
                        # handle optional attachment save (desktop only)
                        attachment_name = ''
                        if attachment_file is not None:
                            if not IS_DESKTOP:
                                st.info("📎 File attachments are only available in the desktop application.")
                            else:
                                try:
                                    os.makedirs('uploads', exist_ok=True)
                                    safe_name = f"user{user['id']}_{int(time.time())}_{attachment_file.name}"
                                    save_path = os.path.join('uploads', safe_name)
                                    with open(save_path, 'wb') as out:
                                        out.write(attachment_file.getbuffer())
                                    attachment_name = safe_name
                                except Exception as e:
                                    st.warning(f"Could not save attachment: {e}")

                        db.create_leave_request(user['id'], str(start_date), str(end_date), reason, leave_type, attachment_name, priority)
                        db.insert_activity_log("leave_request", f"Submitted leave: {start_date} → {end_date} | {leave_type} | Priority: {priority}", user.get('email',''))
                        st.success(t('leave_submitted'))
                        _safe_rerun()
                    except Exception as e:
                        st.error(f"Failed to submit request: {str(e)}")
        st.markdown("</div>", unsafe_allow_html=True)

    # My requests
    st.markdown("---")
    st.subheader(t('my_requests'))
    try:
        df = db.get_leave_requests_by_user(user['id'])
        if df.empty:
            st.info(t('no_requests'))
        else:
            for _, r in df.iterrows():
                status = (r.get('status') or 'Pending').lower()
                badge_class = 'status-pending'
                status_display = r.get('status', 'Pending')
                if status == 'approved':
                    badge_class = 'status-approved'
                    status_display = t('approved')
                elif status == 'rejected':
                    badge_class = 'status-rejected'
                    status_display = t('rejected')
                elif status == 'pending':
                    status_display = t('pending')

                st.markdown("<div class='card request-row'>", unsafe_allow_html=True)
                col_left, col_right = st.columns([3,1])
                with col_left:
                    _priority = r.get('priority', 'Medium') or 'Medium'
                    _priority_colors = {'Urgent': '#ef4444', 'High': '#f97316', 'Medium': '#3b82f6', 'Low': '#6b7280'}
                    _pc = _priority_colors.get(_priority, '#3b82f6')
                    st.markdown(f"**{t('from')}:** {r.get('start_date')}  &nbsp;&nbsp; **{t('to')}:** {r.get('end_date')}", unsafe_allow_html=True)
                    st.markdown(f"**{t('type')}:** {r.get('leave_type', 'Other')} &nbsp;&nbsp; <span style='background:{_pc};color:#fff;padding:2px 10px;border-radius:12px;font-size:12px;font-weight:600;'>{t('priority')}: {_priority}</span>", unsafe_allow_html=True)
                    st.markdown(f"**{t('reason')}:** {r.get('reason')}", unsafe_allow_html=True)
                    # show attachment if present
                    attachment = (r.get('attachment') or '').strip()
                    if attachment:
                        filepath = os.path.join('uploads', attachment)
                        if os.path.exists(filepath):
                            try:
                                with open(filepath, 'rb') as f:
                                    st.download_button(label=f"Download attachment: {attachment}", data=f, file_name=attachment)
                            except Exception:
                                st.markdown(f"Attachment: {attachment}")
                        else:
                            st.markdown(f"Attachment: {attachment}")
                    resp = r.get('admin_response')
                    if resp:
                        st.markdown(f"<span style='color: #e6eef8;'>**Manager response:** {resp}</span>", unsafe_allow_html=True)
                with col_right:
                    st.markdown(f"<span class='status-badge {badge_class}'>{status_display}</span>", unsafe_allow_html=True)
                st.markdown("</div>", unsafe_allow_html=True)
    except Exception as e:
        st.error(f"Could not load your requests: {str(e)}")

elif page_matches(page, 'manage_leaves'):
    st.header(t('manage_leaves'))
    user = st.session_state.get('user')
    if not user or user.get('role') != 'manager':
        st.error(t('must_be_manager_leave'))
        st.stop()

    try:
        df = db.get_all_leave_requests()
        if df.empty:
            st.info(t('no_leave_reqs_found'))
        else:
            # ── Filters ──────────────────────────────────────────────
            fc1, fc2, fc3 = st.columns([2, 2, 3])
            with fc1:
                filter_priority = st.selectbox(t('filter_priority_lbl'), [t('all'), 'Urgent', 'High', 'Medium', 'Low'], key="mgr_leave_priority")
            with fc2:
                filter_status = st.selectbox(t('filter_status_leave_lbl'), [t('all'), t('pending'), t('approved'), t('rejected')], key="mgr_leave_status")
            with fc3:
                search_query = st.text_input(t('search_email_reason'), placeholder=t('type_to_search'), key="mgr_leave_search")
            st.markdown("---")

            # Apply filters (map translated labels back to DB English values)
            _status_to_db = {t('pending'): 'Pending', t('approved'): 'Approved', t('rejected'): 'Rejected'}
            filtered_df = df.copy()
            if filter_priority != t('all'):
                filtered_df = filtered_df[filtered_df['priority'].fillna('Medium') == filter_priority]
            if filter_status != t('all'):
                _db_status = _status_to_db.get(filter_status, filter_status)
                filtered_df = filtered_df[filtered_df['status'].fillna('Pending') == _db_status]
            if search_query.strip():
                q = search_query.strip().lower()
                mask = (
                    filtered_df['user_email'].fillna('').str.lower().str.contains(q) |
                    filtered_df['reason'].fillna('').str.lower().str.contains(q)
                )
                filtered_df = filtered_df[mask]

            if filtered_df.empty:
                st.info(t('no_reqs_filters'))
            else:
                st.caption(t('showing_x_of_y').format(x=len(filtered_df), y=len(df)))

            for _, row in filtered_df.iterrows():
                rid = int(row['id'])
                _p = row.get('priority', 'Medium') or 'Medium'
                _pcolors = {'Urgent': '#ef4444', 'High': '#f97316', 'Medium': '#3b82f6', 'Low': '#6b7280'}
                _pc = _pcolors.get(_p, '#3b82f6')
                _priority_badge = f"<span style='background:{_pc};color:#fff;padding:1px 8px;border-radius:10px;font-size:11px;font-weight:600;'>{_p}</span>"
                with st.expander(f"{t('request_num')} {rid} — {row.get('user_email', '')}  [{_p}]"):
                    st.markdown(f"**{t('priority')}:** {_priority_badge}", unsafe_allow_html=True)
                    st.write(f"{t('from')}: {row.get('start_date')}  {t('to')}: {row.get('end_date')}")
                    st.write(f"{t('type')}: {row.get('leave_type', 'Other')}")
                    st.write(f"{t('reason')}: {row.get('reason')}")
                    # show attachment if present
                    att = (row.get('attachment') or '').strip()
                    if att:
                        att_path = os.path.join('uploads', att)
                        if os.path.exists(att_path):
                            try:
                                with open(att_path, 'rb') as af:
                                    st.download_button(label=f"{t('download_attach')}: {att}", data=af, file_name=att)
                            except Exception:
                                st.write(f"{t('attachment_lbl')}: {att}")
                        else:
                            st.write(f"{t('attachment_lbl')}: {att}")
                    st.write(f"{t('status')}: {row.get('status')}")
                    st.write(f"{t('user_lbl')}: {row.get('user_email')}")
                    with st.form(f"manage_{rid}"):
                        status_list = ["Pending", "Approved", "Rejected"]
                        new_status = st.selectbox(t('set_status'), status_list, index=status_list.index(row.get('status', 'Pending')) if row.get('status') in status_list else 0)
                        response = st.text_area(t('response_user'), value=row.get('admin_response', ''))
                        submitted = st.form_submit_button(t('save'))
                        if submitted:
                            try:
                                db.update_leave_request_status(rid, new_status, response)
                                db.insert_activity_log("leave_status_update", f"Leave request #{rid} → {new_status}", st.session_state.get('user',{}).get('email',''))
                                st.success(t('request_updated'))
                                _safe_rerun()
                            except Exception as e:
                                st.error(f"Failed to update: {str(e)}")
    except Exception as e:
        st.error(f"Error loading requests: {str(e)}")

elif page_matches(page, 'manage_users'):
    st.header(t('manage_users'))
    user = st.session_state.get('user')
    if not user or user.get('role') != 'manager':
        st.error(t('must_be_manager_users_err'))
        st.stop()

    try:
        # Get all records from company_records (same as View Data)
        df = get_cached_records()
        
        if df.empty:
            st.info(t('no_users_found_msg'))
        else:
            # Merge role from users table (by email) so Manage Users shows role column
            try:
                users_df = get_cached_users()
                if not isinstance(users_df, pd.DataFrame):
                    try:
                        users_df = pd.DataFrame(users_df)
                    except Exception:
                        users_df = pd.DataFrame(columns=['email', 'role'])
                if 'email' in users_df.columns and 'role' in users_df.columns:
                    merged = df.merge(users_df[['email', 'role']], on='email', how='left')
                else:
                    merged = df.copy()
            except Exception:
                merged = df.copy()

            # Normalize role display
            if 'role' in merged.columns:
                merged['role'] = merged['role'].fillna('N/A')

            # Display the full data table (hide password column)
            df_display = merged.drop(columns=['password'], errors='ignore')
            st.dataframe(df_display, use_container_width=True, height=400)

            
            
            st.markdown("---")
            st.subheader(t('change_user_role'))

            # Get list of users with email
            users_with_email = df[df['email'].notna()]['email'].tolist()

            if not users_with_email:
                st.warning(t('no_users_email_warn'))
            else:
                cols = st.columns(3)
                with cols[0]:
                    selected_user = st.selectbox(t('select_user'), users_with_email)
                with cols[1]:
                    new_role = st.selectbox(t('new_role'), ["employee", "client", "manager"], format_func=_topt)
                with cols[2]:
                    if st.button(t('update_role')):
                        try:
                            # Get user from users table by email
                            existing_user = db.get_user_by_email(selected_user)
                            if existing_user:
                                ok = db.update_user_role(existing_user['id'], new_role)
                                if ok:
                                    db.insert_activity_log("role_update", f"Changed role of {selected_user} → {new_role}", st.session_state.get('user',{}).get('email',''))
                                    st.success(t('user_role_updated'))
                                    _safe_rerun()
                                else:
                                    st.error(t('failed_update_role_db'))
                            else:
                                # Auto-create a login account for this email so role can be set
                                try:
                                    salt = _generate_salt()
                                    generated_password = secrets.token_urlsafe(10)
                                    password_hash = _hash_password(generated_password, salt)
                                    created = db.create_user(selected_user, password_hash, salt, role=new_role)
                                    if created:
                                        st.success(f"User account created — role: `{new_role}`")
                                        st.info(f"📋 Email: `{selected_user}` | Password: `{generated_password}`")
                                        _safe_rerun()
                                    else:
                                        st.error("Failed to create user account. Please check logs.")
                                except Exception as e:
                                    st.error(f"Failed to create account: {str(e)}")
                        except Exception as e:
                            st.error(f"Failed to update role: {str(e)}")
            
            st.markdown("---")
            st.subheader(t('reset_user_pass'))
            st.info(t('reset_pass_info'))

            if users_with_email:
                col1, col2 = st.columns([2, 1])
                with col1:
                    reset_user_email = st.selectbox(t('select_reset_user'), users_with_email, key="reset_password_user")
                with col2:
                    if st.button(t('reset_pass_btn')):
                        try:
                            # Get user from users table
                            existing_user = db.get_user_by_email(reset_user_email)
                            if existing_user:
                                # Generate new random password
                                import secrets
                                import string
                                new_password = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(12))
                                
                                # Create salt and hash
                                salt = secrets.token_hex(16)
                                hashed = hashlib.sha256((new_password + salt).encode()).hexdigest()
                                
                                # Update password in users table
                                db.update_user_password(existing_user['id'], hashed, salt)
                                
                                # Log the password reset
                                logger.info(f"Password reset for user: {reset_user_email} by manager: {user.get('email')}")
                                
                                # Display new password to manager
                                st.success(t('pass_reset_success'))
                                st.markdown(f"### New Password for **{reset_user_email}**")
                                st.code(new_password, language=None)
                                st.warning(t('pass_save_warning'))
                                
                            else:
                                st.error("User not found in users table.")
                        except Exception as e:
                            st.error(f"Failed to reset password: {str(e)}")
                            logger.error(f"Password reset error for {reset_user_email}: {str(e)}")
    except Exception as e:
        st.error(f"Error loading users: {str(e)}")

# Export Employees (Manager only)
elif page_matches(page, 'export_employees'):
    st.header(f"📥 {t('export_employees_hdr')}")
    user = st.session_state.get('user')
    if not user or user.get('role') != 'manager':
        st.error(t('only_managers_export'))
        st.stop()

    try:
        df = get_cached_records()
        if df.empty:
            st.info(t('no_records_export'))
        else:
            df = df.drop(columns=['password'], errors='ignore')

            st.markdown(f"### {t('select_columns')}")
            cols = df.columns.tolist()
            default_cols = cols if len(cols) <= 10 else [c for c in cols if c in ['id','employee_name','email','department','position','status']]
            selected = st.multiselect(t('columns_label'), options=cols, default=default_cols)

            if len(selected) == 0:
                st.warning(t('select_one_col'))
            else:
                df_sel = df[selected]
                with st.expander(t('preview_columns'), expanded=True):
                    try:
                        show_all = st.checkbox(t('show_all_rows'), value=False, key='export_show_all')
                        if show_all:
                            st.dataframe(df_sel, height=500)
                        else:
                            st.dataframe(df_sel.head(5), height=250)
                    except Exception:
                        st.write(df_sel.head(5))

                st.markdown("---")
                st.markdown(f"### {t('export_format')}")
                fmt = st.selectbox(t('format_label'), ["CSV", "Excel (.xlsx)"], index=0)

                # Prepare export copy and format date/phone-like columns so Excel displays them correctly
                df_export = df_sel.copy()
                for col in df_export.columns:
                    lname = col.lower()
                    # Format date/time-looking columns to a consistent string format
                    if 'date' in lname or 'created' in lname or 'hire' in lname or 'time' in lname:
                        try:
                            df_export[col] = pd.to_datetime(df_export[col], errors='coerce').dt.strftime('%Y-%m-%d')
                            # Replace 'NaT'/'None' with empty string
                            df_export[col] = df_export[col].fillna('')
                        except Exception:
                            pass
                    # Force phone and long-numeric columns to string to avoid scientific notation in Excel
                    if 'phone' in lname or 'tel' in lname or lname in ['phone_number', 'phonenumber']:
                        try:
                            df_export[col] = df_export[col].apply(lambda x: '' if pd.isna(x) else str(x))
                        except Exception:
                            df_export[col] = df_export[col].astype(str)

                if fmt == "CSV":
                    # For CSV, wrap date and phone-like fields as Excel formulas of form ="value"
                    df_csv = df_export.copy()
                    for col in df_csv.columns:
                        lname = col.lower()
                        if 'date' in lname or 'created' in lname or 'hire' in lname or 'time' in lname:
                            # values already formatted as YYYY-MM-DD; wrap to force Excel to treat as text
                            df_csv[col] = df_csv[col].apply(lambda v: f'="{v}"' if v not in [None, ''] else '')
                        if 'phone' in lname or 'tel' in lname or lname in ['phone_number', 'phonenumber']:
                            df_csv[col] = df_csv[col].apply(lambda v: f'="{v}"' if v not in [None, ''] else '')
                        # also prevent large integers from being shown in scientific notation
                        if df_csv[col].dtype.kind in 'iuf' and df_csv[col].astype(str).map(len).max() > 10:
                            df_csv[col] = df_csv[col].apply(lambda v: f'="{v}"' if pd.notna(v) else '')

                    csv_bytes = df_csv.to_csv(index=False).encode('utf-8')
                    st.download_button(t('download_csv'), data=csv_bytes, file_name="employees.csv", mime='text/csv')
                else:
                    # Try to write Excel using openpyxl or xlsxwriter; fall back to CSV with guidance
                    from io import BytesIO
                    output = BytesIO()
                    excel_written = False
                    try:
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            df_export.to_excel(writer, index=False, sheet_name='Employees')
                        excel_written = True
                    except ModuleNotFoundError:
                        try:
                            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                                df_export.to_excel(writer, index=False, sheet_name='Employees')
                            excel_written = True
                        except Exception:
                            excel_written = False

                    if excel_written:
                        # Try to adjust column widths so Excel shows dates/numbers properly
                        try:
                            output.seek(0)
                            # If using openpyxl engine, we already wrote to output via ExcelWriter
                            # Re-open with openpyxl to adjust column widths
                            from openpyxl import load_workbook
                            from openpyxl.utils import get_column_letter
                            wb = load_workbook(filename=BytesIO(output.getvalue()))
                            ws = wb.active
                            for idx, col in enumerate(df_export.columns, start=1):
                                max_len = max(df_export[col].astype(str).map(len).max(), len(str(col)))
                                ws.column_dimensions[get_column_letter(idx)].width = max_len + 3
                            out2 = BytesIO()
                            wb.save(out2)
                            data = out2.getvalue()
                        except Exception:
                            # Fallback: return the original bytes
                            data = output.getvalue()
                        st.download_button(t('download_excel'), data=data, file_name="employees.xlsx", mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    else:
                        st.error('Excel export requires `openpyxl` or `xlsxwriter`. Install it in your environment: `pip install openpyxl`')
                        csv_bytes = df_export.to_csv(index=False).encode('utf-8')
                        st.download_button(t('download_csv_fallback'), data=csv_bytes, file_name="employees.csv", mime='text/csv')
    except Exception as e:
        st.error(f"Error exporting employees: {str(e)}")

# Export Shipments (Employee / Manager)
elif page_matches(page, 'export_shipments'):
    st.header(f"📥 {t('export_shipments_hdr')}")
    user = st.session_state.get('user')
    if not user or user.get('role') not in ['employee', 'manager']:
        st.error(t('only_emp_mgr_export'))
        st.stop()

    try:
        df = get_cached_shipments()
        if df.empty:
            st.info(t('no_shipments_export'))
        else:
            df = df.drop(columns=['internal_notes'], errors='ignore')

            st.markdown(f"### {t('select_columns')}")
            cols = df.columns.tolist()
            default_cols = cols if len(cols) <= 10 else [c for c in cols if c in ['id','number','origin','destination','status','created_at']]
            selected = st.multiselect(t('columns_label'), options=cols, default=default_cols)

            if len(selected) == 0:
                st.warning(t('select_one_col'))
            else:
                df_sel = df[selected]
                with st.expander(t('preview_columns'), expanded=True):
                    try:
                        show_all = st.checkbox(t('show_all_rows'), value=False, key='export_shipments_show_all')
                        if show_all:
                            st.dataframe(df_sel, height=500)
                        else:
                            st.dataframe(df_sel.head(5), height=250)
                    except Exception:
                        st.write(df_sel.head(5))

                st.markdown("---")
                st.markdown(f"### {t('export_format')}")
                fmt = st.selectbox(t('format_label'), ["CSV", "Excel (.xlsx)"], index=0)

                # Prepare export copy and format date/phone-like columns so Excel displays them correctly
                df_export = df_sel.copy()
                for col in df_export.columns:
                    lname = col.lower()
                    if 'date' in lname or 'created' in lname or 'depart' in lname or 'arrival' in lname or 'time' in lname:
                        try:
                            df_export[col] = pd.to_datetime(df_export[col], errors='coerce').dt.strftime('%Y-%m-%d')
                            df_export[col] = df_export[col].fillna('')
                        except Exception:
                            pass
                    if 'phone' in lname or 'tel' in lname:
                        try:
                            df_export[col] = df_export[col].apply(lambda x: '' if pd.isna(x) else str(x))
                        except Exception:
                            df_export[col] = df_export[col].astype(str)

                if fmt == "CSV":
                    df_csv = df_export.copy()
                    for col in df_csv.columns:
                        lname = col.lower()
                        if 'date' in lname or 'created' in lname or 'depart' in lname or 'arrival' in lname or 'time' in lname:
                            df_csv[col] = df_csv[col].apply(lambda v: f'="{v}"' if v not in [None, ''] else '')
                        if 'phone' in lname or 'tel' in lname:
                            df_csv[col] = df_csv[col].apply(lambda v: f'="{v}"' if v not in [None, ''] else '')
                        if df_csv[col].dtype.kind in 'iuf' and df_csv[col].astype(str).map(len).max() > 10:
                            df_csv[col] = df_csv[col].apply(lambda v: f'="{v}"' if pd.notna(v) else '')

                    csv_bytes = df_csv.to_csv(index=False).encode('utf-8')
                    st.download_button(t('download_csv'), data=csv_bytes, file_name="shipments.csv", mime='text/csv')
                else:
                    from io import BytesIO
                    output = BytesIO()
                    excel_written = False
                    try:
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            df_export.to_excel(writer, index=False, sheet_name='Shipments')
                        excel_written = True
                    except ModuleNotFoundError:
                        try:
                            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                                df_export.to_excel(writer, index=False, sheet_name='Shipments')
                            excel_written = True
                        except Exception:
                            excel_written = False

                    if excel_written:
                        try:
                            from openpyxl import load_workbook
                            from openpyxl.utils import get_column_letter
                            out_bytes = BytesIO(output.getvalue())
                            wb = load_workbook(filename=out_bytes)
                            ws = wb.active
                            for idx, col in enumerate(df_export.columns, start=1):
                                max_len = max(df_export[col].astype(str).map(len).max(), len(str(col)))
                                ws.column_dimensions[get_column_letter(idx)].width = max_len + 3
                            out2 = BytesIO()
                            wb.save(out2)
                            data = out2.getvalue()
                        except Exception:
                            data = output.getvalue()
                        st.download_button(t('download_excel'), data=data, file_name="shipments.xlsx", mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    else:
                        st.error('Excel export requires `openpyxl` or `xlsxwriter`. Install it in your environment: `pip install openpyxl`')
                        csv_bytes = df_export.to_csv(index=False).encode('utf-8')
                        st.download_button("⬇️ Download CSV (fallback)", data=csv_bytes, file_name="shipments.csv", mime='text/csv')
    except Exception as e:
        st.error(f"Error exporting shipments: {str(e)}")

# Shipment Management Pages
elif page_matches(page, 'manage_shipments'):
    st.header(t('manage_shipments'))
    user = st.session_state.get('user')
    if not user or user.get('role') not in ['manager', 'employee']:
        st.error(t('only_mgr_emp_ships'))
        st.stop()

    try:
        # Fetch data from database with caching
        df = get_cached_shipments()
        
        if df.empty:
            st.info(t('no_shipments_start'))
        else:
            # Filter options
            col1, col2, col3 = st.columns(3)
            with col1:
                shipment_types = ['All'] + list(df['type'].unique())
                selected_type = st.selectbox(t('filter_by_type_lbl'), shipment_types)
            with col2:
                statuses = ['All'] + list(df['status'].unique())
                selected_status = st.selectbox(t('filter_by_status_lbl2'), statuses)
            with col3:
                search_term = st.text_input(t('search_lbl2'), placeholder=t('search_ship_ph'))
            
            # Apply filters
            filtered_df = df.copy()
            if selected_type != 'All':
                filtered_df = filtered_df[filtered_df['type'] == selected_type]
            if selected_status != 'All':
                filtered_df = filtered_df[filtered_df['status'] == selected_status]
            if search_term:
                filtered_df = filtered_df[
                    filtered_df['shipment_number'].str.contains(search_term, case=False, na=False) |
                    filtered_df['client_email'].str.contains(search_term, case=False, na=False)
                ]
            
            st.dataframe(filtered_df, width='stretch')
            
            # Shipment details and management
            st.markdown("---")
            st.subheader(t('ship_details_mgmt'))
            
            if not filtered_df.empty:
                shipment_nums = filtered_df['shipment_number'].tolist()
                selected_shipment = st.selectbox(t('select_shipment_lbl'), shipment_nums)
                
                # Get fresh data from database for selected shipment
                fresh_df = db.get_all_shipments()
                ship_data = fresh_df[fresh_df['shipment_number'] == selected_shipment].iloc[0]
                
                tab1, tab2, tab3, tab4 = st.tabs([t('tab_details'), f"📦 {t('tab_cargo_items')}", f"🗺️ {t('tab_tracking')}", f"📄 {t('tab_documents')}"])
                
                with tab1:
                    col_a, col_b = st.columns(2)
                    with col_a:
                        st.write(f"**{t('shipment_number')}:** {ship_data['shipment_number']}")
                        st.write(f"**Type:** {ship_data['type']}")
                        st.write(f"**{t('origin')}:** {ship_data['origin_country']}")
                        st.write(f"**{t('destination')}:** {ship_data['destination_country']}")
                        st.write(f"**Client:** {ship_data['client_email']}")
                    with col_b:
                        st.write(f"**{t('departure_date')}:** {ship_data['departure_date']}")
                        st.write(f"**{t('expected_arrival')}:** {ship_data['expected_arrival']}")
                        st.write(f"**{t('actual_arrival')}:** {ship_data.get('actual_arrival', 'N/A')}")
                        st.write(f"**{t('total_weight')}:** {ship_data['total_weight']} kg")
                        st.write(f"**{t('total_value')}:** {ship_data['currency']} {ship_data['total_value']:,.2f}")
                    
                    st.markdown("---")
                    st.subheader(t('update_status_hdr'))
                    with st.form("update_status_form"):
                        status_options = ["Pending", "In Transit", "Customs", "Delivered", "Cancelled"]
                        new_status = st.selectbox(t('shipment_status'), status_options, 
                                                index=status_options.index(ship_data['status']) if ship_data['status'] in status_options else 0)
                        actual_arrival_date = st.date_input(t('actual_arrival'), value=None)
                        customs_check = st.checkbox(t('customs_cleared'), value=bool(ship_data.get('customs_cleared', 0)))

                        if st.form_submit_button("Update Status"):
                            try:
                                # Update shipment status
                                db.update_shipment_status(int(ship_data['id']), new_status, 
                                                        str(actual_arrival_date) if actual_arrival_date else None)
                                # Update customs status
                                db.update_customs_status(int(ship_data['id']), customs_check)

                                st.success(t('status_updated_success'))
                                # Clear cached shipments and rerun so UI shows fresh data
                                _safe_rerun()
                            except Exception as e:
                                st.error(f"❌ Error: {str(e)}")
                
                with tab2:
                    st.subheader(t('cargo_items'))
                    cargo_df = db.get_cargo_items_by_shipment(int(ship_data['id']))
                    if not cargo_df.empty:
                        st.dataframe(cargo_df, width='stretch')
                        
                        st.markdown("---")
                        st.markdown("###  Manage Cargo Items")
                        # Edit and Delete cargo items
                        col_edit, col_delete = st.columns(2)
                        
                        with col_edit:
                            st.markdown("####  Edit Cargo Item")
                            item_to_edit = st.selectbox(t('select_item_edit'), 
                                [(f"{row['id']} - {row['item_name']}", row['id']) for _, row in cargo_df.iterrows()],
                                format_func=lambda x: x[0], key="edit_cargo_select")
                            
                            if item_to_edit:
                                selected_id = item_to_edit[1]
                                selected_item = cargo_df[cargo_df['id'] == selected_id].iloc[0]
                                
                                with st.form("edit_cargo_form"):
                                    edit_item_name = st.text_input(t('item_name_lbl'), value=selected_item['item_name'])
                                    col_e1, col_e2 = st.columns(2)
                                    with col_e1:
                                        edit_quantity = st.number_input(t('quantity_lbl'), min_value=1, value=int(selected_item['quantity']))
                                        edit_weight = st.number_input(t('weight_kg_lbl'), min_value=0.0, value=float(selected_item['weight']))
                                    with col_e2:
                                        edit_unit = st.selectbox("Unit", ["pcs", "kg", "ton", "box", "container"],
                                            index=["pcs", "kg", "ton", "box", "container"].index(selected_item['unit']) 
                                            if selected_item['unit'] in ["pcs", "kg", "ton", "box", "container"] else 0)
                                        edit_value = st.number_input(t('value_lbl'), min_value=0.0, value=float(selected_item['value']))
                                    edit_description = st.text_area("Description", value=selected_item.get('description', ''))
                                    edit_hs_code = st.text_input(t('hs_code_lbl'), value=selected_item.get('hs_code', ''))
                                    
                                    if st.form_submit_button("💾 Save Changes", width='stretch'):
                                        try:
                                            db.update_cargo_item(
                                                item_id=selected_id,
                                                item_name=edit_item_name,
                                                quantity=edit_quantity,
                                                weight=edit_weight,
                                                unit=edit_unit,
                                                value=edit_value,
                                                description=edit_description,
                                                hs_code=edit_hs_code
                                            )
                                            st.success(t('cargo_item_updated'))
                                            _safe_rerun()
                                        except Exception as e:
                                            st.error(f"❌ Error: {str(e)}")
                        
                        with col_delete:
                            st.markdown("#### 🗑️ Delete Cargo Item")
                            item_to_delete = st.selectbox("Select item to delete:", 
                                [(f"{row['id']} - {row['item_name']}", row['id']) for _, row in cargo_df.iterrows()],
                                format_func=lambda x: x[0], key="delete_cargo_select")
                            
                            if item_to_delete:
                                delete_id = item_to_delete[1]
                                delete_item = cargo_df[cargo_df['id'] == delete_id].iloc[0]
                                
                                st.markdown("**Item Details:**")
                                st.write(f" **Item:** {delete_item['item_name']}")
                                st.write(f" **Quantity:** {delete_item['quantity']} {delete_item['unit']}")
                                st.write(f" **Weight:** {delete_item['weight']} kg")
                                st.write(f" **Value:** ${delete_item['value']:,.2f}")
                                
                                st.warning(t('action_undone_warn'))
                                
                                if st.button("🗑️ Confirm Delete", type="primary", width='stretch'):
                                    try:
                                        db.delete_cargo_item(delete_id)
                                        st.success(t('item_deleted_success'))
                                        _safe_rerun()
                                    except Exception as e:
                                        st.error(f"❌ Error: {str(e)}")
                    else:
                        st.info(t('no_cargo_items'))
                    
                    st.markdown("---")
                    st.subheader(f"➕ {t('add_cargo')}")
                    with st.form("add_cargo_form"):
                        c1, c2 = st.columns(2)
                        with c1:
                            item_name = st.text_input(t('item_name') + " *")
                            quantity = st.number_input(t('quantity') + " *", min_value=1, value=1)
                            weight = st.number_input(t('weight') + " *", min_value=0.0, value=0.0)
                        with c2:
                            unit = st.selectbox(t('unit_lbl') + " *", ["pcs", "kg", "ton", "box", "container"])
                            value = st.number_input(t('value') + " *", min_value=0.0, value=0.0)
                            description = st.text_area(t('description'))
                        hs_code = st.text_input(t('hs_code_optional'))
                        
                        if st.form_submit_button(f"➕ {t('add_cargo')}", width='stretch'):
                            if item_name:
                                try:
                                    db.add_cargo_item(
                                        int(ship_data['id']), 
                                        item_name, 
                                        description, 
                                        int(quantity), 
                                        unit, 
                                        float(weight), 
                                        float(value), 
                                        hs_code
                                    )
                                    st.success(t('cargo_item_added'))
                                    _safe_rerun()
                                except Exception as e:
                                    st.error(f"❌ Error: {str(e)}")
                            else:
                                st.error("⚠️ Please enter item name")
                
                with tab3:
                    st.subheader(t('tracking'))
                    tracking_df = db.get_tracking_updates(int(ship_data['id']))
                    if not tracking_df.empty:
                        for _, row in tracking_df.iterrows():
                            status_class = row['status'].lower().replace(' ', '-')
                            notes_html = f"<div style='color: #b8bcc4; margin-top: 5px;'>📝 {row['notes']}</div>" if row.get('notes') else ''
                            updated_by = row.get('updated_by_email', 'System')
                            
                            st.markdown(f"""
                            <div style='background: #262730; border: 1px solid #262730; padding: 15px; border-radius: 8px; margin-bottom: 10px;'>
                                <div style='color: #e6eef8;'><strong>📅 {row['update_date']}</strong> - 📍 {row['location']}</div>
                                <div style='margin: 8px 0;'><span class='status-badge status-{status_class}'>{row['status']}</span></div>
                                {notes_html}
                                👤 Updated by: {updated_by}
                            </div>
                            """, unsafe_allow_html=True)
                    else:
                        st.info(t('no_tracking_yet'))
                    
                    st.markdown("---")
                    st.subheader(t('update_tracking'))
                    with st.form("add_tracking_form"):
                        location = st.text_input(t('location') + " *", placeholder="e.g., Port of Jeddah")
                        track_status = st.selectbox(t('status') + " *", ["Departed", "In Transit", "Arrived at Port", "Customs", "Out for Delivery", "Delivered"], format_func=_topt)
                        track_date = st.date_input("Update Date *", value=date.today())
                        track_notes = st.text_area("Notes", placeholder="Optional notes about this update...")
                        
                        if st.form_submit_button("📍 Update", width='stretch'):
                            if location:
                                try:
                                    db.add_tracking_update(int(ship_data['id']), location, track_status, 
                                                          track_notes, str(track_date), user['id'])
                                    st.success(t('tracking_added'))
                                    _safe_rerun()
                                except Exception as e:
                                    st.error(f"❌ Error: {str(e)}")
                            else:
                                st.error(t('please_enter_location'))
                
                with tab4:
                    st.subheader(t('documents'))
                    docs_df = db.get_shipment_documents(int(ship_data['id']))
                    if not docs_df.empty:
                        for _, doc in docs_df.iterrows():
                            st.markdown(f"""
                            <div class='card'>
                                <strong>{doc['document_type']}</strong><br>
                                File: {doc['file_path']}<br>
                                Uploaded by: {doc.get('uploaded_by_email', 'Unknown')}<br>
                                Date: {doc['created_at']}<br>
                                {f"Notes: {doc['notes']}" if doc['notes'] else ''}
                            </div>
                            """, unsafe_allow_html=True)
                    else:
                        st.info(t('no_documents_yet'))
                    
                    st.markdown("---")
                    st.subheader(t('upload_document'))
                    with st.form("upload_doc_form"):
                        doc_type = st.selectbox(t('doc_type_lbl'),
                            ["Invoice", "Bill of Lading", "Customs Declaration", "Certificate of Origin", "Packing List", "Other"], format_func=_topt)
                        uploaded_file = st.file_uploader(t('choose_file_lbl'))
                        doc_notes = st.text_area("Notes")
                        
                        if st.form_submit_button("Upload"):
                            if uploaded_file:
                                if not IS_DESKTOP:
                                    st.info("📎 Document upload is only available in the desktop application.")
                                else:
                                    try:
                                        os.makedirs("shipment_documents", exist_ok=True)
                                        file_path = os.path.join("shipment_documents",
                                                                f"{ship_data['shipment_number']}_{uploaded_file.name}")
                                        with open(file_path, "wb") as f:
                                            f.write(uploaded_file.getbuffer())
                                        db.add_shipment_document(int(ship_data['id']), doc_type, file_path, user['id'], doc_notes)
                                        st.success(t('document_uploaded'))
                                        _safe_rerun()
                                    except Exception as e:
                                        st.error(f"Error: {str(e)}")
                            else:
                                st.error(t('please_select_file'))
    except Exception as e:
        st.error(f"Error loading shipments: {str(e)}")

elif page_matches(page, 'add_shipment'):
    st.header(t('add_shipment'))
    user = st.session_state.get('user')
    if not user or user.get('role') not in ['manager', 'employee']:
        st.error("You must be a manager or employee to add shipments.")
        st.stop()

    with st.form("add_shipment_form"):
        col1, col2 = st.columns(2)
        
        with col1:
            shipment_number = st.text_input(t('shipment_number') + " *", placeholder="SH-2025-001")
            shipment_type = st.selectbox(t('type_field') + " *", ["Import", "Export"], format_func=_topt)
            
            # Country dropdown list (alphabetically sorted)
            countries = [
                "Algeria", "Argentina", "Australia", "Austria", "Bahrain", "Belgium", "Brazil",
                "Canada", "China", "Denmark", "Egypt", "France", "Germany", "India", "Iraq",
                "Italy", "Japan", "Jordan", "Kuwait", "Lebanon", "Libya", "Mexico", "Morocco",
                "Netherlands", "Norway", "Oman", "Palestine", "Poland", "Qatar", "Russia",
                "Saudi Arabia", "South Korea", "Spain", "Sudan", "Sweden", "Switzerland", "Syria",
                "Tunisia", "Turkey", "UAE", "United Kingdom", "United States", "Yemen"
            ]
            
            origin_country = st.selectbox(t('origin') + " *", countries)
            destination_country = st.selectbox(t('destination') + " *", countries)
            departure_date = st.date_input(t('departure_date') + " *", value=date.today())
        
        with col2:
            # Get all clients for selection
            try:
                users_df = db.get_all_users()
                clients = users_df[users_df['role'] == 'client']
                if clients.empty:
                    st.warning(t('no_clients_warning'))
                    client_options = []
                else:
                    client_options = clients['email'].tolist()
            except Exception as e:
                st.error(f"Error loading users: {str(e)}")
                client_options = []
            
            if client_options:
                client_email = st.selectbox(t('client') + " *", client_options)
            else:
                st.error(t('no_clients_add_ship'))
                client_email = None

            expected_arrival = st.date_input(t('expected_arrival') + " *")
            total_weight = st.number_input(t('total_weight') + " *", min_value=0.0, value=0.0)
            total_value = st.number_input(t('total_value') + " *", min_value=0.0, value=0.0)
            currency = st.selectbox(t('currency_lbl'), ["USD", "EUR", "GBP", "TRY"])
        
        notes = st.text_area("Notes")
        
        submitted = st.form_submit_button("➕ Add Shipment", width='stretch')
        
        if submitted:
            if not client_options:
                st.error(t('no_clients_add_ship'))
            elif shipment_number and origin_country and destination_country and client_email:
                try:
                    # Get client ID
                    client_data = users_df[users_df['email'] == client_email].iloc[0]
                    client_id = int(client_data['id'])
                    
                    st.write(f"Debug: Creating shipment for client_id={client_id}")  # Debug
                    
                    shipment_id = db.create_shipment(
                        shipment_number=shipment_number,
                        client_id=client_id,
                        shipment_type=shipment_type,
                        origin_country=origin_country,
                        destination_country=destination_country,
                        departure_date=str(departure_date),
                        expected_arrival=str(expected_arrival),
                        total_weight=total_weight,
                        total_value=total_value,
                        currency=currency,
                        notes=notes
                    )
                    
                    st.write(f"Debug: Shipment ID returned: {shipment_id}")  # Debug
                    
                    if shipment_id:
                        # store for one-time display after rerun
                        try:
                            st.session_state['last_shipment_created'] = {'id': shipment_id, 'number': shipment_number.strip()}
                        except Exception:
                            pass
                        _safe_rerun()
                    else:
                        st.error("❌ Failed to create shipment. The shipment number might already exist.")
                except Exception as e:
                    st.error(f"❌ Error creating shipment: {str(e)}")
                    import traceback
                    st.code(traceback.format_exc())
            else:
                st.error(f"⚠️ {t('fill_required')} (*)")

elif page_matches(page, 'client_dashboard'):
    user = st.session_state.get('user') or {}
    _cid   = user.get('id')
    _cemail= user.get('email','')
    _cname_row = _q("SELECT full_name FROM users WHERE id=:uid", {"uid": _cid})
    _cname = str(_cname_row.iloc[0]["full_name"]) if (_cname_row is not None and not _cname_row.empty and _cname_row.iloc[0]["full_name"]) else _cemail

    st.markdown(f"## 👋 {t('welcome_back')}, **{_cname}**")
    st.markdown("---")

    # ── Metric cards ──────────────────────────────────────────────────────────
    _shp_df  = _q("SELECT status FROM shipments WHERE client_id=:ci", {"ci": _cid})
    _tkt_df  = _q("SELECT status FROM cs_tickets WHERE client_id=:ci", {"ci": _cid})
    _inv_df  = _q("SELECT status FROM finance_invoices WHERE client_email=:ce", {"ce": _cemail})
    _off_df  = _q("SELECT status FROM sales_offers WHERE client_email=:ce AND status='Sent'", {"ce": _cemail})
    _unread_msgs = get_cached_unread_messages(_cid) if _cid else 0
    _unread_tkts = get_cached_unread_tickets('client', _cid) if _cid else 0

    _active_shp  = int((_shp_df['status'].isin(['Pending','In Transit','At Customs'])).sum()) if not _shp_df.empty else 0
    _open_tkt    = int((_tkt_df['status'].isin(['Open','In Progress'])).sum()) if not _tkt_df.empty else 0
    _unpaid_inv  = int((_inv_df['status'] == 'Pending').sum()) if not _inv_df.empty else 0
    _new_offers  = len(_off_df) if not _off_df.empty else 0

    mc1, mc2, mc3, mc4, mc5 = st.columns(5)
    mc1.metric(t('active_shipments'),  _active_shp)
    mc2.metric(t('open_tickets'),      f"{_open_tkt}" + (f" 🔴" if _unread_tkts else ""))
    mc3.metric(t('unpaid_invoices'),   _unpaid_inv)
    mc4.metric(t('new_offers'),        _new_offers)
    mc5.metric(t('unread_messages'),   f"{_unread_msgs}" + (" 🔴" if _unread_msgs else ""))

    st.markdown("---")

    STATUS_COLOR = {"Delivered":"🟢","In Transit":"🟡","Pending":"🔵","At Customs":"🟠","Cancelled":"🔴"}
    TKT_COLOR    = {"Open":"🔴","In Progress":"🟡","Resolved":"🟢","Closed":"⚫"}
    OFF_COLOR    = {"Sent":"🟡","Accepted":"🟢","Rejected":"🔴","Draft":"⚫"}

    # ── Recent Shipments ──────────────────────────────────────────────────────
    st.subheader(f"🚢 {t('recent_shipments')}")
    _rshp = _q(
        "SELECT shipment_number, status, origin, destination, expected_arrival "
        "FROM shipments WHERE client_id=:ci ORDER BY id DESC LIMIT 5", {"ci": _cid}
    )
    if _rshp is not None and not _rshp.empty:
        for _, r in _rshp.iterrows():
            ic  = STATUS_COLOR.get(str(r["status"]), "⚪")
            eta = str(r.get("expected_arrival",""))[:10] if r.get("expected_arrival") else "—"
            sa, sb, sc, sd = st.columns([2, 2, 3, 2])
            sa.markdown(f"**{r['shipment_number']}**")
            sb.markdown(f"{ic} {r['status']}")
            sc.markdown(f"📍 {r.get('origin','—')} → {r.get('destination','—')}")
            sd.markdown(f"📅 {eta}")
    else:
        st.caption(t('no_recent_activity'))

    st.markdown("---")

    # ── Recent Tickets ────────────────────────────────────────────────────────
    st.subheader(f"🎫 {t('recent_tickets')}")
    _rtkt = _q(
        "SELECT ticket_number, subject, status, priority FROM cs_tickets "
        "WHERE client_id=:ci ORDER BY created_at DESC LIMIT 4", {"ci": _cid}
    )
    if _rtkt is not None and not _rtkt.empty:
        for _, r in _rtkt.iterrows():
            ic = TKT_COLOR.get(str(r["status"]), "⚪")
            ta, tb, tc = st.columns([2, 5, 2])
            ta.markdown(f"**{r['ticket_number']}**")
            tb.markdown(str(r['subject'])[:60])
            tc.markdown(f"{ic} {r['status']}")
    else:
        st.caption(t('no_recent_activity'))

    st.markdown("---")

    # ── Recent Offers ─────────────────────────────────────────────────────────
    st.subheader(f"📋 {t('recent_offers')}")
    _roff = _q(
        "SELECT offer_number, freight_type, origin, destination, total_value, status "
        "FROM sales_offers WHERE client_email=:ce ORDER BY created_at DESC LIMIT 4", {"ce": _cemail}
    )
    if _roff is not None and not _roff.empty:
        for _, r in _roff.iterrows():
            ic  = OFF_COLOR.get(str(r["status"]), "⚪")
            val = f"${float(r['total_value']):,.0f}" if r.get('total_value') else "—"
            oa, ob, oc, od = st.columns([2, 3, 2, 2])
            oa.markdown(f"**{r['offer_number']}**")
            ob.markdown(f"📍 {r.get('origin','—')} → {r.get('destination','—')}")
            oc.markdown(f"💰 {val}")
            od.markdown(f"{ic} {r['status']}")
    else:
        st.caption(t('no_recent_activity'))

elif page_matches(page, 'my_shipments'):
    st.header(t('my_shipments'))
    user = st.session_state.get('user')
    if not user or user.get('role') != 'client':
        st.error(t('page_for_clients'))
        st.stop()

    try:
        df = _get_shipments_by_client(user['id'])
        if df.empty:
            st.info(t('no_shipments_client'))
        else:
            # Search & filter
            _sf1, _sf2 = st.columns([3, 2])
            _shp_search = _sf1.text_input("🔍 " + t('search'), placeholder=t('shipment_number'), key="shp_search", label_visibility="collapsed")
            _shp_statuses = ["All"] + sorted(df["status"].dropna().unique().tolist())
            _shp_status_f = _sf2.selectbox(t('status'), _shp_statuses, key="shp_status_f", format_func=_topt, label_visibility="collapsed")
            _df_filtered = df.copy()
            if _shp_search:
                _sq = _shp_search.lower()
                _mask = (
                    _df_filtered["shipment_number"].str.lower().str.contains(_sq, na=False) |
                    _df_filtered["origin_country"].str.lower().str.contains(_sq, na=False) |
                    _df_filtered["destination_country"].str.lower().str.contains(_sq, na=False)
                )
                _df_filtered = _df_filtered[_mask]
            if _shp_status_f != "All":
                _df_filtered = _df_filtered[_df_filtered["status"] == _shp_status_f]
            st.caption(f"{len(_df_filtered)} / {len(df)} {t('tickets_found')}")
            st.dataframe(_df_filtered, width='stretch')

            st.markdown("---")
            st.subheader(t('ship_details_hdr'))

            shipment_nums = df['shipment_number'].tolist()
            selected_shipment = st.selectbox(t('select_your_shipment'), shipment_nums, key="shp_sel_box")

            ship_data = df[df['shipment_number'] == selected_shipment].iloc[0]

            col1, col2 = st.columns(2)
            with col1:
                st.metric(t('shipment_status'), ship_data['status'])
                st.write(f"**{t('origin')}:** {ship_data['origin_country']}")
                st.write(f"**{t('destination')}:** {ship_data['destination_country']}")
                st.write(f"**Type:** {ship_data['type']}")
            with col2:
                st.metric(t('total_value'), f"{ship_data['currency']} {ship_data['total_value']:,.2f}")
                st.write(f"**{t('departure_date')}:** {ship_data['departure_date']}")
                st.write(f"**{t('expected_arrival')}:** {ship_data['expected_arrival']}")
                st.write(f"**{t('customs_cleared')}:** {'Yes' if ship_data.get('customs_cleared') else 'No'}")

            # Show cargo items
            st.markdown("---")
            st.subheader(t('cargo_items'))
            try:
                cargo_df = _get_cargo_by_shipment(int(ship_data['id']))
                if not cargo_df.empty:
                    st.dataframe(cargo_df[['item_name', 'quantity', 'unit', 'weight', 'value']], width='stretch')
                else:
                    st.info(t('no_cargo_listed'))
            except Exception as _ce:
                st.warning(f"Could not load cargo items: {_ce}")

            # Show tracking
            st.markdown("---")
            st.subheader(t('tracking'))
            tracking_df = _q("SELECT location, status, notes, created_at FROM tracking_updates WHERE shipment_id=:sid ORDER BY created_at DESC", {"sid": int(ship_data['id'])})
            if not tracking_df.empty:
                for _, row in tracking_df.iterrows():
                    st.markdown(f"""
                    <div class='card'>
                        <strong>{str(row['created_at'])[:16]}</strong> - {row.get('location','')}<br>
                        <span class='status-badge status-{str(row['status']).lower().replace(' ', '-')}'>{row['status']}</span><br>
                        {row['notes'] if row.get('notes') else ''}
                    </div>
                    """, unsafe_allow_html=True)
            else:
                st.info(t('no_tracking_yet'))

            # Rating section for delivered shipments
            if str(ship_data.get("status","")) == "Delivered":
                st.markdown("---")
                st.subheader(t('rate_shipment'))
                existing_fb = db.fetch_dataframe(
                    "SELECT id FROM cs_feedback WHERE client_id=:ci AND shipment_ref=:sr",
                    {"ci": user["id"], "sr": selected_shipment}
                )
                if not existing_fb.empty:
                    st.success("You have already rated this shipment. Thank you!")
                else:
                    with st.form(f"rate_form_{selected_shipment}", clear_on_submit=True):
                        rf1, rf2 = st.columns(2)
                        rating_v  = rf1.select_slider("Your Rating *", options=[1,2,3,4,5], value=5,
                                                      format_func=lambda x: f"{'⭐'*x} ({x}/5)")
                        cat_v     = rf2.selectbox(t('category_label'), ["Overall","Delivery Time","Service Quality",
                                                               "Documentation","Communication","Pricing"], format_func=_topt)
                        comment_v = st.text_area("Comment (optional)", height=80)
                        if st.form_submit_button("Submit Rating", type="primary"):
                            db.add_cs_feedback(user.get("email",""), user.get("email",""),
                                               user["id"], rating_v, cat_v, comment_v, selected_shipment)
                            st.success("Thank you for your feedback!")
                            st.rerun()
    except Exception as e:
        st.error(f"Error loading your shipments: {str(e)}")

elif page_matches(page, 'rate_service'):
    st.header("⭐ " + t('rate_service'))
    st.caption(t('rate_service_caption'))
    st.markdown("---")

    user = st.session_state.get("user")
    if not user:
        st.error(t('login_required_msg'))
        st.stop()

    uid = user["id"]
    client_email = user.get("email","")

    # Map DB category names to translated labels
    _cat_map = {
        'Overall':          t('cat_overall'),
        'Delivery Time':    t('cat_delivery_time'),
        'Service Quality':  t('cat_service_quality'),
        'Documentation':    t('cat_documentation'),
        'Communication':    t('cat_communication'),
        'Pricing':          t('cat_pricing'),
    }
    # Reverse map: translated label → DB value (for form submission)
    _cat_rev = {v: k for k, v in _cat_map.items()}
    _cat_options = list(_cat_map.values())

    # Show all delivered shipments for this client
    try:
        my_ships = db.fetch_dataframe(
            "SELECT shipment_number, origin, destination, actual_arrival FROM shipments WHERE client_id=:ci AND status='Delivered' ORDER BY actual_arrival DESC",
            {"ci": uid}
        )
    except Exception:
        my_ships = pd.DataFrame()

    if my_ships.empty:
        st.info(t('no_delivered_ships'))
    else:
        # Check which ones already rated
        try:
            already_rated = db.fetch_dataframe(
                "SELECT shipment_ref FROM cs_feedback WHERE client_id=:ci",
                {"ci": uid}
            )
            rated_set = set(already_rated["shipment_ref"].tolist()) if not already_rated.empty else set()
        except Exception:
            rated_set = set()

        pending = my_ships[~my_ships["shipment_number"].isin(rated_set)]
        rated   = my_ships[my_ships["shipment_number"].isin(rated_set)]

        if not pending.empty:
            st.subheader(t('ships_awaiting_rating').format(n=len(pending)))
            for _, sh in pending.iterrows():
                with st.expander(f"📦 {sh['shipment_number']}  |  {sh['origin']} → {sh['destination']}  |  {t('delivered_lbl')}: {str(sh['actual_arrival'])[:10]}"):
                    with st.form(f"rate_{sh['shipment_number']}", clear_on_submit=True):
                        rb1, rb2 = st.columns(2)
                        r_rating  = rb1.select_slider(t('rating_lbl'), options=[1,2,3,4,5], value=5,
                                                      format_func=lambda x: f"{'⭐'*x}  ({x}/5)")
                        r_cat_tr  = rb2.selectbox(t('what_rating'), _cat_options)
                        r_comment = st.text_area(t('comment_optional'), height=80)
                        if st.form_submit_button(t('submit_rating'), type="primary"):
                            r_cat_db = _cat_rev.get(r_cat_tr, r_cat_tr)
                            db.add_cs_feedback(client_email, client_email, uid,
                                               r_rating, r_cat_db, r_comment, sh["shipment_number"])
                            st.success(t('rating_submitted'))
                            st.rerun()
        else:
            st.success(t('all_ships_rated'))

        if not rated.empty:
            st.markdown("---")
            st.subheader(t('already_rated').format(n=len(rated)))
            try:
                fb_df = db.fetch_dataframe(
                    "SELECT id, shipment_ref, rating, category, comment, created_at FROM cs_feedback WHERE client_id=:ci ORDER BY created_at DESC",
                    {"ci": uid}
                )
                STAR = {5:"⭐⭐⭐⭐⭐",4:"⭐⭐⭐⭐",3:"⭐⭐⭐",2:"⭐⭐",1:"⭐"}
                for _, fb in fb_df.iterrows():
                    stars = STAR.get(int(fb["rating"]),"⭐")
                    _cat_display = _cat_map.get(fb['category'], fb['category'])
                    _fb_edit_key = f"editing_fb_{fb['id']}"
                    _fb_del_key  = f"confirm_del_fb_{fb['id']}"

                    with st.container():
                        if not st.session_state.get(_fb_edit_key, False):
                            # View row
                            rv1, rv2, rv3 = st.columns([5, 1.5, 1.5])
                            _date_str = str(fb['created_at'])[:10]
                            _date_badge = f"<span style='background:#1d4ed8;color:#fff;padding:2px 8px;border-radius:10px;font-size:0.78em;font-weight:500'>{_date_str}</span>"
                            rv1.markdown(f"{stars}  **{fb['shipment_ref']}** — {_cat_display} &nbsp;{_date_badge}", unsafe_allow_html=True)
                            if rv2.button(f"✏️ {t('edit_rating_btn')}", key=f"fb_edit_btn_{fb['id']}"):
                                st.session_state[_fb_edit_key] = True
                                st.rerun()
                            if rv3.button(f"🗑️ {t('delete_rating_btn')}", key=f"fb_del_btn_{fb['id']}"):
                                st.session_state[_fb_del_key] = True
                                st.rerun()
                            if fb.get("comment"):
                                st.markdown(f"<span style='color:#9ca3af;font-size:0.9em'>{fb['comment']}</span>", unsafe_allow_html=True)

                            # Delete confirmation
                            if st.session_state.get(_fb_del_key, False):
                                st.warning(t('confirm_delete_rating'))
                                dc1, dc2, _ = st.columns([1, 1, 5])
                                if dc1.button(f"✅ {t('delete_rating_btn')}", key=f"fb_del_confirm_{fb['id']}", type="primary"):
                                    if db.delete_cs_feedback(fb["id"]):
                                        st.session_state.pop(_fb_del_key, None)
                                        st.warning(t('rating_deleted_ok'))
                                        st.rerun()
                                if dc2.button(f"❌ {t('cancel_btn')}", key=f"fb_del_cancel_{fb['id']}"):
                                    st.session_state.pop(_fb_del_key, None)
                                    st.rerun()
                        else:
                            # Edit form
                            with st.form(f"edit_fb_form_{fb['id']}", clear_on_submit=False):
                                ef1, ef2 = st.columns(2)
                                cur_rating = int(fb["rating"])
                                cur_cat_tr = _cat_map.get(fb['category'], fb['category'])
                                new_r_rating = ef1.select_slider(t('rating_lbl'), options=[1,2,3,4,5], value=cur_rating,
                                                                 format_func=lambda x: f"{'⭐'*x}  ({x}/5)")
                                new_r_cat_tr = ef2.selectbox(t('what_rating'), _cat_options,
                                                             index=_cat_options.index(cur_cat_tr) if cur_cat_tr in _cat_options else 0)
                                new_r_comment = st.text_area(t('comment_optional'), value=str(fb.get("comment","") or ""), height=80)
                                ef_s1, ef_s2 = st.columns(2)
                                if ef_s1.form_submit_button(f"💾 {t('save_ticket_btn')}", type="primary"):
                                    new_r_cat_db = _cat_rev.get(new_r_cat_tr, new_r_cat_tr)
                                    if db.edit_cs_feedback(fb["id"], new_r_rating, new_r_cat_db, new_r_comment):
                                        st.session_state.pop(_fb_edit_key, None)
                                        st.success(t('rating_edited_success'))
                                        st.rerun()
                                if ef_s2.form_submit_button(f"❌ {t('cancel_btn')}"):
                                    st.session_state.pop(_fb_edit_key, None)
                                    st.rerun()
                        st.divider()
            except Exception:
                pass

elif page_matches(page, 'my_tickets'):
    st.header("🎫 " + t('my_tickets'))
    st.markdown("---")

    user = st.session_state.get("user")
    if not user:
        st.error(t('login_required_msg'))
    else:
        client_id    = user["id"]
        client_email = user.get("email", "")
        client_name  = user.get("email", "")

        CATEGORIES = ["Shipment Inquiry","Document Request","Customs Issue","Complaint","Rate Request","General Inquiry","New Shipment Request","Add Cargo to Shipment"]
        PRIORITIES  = ["Low","Medium","High","Urgent"]
        STATUSES    = ["Open","In Progress","Resolved","Closed"]
        PRIORITY_ICON = {"Urgent":"🔴","High":"🟠","Medium":"🟡","Low":"🟢"}
        STATUS_ICON   = {"Open":"🔴","In Progress":"🟡","Resolved":"🟢","Closed":"⚫"}

        tab_my, tab_new = st.tabs([t('tab_all_tickets'), f"➕ {t('tab_new_ticket')}"])

        with tab_my:
            my_tkt_df = db.fetch_dataframe(
                "SELECT * FROM cs_tickets WHERE client_id=:ci ORDER BY created_at DESC",
                {"ci": client_id}
            )
            if not my_tkt_df.empty:
                unread_map = db.get_all_unread_counts("client", client_id)
                # Search & filter
                _tf1, _tf2, _tf3 = st.columns([3, 2, 2])
                _tkt_search  = _tf1.text_input("🔍", placeholder=t('subject'), key="tkt_search", label_visibility="collapsed")
                _tkt_stat_f  = _tf2.selectbox(t('status'),   ["All"]+STATUSES,    key="tkt_sf", format_func=_topt, label_visibility="collapsed")
                _tkt_cat_f   = _tf3.selectbox(t('category_label'), ["All"]+CATEGORIES, key="tkt_cf", format_func=_topt, label_visibility="collapsed")
                _tkt_display = my_tkt_df.copy()
                if _tkt_search:
                    _tkt_display = _tkt_display[_tkt_display["subject"].str.contains(_tkt_search, case=False, na=False)]
                if _tkt_stat_f != "All":
                    _tkt_display = _tkt_display[_tkt_display["status"] == _tkt_stat_f]
                if _tkt_cat_f != "All":
                    _tkt_display = _tkt_display[_tkt_display["category"] == _tkt_cat_f]
                st.caption(f"{len(_tkt_display)} / {len(my_tkt_df)} {t('tickets_found')}")
                for _, row in _tkt_display.iterrows():
                    tid = int(row["id"])
                    s_icon  = STATUS_ICON.get(str(row["status"]),"⚪")
                    p_icon  = PRIORITY_ICON.get(str(row["priority"]),"⚪")
                    unread  = unread_map.get(tid, 0)
                    notif   = f"  🔴 {unread} new" if unread > 0 else ""
                    with st.expander(f"{s_icon} {row['ticket_number']}  |  {p_icon} {row['priority']}  —  {row['subject'][:55]}{notif}"):
                        r1,r2,r3 = st.columns(3)
                        r1.write(f"**{t('status')}:** {_topt(str(row['status']))}")
                        r2.write(f"**{t('category_label')}:** {_topt(str(row['category']))}")
                        r3.write(f"**{t('created_lbl')}:** {str(row['created_at'])[:10]}")
                        if row.get("shipment_ref"):
                            st.write(f"**{t('shipment_ref')}:** {row['shipment_ref']}")
                        st.markdown(f"**{t('description')}:**")
                        st.markdown(f"<div style='background:rgba(120,120,180,0.12);border-left:3px solid rgba(120,120,220,0.5);border-radius:6px;padding:10px 14px;margin:4px 0'>{row['description']}</div>", unsafe_allow_html=True)
                        if row.get("resolution"):
                            st.markdown(f"<div style='background:rgba(34,197,94,0.12);border-left:3px solid rgba(34,197,94,0.5);border-radius:6px;padding:10px 14px;margin:4px 0'><b>{t('resolution_lbl')}:</b> {row['resolution']}</div>", unsafe_allow_html=True)
                        if row["status"] in ("Resolved","Closed"):
                            st.caption(f"{t('resolved_at_lbl')}: {str(row.get('resolved_at',''))[:16]}")

                        st.markdown("---")
                        # ── Conversation toggle button ──
                        conv_key = f"show_conv_client_{tid}"
                        btn_lbl  = (f"💬 {t('ticket_conversation')}  🔴 {unread} {t('unread_count')}"
                                    if unread > 0 else f"💬 {t('ticket_conversation')}")
                        if st.button(btn_lbl, key=f"conv_btn_c_{tid}"):
                            st.session_state[conv_key] = not st.session_state.get(conv_key, False)

                        if st.session_state.get(conv_key, False):
                            if unread > 0:
                                db.mark_ticket_replies_read(tid, "client")
                            replies_df = db.get_ticket_replies(tid)
                            if replies_df.empty:
                                st.caption(t('no_replies_yet'))
                            else:
                                for _, rpl in replies_df.iterrows():
                                    is_cs  = str(rpl["sender_role"]) == "cs"
                                    slbl   = t('reply_from_cs') if is_cs else t('reply_from_client')
                                    bg     = "rgba(59,130,246,0.15)" if is_cs else "rgba(34,197,94,0.12)"
                                    border = "rgba(59,130,246,0.45)" if is_cs else "rgba(34,197,94,0.4)"
                                    align  = "left" if is_cs else "right"
                                    st.markdown(
                                        f"<div style='background:{bg};border-left:3px solid {border};"
                                        f"border-radius:8px;padding:8px 12px;margin:4px 0;text-align:{align}'>"
                                        f"<small><b>{slbl}</b> · {str(rpl['created_at'])[:16]}</small><br>{rpl['message']}"
                                        f"</div>", unsafe_allow_html=True
                                    )
                            with st.form(f"client_reply_form_{tid}", clear_on_submit=True):
                                rpl_msg = st.text_area(t('type_reply'), height=70, label_visibility="collapsed", placeholder=t('type_reply'))
                                if st.form_submit_button(f"📤 {t('send_reply_btn')}", type="primary"):
                                    if not rpl_msg.strip():
                                        st.error(t('reply_required'))
                                    else:
                                        db.add_ticket_reply(tid, "client", client_email, rpl_msg.strip())
                                        st.success(t('reply_sent'))
                                        st.rerun()
            else:
                st.info(t('no_tickets_found'))

        with tab_new:
            st.subheader(t('create_new_ticket'))
            my_ships_df = db.fetch_dataframe(
                "SELECT shipment_number FROM shipments WHERE client_id=:ci ORDER BY id DESC LIMIT 30",
                {"ci": client_id}
            )
            ship_opts_c = ["None"] + (my_ships_df["shipment_number"].tolist() if not my_ships_df.empty else [])

            # Category selector outside form for dynamic fields
            category_c = st.selectbox(t('category_label') + " *", CATEGORIES, format_func=_topt, key="client_ticket_cat")

            if category_c == "New Shipment Request":
                st.info("🚢 Fill in the shipment details below. Our team will review and create the shipment for you.")
                with st.form("client_new_ticket_shipment", clear_on_submit=True):
                    subject_c = st.text_input(t('subject') + " *", placeholder="e.g. New Sea Shipment from Istanbul to Dubai")
                    ns1, ns2 = st.columns(2)
                    ns_origin = ns1.text_input("Origin *", placeholder="e.g. Istanbul, Turkey")
                    ns_dest   = ns2.text_input("Destination *", placeholder="e.g. Dubai, UAE")
                    ns3, ns4 = st.columns(2)
                    ns_type   = ns3.selectbox("Shipment Type *", ["Sea","Air","Land"])
                    ns_date   = ns4.date_input("Desired Departure Date")
                    ns5, ns6  = st.columns(2)
                    ns_weight = ns5.number_input("Estimated Weight (kg)", min_value=0.0, step=0.5)
                    ns_value  = ns6.number_input("Estimated Value (USD)", min_value=0.0, step=1.0)
                    ns_notes  = st.text_area("Additional Notes", height=80)
                    if st.form_submit_button(t('create_ticket_btn'), type="primary"):
                        if not subject_c or not ns_origin or not ns_dest:
                            st.error("Subject, Origin, and Destination are required.")
                        else:
                            desc_c = (f"[New Shipment Request]\n"
                                      f"Type: {ns_type}\n"
                                      f"Origin: {ns_origin}\n"
                                      f"Destination: {ns_dest}\n"
                                      f"Desired Departure: {ns_date}\n"
                                      f"Est. Weight: {ns_weight} kg\n"
                                      f"Est. Value: {ns_value} USD\n"
                                      f"Notes: {ns_notes or '—'}")
                            tn_c = f"TKT-{datetime.now().strftime('%Y%m%d%H%M%S')}"
                            ok_c = db.create_cs_ticket(tn_c, client_name, client_email, client_id,
                                                       category_c, subject_c, desc_c, "High", None)
                            if ok_c:
                                db.insert_activity_log("Ticket Created", f"Client ticket {tn_c} submitted", client_email)
                                st.success(t('ticket_submitted'))
                                st.rerun()
                            else:
                                st.error(t('failed_create_ticket_err'))

            elif category_c == "Add Cargo to Shipment":
                st.info("📦 Select an existing shipment and describe the cargo you want to add.")
                with st.form("client_new_ticket_cargo", clear_on_submit=True):
                    subject_c  = st.text_input(t('subject') + " *", placeholder="e.g. Add extra cargo to SHP-2026-001")
                    ac1, ac2   = st.columns(2)
                    ac_ship    = ac1.selectbox("Shipment *", [s for s in ship_opts_c if s != "None"] or ["— No shipments —"])
                    ac_item    = ac2.text_input("Item / Cargo Name *", placeholder="e.g. Electronic components")
                    ac3, ac4, ac5 = st.columns(3)
                    ac_qty     = ac3.number_input("Quantity", min_value=1, step=1, value=1)
                    ac_weight  = ac4.number_input("Weight (kg)", min_value=0.0, step=0.5)
                    ac_value   = ac5.number_input("Value (USD)", min_value=0.0, step=1.0)
                    ac_desc    = st.text_area("Cargo Description", height=80)
                    if st.form_submit_button(t('create_ticket_btn'), type="primary"):
                        no_ships = not my_ships_df.empty and ac_ship not in my_ships_df["shipment_number"].tolist()
                        if not subject_c or not ac_item or my_ships_df.empty:
                            st.error("Subject, shipment, and item name are required." if not my_ships_df.empty else t('no_shipments_client'))
                        else:
                            desc_c = (f"[Add Cargo Request]\n"
                                      f"Shipment: {ac_ship}\n"
                                      f"Item: {ac_item}\n"
                                      f"Quantity: {ac_qty}\n"
                                      f"Weight: {ac_weight} kg\n"
                                      f"Value: {ac_value} USD\n"
                                      f"Description: {ac_desc or '—'}")
                            tn_c  = f"TKT-{datetime.now().strftime('%Y%m%d%H%M%S')}"
                            sr_c  = ac_ship
                            ok_c  = db.create_cs_ticket(tn_c, client_name, client_email, client_id,
                                                        category_c, subject_c, desc_c, "Medium", sr_c)
                            if ok_c:
                                db.insert_activity_log("Ticket Created", f"Client ticket {tn_c} submitted", client_email)
                                st.success(t('ticket_submitted'))
                                st.rerun()
                            else:
                                st.error(t('failed_create_ticket_err'))

            else:
                with st.form("client_new_ticket_form", clear_on_submit=True):
                    ship_ref_c = st.selectbox(t('shipment_ref'), ship_opts_c)
                    subject_c  = st.text_input(t('subject') + " *")
                    desc_c     = st.text_area(t('description') + " *", height=120)
                    if st.form_submit_button(t('create_ticket_btn'), type="primary"):
                        if not subject_c or not desc_c:
                            st.error(t('client_subj_desc_required'))
                        else:
                            tn_c = f"TKT-{datetime.now().strftime('%Y%m%d%H%M%S')}"
                            sr_c = None if ship_ref_c == "None" else ship_ref_c
                            ok_c = db.create_cs_ticket(tn_c, client_name, client_email, client_id,
                                                       category_c, subject_c, desc_c, "Medium", sr_c)
                            if ok_c:
                                db.insert_activity_log("Ticket Created", f"Client ticket {tn_c} submitted", client_email)
                                st.success(t('ticket_submitted'))
                                st.rerun()
                            else:
                                st.error(t('failed_create_ticket_err'))

elif page_matches(page, 'my_invoices'):
    st.header("🧾 " + t('my_invoices'))
    st.markdown("---")

    user = st.session_state.get("user")
    if not user:
        st.error(t('login_required_msg'))
    else:
        client_email = user.get("email", "")
        inv_df = db.fetch_dataframe(
            "SELECT * FROM finance_invoices WHERE client_email=:ce ORDER BY created_at DESC",
            {"ce": client_email}
        )
        if not inv_df.empty:
            STATUS_COLORS = {"Paid":"🟢","Draft":"⚪","Sent":"🔵","Overdue":"🔴","Cancelled":"⬛","Unpaid":"🟡"}
            total_paid    = float(inv_df[inv_df["status"]=="Paid"]["total"].sum())
            total_pending = float(inv_df[inv_df["status"].isin(["Sent","Unpaid","Overdue"])]["total"].sum())
            mi1, mi2, mi3 = st.columns(3)
            mi1.metric(t('total_invoices_lbl'), len(inv_df))
            mi2.metric(t('paid_lbl'), f"${total_paid:,.2f}")
            mi3.metric(t('outstanding_lbl'), f"${total_pending:,.2f}")
            st.markdown("---")
            for _, row in inv_df.iterrows():
                icon = STATUS_COLORS.get(str(row["status"]), "⚪")
                with st.expander(f"{icon} {row['invoice_number']}  |  {row.get('shipment_ref','—')}  |  ${row['total']:,.2f}  |  {row['status']}"):
                    # Header info
                    h1, h2, h3 = st.columns(3)
                    h1.write(f"**{t('issue_date_lbl')}:** {str(row['issue_date'])[:10]}")
                    h2.write(f"**{t('due_date_lbl')}:** {str(row['due_date'])[:10]}")
                    h3.write(f"**{t('shipment_ref_lbl')}:** {row.get('shipment_ref','—') or '—'}")
                    st.markdown("---")
                    # Full cost breakdown
                    st.markdown(f"**{t('cost_breakdown')}**")
                    cargo_val  = float(row.get("cargo_value", 0) or 0)
                    freight    = float(row.get("freight_charge", 0) or 0)
                    handling   = float(row.get("handling_fee", 0) or 0)
                    insurance  = float(row.get("insurance_fee", 0) or 0)
                    tax_amt    = float(row.get("tax_amount", 0) or 0)
                    total      = float(row.get("total", 0) or 0)
                    services   = freight + handling + insurance

                    breakdown_data = {
                        t('amount_usd_col'): [
                            f"${cargo_val:,.2f}",
                            f"${freight:,.2f}",
                            f"${handling:,.2f}",
                            f"${insurance:,.2f}",
                            f"${services:,.2f}",
                            f"${tax_amt:,.2f}",
                            f"**${total:,.2f}**",
                        ]
                    }
                    _idx = [
                        t('cargo_value_lbl'),
                        t('freight_charge_lbl'),
                        t('handling_fee_lbl'),
                        t('insurance_lbl'),
                        t('services_subtotal_lbl'),
                        f"{t('vat_lbl')} ({row['tax_rate']}%)",
                        f"🔷 {t('grand_total_lbl')}",
                    ]
                    import pandas as _pd
                    st.table(_pd.DataFrame(breakdown_data, index=_idx))
                    if row.get("description"):
                        st.caption(f"{t('description')}: {row['description']}")
                    st.markdown("---")
                    _pdf_bytes = generate_invoice_pdf(row.to_dict())
                    if _pdf_bytes:
                        st.download_button(
                            label=f"⬇️ {t('download_pdf')}",
                            data=_pdf_bytes,
                            file_name=f"{row['invoice_number']}.pdf",
                            mime="application/pdf",
                            key=f"dl_inv_{row['id']}"
                        )
        else:
            st.info(t('no_invoices'))

elif page_matches(page, 'my_offers'):
    st.header("📋 " + t('my_offers'))
    st.markdown("---")

    user = st.session_state.get("user")
    if not user:
        st.error(t('login_required_msg'))
    else:
        client_email = user.get("email", "")
        off_df = db.fetch_dataframe(
            "SELECT * FROM sales_offers WHERE client_email=:ce AND status != 'Draft' ORDER BY created_at DESC",
            {"ce": client_email}
        )

        STATUS_ICON_O = {"Sent":"🔵","Accepted":"🟢","Rejected":"🔴","Expired":"⚫"}

        if not off_df.empty:
            oa1, oa2, oa3 = st.columns(3)
            oa1.metric(t('total_offers'), len(off_df))
            oa2.metric(t('accepted_value'), f"${float(off_df[off_df['status']=='Accepted']['total_value'].sum() or 0):,.0f}")
            oa3.metric(t('pending_sent'),   f"${float(off_df[off_df['status']=='Sent']['total_value'].sum() or 0):,.0f}")
            st.markdown("---")

            for _, row in off_df.iterrows():
                icon    = STATUS_ICON_O.get(str(row["status"]), "⚪")
                val_str = f"${float(row['total_value'] or 0):,.0f} {row.get('currency','USD')}" if row.get("total_value") else "TBD"
                with st.expander(f"{icon} {row['offer_number']}  |  {val_str}  |  {row['status']}  |  {t('valid_until_lbl')} {str(row.get('validity_date','—'))[:10]}"):
                    oc1, oc2 = st.columns(2)
                    oc1.write(f"**{t('freight_lbl')}** {row.get('freight_type','—')}")
                    oc2.write(f"**{t('route_lbl')}** {row.get('origin','—')} → {row.get('destination','—')}")
                    oc3, oc4 = st.columns(2)
                    oc3.write(f"**{t('commodity_lbl')}** {row.get('commodity','—')}")
                    oc4.write(f"**{t('valid_until_lbl')}** {str(row.get('validity_date','—'))[:10]}")
                    wt  = row.get("weight_kg")
                    vol = row.get("volume_cbm")
                    ow1, ow2 = st.columns(2)
                    ow1.write(f"**{t('weight_lbl')}** {f'{float(wt):,.1f} kg' if wt else '—'}")
                    ow2.write(f"**{t('volume_lbl')}** {f'{float(vol):,.2f} CBM' if vol else '—'}")
                    if row.get("notes"):
                        st.caption(row["notes"])

                    if str(row["status"]) == "Sent":
                        st.markdown("---")
                        btn_acc, btn_rej, _ = st.columns([1, 1, 4])
                        if btn_acc.button(f"✅ {t('accept_offer')}", key=f"acc_{row['id']}", type="primary"):
                            if db.update_offer_status(int(row["id"]), "Accepted"):
                                # Auto-create a Deal in Sales
                                _deal_title = f"{row.get('origin','—')} → {row.get('destination','—')} ({row.get('freight_type','—')})"
                                db.add_deal(
                                    title        = _deal_title,
                                    client_name  = client_email,
                                    value        = float(row.get("total_value") or 0),
                                    currency     = str(row.get("currency","USD")),
                                    stage        = "Proposal",
                                    probability  = 80,
                                    close_date   = str(row.get("validity_date", date.today())),
                                    freight_type = str(row.get("freight_type","")),
                                    origin       = str(row.get("origin","")),
                                    destination  = str(row.get("destination","")),
                                    notes        = f"Auto-created from accepted offer {row.get('offer_number','')}",
                                )
                                st.success(t('offer_accepted'))
                                st.rerun()
                        if btn_rej.button(f"❌ {t('reject_offer')}", key=f"rej_{row['id']}", type="secondary"):
                            if db.update_offer_status(int(row["id"]), "Rejected"):
                                st.warning(t('offer_rejected'))
                                st.rerun()
        else:
            st.info(t('no_offers_client'))

elif page_matches(page, 'track_shipment'):
    st.header(t('track_shipment'))

    shipment_number = st.text_input(t('enter_shipment_num_lbl'), placeholder='SH-2025-001')
    
    if st.button(t('track_btn')):
        if shipment_number:
            try:
                df = get_cached_shipments()
                ship_data = df[df['shipment_number'] == shipment_number]

                if ship_data.empty:
                    st.error(t('ship_not_found'))
                else:
                    ship = ship_data.iloc[0]

                    # Check if user is client and can only view their own shipments
                    user = st.session_state.get('user')
                    if user and user.get('role') == 'client' and ship['client_id'] != user['id']:
                        st.error(t('own_ships_only'))
                        st.stop()

                    st.success(t('ship_found').format(n=shipment_number))

                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric(t('status'), ship['status'])
                    with col2:
                        st.metric(t('origin_lbl'), ship['origin_country'])
                    with col3:
                        st.metric(t('destination_lbl'), ship['destination_country'])

                    st.markdown("---")
                    st.subheader(t('tracking_history'))

                    tracking_df = db.get_tracking_updates(ship['id'])
                    if not tracking_df.empty:
                        for _, row in tracking_df.iterrows():
                            st.markdown(f"""
                            <div class='card'>
                                <strong>📍 {row['location']}</strong><br>
                                <span class='status-badge status-{row['status'].lower().replace(' ', '-')}'>{row['status']}</span><br>
                                {t('date_lbl')}: {row['update_date']}<br>
                                {row['notes'] if row['notes'] else ''}
                            </div>
                            """, unsafe_allow_html=True)
                    else:
                        st.info(t('no_tracking_updates'))
            except Exception as e:
                st.error(f"Error: {str(e)}")
        else:
            st.warning(t('enter_ship_number'))

elif page_matches(page, 'shipment_analytics'):
    st.header(t('shipment_analytics'))
    user = st.session_state.get('user')
    if not user or user.get('role') not in ['manager', 'employee']:
        st.error("You must be a manager or employee to view analytics.")
        st.stop()

    try:
        stats = db.get_shipment_statistics()
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Total Shipments", stats['total_shipments'])
        with col2:
            st.metric("Imports", stats['total_imports'])
        with col3:
            st.metric("Exports", stats['total_exports'])
        with col4:
            st.metric("In Transit", stats['in_transit'])
        
        st.metric("Total Value", f"${stats['total_value']:,.2f}")
        
        # Charts
        df = db.get_all_shipments()
        if not df.empty:
            st.markdown("---")
            col_a, col_b = st.columns(2)
            
            with col_a:
                st.subheader("Shipments by Type")
                type_counts = df['type'].value_counts()
                st.bar_chart(type_counts)
            
            with col_b:
                st.subheader("Shipments by Status")
                status_counts = df['status'].value_counts()
                st.bar_chart(status_counts)
            
            st.markdown("---")
            st.subheader("Recent Shipments")
            st.dataframe(df.head(10), width='stretch')
    except Exception as e:
        st.error(f"Error loading analytics: {str(e)}")

elif page_matches(page, 'cargo_requests'):
    st.header(t('cargo_requests'))
    user = st.session_state.get('user')
    if not user or user.get('role') != 'client':
        st.error(t('page_for_clients'))
        st.stop()

    try:
        st.subheader(t('request_cargo_change'))
        
        # Get all cargo items from database
        conn = db.get_connection()
        cargo_df = pd.read_sql_query('SELECT * FROM cargo_items', conn)
        conn.close()
        
        if not cargo_df.empty:
            with st.form("cargo_request_form"):
                # Select cargo item directly (no shipment selection needed)
                cargo_options = [(f"{row['item_name']} - {row['quantity']} {row['unit']}", row['id']) 
                               for _, row in cargo_df.iterrows()]
                selected_cargo = st.selectbox(t('select_cargo_item'), cargo_options, format_func=lambda x: x[0])
                
                # Request type
                request_type = st.selectbox(t('request_type'), [t('modify'), t('remove')])
                
                # Reason
                reason = st.text_area(t('request_reason'), placeholder="Please explain why you need this change...")
                
                submitted = st.form_submit_button(t('submit'))
                
                if submitted:
                    if reason:
                        try:
                            cargo_id = selected_cargo[1]
                            db.create_cargo_request(cargo_id, user['id'], request_type, reason)
                            st.success(t('request_submitted_success'))
                            _safe_rerun()
                        except Exception as e:
                            st.error(f"Error: {str(e)}")
                    else:
                        st.error("Please provide a reason for your request.")
        else:
            st.info("No cargo items available for requests.")
        
        # Show client's requests
        st.markdown("---")
        st.subheader(t('my_cargo_requests'))
        requests_df = db.get_cargo_requests_by_client(user['id'])
        
        if requests_df.empty:
            st.info("You have no cargo requests.")
        else:
            for _, req in requests_df.iterrows():
                status = req.get('status', 'Pending')
                badge_class = 'status-pending'
                if status == 'Approved':
                    badge_class = 'status-approved'
                elif status == 'Rejected':
                    badge_class = 'status-rejected'
                
                st.markdown("<div class='card'>", unsafe_allow_html=True)
                col1, col2 = st.columns([3, 1])
                with col1:
                    st.write(f"**Shipment:** {req.get('shipment_number', 'N/A')}")
                    st.write(f"**Item:** {req.get('item_name', 'N/A')}")
                    st.write(f"**Request Type:** {req.get('request_type', 'N/A')}")
                    st.write(f"**Reason:** {req.get('reason', 'N/A')}")
                    if req.get('employee_response'):
                        st.write(f"**Employee Response:** {req.get('employee_response')}")
                    st.write(f"**Date:** {req.get('created_at', 'N/A')}")
                with col2:
                    st.markdown(f"<span class='status-badge {badge_class}'>{status}</span>", unsafe_allow_html=True)
                st.markdown("</div>", unsafe_allow_html=True)
    except Exception as e:
        st.error(f"Error: {str(e)}")

elif page_matches(page, 'manage_cargo_requests'):
    st.header(t('manage_cargo_requests'))
    user = st.session_state.get('user')
    if not user or user.get('role') not in ['manager', 'employee']:
        st.error("You must be a manager or employee to manage cargo requests.")
        st.stop()

    try:
        requests_df = db.get_all_cargo_requests()
        
        if requests_df.empty:
            st.info("No cargo requests found.")
        else:
            # Filter options
            col1, col2 = st.columns(2)
            with col1:
                status_filter = st.selectbox(t('filter_status'), ["All", "Pending", "Approved", "Rejected"], format_func=_topt)
            with col2:
                type_filter = st.selectbox(t('filter_by_type'), ["All", t('modify'), t('remove')])
            
            # Apply filters
            filtered_df = requests_df.copy()
            if status_filter != "All":
                filtered_df = filtered_df[filtered_df['status'] == status_filter]
            if type_filter != "All":
                filtered_df = filtered_df[filtered_df['request_type'] == type_filter]
            
            st.markdown("---")
            
            if filtered_df.empty:
                st.info("No requests match the selected filters.")
            else:
                for _, req in filtered_df.iterrows():
                    request_id = int(req['id'])
                    
                    with st.expander(f"Request #{request_id} - {req.get('client_email', 'Unknown')} - {req.get('status', 'Pending')}"):
                        col_info, col_action = st.columns([2, 1])
                        
                        with col_info:
                            st.write(f"**Client:** {req.get('client_email', 'Unknown')}")
                            st.write(f"**Shipment:** {req.get('shipment_number', 'N/A')}")
                            st.write(f"**Cargo Item:** {req.get('item_name', 'N/A')}")
                            st.write(f"**Request Type:** {req.get('request_type', 'N/A')}")
                            st.write(f"**Reason:** {req.get('reason', 'N/A')}")
                            st.write(f"**Date:** {req.get('created_at', 'N/A')}")
                            st.write(f"**Current Status:** {req.get('status', 'Pending')}")
                        
                        with col_action:
                            if req.get('status') == 'Pending':
                                with st.form(f"manage_request_{request_id}"):
                                    new_status = st.selectbox(t('action_lbl'), ["Pending", "Approved", "Rejected"],
                                                            key=f"status_{request_id}", format_func=_topt)
                                    response = st.text_area("Response to client:", key=f"response_{request_id}")
                                    
                                    if st.form_submit_button(t('save')):
                                        try:
                                            db.update_cargo_request_status(request_id, new_status, response)
                                            
                                            # If approved and request is to remove, delete the cargo item
                                            if new_status == "Approved" and req.get('request_type') == t('remove'):
                                                db.delete_cargo_item(req['cargo_item_id'])
                                            
                                            st.success("Request updated successfully!")
                                            _safe_rerun()
                                        except Exception as e:
                                            st.error(f"Error: {str(e)}")
                            else:
                                st.info(f"Status: {req.get('status')}")
                                if req.get('employee_response'):
                                    st.write(f"**Response:** {req.get('employee_response')}")
    except Exception as e:
        st.error(f"Error loading requests: {str(e)}")

elif page_matches(page, 'edit_shipment'):
    st.header(t('edit_shipment'))
    user = st.session_state.get('user')
    if not user or user.get('role') not in ['manager', 'employee']:
        st.error("Only managers and employees can edit shipments.")
        st.stop()

    # Clear any session state cache
    if 'last_edited_shipment' not in st.session_state:
        st.session_state.last_edited_shipment = None

    try:
        # Get fresh data
        df = db.get_all_shipments()
        
        if df.empty:
            st.info(t('no_shipments_found'))
        else:
            # Select shipment to edit
            shipment_nums = df['shipment_number'].tolist()
            selected_shipment = st.selectbox(t('select_shipment_edit'), shipment_nums)
            
            # Always get fresh data for selected shipment
            fresh_df = db.get_all_shipments()
            ship_data = fresh_df[fresh_df['shipment_number'] == selected_shipment].iloc[0]
            
            st.markdown("---")
            st.subheader(f"Edit Shipment: {selected_shipment}")
            
            # Show current values
            with st.expander("Current Values", expanded=False):
                st.write(f"Type: {ship_data['type']}")
                st.write(f"Origin: {ship_data['origin_country']}")
                st.write(f"Destination: {ship_data['destination_country']}")
                st.write(f"Weight: {ship_data['total_weight']} kg")
                st.write(f"Value: {ship_data['currency']} {ship_data['total_value']}")
            
            with st.form("edit_shipment_form", clear_on_submit=False):
                col1, col2 = st.columns(2)
                
                with col1:
                    shipment_type = st.selectbox(t('type'), ["Import", "Export"], 
                                                index=0 if ship_data['type'] == 'Import' else 1, format_func=_topt)
                    origin = st.text_input(t('origin'), value=ship_data['origin_country'])
                    destination = st.text_input(t('destination'), value=ship_data['destination_country'])
                    departure = st.date_input(t('departure_date'), 
                                             value=pd.to_datetime(ship_data['departure_date']).date())
                
                with col2:
                    expected = st.date_input(t('expected_arrival'), 
                                            value=pd.to_datetime(ship_data['expected_arrival']).date())
                    weight = st.number_input(t('total_weight'), value=float(ship_data['total_weight']), min_value=0.0)
                    value_amount = st.number_input(t('total_value'), value=float(ship_data['total_value']), min_value=0.0)
                    currency = st.selectbox(t('currency_lbl'), ["USD", "EUR", "TRY"],
                                           index=["USD", "EUR", "TRY"].index(ship_data['currency']))
        
                submitted = st.form_submit_button("💾 Save Changes")
                
                if submitted:
                    try:
                        with db.get_connection() as conn:
                            conn.execute(text('''
                                UPDATE shipments
                                SET shipment_type=:shipment_type,
                                    origin=:origin,
                                    destination=:destination,
                                    departure_date=:departure_date,
                                    expected_arrival=:expected_arrival,
                                    total_weight=:total_weight,
                                    total_value=:total_value,
                                    currency=:currency
                                WHERE id=:id
                            '''), {
                                "shipment_type": shipment_type,
                                "origin": origin,
                                "destination": destination,
                                "departure_date": str(departure),
                                "expected_arrival": str(expected),
                                "total_weight": weight,
                                "total_value": value_amount,
                                "currency": currency,
                                "id": int(ship_data['id'])
                            })
                            try:
                                conn.commit()
                            except Exception:
                                pass
                        
                        conn.commit()
                        conn.close()
                        
                        # Store in session state
                        st.session_state.last_edited_shipment = selected_shipment
                        
                        st.success(f"✅ Shipment {selected_shipment} updated successfully!")
                        st.info("Please refresh the page to see changes in the table.")
                        _safe_rerun()
                        
                    except Exception as e:
                        st.error(f"❌ Error: {str(e)}")
                        import traceback
                        st.code(traceback.format_exc())
    
    except Exception as e:
        st.error(f"Error: {str(e)}")
        import traceback
        st.code(traceback.format_exc())

elif page_matches(page, 'delete_shipment'):
    st.header(t('delete_shipment'))
    user = st.session_state.get('user')
    if not user or user.get('role') not in ['manager', 'employee']:
        st.error("Only managers and employees can delete shipments.")
        st.stop()

    try:
        df = db.get_all_shipments()
        
        if df.empty:
            st.info(t('no_shipments_found'))
        else:
            # Select shipment to delete
            shipment_nums = df['shipment_number'].tolist()
            selected_shipment = st.selectbox(t('select_shipment_delete'), shipment_nums)
            
            ship_data = df[df['shipment_number'] == selected_shipment].iloc[0]
            
            st.markdown("---")
            st.warning(f"⚠️ You are about to delete shipment: **{selected_shipment}**")
            
            # Show shipment details
            col1, col2 = st.columns(2)
            with col1:
                st.write(f"**Type:** {ship_data['type']}")
                st.write(f"**Origin:** {ship_data['origin_country']}")
                st.write(f"**Destination:** {ship_data['destination_country']}")
            with col2:
                st.write(f"**Status:** {ship_data['status']}")
                st.write(f"**Client:** {ship_data['client_email']}")
                st.write(f"**Departure:** {ship_data['departure_date']}")
            
            st.markdown("---")
            
            confirm = st.checkbox("I confirm I want to delete this shipment")
            
            if st.button("🗑️ Delete Shipment", type="primary", disabled=not confirm):
                try:
                    with db.get_connection() as conn:
                        conn.execute(text('DELETE FROM shipments WHERE id = :id'), {"id": int(ship_data['id'])})
                        try:
                            conn.commit()
                        except Exception:
                            pass
                    st.success(f"✅ Shipment {selected_shipment} deleted successfully!")
                    _safe_rerun()
                except Exception as e:
                    st.error(f"❌ Error deleting shipment: {str(e)}")
    
    except Exception as e:
        st.error(f"Error: {str(e)}")


# Department dashboards and features
elif page_matches(page, 'finance_dashboard'):
    st.header(t('finance_dashboard'))
    st.caption(f"{t('last_updated')}: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    st.markdown("---")

    summary = db.get_finance_summary()

    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric(t('total_revenue'),    f"${summary['total_revenue']:,.0f}")
    k2.metric(t('total_expenses'),   f"${summary['total_expenses']:,.0f}")
    k3.metric(t('net_profit'),       f"${summary['net_profit']:,.0f}",
              delta=f"{(summary['net_profit']/summary['total_revenue']*100):.1f}% {t('margin_lbl')}" if summary['total_revenue'] else None)
    k4.metric(t('outstanding'),      f"${summary['outstanding']:,.0f}")
    k5.metric(t('overdue'),          f"${summary['overdue']:,.0f}",
              delta_color="inverse" if summary['overdue'] > 0 else "normal",
              delta=t('needs_action') if summary['overdue'] > 0 else None)

    st.markdown("---")
    rev_df, exp_df = db.get_monthly_financials()

    col_l, col_r = st.columns(2)
    with col_l:
        if not rev_df.empty or not exp_df.empty:
            merged = pd.merge(rev_df, exp_df, on="month", how="outer").fillna(0).sort_values("month")
            merged["profit"] = merged["revenue"] - merged["expenses"]
            fig_re = px.bar(merged, x="month", y=["revenue","expenses"],
                            labels={"value":"USD","month":"Month","variable":"Type"},
                            color_discrete_map={"revenue":"#22c55e","expenses":"#ef4444"},
                            barmode="group", height=300, title=t('chart_rev_vs_exp'))
            fig_re.update_layout(margin=dict(l=0,r=0,t=30,b=0))
            st.plotly_chart(fig_re, use_container_width=True)
        else:
            st.info(t('no_financial_data_short'))

    with col_r:
        inv_df = db.get_invoices()
        if not inv_df.empty:
            status_counts = inv_df["status"].value_counts().reset_index()
            status_counts.columns = [t('status'), "Count"]
            color_map = {"Paid":"#22c55e","Draft":"#94a3b8","Sent":"#3b82f6","Overdue":"#ef4444","Cancelled":"#6b7280","Unpaid":"#f59e0b"}
            fig_pie = px.pie(status_counts, names=t('status'), values="Count",
                             color=t('status'), color_discrete_map=color_map,
                             title=t('chart_invoice_status'), height=300)
            fig_pie.update_layout(margin=dict(l=0,r=0,t=30,b=0))
            st.plotly_chart(fig_pie, use_container_width=True)
        else:
            st.info(t('no_invoices_yet'))

    st.markdown("---")
    st.subheader(t('recent_invoices'))
    inv_df = db.get_invoices()
    if not inv_df.empty:
        disp = inv_df[["invoice_number","client_name","shipment_ref","total","status","due_date"]].head(8)
        disp.columns = [t('col_invoice'), t('col_client'), t('shipment_ref'), t('col_total_usd'), t('status'), t('col_due_date')]
        st.dataframe(disp, use_container_width=True, hide_index=True)
    else:
        st.info(t('no_invoices'))

elif page_matches(page, 'invoices'):
    st.header("🧾 " + t('invoices'))
    st.markdown("---")
    tab_list, tab_new = st.tabs([t('tab_all_invoices'), t('tab_new_invoice')])

    with tab_list:
        status_f = st.selectbox(t('filter_by_status'), ["All","Draft","Sent","Paid","Unpaid","Overdue","Cancelled"], key="inv_filter", format_func=_topt)
        inv_df = db.get_invoices(status_f)
        if not inv_df.empty:
            status_colors = {"Paid":"🟢","Draft":"⚪","Sent":"🔵","Overdue":"🔴","Cancelled":"⬛","Unpaid":"🟡"}
            for _, row in inv_df.iterrows():
                icon = status_colors.get(row["status"], "⚪")
                with st.expander(f"{icon} {row['invoice_number']} — {row['client_name']}  |  ${row['total']:,.2f}  |  {row['status']}"):
                    c1, c2, c3 = st.columns(3)
                    c1.write(f"**{t('shipment_ref_lbl')}:** {row.get('shipment_ref','—')}")
                    c2.write(f"**{t('issue_date_lbl')}:** {row['issue_date']}")
                    c3.write(f"**{t('due_date_lbl')}:** {row['due_date']}")
                    c1b, c2b, c3b = st.columns(3)
                    c1b.write(f"**{t('amount_lbl')}:** ${row['amount']:,.2f}")
                    c2b.write(f"**{t('tax_lbl')} ({row['tax_rate']}%):** ${row['tax_amount']:,.2f}")
                    c3b.write(f"**{t('total_lbl')}:** ${row['total']:,.2f}")
                    if row.get("description"):
                        st.write(f"**{t('description')}:** {row['description']}")
                    st.markdown("---")
                    btn1, btn2, btn3, btn4 = st.columns(4)
                    if row["status"] != "Paid":
                        if btn1.button(t('mark_sent'),    key=f"sent_{row['id']}"):
                            db.update_invoice_status(row["id"], "Sent"); st.rerun()
                        if btn2.button(t('mark_paid'),    key=f"paid_{row['id']}"):
                            db.update_invoice_status(row["id"], "Paid"); st.rerun()
                        if btn3.button(t('mark_overdue'), key=f"over_{row['id']}"):
                            db.update_invoice_status(row["id"], "Overdue"); st.rerun()
                    if btn4.button(t('delete_btn'), key=f"delinv_{row['id']}", type="secondary"):
                        db.delete_invoice(row["id"]); st.rerun()
        else:
            st.info(t('no_invoices'))

    with tab_new:
        st.subheader(t('create_new_invoice'))
        inv_clients_df = db.fetch_dataframe("SELECT id, email FROM users WHERE role='client' ORDER BY email")
        client_opts_inv = {r["email"]: r["id"] for _, r in inv_clients_df.iterrows()} if not inv_clients_df.empty else {}
        sel_inv_client = st.selectbox(t('client') + " *", ["— Select —"] + list(client_opts_inv.keys()), key="inv_client_sel")
        with st.form("new_invoice_form", clear_on_submit=True):
            ni1, ni2 = st.columns(2)
            client_name  = ni1.text_input(t('client_name_field'))
            shipment_ref = ni2.text_input(t('shipment_ref'))
            ni3, ni4 = st.columns(2)
            amount   = ni3.number_input(t('amount_usd'), min_value=0.0, step=100.0)
            tax_rate = ni4.number_input(t('tax_rate'), min_value=0.0, max_value=50.0, value=18.0, step=0.5)
            ni5, ni6 = st.columns(2)
            due_date    = ni5.date_input(t('due_date'))
            description = ni6.text_area(t('desc_services'), height=80)
            submitted = st.form_submit_button(t('create_invoice_btn'), type="primary")
            if submitted:
                if sel_inv_client == "— Select —" or not client_name or amount <= 0:
                    st.error(t('client_amount_required'))
                else:
                    inv_num      = f"INV-{datetime.now().strftime('%Y%m%d%H%M%S')}"
                    uid          = st.session_state.get("user", {}).get("id", 1)
                    client_email = sel_inv_client
                    ok = db.create_invoice(inv_num, client_name, client_email, shipment_ref,
                                           amount, tax_rate, str(due_date), description, uid)
                    if ok:
                        st.success(t('invoice_created'))
                    else:
                        st.error(t('failed_create_invoice_err'))

elif page_matches(page, 'payments'):
    st.header("💳 " + t('payments'))
    st.markdown("---")
    tab_list, tab_new = st.tabs([t('tab_payment_history'), t('tab_record_payment')])

    with tab_list:
        pay_df = db.get_payments()
        if not pay_df.empty:
            disp = pay_df[["payment_date","invoice_number","client_name","amount","method","reference_no"]].copy()
            disp.columns = [t('col_date'), t('col_invoice'), t('col_client'), t('col_amount_usd'), t('col_method'), t('col_reference')]
            disp["Amount (USD)"] = disp["Amount (USD)"].apply(lambda x: f"${x:,.2f}")
            st.dataframe(disp, use_container_width=True, hide_index=True)

            st.markdown("---")
            pm1, pm2 = st.columns(2)
            with pm1:
                method_counts = pay_df["method"].value_counts().reset_index()
                method_counts.columns = ["Method","Count"]
                fig_m = px.pie(method_counts, names="Method", values="Count",
                               title=t('chart_payments_method'), height=280)
                fig_m.update_layout(margin=dict(l=0,r=0,t=30,b=0))
                st.plotly_chart(fig_m, use_container_width=True)
            with pm2:
                pay_df["payment_date"] = pd.to_datetime(pay_df["payment_date"])
                pay_df["month"] = pay_df["payment_date"].dt.to_period("M").astype(str)
                monthly = pay_df.groupby("month")["amount"].sum().reset_index()
                fig_pm = px.bar(monthly, x="month", y="amount",
                                labels={"amount":"USD","month":"Month"},
                                color_discrete_sequence=["#22c55e"],
                                title=t('chart_monthly_payments'), height=280)
                fig_pm.update_layout(margin=dict(l=0,r=0,t=30,b=0))
                st.plotly_chart(fig_pm, use_container_width=True)
        else:
            st.info(t('no_payments'))

    with tab_new:
        st.subheader(t('record_payment'))
        inv_df = db.get_invoices()
        unpaid = inv_df[inv_df["status"].isin(["Sent","Unpaid","Overdue","Draft"])] if not inv_df.empty else pd.DataFrame()
        if unpaid.empty:
            st.info(t('no_unpaid_invoices'))
        else:
            with st.form("record_payment_form", clear_on_submit=True):
                inv_options = {f"{r['invoice_number']} — {r['client_name']} (${r['total']:,.2f})": r["id"]
                               for _, r in unpaid.iterrows()}
                selected_inv = st.selectbox(t('select_invoice'), list(inv_options.keys()))
                rp1, rp2 = st.columns(2)
                pay_amount  = rp1.number_input(t('amount_received'), min_value=0.0, step=100.0)
                pay_date    = rp2.date_input(t('payment_date'))
                rp3, rp4 = st.columns(2)
                method   = rp3.selectbox(t('payment_method'), ["Bank Transfer","Letter of Credit","Cash","Credit Card","Cheque"], format_func=_topt)
                ref_no   = rp4.text_input(t('ref_no'))
                notes    = st.text_area(t('notes_desc'), height=70)
                if st.form_submit_button(t('record_payment_btn'), type="primary"):
                    if pay_amount <= 0:
                        st.error(t('amount_gt_zero'))
                    else:
                        inv_id = inv_options[selected_inv]
                        ok = db.add_payment(inv_id, pay_amount, str(pay_date), method, ref_no, notes)
                        if ok:
                            st.success(t('payment_recorded'))
                        else:
                            st.error(t('failed_record_payment_err'))

elif page_matches(page, 'expenses'):
    st.header("💸 " + t('expenses'))
    st.markdown("---")
    tab_list, tab_add = st.tabs([t('tab_all_expenses'), t('tab_add_expense')])

    EXPENSE_CATS = ["Cargo Cost","Customs Duty","Freight Fee","Fuel & Transport",
                    "Staff Salary","Office Rent","Insurance","Port Handling","Other"]

    with tab_list:
        ef1, ef2 = st.columns([2,1])
        status_fe = ef1.selectbox(t('filter_by_status'), ["All","Pending","Approved","Rejected"], key="exp_status", format_func=_topt)
        cat_fe    = ef2.selectbox(t('filter_by_cat'), ["All"] + EXPENSE_CATS, key="exp_cat", format_func=_topt)

        exp_df = db.get_expenses(status_fe if status_fe != "All" else None)
        if not exp_df.empty and cat_fe != "All":
            exp_df = exp_df[exp_df["category"] == cat_fe]

        if not exp_df.empty:
            total_shown = exp_df["amount"].sum()
            st.metric(t('total_filtered'), f"${total_shown:,.2f}")

            status_icons = {"Pending":"🟡","Approved":"🟢","Rejected":"🔴"}
            for _, row in exp_df.iterrows():
                icon = status_icons.get(row["status"],"⚪")
                with st.expander(f"{icon} {row['category']}  |  ${row['amount']:,.2f}  |  {row['expense_date']}"):
                    ec1, ec2 = st.columns(2)
                    ec1.write(f"**{t('vendor_lbl')}:** {row.get('vendor','—')}")
                    ec2.write(f"**{t('receipt_ref_lbl')}:** {row.get('receipt_ref','—')}")
                    st.write(f"**{t('description')}:** {row['description']}")
                    st.write(f"**{t('status')}:** {row['status']}")
                    if row["status"] == "Pending":
                        ba1, ba2 = st.columns(2)
                        if ba1.button(t('approve_btn'), key=f"appexp_{row['id']}", type="primary"):
                            db.update_expense_status(row["id"], "Approved"); st.rerun()
                        if ba2.button(t('reject_btn'),  key=f"rejexp_{row['id']}", type="secondary"):
                            db.update_expense_status(row["id"], "Rejected"); st.rerun()

            st.markdown("---")
            cat_sum = exp_df[exp_df["status"]=="Approved"].groupby("category")["amount"].sum().reset_index()
            if not cat_sum.empty:
                fig_cat = px.pie(cat_sum, names="category", values="amount",
                                 title=t('chart_approved_expenses'), height=300)
                fig_cat.update_layout(margin=dict(l=0,r=0,t=30,b=0))
                st.plotly_chart(fig_cat, use_container_width=True)
        else:
            st.info(t('no_expenses_found'))

    with tab_add:
        st.subheader(t('add_new_expense'))
        with st.form("add_expense_form", clear_on_submit=True):
            ae1, ae2 = st.columns(2)
            category   = ae1.selectbox(t('category_field'), EXPENSE_CATS, format_func=_topt)
            amount_e   = ae2.number_input(t('amount_usd'), min_value=0.0, step=50.0)
            ae3, ae4 = st.columns(2)
            exp_date   = ae3.date_input(t('expense_date'))
            vendor     = ae4.text_input(t('vendor_supplier'))
            ae5, ae6 = st.columns(2)
            receipt_r  = ae5.text_input(t('receipt_ref'))
            description_e = ae6.text_area(t('description') + " *", height=80)
            if st.form_submit_button(t('add_expense_btn'), type="primary"):
                if not description_e or amount_e <= 0:
                    st.error(t('fill_required'))
                else:
                    uid = st.session_state.get("user_id", 1)
                    ok  = db.add_expense(category, description_e, amount_e, str(exp_date), vendor, receipt_r, uid)
                    if ok:
                        st.success(t('expense_added'))
                    else:
                        st.error(t('failed_add_expense_err'))

elif page_matches(page, 'financial_reports'):
    st.header("📊 " + t('financial_reports'))
    st.markdown("---")

    # ── Load data ─────────────────────────────────────────────────────────────
    summary   = db.get_finance_summary()
    rev_df, exp_df = db.get_monthly_financials()
    inv_all   = db.get_invoices()
    exp_all   = db.get_expenses()

    total_rev  = summary.get('total_revenue', 0)
    total_exp  = summary.get('total_expenses', 0)
    net_profit = summary.get('net_profit', 0)
    outstanding = summary.get('outstanding', 0)
    overdue    = summary.get('overdue', 0)
    total_inv  = summary.get('total_invoices', 0)
    paid_count = summary.get('paid_count', 0)
    collection_rate = round((paid_count / total_inv * 100) if total_inv else 0, 1)

    # ── KPI Cards ─────────────────────────────────────────────────────────────
    k1,k2,k3,k4,k5 = st.columns(5)
    k1.metric("💰 " + t('total_revenue'),  f"${total_rev:,.0f}")
    k2.metric("📤 " + t('total_expenses'), f"${total_exp:,.0f}")
    k3.metric("📈 " + t('net_profit'),
              f"${net_profit:,.0f}",
              delta=f"{round(net_profit/total_rev*100,1)}% margin" if total_rev else None)
    k4.metric("⏳ Outstanding", f"${outstanding:,.0f}",
              delta=f"⚠️ ${overdue:,.0f} overdue" if overdue else None,
              delta_color="inverse")
    k5.metric("✅ Collection Rate",
              f"{collection_rate}%",
              delta=f"{paid_count}/{total_inv} invoices paid")

    st.markdown("---")

    # ── Monthly Trend: Revenue / Expenses / Net Profit ─────────────────────────
    if not rev_df.empty or not exp_df.empty:
        merged = pd.merge(rev_df, exp_df, on="month", how="outer").fillna(0).sort_values("month")
        merged["profit"] = merged["revenue"] - merged["expenses"]

        import plotly.graph_objects as go
        fig_trend = go.Figure()
        fig_trend.add_bar(x=merged["month"], y=merged["revenue"],  name="Revenue",  marker_color="#22c55e")
        fig_trend.add_bar(x=merged["month"], y=merged["expenses"], name="Expenses", marker_color="#ef4444")
        fig_trend.add_scatter(x=merged["month"], y=merged["profit"],
                              name="Net Profit", mode="lines+markers",
                              line=dict(color="#3b82f6", width=3),
                              marker=dict(size=8))
        fig_trend.update_layout(
            title="📅 Monthly Revenue vs Expenses vs Net Profit",
            barmode="group", height=340,
            legend=dict(orientation="h", y=1.1),
            margin=dict(l=0,r=0,t=50,b=0),
            yaxis=dict(tickprefix="$", tickformat=",.0f"),
            xaxis_title="Month"
        )
        st.plotly_chart(fig_trend, use_container_width=True)

    st.markdown("---")

    # ── Row 2: Invoice Status Donut | Expense by Category ─────────────────────
    col_a, col_b = st.columns(2)

    with col_a:
        st.subheader("🧾 Invoice Status Breakdown")
        if not inv_all.empty:
            status_grp = inv_all.groupby("status").agg(
                count=("id","count"), amount=("total","sum")
            ).reset_index()
            STATUS_CLR = {"Paid":"#22c55e","Pending":"#f59e0b","Overdue":"#ef4444","Cancelled":"#6b7280"}
            colors = [STATUS_CLR.get(s,"#94a3b8") for s in status_grp["status"]]
            fig_donut = go.Figure(go.Pie(
                labels=status_grp["status"],
                values=status_grp["amount"],
                hole=0.55,
                marker=dict(colors=colors),
                textinfo="label+percent",
                hovertemplate="%{label}<br>$%{value:,.0f}<br>%{percent}<extra></extra>"
            ))
            fig_donut.update_layout(
                height=300, margin=dict(l=0,r=0,t=10,b=0),
                annotations=[dict(text=f"${total_rev:,.0f}<br>Total", x=0.5, y=0.5,
                                  font_size=13, showarrow=False)]
            )
            st.plotly_chart(fig_donut, use_container_width=True)
            # Summary table below donut
            status_grp["amount"] = status_grp["amount"].apply(lambda x: f"${x:,.0f}")
            status_grp.columns = ["Status","Count","Amount"]
            st.dataframe(status_grp, use_container_width=True, hide_index=True)

    with col_b:
        st.subheader("📤 Expenses by Category")
        if not exp_all.empty:
            cat_sum = exp_all.groupby("category")["amount"].sum().reset_index().sort_values("amount")
            fig_cat = go.Figure(go.Bar(
                x=cat_sum["amount"], y=cat_sum["category"],
                orientation="h",
                marker=dict(
                    color=cat_sum["amount"],
                    colorscale="Reds",
                    showscale=False
                ),
                text=cat_sum["amount"].apply(lambda x: f"${x:,.0f}"),
                textposition="outside"
            ))
            fig_cat.update_layout(
                height=300, margin=dict(l=0,r=0,t=10,b=0),
                xaxis=dict(tickprefix="$", tickformat=",.0f"),
                yaxis_title=None
            )
            st.plotly_chart(fig_cat, use_container_width=True)

    st.markdown("---")

    # ── Row 3: Top Clients | Monthly Invoiced vs Collected ─────────────────────
    col_c, col_d = st.columns(2)

    with col_c:
        st.subheader("🏆 Top Clients by Revenue")
        if not inv_all.empty:
            paid_inv = inv_all[inv_all["status"] == "Paid"]
            if not paid_inv.empty:
                top_clients = (
                    paid_inv.groupby("client_name")["total"]
                    .sum().reset_index()
                    .sort_values("total", ascending=False)
                    .head(6)
                )
                fig_top = px.bar(
                    top_clients, x="total", y="client_name",
                    orientation="h",
                    color="total",
                    color_continuous_scale="Greens",
                    labels={"total":"Revenue (USD)","client_name":"Client"},
                    height=280
                )
                fig_top.update_coloraxes(showscale=False)
                fig_top.update_layout(
                    margin=dict(l=0,r=0,t=0,b=0),
                    xaxis=dict(tickprefix="$", tickformat=",.0f"),
                    yaxis_title=None
                )
                fig_top.update_traces(
                    text=top_clients["total"].apply(lambda x: f"${x:,.0f}"),
                    textposition="outside"
                )
                st.plotly_chart(fig_top, use_container_width=True)

    with col_d:
        st.subheader("📥 Invoiced vs Collected by Month")
        if not inv_all.empty and "issue_date" in inv_all.columns:
            inv_all["month"] = pd.to_datetime(inv_all["issue_date"]).dt.strftime("%Y-%m")
            inv_month = inv_all.groupby("month").agg(
                invoiced=("total","sum"),
                collected=("total", lambda x: x[inv_all.loc[x.index,"status"]=="Paid"].sum())
            ).reset_index().sort_values("month")

            fig_coll = go.Figure()
            fig_coll.add_bar(x=inv_month["month"], y=inv_month["invoiced"],
                             name="Invoiced", marker_color="#93c5fd")
            fig_coll.add_bar(x=inv_month["month"], y=inv_month["collected"],
                             name="Collected", marker_color="#22c55e")
            fig_coll.update_layout(
                barmode="overlay", height=280,
                margin=dict(l=0,r=0,t=0,b=0),
                legend=dict(orientation="h", y=1.1),
                yaxis=dict(tickprefix="$", tickformat=",.0f")
            )
            st.plotly_chart(fig_coll, use_container_width=True)

    st.markdown("---")

    # ── P&L Summary Table ──────────────────────────────────────────────────────
    st.subheader("📋 " + t('profit_loss') + " — Monthly Summary")
    if not merged.empty:
        tbl = merged.copy()
        tbl["margin"] = tbl.apply(
            lambda r: f"{round(r['profit']/r['revenue']*100,1)}%" if r['revenue'] > 0 else "—", axis=1
        )
        tbl["revenue"]  = tbl["revenue"].apply(lambda x: f"${x:,.0f}")
        tbl["expenses"] = tbl["expenses"].apply(lambda x: f"${x:,.0f}")
        tbl["profit"]   = tbl["profit"].apply(lambda x: f"${x:,.0f}")
        tbl.columns     = ["Month","Revenue","Expenses","Net Profit","Margin"]
        st.dataframe(tbl, use_container_width=True, hide_index=True)

elif page_matches(page, 'marketing_dashboard'):
    st.header(t('marketing_dashboard'))
    st.caption(f"{t('last_updated')}: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    st.markdown("---")

    try:
        campaigns_df = db.fetch_dataframe("SELECT * FROM marketing_campaigns")
        total_campaigns  = len(campaigns_df)
        active_campaigns = len(campaigns_df[campaigns_df["status"] == "Active"]) if not campaigns_df.empty else 0
        total_budget     = campaigns_df["budget"].sum() if not campaigns_df.empty else 0
        total_spent      = campaigns_df["spent"].sum()  if not campaigns_df.empty else 0
    except Exception:
        total_campaigns = active_campaigns = total_budget = total_spent = 0
        campaigns_df = pd.DataFrame()

    try:
        metrics_df = db.fetch_dataframe("SELECT * FROM marketing_campaign_metrics")
        total_impressions  = int(metrics_df["impressions"].sum())  if not metrics_df.empty else 0
        total_clicks       = int(metrics_df["clicks"].sum())       if not metrics_df.empty else 0
        total_conversions  = int(metrics_df["conversions"].sum())  if not metrics_df.empty else 0
        total_revenue      = metrics_df["revenue"].sum()           if not metrics_df.empty else 0
    except Exception:
        total_impressions = total_clicks = total_conversions = total_revenue = 0
        metrics_df = pd.DataFrame()

    try:
        posts_df = db.fetch_dataframe("SELECT * FROM marketing_social_posts")
        total_posts       = len(posts_df)
        scheduled_posts   = len(posts_df[posts_df["status"] == "Scheduled"]) if not posts_df.empty else 0
    except Exception:
        total_posts = scheduled_posts = 0
        posts_df = pd.DataFrame()

    # KPI Cards
    c1, c2, c3, c4 = st.columns(4)
    c1.metric(t('active_campaigns_kpi'), active_campaigns, t('x_total').format(n=total_campaigns))
    c2.metric(t('total_budget_kpi'),     f"${total_budget:,.0f}", t('x_spent').format(n=f"{total_spent:,.0f}"))
    c3.metric(t('conversions'),          total_conversions, t('x_clicks').format(n=total_clicks))
    c4.metric(t('scheduled_posts_kpi'),  scheduled_posts,  t('x_total').format(n=total_posts))

    st.markdown("---")
    col_l, col_r = st.columns(2)

    with col_l:
        st.subheader(t('campaign_performance'))
        if not metrics_df.empty and not campaigns_df.empty:
            merged = metrics_df.merge(campaigns_df[["id","name"]], left_on="campaign_id", right_on="id", how="left")
            grouped = merged.groupby("name")[["impressions","clicks","conversions"]].sum().reset_index()
            fig = px.bar(grouped, x="name", y=["impressions","clicks","conversions"],
                         barmode="group", height=280,
                         color_discrete_map={"impressions":"#3b82f6","clicks":"#22c55e","conversions":"#f59e0b"})
            fig.update_layout(margin=dict(l=0,r=0,t=10,b=0), xaxis_title="", legend_title="")
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info(t('no_campaign_metrics'))

    with col_r:
        st.subheader(t('budget_vs_spent'))
        if not campaigns_df.empty:
            bdf = campaigns_df[campaigns_df["budget"] > 0][["name","budget","spent"]].copy()
            if not bdf.empty:
                fig2 = px.bar(bdf, x="name", y=["budget","spent"], barmode="group", height=280,
                              color_discrete_map={"budget":"#6366f1","spent":"#ef4444"})
                fig2.update_layout(margin=dict(l=0,r=0,t=10,b=0), xaxis_title="", legend_title="")
                st.plotly_chart(fig2, use_container_width=True)
            else:
                st.info(t('no_budget_data'))
        else:
            st.info(t('no_campaigns'))

    st.markdown("---")
    st.subheader(t('recent_campaigns'))
    if not campaigns_df.empty:
        show = campaigns_df[["name","type","status","budget","spent","start_date","end_date"]].copy()
        show.columns = [t('name_lbl'), t('type'), t('status'), t('budget_lbl'), t('spent_lbl'), t('start_date_field'), t('end_date_field')]
        st.dataframe(show, use_container_width=True, hide_index=True)
    else:
        st.info(t('no_campaigns_go'))

elif page_matches(page, 'campaigns'):
    st.header(t('campaigns'))
    user = st.session_state["user"]
    tab_list, tab_new, tab_metrics = st.tabs([t('tab_all_campaigns'), f"➕ {t('tab_new_campaign')}", f"📊 {t('tab_add_metrics')}"])

    with tab_list:
        try:
            df = db.fetch_dataframe("SELECT * FROM marketing_campaigns ORDER BY created_at DESC")
            if df.empty:
                st.info(t('no_campaigns'))
            else:
                status_filter = st.selectbox(t('filter_by_status_lbl'), ["All","Draft","Active","Paused","Completed"], format_func=_topt)
                if status_filter != "All":
                    df = df[df["status"] == status_filter]
                for _, row in df.iterrows():
                    icon = {"Draft":"📝","Active":"🟢","Paused":"🟡","Completed":"✅"}.get(row["status"],"⚪")
                    with st.expander(f"{icon} {row['name']} | {row['type']} | {t('budget_lbl')}: ${row['budget']:,.0f}"):
                        c1,c2,c3 = st.columns(3)
                        c1.metric(t('budget_lbl'),  f"${row['budget']:,.0f}")
                        c2.metric(t('spent_lbl'),   f"${row['spent']:,.0f}")
                        c3.metric(t('status'),  row["status"])
                        st.write(f"**{t('period_lbl')}:** {row['start_date']} → {row['end_date']}")
                        st.write(f"**{t('target_aud_lbl')}:** {row['target_audience']}")
                        st.write(f"**{t('description')}:** {row['description']}")
                        col_s, col_d = st.columns([3,1])
                        with col_s:
                            new_status = st.selectbox(t('change_status'),
                                ["Draft","Active","Paused","Completed"],
                                index=["Draft","Active","Paused","Completed"].index(row["status"]),
                                key=f"cst_{row['id']}", format_func=_topt)
                            if st.button(t('update_btn'), key=f"cupd_{row['id']}"):
                                with db.get_connection() as conn:
                                    conn.execute(text("UPDATE marketing_campaigns SET status=:s WHERE id=:id"), {"s":new_status,"id":row["id"]})
                                    conn.commit()
                                st.success(t('updated_label'))
                                st.rerun()
                        with col_d:
                            if st.button(f"🗑️ {t('delete_action')}", key=f"cdel_{row['id']}"):
                                with db.get_connection() as conn:
                                    conn.execute(text("DELETE FROM marketing_campaigns WHERE id=:id"), {"id":row["id"]})
                                    conn.commit()
                                st.warning(t('deleted_label'))
                                st.rerun()
        except Exception as e:
            st.error(f"Error: {e}")

    with tab_new:
        with st.form("new_campaign_form", clear_on_submit=True):
            name   = st.text_input(t('campaign_name'))
            c_type = st.selectbox(t('type_field'), ["Email","Social Media","SEO","PPC","Content","Event","Other"], format_func=_topt)
            col1,col2 = st.columns(2)
            budget     = col1.number_input(t('budget_field'), min_value=0.0, step=100.0)
            status_new = col2.selectbox(t('status'), ["Draft","Active"], format_func=_topt)
            col3,col4 = st.columns(2)
            start_date = col3.date_input(t('start_date_field'))
            end_date   = col4.date_input(t('end_date_field'))
            audience   = st.text_input(t('target_audience'))
            desc       = st.text_area(t('description'))
            submitted  = st.form_submit_button(t('create_campaign_btn'))
            if submitted:
                if not name.strip():
                    st.error(t('campaign_name_required'))
                else:
                    try:
                        with db.get_connection() as conn:
                            conn.execute(text(
                                "INSERT INTO marketing_campaigns (name,type,status,budget,start_date,end_date,target_audience,description,created_by) "
                                "VALUES (:name,:type,:status,:budget,:start,:end,:audience,:desc,:uid)"
                            ), {"name":name.strip(),"type":c_type,"status":status_new,"budget":budget,
                                "start":start_date,"end":end_date,"audience":audience,"desc":desc,"uid":user["id"]})
                            conn.commit()
                        st.success(t('campaign_created_ok').format(name=name))
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error: {e}")

    with tab_metrics:
        try:
            camps = db.fetch_dataframe("SELECT id, name FROM marketing_campaigns")
            if camps.empty:
                st.info(t('create_campaign_first'))
            else:
                with st.form("metrics_form", clear_on_submit=True):
                    camp_choice = st.selectbox(t('campaigns'), camps["name"].tolist())
                    metric_date = st.date_input(t('date_field'))
                    m1,m2,m3,m4 = st.columns(4)
                    impressions  = m1.number_input(t('impressions'),  min_value=0, step=100)
                    clicks       = m2.number_input(t('clicks'),       min_value=0, step=10)
                    conversions  = m3.number_input(t('conversions'),  min_value=0, step=1)
                    revenue      = m4.number_input(t('revenue_field'),  min_value=0.0, step=10.0)
                    if st.form_submit_button(t('save_metrics')):
                        camp_id = int(camps[camps["name"]==camp_choice]["id"].iloc[0])
                        with db.get_connection() as conn:
                            conn.execute(text(
                                "INSERT INTO marketing_campaign_metrics (campaign_id,metric_date,impressions,clicks,conversions,revenue) "
                                "VALUES (:cid,:date,:imp,:clk,:conv,:rev)"
                            ), {"cid":camp_id,"date":metric_date,"imp":impressions,"clk":clicks,"conv":conversions,"rev":revenue})
                            conn.execute(text("UPDATE marketing_campaigns SET spent=spent+:rev WHERE id=:id"), {"rev":revenue,"id":camp_id})
                            conn.commit()
                        st.success(t('metrics_saved'))
                        st.rerun()
        except Exception as e:
            st.error(f"Error: {e}")

elif page_matches(page, 'marketing_analytics'):
    st.header(t('marketing_analytics'))
    st.markdown("---")

    try:
        metrics_df   = db.fetch_dataframe("SELECT m.*, c.name as campaign_name FROM marketing_campaign_metrics m LEFT JOIN marketing_campaigns c ON m.campaign_id=c.id ORDER BY m.metric_date")
        campaigns_df = db.fetch_dataframe("SELECT * FROM marketing_campaigns")
        competitors_df = db.fetch_dataframe("SELECT * FROM marketing_competitors")
    except Exception:
        metrics_df = campaigns_df = competitors_df = pd.DataFrame()

    tab_perf, tab_audience, tab_comp = st.tabs([f"📈 {t('tab_performance')}", f"👥 {t('tab_customer_behavior')}", f"🏆 {t('tab_competitors')}"])

    with tab_perf:
        if not metrics_df.empty:
            metrics_df["metric_date"] = pd.to_datetime(metrics_df["metric_date"])
            metrics_df["ctr"] = (metrics_df["clicks"] / metrics_df["impressions"].replace(0,1) * 100).round(2)
            metrics_df["cvr"] = (metrics_df["conversions"] / metrics_df["clicks"].replace(0,1) * 100).round(2)

            t1,t2,t3,t4 = st.columns(4)
            t1.metric(t('total_impressions_kpi'), f"{int(metrics_df['impressions'].sum()):,}")
            t2.metric(t('total_clicks_kpi'),      f"{int(metrics_df['clicks'].sum()):,}")
            t3.metric(t('avg_ctr'),               f"{metrics_df['ctr'].mean():.2f}%")
            t4.metric(t('total_revenue_kpi'),     f"${metrics_df['revenue'].sum():,.0f}")

            st.markdown("---")
            camp_options = [t('all')] + metrics_df["campaign_name"].dropna().unique().tolist()
            selected_camp = st.selectbox(t('filter_by_campaign'), camp_options)
            plot_df = metrics_df if selected_camp == t('all') else metrics_df[metrics_df["campaign_name"]==selected_camp]

            ch_l, ch_r = st.columns(2)
            with ch_l:
                fig_imp = px.line(plot_df, x="metric_date", y="impressions",
                                  labels={"impressions": t('impressions_lbl'), "metric_date": t('date_lbl')},
                                  color_discrete_sequence=["#3b82f6"], height=270,
                                  title=t('impressions_over_time'))
                fig_imp.update_layout(margin=dict(l=0,r=0,t=30,b=0))
                st.plotly_chart(fig_imp, use_container_width=True)
            with ch_r:
                fig_cc = px.line(plot_df, x="metric_date", y=["clicks","conversions"],
                                 labels={"value": t('count_lbl'), "metric_date": t('date_lbl'), "variable": t('metric_lbl')},
                                 color_discrete_map={"clicks":"#22c55e","conversions":"#f59e0b"},
                                 height=270, title=t('clicks_conv_over_time'))
                fig_cc.update_layout(margin=dict(l=0,r=0,t=30,b=0))
                st.plotly_chart(fig_cc, use_container_width=True)

            col_l, col_r = st.columns(2)
            with col_l:
                fig2 = px.bar(plot_df, x="metric_date", y="revenue", title=t('daily_revenue'), height=260,
                              color_discrete_sequence=["#6366f1"])
                fig2.update_layout(margin=dict(l=0,r=0,t=30,b=0))
                st.plotly_chart(fig2, use_container_width=True)
            with col_r:
                if not campaigns_df.empty:
                    type_counts = campaigns_df["type"].value_counts().reset_index()
                    type_counts.columns = [t('type'), t('count_lbl')]
                    fig3 = px.pie(type_counts, names=t('type'), values=t('count_lbl'), title=t('campaigns_by_type'), height=260)
                    fig3.update_layout(margin=dict(l=0,r=0,t=30,b=0))
                    st.plotly_chart(fig3, use_container_width=True)
        else:
            st.info(t('no_metrics_campaigns'))

    with tab_audience:
        st.subheader(t('customer_behavior'))
        if not metrics_df.empty:
            col_l, col_r = st.columns(2)
            with col_l:
                ctr_df = metrics_df.groupby("campaign_name")[["clicks","impressions"]].sum().reset_index()
                ctr_df["CTR %"] = (ctr_df["clicks"] / ctr_df["impressions"].replace(0,1) * 100).round(2)
                fig4 = px.bar(ctr_df, x="campaign_name", y="CTR %", title=t('ctr_by_campaign'),
                              color="CTR %", color_continuous_scale="Blues", height=280)
                fig4.update_layout(margin=dict(l=0,r=0,t=30,b=0), xaxis_title="")
                st.plotly_chart(fig4, use_container_width=True)
            with col_r:
                cvr_df = metrics_df.groupby("campaign_name")[["conversions","clicks"]].sum().reset_index()
                cvr_df["CVR %"] = (cvr_df["conversions"] / cvr_df["clicks"].replace(0,1) * 100).round(2)
                fig5 = px.bar(cvr_df, x="campaign_name", y="CVR %", title=t('cvr_by_campaign'),
                              color="CVR %", color_continuous_scale="Greens", height=280)
                fig5.update_layout(margin=dict(l=0,r=0,t=30,b=0), xaxis_title="")
                st.plotly_chart(fig5, use_container_width=True)
        else:
            st.info(t('no_data_available_yet'))

    with tab_comp:
        st.subheader(t('competitor_analysis'))
        col_list, col_add = st.columns([2,1])
        with col_list:
            if not competitors_df.empty:
                for _, row in competitors_df.iterrows():
                    with st.expander(f" {row['name']} — {row['website']}"):
                        st.write(f"🟢 **{t('strengths_field')}:** {row['strengths']}")
                        st.write(f"🔴 **{t('weaknesses_field')}:** {row['weaknesses']}")
                        st.write(f" **{t('notes_desc')}:** {row['notes']}")
                        if st.button(f"🗑️ {t('delete_action')}", key=f"comp_del_{row['id']}"):
                            with db.get_connection() as conn:
                                conn.execute(text("DELETE FROM marketing_competitors WHERE id=:id"), {"id":row["id"]})
                                conn.commit()
                            st.rerun()
            else:
                st.info(t('no_competitors'))
        with col_add:
            with st.form("add_competitor", clear_on_submit=True):
                st.write(f"**{t('add_competitor')}**")
                comp_name = st.text_input(t('name_field'))
                comp_web  = st.text_input(t('website_field'))
                comp_str  = st.text_area(t('strengths_field'),  height=80)
                comp_weak = st.text_area(t('weaknesses_field'), height=80)
                comp_note = st.text_area(t('notes_desc'),      height=60)
                if st.form_submit_button(t('add_record_btn')):
                    if comp_name.strip():
                        with db.get_connection() as conn:
                            conn.execute(text(
                                "INSERT INTO marketing_competitors (name,website,strengths,weaknesses,notes) VALUES (:n,:w,:s,:wk,:nt)"
                            ), {"n":comp_name,"w":comp_web,"s":comp_str,"wk":comp_weak,"nt":comp_note})
                            conn.commit()
                        st.success(t('added_label'))
                        st.rerun()

elif page_matches(page, 'social_media'):
    st.header("📱 " + t('social_media'))
    user = st.session_state["user"]
    tab_feed, tab_new, tab_stats = st.tabs([f"📢 {t('tab_posts')}", f"✏️ {t('tab_new_post')}", f"📊 {t('tab_stats')}"])

    with tab_feed:
        try:
            posts_df = db.fetch_dataframe("SELECT * FROM marketing_social_posts ORDER BY created_at DESC")
            if posts_df.empty:
                st.info(t('no_posts_yet'))
            else:
                plat_filter = st.selectbox(t('platform_field'), ["All","Instagram","Facebook","Twitter","LinkedIn","TikTok"])
                stat_filter = st.selectbox(t('status'), ["All","Draft","Scheduled","Published"], format_func=_topt)
                filtered = posts_df.copy()
                if plat_filter != "All": filtered = filtered[filtered["platform"]==plat_filter]
                if stat_filter != "All": filtered = filtered[filtered["status"]==stat_filter]
                for _, row in filtered.iterrows():
                    plat_icon = {"Instagram":"📸","Facebook":"🅰️","Twitter":"🐦","LinkedIn":"💼","TikTok":"🎵"}.get(row["platform"],"📱")
                    stat_icon = {"Draft":"📝","Scheduled":"🕐","Published":"🟢"}.get(row["status"],"⚪")
                    with st.expander(f"{plat_icon} {row['platform']} | {stat_icon} {row['status']} | {str(row['created_at'])[:10]}"):
                        st.write(row["content"])
                        if row["status"] == "Published":
                            m1,m2,m3 = st.columns(3)
                            m1.metric(f"👍 {t('likes_lbl')}",    row["likes"])
                            m2.metric(f"🔄 {t('shares_lbl')}",   row["shares"])
                            m3.metric(f"💬 {t('comments_lbl')}", row["comments"])
                        if row["status"] in ["Draft","Scheduled"]:
                            col_s, col_d = st.columns([3,1])
                            with col_s:
                                new_stat = st.selectbox(t('status'), ["Draft","Scheduled","Published"],
                                    index=["Draft","Scheduled","Published"].index(row["status"]),
                                    key=f"pst_{row['id']}", format_func=_topt)
                                if st.button(t('update_btn'), key=f"pupd_{row['id']}"):
                                    pub_at = "NOW()" if new_stat == "Published" else "NULL"
                                    with db.get_connection() as conn:
                                        conn.execute(text(f"UPDATE marketing_social_posts SET status=:s, published_at=IF(:s='Published',NOW(),NULL) WHERE id=:id"),
                                            {"s":new_stat,"id":row["id"]})
                                        conn.commit()
                                    st.rerun()
                            with col_d:
                                if st.button("🗑️", key=f"pdel_{row['id']}"):
                                    with db.get_connection() as conn:
                                        conn.execute(text("DELETE FROM marketing_social_posts WHERE id=:id"),{"id":row["id"]})
                                        conn.commit()
                                    st.rerun()
        except Exception as e:
            st.error(f"Error: {e}")

    with tab_new:
        with st.form("new_post_form", clear_on_submit=True):
            platform   = st.selectbox(t('platform_field'), ["Instagram","Facebook","Twitter","LinkedIn","TikTok"])
            content    = st.text_area(t('content_field'), height=150, placeholder=t('write_post_ph'))
            col1,col2 = st.columns(2)
            post_status = col1.selectbox(t('status'), ["Draft","Scheduled"], format_func=_topt)
            sched_at    = col2.date_input(t('schedule_date'))
            try:
                camps = db.fetch_dataframe("SELECT id,name FROM marketing_campaigns WHERE status='Active'")
                camp_names = ["None"] + camps["name"].tolist()
            except Exception:
                camp_names = ["None"]
                camps = pd.DataFrame()
            camp_sel = st.selectbox(t('link_to_campaign'), camp_names)
            if st.form_submit_button(t('save_post')):
                if not content.strip():
                    st.error(t('content_required'))
                else:
                    camp_id = None
                    if camp_sel != "None" and not camps.empty:
                        camp_id = int(camps[camps["name"]==camp_sel]["id"].iloc[0])
                    try:
                        with db.get_connection() as conn:
                            conn.execute(text(
                                "INSERT INTO marketing_social_posts (platform,content,status,scheduled_at,campaign_id,created_by) "
                                "VALUES (:plat,:content,:status,:sched,:camp,:uid)"
                            ), {"plat":platform,"content":content.strip(),"status":post_status,
                                "sched":sched_at,"camp":camp_id,"uid":user["id"]})
                            conn.commit()
                        st.success(t('post_saved'))
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error: {e}")

    with tab_stats:
        try:
            posts_df = db.fetch_dataframe("SELECT * FROM marketing_social_posts WHERE status='Published'")
            if posts_df.empty:
                st.info(t('no_published_posts'))
            else:
                col_l, col_r = st.columns(2)
                with col_l:
                    by_plat = posts_df.groupby("platform")[["likes","shares","comments"]].sum().reset_index()
                    fig = px.bar(by_plat, x="platform", y=["likes","shares","comments"],
                                 barmode="group", title=t('engagement_by_platform'), height=280,
                                 color_discrete_map={"likes":"#ec4899","shares":"#3b82f6","comments":"#f59e0b"})
                    fig.update_layout(margin=dict(l=0,r=0,t=30,b=0), xaxis_title="")
                    st.plotly_chart(fig, use_container_width=True)
                with col_r:
                    plat_counts = posts_df["platform"].value_counts().reset_index()
                    plat_counts.columns = [t('platform_lbl'), t('posts_lbl')]
                    fig2 = px.pie(plat_counts, names=t('platform_lbl'), values=t('posts_lbl'),
                                  title=t('posts_distribution'), height=280)
                    fig2.update_layout(margin=dict(l=0,r=0,t=30,b=0))
                    st.plotly_chart(fig2, use_container_width=True)

                st.subheader(t('update_engagement'))
                sel_post = st.selectbox(t('select_post'), posts_df.apply(lambda r: f"[{r['platform']}] {str(r['content'])[:60]}", axis=1).tolist())
                post_idx = posts_df.index[posts_df.apply(lambda r: f"[{r['platform']}] {str(r['content'])[:60]}", axis=1) == sel_post].tolist()
                if post_idx:
                    row = posts_df.loc[post_idx[0]]
                    e1,e2,e3 = st.columns(3)
                    new_likes    = e1.number_input(t('likes_lbl'),    min_value=0, value=int(row["likes"]))
                    new_shares   = e2.number_input(t('shares_lbl'),   min_value=0, value=int(row["shares"]))
                    new_comments = e3.number_input(t('comments_lbl'), min_value=0, value=int(row["comments"]))
                    if st.button(t('update_engagement_btn')):
                        with db.get_connection() as conn:
                            conn.execute(text("UPDATE marketing_social_posts SET likes=:l,shares=:s,comments=:c WHERE id=:id"),
                                {"l":new_likes,"s":new_shares,"c":new_comments,"id":int(row["id"])})
                            conn.commit()
                        st.success(t('updated_label'))
                        st.rerun()
        except Exception as e:
            st.error(f"Error: {e}")

elif page_matches(page, 'it_dashboard'):
    import time as _time, socket
    try:
        import psutil
    except Exception:
        psutil = None

    cpu_val = ram_val = disk_val = resp_val = 0

    if psutil is None:
        st.warning(t('psutil_missing'))
        server_stats = []
        logs = []
    else:
        # قراءة مباشرة من psutil للقيم الحالية
        cpu_val  = psutil.cpu_percent(interval=0.5)
        ram_val  = psutil.virtual_memory().percent
        disk_val = psutil.disk_usage("/").percent

        # قياس زمن استجابة قاعدة البيانات
        t0 = _time.time()
        try:
            with db.get_connection():
                pass
        except Exception:
            pass
        resp_val = round((_time.time() - t0) * 1000, 1)

        # حفظ snapshot في MySQL — فوراً في أول زيارة، ثم كل 30 ثانية
        if "it_last_snapshot" not in st.session_state or (_time.time() - st.session_state.get("it_last_snapshot", 0)) > 30:
            nio = psutil.net_io_counters()
            db.insert_performance_snapshot(cpu_val, ram_val, disk_val,
                                           round(nio.bytes_sent/1024/1024, 2),
                                           round(nio.bytes_recv/1024/1024, 2),
                                           resp_val)
            hostname = socket.gethostname()
            try:
                ip = socket.gethostbyname(hostname)
            except Exception:
                ip = "127.0.0.1"
            uptime_seconds = int(_time.time() - psutil.boot_time())
            status = "Online" if cpu_val < 90 and ram_val < 90 else "Degraded"
            db.insert_server_snapshot(hostname, status, cpu_val, ram_val, disk_val, uptime_seconds, ip)
            st.session_state["it_last_snapshot"] = _time.time()

        server_stats = db.get_latest_server_status()
        logs         = db.get_recent_activity_log(8)

    try:
        suspicious_count = int(db.fetch_dataframe(
            "SELECT COUNT(*) as cnt FROM it_security_events WHERE event_type='failed_login' AND timestamp >= NOW() - INTERVAL 7 DAY"
        ).iloc[0]["cnt"])
    except Exception:
        suspicious_count = 0

    try:
        users_df      = db.fetch_dataframe("SELECT id, created_at FROM users")
        num_users     = len(users_df)
        num_new_users = int((users_df["created_at"] >= (pd.Timestamp.now() - pd.Timedelta(days=30))).sum()) if "created_at" in users_df.columns else 0
    except Exception:
        num_users, num_new_users = 0, 0

    try:
        open_tickets = int(db.fetch_dataframe("SELECT COUNT(*) as cnt FROM support_tickets WHERE status != 'Closed'").iloc[0]["cnt"])
    except Exception:
        open_tickets = 0

    st.header("💻 " + t('it_dashboard'))
    st.caption(f"{t('last_updated')}: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    st.markdown("---")

    c1, c2, c3, c4 = st.columns(4)
    c1.metric(t('cpu_usage'),        f"{cpu_val}%")
    c2.metric(t('ram_usage'),        f"{ram_val}%")
    c3.metric(t('db_response_time'), f"{resp_val} ms")
    c4.metric(t('open_tickets'),     open_tickets)

    st.markdown("---")
    perf_history = db.get_performance_history(20)
    col_left, col_right = st.columns(2)

    with col_left:
        st.subheader(t('system_performance'))
        if perf_history:
            ph_df = pd.DataFrame(perf_history)
            ph_df["timestamp"] = pd.to_datetime(ph_df["timestamp"])
            fig_perf = px.line(ph_df, x="timestamp", y=["cpu", "ram", "disk"],
                               labels={"value": "%", "timestamp": "Time", "variable": "Metric"},
                               color_discrete_map={"cpu": "#3b82f6", "ram": "#22c55e", "disk": "#f59e0b"})
            fig_perf.update_layout(margin=dict(l=0, r=0, t=30, b=0), height=260)
            st.plotly_chart(fig_perf, use_container_width=True)
        else:
            st.info(t('no_perf_data'))

    with col_right:
        online_count  = sum(1 for s in server_stats if s.get("is_online"))
        offline_count = len(server_stats) - online_count
        st.subheader(f"Devices ({online_count} 🟢 Online / {offline_count} 🔴 Offline)")
        if server_stats:
            for s in server_stats:
                cpu_s    = s.get("cpu", 0) or 0
                name_s   = s.get("server_name", "Unknown")
                is_online = s.get("is_online", False)
                last_seen = s.get("last_seen_seconds", 0)
                ip_s     = s.get("ip_address", "")

                if is_online:
                    uptime_h = round((s.get("uptime_seconds") or 0) / 3600, 1)
                    cpu_color = "green" if cpu_s < 70 else "orange" if cpu_s < 90 else "red"
                    st.markdown(f"🟢 **{name_s}** `{ip_s}`")
                    st.markdown(f"&nbsp;&nbsp;&nbsp;CPU: `{cpu_s:.1f}%` | RAM: `{s.get('ram',0):.1f}%` | Disk: `{s.get('disk',0):.1f}%` | Uptime: `{uptime_h}h`")
                    st.progress(min(int(cpu_s), 100))
                else:
                    mins_ago = last_seen // 60
                    st.markdown(f"🔴 **{name_s}** `{ip_s}` — *Offline since {mins_ago} min ago*")
                st.markdown("---")
        else:
            st.info(t('no_devices_registered'))

    st.markdown("---")
    col_u, col_sec = st.columns(2)
    with col_u:
        st.subheader(t('users'))
        st.metric(t('total_users'),    num_users)
        st.metric(t('new_this_month'), num_new_users)
    with col_sec:
        st.subheader(t('security_last7'))
        st.metric(t('failed_logins'), suspicious_count)

    st.markdown("---")
    st.subheader(t('recent_activity_log'))
    if logs:
        log_df = pd.DataFrame(logs)
        show_cols = [c for c in ["timestamp", "event_type", "description", "user_email"] if c in log_df.columns]
        log_df = log_df[show_cols]
        log_df.columns = ["Time", "Event", "Description", "User"][:len(show_cols)]
        st.dataframe(log_df, use_container_width=True, hide_index=True)
    else:
        st.info(t('no_activity'))

elif page_matches(page, 'system_management'):
    try:
        import psutil
    except Exception:
        psutil = None
    import time as _time
    st.header("💻 " + t('system_management'))
    st.markdown("---")

    if psutil is None:
        st.error(t('no_psutil'))
        st.stop()

    cpu_live  = psutil.cpu_percent(interval=0.5)
    ram_live  = psutil.virtual_memory()
    disk_live = psutil.disk_usage("/")
    net_live  = psutil.net_io_counters()

    st.subheader(t('live_metrics'))
    m1, m2, m3 = st.columns(3)
    m1.metric(t('cpu_usage'), f"{cpu_live:.1f}%")
    m2.metric(t('ram_usage'),  f"{ram_live.percent:.1f}%",
              f"{ram_live.used // (1024**3):.1f} GB / {ram_live.total // (1024**3):.1f} GB")
    m3.metric("Disk Used", f"{disk_live.percent:.1f}%",
              f"{disk_live.used // (1024**3):.1f} GB / {disk_live.total // (1024**3):.1f} GB")

    st.markdown("---")
    perf_history = db.get_performance_history(30)
    if perf_history:
        st.subheader(t('perf_history'))
        ph_df = pd.DataFrame(perf_history)
        ph_df["timestamp"] = pd.to_datetime(ph_df["timestamp"])
        fig_area = px.area(ph_df, x="timestamp", y=["cpu", "ram", "disk"],
                           labels={"value": "%", "timestamp": "Time", "variable": "Metric"},
                           color_discrete_map={"cpu": "#3b82f6", "ram": "#22c55e", "disk": "#f59e0b"})
        fig_area.update_layout(margin=dict(l=0, r=0, t=30, b=0), height=300)
        st.plotly_chart(fig_area, use_container_width=True)

    st.markdown("---")
    server_stats = db.get_latest_server_status()
    st.subheader(t('server_status'))
    if server_stats:
        srv_df = pd.DataFrame(server_stats)
        cols_to_show = [c for c in ["server_name","status","cpu","ram","disk","uptime_seconds","ip_address","timestamp"] if c in srv_df.columns]
        srv_df = srv_df[cols_to_show].copy()
        if "uptime_seconds" in srv_df.columns:
            srv_df["uptime_hours"] = (srv_df["uptime_seconds"] / 3600).round(1)
            srv_df = srv_df.drop(columns=["uptime_seconds"])
        srv_df.columns = [c.replace("_", " ").title() for c in srv_df.columns]
        st.dataframe(srv_df, use_container_width=True, hide_index=True)
    else:
        st.info(t('no_server_data_yet'))

    st.markdown("---")
    st.subheader(t('network_io'))
    n1, n2 = st.columns(2)
    n1.metric(t('total_sent'),     f"{net_live.bytes_sent / (1024**3):.2f} GB")
    n2.metric(t('total_received'), f"{net_live.bytes_recv / (1024**3):.2f} GB")

elif page_matches(page, 'support_tickets'):
    st.header("🎫 " + t('support_tickets'))
    user = st.session_state["user"]
    tab1, tab2 = st.tabs([f"➕ {t('tab_new_ticket')}", t('tab_all_tickets')])

    with tab1:
        with st.form("new_ticket_form", clear_on_submit=True):
            t_title = st.text_input(t('ticket_title_lbl'))
            t_desc  = st.text_area(t('ticket_desc_lbl'))
            submitted = st.form_submit_button(t('submit_ticket_btn'))
            if submitted:
                if not t_title.strip() or not t_desc.strip():
                    st.error(t('fill_all_fields'))
                else:
                    try:
                        with db.get_connection() as conn:
                            conn.execute(
                                text("INSERT INTO support_tickets (user_id, title, description, status, created_at) VALUES (:uid, :title, :desc, 'Open', NOW())"),
                                {"uid": user["id"], "title": t_title.strip(), "desc": t_desc.strip()}
                            )
                            conn.commit()
                        db.insert_activity_log("Ticket Created", f"New ticket: {t_title.strip()}", user.get("email", ""))
                        st.success(t('ticket_submitted'))
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error: {e}")

    with tab2:
        try:
            tickets_df = db.fetch_dataframe(
                "SELECT t.id, t.title, t.description, t.status, t.created_at FROM support_tickets t ORDER BY t.created_at DESC"
            )
            if tickets_df.empty:
                st.info(t('no_tickets_found'))
            else:
                status_filter = st.selectbox(t('filter_by_status'), ["All", "Open", "In Progress", "Closed"], format_func=_topt)
                if status_filter != "All":
                    tickets_df = tickets_df[tickets_df["status"] == status_filter]
                for _, row in tickets_df.iterrows():
                    status_icon = {"Open": "🔴", "In Progress": "🟡", "Closed": "✅"}.get(row["status"], "⚪")
                    with st.expander(f"{status_icon} [{row['id']}] {row['title']} — {row['created_at']}"):
                        st.write(f"**Description:** {row['description']}")
                        col_s, col_d = st.columns([3, 1])
                        with col_s:
                            new_status = st.selectbox(
                                t('change_status_lbl'),
                                ["Open", "In Progress", "Closed"],
                                index=["Open", "In Progress", "Closed"].index(row["status"]),
                                key=f"sel_{row['id']}"
                            , format_func=_topt)
                            if st.button(t('update_btn'), key=f"upd_{row['id']}"):
                                if db.update_ticket_status(row["id"], new_status):
                                    db.insert_activity_log("Ticket Updated", f"Ticket #{row['id']} → {new_status}", user.get("email", ""))
                                    st.success(t('status_updated'))
                                    st.rerun()
                        with col_d:
                            if st.button("🗑️ Delete", key=f"del_{row['id']}"):
                                if db.delete_ticket(row["id"]):
                                    db.insert_activity_log("Ticket Deleted", f"Ticket #{row['id']} deleted", user.get("email", ""))
                                    st.warning(t('ticket_deleted'))
                                    st.rerun()
        except Exception as e:
            st.error(f"Error loading tickets: {e}")

elif page_matches(page, 'security'):
    st.header("🔒 " + t('security'))
    st.markdown("---")
    tab_events, tab_users = st.tabs([f"🚨 {t('tab_security_events')}", f"👥 {t('tab_user_activity')}"])

    with tab_events:
        st.subheader(t('security_events_log'))
        events = db.get_all_security_events(100)
        if events:
            ev_df = pd.DataFrame(events)
            show_cols = [c for c in ["timestamp", "event_type", "email", "description", "ip_address"] if c in ev_df.columns]
            ev_df = ev_df[show_cols].copy()
            ev_df.columns = [c.replace("_", " ").title() for c in show_cols]
            type_col = "Event Type" if "Event Type" in ev_df.columns else ev_df.columns[1]
            filter_type = st.selectbox(t('filter_event_type'), ["All"] + list(ev_df[type_col].dropna().unique()))
            if filter_type != "All":
                ev_df = ev_df[ev_df[type_col] == filter_type]
            st.dataframe(ev_df, use_container_width=True, hide_index=True)
            if not ev_df.empty:
                count_df = ev_df.groupby(type_col).size().reset_index(name=t('count_lbl'))
                fig_sec = px.bar(count_df, x=type_col, y=t('count_lbl'), color=type_col,
                                 title=t('events_by_type'), height=300)
                fig_sec.update_layout(margin=dict(l=0, r=0, t=40, b=0), showlegend=False)
                st.plotly_chart(fig_sec, use_container_width=True)
        else:
            st.info(t('no_security_events_yet'))

    with tab_users:
        st.subheader(t('registered_users'))
        try:
            users_full = db.fetch_dataframe("SELECT id, email, role, created_at FROM users ORDER BY created_at DESC")
            if not users_full.empty:
                st.dataframe(users_full, use_container_width=True, hide_index=True)
                role_counts = users_full["role"].value_counts().reset_index()
                role_counts.columns = [t('role_col'), t('count_lbl')]
                fig_roles = px.pie(role_counts, names=t('role_col'), values=t('count_lbl'), title=t('users_by_role'), height=300)
                fig_roles.update_layout(margin=dict(l=0, r=0, t=40, b=0))
                st.plotly_chart(fig_roles, use_container_width=True)
            else:
                st.info(t('no_users_found_msg'))
        except Exception as e:
            st.error(f"Error loading users: {e}")

elif page_matches(page, 'logistics_dashboard'):
    st.header(t('logistics_dashboard'))
    st.caption(f"{t('last_updated')}: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    st.markdown("---")

    summary = db.get_logistics_summary()

    k1,k2,k3,k4,k5 = st.columns(5)
    k1.metric(t('total_shipments'),   summary["total"])
    k2.metric(t('in_transit'),        summary["in_transit"])
    k3.metric(t('delivered_month'), summary["delivered_this_month"])
    k4.metric(t('pending_customs'), summary["pending"])
    k5.metric(t('overdue'),           summary["overdue"],
              delta=t('needs_action') if summary["overdue"] > 0 else None,
              delta_color="inverse" if summary["overdue"] > 0 else "normal")

    st.markdown("---")
    col_l, col_r = st.columns(2)

    with col_l:
        df_all = db.get_shipments_logistics()
        if not df_all.empty:
            sc = df_all["status"].value_counts().reset_index()
            sc.columns = [t('status'), t('count_lbl')]
            color_map = {"Delivered":"#22c55e","In Transit":"#3b82f6","Pending":"#94a3b8",
                         "Customs":"#f59e0b","Overdue":"#ef4444","Confirmed":"#6366f1"}
            fig_s = px.pie(sc, names=t('status'), values=t('count_lbl'), color=t('status'),
                           color_discrete_map=color_map, title=t('shipments_by_status'), height=300)
            fig_s.update_layout(margin=dict(l=0,r=0,t=30,b=0))
            st.plotly_chart(fig_s, use_container_width=True)

    with col_r:
        vol_df = db.get_monthly_shipment_volume()
        if not vol_df.empty:
            fig_v = px.bar(vol_df, x="month", y="count", color="type",
                           labels={"count": t('shipments_lbl'), "month": t('month_lbl'), "type": t('type')},
                           color_discrete_map={"Export":"#22c55e","Import":"#3b82f6"},
                           barmode="group", title=t('monthly_ship_vol'), height=300)
            fig_v.update_layout(margin=dict(l=0,r=0,t=30,b=0))
            st.plotly_chart(fig_v, use_container_width=True)

    st.markdown("---")
    st.subheader(t('recent_shipments'))
    df_all = db.get_shipments_logistics()
    if not df_all.empty:
        disp_cols = ["shipment_number","shipment_type","freight_mode","origin","destination",
                     "carrier_name","status","departure_date","expected_arrival"]
        disp = df_all[disp_cols].head(10).copy()
        disp.columns = [t('shipment_hash'), t('type'), t('mode_lbl'), t('origin_lbl'), t('destination_lbl'), t('carrier_lbl'), t('status'), t('departure_lbl'), t('eta_lbl')]
        st.dataframe(disp, use_container_width=True, hide_index=True)

elif page_matches(page, 'shipments'):
    st.header("🚢 " + t('shipments'))
    st.markdown("---")
    tab_all, tab_new, tab_track = st.tabs([t('tab_all_shipments'), t('tab_new_shipment'), t('tab_tracking_timeline')])

    STATUS_OPTS = ["All","Pending","Confirmed","In Transit","Customs","Delivered","Overdue"]
    MODE_OPTS   = ["All","Sea","Air","Land"]
    TYPE_OPTS   = ["All","Export","Import"]
    STATUS_ICON = {"Delivered":"🟢","In Transit":"🔵","Pending":"⚪","Customs":"🟡",
                   "Overdue":"🔴","Confirmed":"🟣"}

    with tab_all:
        f1,f2,f3 = st.columns(3)
        sf = f1.selectbox(t('status'),   STATUS_OPTS, key="sh_sf", format_func=_topt)
        mf = f2.selectbox(t('mode_lbl'), MODE_OPTS,   key="sh_mf", format_func=_topt)
        tf = f3.selectbox(t('type_field'), TYPE_OPTS, key="sh_tf", format_func=_topt)
        df_sh = db.get_shipments_logistics(
            sf if sf!="All" else None,
            mf if mf!="All" else None,
            tf if tf!="All" else None
        )
        if not df_sh.empty:
            st.caption(f"{len(df_sh)} shipment(s) found")
            for _, row in df_sh.iterrows():
                icon = STATUS_ICON.get(str(row["status"]),"⚪")
                mode_icon = "✈️" if row["freight_mode"]=="Air" else "🚢" if row["freight_mode"]=="Sea" else "🚛"
                with st.expander(f"{icon} {row['shipment_number']} — {row['origin']} → {row['destination']}  |  {mode_icon} {row['freight_mode']}  |  {row['status']}"):
                    c1,c2,c3,c4 = st.columns(4)
                    c1.write(f"**Type:** {row['shipment_type']}")
                    c2.write(f"**Carrier:** {row.get('carrier_name','Not assigned') or 'Not assigned'}")
                    c3.write(f"**Container:** {row.get('container_type','—') or '—'}")
                    c4.write(f"**Incoterms:** {row.get('incoterms','—') or '—'}")
                    c5,c6,c7,c8 = st.columns(4)
                    c5.write(f"**Departure:** {row['departure_date']}")
                    c6.write(f"**ETA:** {row['expected_arrival']}")
                    c7.write(f"**Weight:** {row['total_weight']:,.0f} kg")
                    c8.write(f"**Value:** ${row['total_value']:,.0f}")
                    if row.get("notes"):
                        st.info(row["notes"])

                    # Cargo items
                    cargo_df = db.get_cargo_items_for_shipment(row["id"])
                    if not cargo_df.empty:
                        st.markdown("**Cargo Items:**")
                        cargo_df.columns = ["Item","Description","Qty","Unit","Weight(kg)","Value(USD)","HS Code"]
                        st.dataframe(cargo_df, use_container_width=True, hide_index=True)

                    # Tracking history
                    track_df = db.get_tracking_history(row["id"])
                    if not track_df.empty:
                        st.markdown("**Tracking History:**")
                        for _, tr in track_df.iterrows():
                            st.markdown(f"- **{tr['update_date']}** — {tr['location']} → `{tr['status']}` {('• '+str(tr['notes'])) if tr.get('notes') else ''}")

                    st.markdown("---")
                    b1,b2,b3,b4,b5 = st.columns(5)
                    new_statuses = [s for s in ["Confirmed","In Transit","Customs","Delivered","Overdue"] if s != row["status"]]
                    sel_status = b1.selectbox(t('change_status_lbl2'), new_statuses, key=f"sel_st_{row['id']}")
                    if b2.button("Update", key=f"upd_st_{row['id']}", type="primary"):
                        db.update_shipment_status_logistics(row["id"], sel_status)
                        st.rerun()

                    with b3.popover("Add Tracking Update"):
                        loc_in  = st.text_input(t('location'), key=f"loc_{row['id']}")
                        stat_in = st.selectbox(t('status'), ["Departed","In Transit","Arrived","Customs","Delivered"], key=f"tst_{row['id']}")
                        note_in = st.text_area(t('notes_desc'), height=60, key=f"tnote_{row['id']}")
                        if st.button(t('save_update'), key=f"savt_{row['id']}", type="primary"):
                            uid = st.session_state.get("user_id",1)
                            db.add_tracking_update_logistics(row["id"], loc_in, stat_in, note_in, uid)
                            st.rerun()

                    with b4.popover("➕ Add Cargo"):
                        ci_name  = st.text_input("Item Name *", key=f"ci_name_{row['id']}")
                        ci_desc  = st.text_input("Description", key=f"ci_desc_{row['id']}")
                        ci_col1, ci_col2 = st.columns(2)
                        ci_qty   = ci_col1.number_input("Quantity *", min_value=1, value=1, key=f"ci_qty_{row['id']}")
                        ci_unit  = ci_col2.selectbox("Unit", ["pcs","kg","ton","m³","carton","pallet","box"], key=f"ci_unit_{row['id']}")
                        ci_col3, ci_col4 = st.columns(2)
                        ci_wt    = ci_col3.number_input("Weight (kg)", min_value=0.0, value=0.0, step=0.1, key=f"ci_wt_{row['id']}")
                        ci_val   = ci_col4.number_input("Value (USD)", min_value=0.0, value=0.0, step=1.0, key=f"ci_val_{row['id']}")
                        ci_hs    = st.text_input("HS Code", key=f"ci_hs_{row['id']}")
                        if st.button("Save Cargo Item", key=f"ci_save_{row['id']}", type="primary"):
                            if ci_name.strip():
                                ok = db.add_cargo_item(
                                    shipment_id=row["id"],
                                    item_name=ci_name.strip(),
                                    description=ci_desc.strip(),
                                    quantity=int(ci_qty),
                                    unit=ci_unit,
                                    weight=float(ci_wt),
                                    value=float(ci_val),
                                    hs_code=ci_hs.strip()
                                )
                                if ok:
                                    st.success("Cargo item added!")
                                    st.rerun()
                                else:
                                    st.error("Failed to save cargo item.")
                            else:
                                st.warning("Item name is required.")
        else:
            st.info(t('no_shipments_found'))

    with tab_new:
        st.subheader(t('create_new_shipment'))
        clients_df = db.fetch_dataframe("SELECT id, email FROM users WHERE role='client' ORDER BY email")

        COUNTRIES = [
            "Afghanistan","Albania","Algeria","Angola","Argentina","Armenia","Australia","Austria",
            "Azerbaijan","Bahrain","Bangladesh","Belarus","Belgium","Bolivia","Bosnia and Herzegovina",
            "Brazil","Bulgaria","Cambodia","Cameroon","Canada","Chile","China","Colombia","Croatia",
            "Cyprus","Czech Republic","Denmark","Ecuador","Egypt","Ethiopia","Finland","France",
            "Georgia","Germany","Ghana","Greece","Hungary","India","Indonesia","Iran","Iraq","Ireland",
            "Israel","Italy","Japan","Jordan","Kazakhstan","Kenya","Kuwait","Lebanon","Libya","Malaysia",
            "Mexico","Morocco","Netherlands","New Zealand","Nigeria","Norway","Oman","Pakistan","Peru",
            "Philippines","Poland","Portugal","Qatar","Romania","Russia","Saudi Arabia","Senegal",
            "Serbia","Singapore","Slovakia","South Africa","South Korea","Spain","Sri Lanka","Sudan",
            "Sweden","Switzerland","Syria","Taiwan","Thailand","Tunisia","Turkey","UAE","Ukraine",
            "United Kingdom","United States","Uzbekistan","Venezuela","Vietnam","Yemen",
        ]

        with st.form("new_shipment_form", clear_on_submit=True):
            ns1,ns2 = st.columns(2)
            sh_type  = ns1.selectbox(t('type') + " *", ["Export","Import"], format_func=_topt)
            sh_mode  = ns2.selectbox(t('freight_mode'), ["Sea","Air","Land"], format_func=_topt)
            ns3,ns4 = st.columns(2)
            origin   = ns3.selectbox(t('origin_country'), COUNTRIES, index=COUNTRIES.index("Turkey"))
            dest     = ns4.selectbox(t('dest_country'), COUNTRIES, index=COUNTRIES.index("UAE"))
            ns5,ns6 = st.columns(2)
            dep_date = ns5.date_input(t('departure_date'))
            eta_date = ns6.date_input(t('expected_arrival_field'))
            ns7,ns8 = st.columns(2)
            carrier  = ns7.text_input(t('carrier_name_field'))
            cont_type= ns8.selectbox(t('container_type'), ["FCL 20GP","FCL 40HC","LCL","Air Pallet","Bulk","Reefer FCL","Other"])
            ns9,ns10 = st.columns(2)
            incoterm = ns9.selectbox(t('incoterms_field'), ["FOB","CIF","EXW","DDP","CFR","DAP","FCA"])
            client_opts = {r["email"]: r["id"] for _,r in clients_df.iterrows()} if not clients_df.empty else {}
            sel_client = ns10.selectbox(t('client'), list(client_opts.keys()) if client_opts else ["No clients"])
            ns11,ns12 = st.columns(2)
            weight   = ns11.number_input(t('total_weight_field'), min_value=0.0, step=100.0)
            value    = ns12.number_input(t('total_value_field'), min_value=0.0, step=1000.0)
            notes_ns = st.text_area(t('notes_desc'), height=70)
            if st.form_submit_button(t('create_shipment_btn'), type="primary"):
                if origin == dest:
                    st.error(t('origin_dest_same'))
                else:
                    sn  = f"SHP-{datetime.now().strftime('%Y-%m%d%H%M%S')}"
                    cid = client_opts.get(sel_client, list(client_opts.values())[0]) if client_opts else 6
                    ok  = db.create_logistics_shipment(
                        sn, cid, sh_type, sh_mode, origin, dest, carrier or None,
                        cont_type, incoterm, str(dep_date), str(eta_date), weight, value, notes_ns
                    )
                    if ok:
                        st.success(f"Shipment {sn} created.")
                    else:
                        st.error(t('failed_create_shipment_err'))

    with tab_track:
        st.subheader(t('shipment_tracking'))
        df_sh_t = db.get_shipments_logistics()
        if not df_sh_t.empty:
            sh_opts = {f"{r['shipment_number']} — {r['origin']} → {r['destination']}": r["id"]
                       for _,r in df_sh_t.iterrows()}
            sel_sh = st.selectbox(t('select_shipment'), list(sh_opts.keys()))
            sh_id  = sh_opts[sel_sh]
            track_df = db.get_tracking_history(sh_id)
            row_sh   = df_sh_t[df_sh_t["id"]==sh_id].iloc[0]

            tc1,tc2,tc3 = st.columns(3)
            tc1.metric("Status",   row_sh["status"])
            tc2.metric("Mode",     row_sh["freight_mode"])
            tc3.metric("Carrier",  str(row_sh.get("carrier_name","—") or "—"))

            st.markdown("---")
            if not track_df.empty:
                for i, tr in track_df.iterrows():
                    step_icon = "🟢" if str(tr["status"]) in ["Delivered","Arrived"] else "🔵" if str(tr["status"])=="In Transit" else "🟡" if str(tr["status"])=="Customs" else "⚫"
                    st.markdown(f"{step_icon} **{tr['update_date']}** — **{tr['location']}**")
                    st.markdown(f"&nbsp;&nbsp;&nbsp;&nbsp;`{tr['status']}` {('— '+str(tr['notes'])) if tr.get('notes') else ''}")
            else:
                st.info(t('no_tracking_ship'))
        else:
            st.info(t('no_ship_available'))

elif page_matches(page, 'routes'):
    st.header("🗺️ " + t('routes'))
    st.markdown("---")
    tab_list, tab_add = st.tabs([t('tab_active_routes'), t('tab_add_route')])

    with tab_list:
        rf1, rf2 = st.columns(2)
        mode_rf = rf1.selectbox(t('filter_mode'), ["All","Sea","Air","Land"], key="rt_mode", format_func=_topt)
        routes_df = db.get_all_routes()
        if not routes_df.empty:
            if mode_rf != "All":
                routes_df = routes_df[routes_df["freight_mode"]==mode_rf]
            for _, r in routes_df.iterrows():
                mode_icon = "✈️" if r["freight_mode"]=="Air" else "🚢" if r["freight_mode"]=="Sea" else "🚛"
                with st.expander(f"{mode_icon} {r['route_name']}  |  {r['transit_days']} days  |  {r['frequency']}"):
                    rc1,rc2,rc3 = st.columns(3)
                    rc1.write(f"**Origin:** {r['origin_port']} ({r['origin_country']})")
                    rc2.write(f"**Destination:** {r['destination_port']} ({r['destination_country']})")
                    rc3.write(f"**Carrier:** {r.get('carrier','—') or '—'}")
                    rc4,rc5 = st.columns(2)
                    rc4.write(f"**Transit:** {r['transit_days']} days")
                    rc5.write(f"**Frequency:** {r['frequency']}")
                    if r.get("notes"):
                        st.caption(r["notes"])
        else:
            st.info(t('no_routes_found'))

    with tab_add:
        st.subheader(t('add_new_route'))
        COUNTRIES_RT = [
            "Afghanistan","Albania","Algeria","Angola","Argentina","Armenia","Australia","Austria",
            "Azerbaijan","Bahrain","Bangladesh","Belarus","Belgium","Bolivia","Bosnia and Herzegovina",
            "Brazil","Bulgaria","Cambodia","Cameroon","Canada","Chile","China","Colombia","Croatia",
            "Cyprus","Czech Republic","Denmark","Ecuador","Egypt","Ethiopia","Finland","France",
            "Georgia","Germany","Ghana","Greece","Hungary","India","Indonesia","Iran","Iraq","Ireland",
            "Israel","Italy","Japan","Jordan","Kazakhstan","Kenya","Kuwait","Lebanon","Libya","Malaysia",
            "Mexico","Morocco","Netherlands","New Zealand","Nigeria","Norway","Oman","Pakistan","Peru",
            "Philippines","Poland","Portugal","Qatar","Romania","Russia","Saudi Arabia","Senegal",
            "Serbia","Singapore","Slovakia","South Africa","South Korea","Spain","Sri Lanka","Sudan",
            "Sweden","Switzerland","Syria","Taiwan","Thailand","Tunisia","Turkey","UAE","Ukraine",
            "United Kingdom","United States","Uzbekistan","Venezuela","Vietnam","Yemen",
        ]
        with st.form("add_route_form", clear_on_submit=True):
            ar1,ar2 = st.columns(2)
            route_name  = ar1.text_input(t('route_name'))
            freight_mod = ar2.selectbox(t('freight_mode'), ["Sea","Air","Land"], format_func=_topt)
            ar3,ar4 = st.columns(2)
            orig_port   = ar3.text_input(t('origin_port'))
            orig_ctry   = ar4.selectbox(t('origin_country'), COUNTRIES_RT, index=COUNTRIES_RT.index("Turkey"))
            ar5,ar6 = st.columns(2)
            dest_port   = ar5.text_input(t('dest_port'))
            dest_ctry   = ar6.selectbox(t('dest_country'), COUNTRIES_RT, index=COUNTRIES_RT.index("UAE"))
            ar7,ar8,ar9 = st.columns(3)
            carrier_r   = ar7.text_input(t('carrier_name_field'))
            transit_d   = ar8.number_input(t('transit_days'), min_value=1, max_value=90, value=14)
            freq_r      = ar9.text_input(t('frequency'))
            notes_r     = st.text_area(t('notes_desc'), height=70)
            if st.form_submit_button(t('add_route_btn'), type="primary"):
                if not route_name or not orig_port or not dest_port:
                    st.error(t('route_required'))
                else:
                    ok = db.add_route(route_name, orig_port, orig_ctry, dest_port, dest_ctry,
                                      freight_mod, carrier_r, transit_d, freq_r, notes_r)
                    if ok:
                        st.success(f"Route '{route_name}' added.")
                    else:
                        st.error(t('failed_add_route_err'))

elif page_matches(page, 'delivery_assignments'):
    st.header(t('delivery_assignments'))
    st.markdown("---")
    tab_unassigned, tab_active, tab_done = st.tabs([t('tab_unassigned'), t('tab_active_assignments'), t('tab_completed')])

    df_all_a = db.get_shipments_logistics()

    with tab_unassigned:
        if not df_all_a.empty:
            unassigned = df_all_a[df_all_a["carrier_name"].isna() | (df_all_a["carrier_name"]=="")]
        else:
            unassigned = pd.DataFrame()

        if not unassigned.empty:
            st.caption(f"{len(unassigned)} shipment(s) awaiting carrier assignment")
            for _, row in unassigned.iterrows():
                mode_icon = "✈️" if row["freight_mode"]=="Air" else "🚢"
                with st.expander(f"⚪ {row['shipment_number']} — {row['origin']} → {row['destination']}  |  {mode_icon} {row['freight_mode']}"):
                    ua1,ua2 = st.columns(2)
                    ua1.write(f"**Type:** {row['shipment_type']}")
                    ua2.write(f"**Departure:** {row['departure_date']}")
                    ua3,ua4 = st.columns(2)
                    ua3.write(f"**Weight:** {row['total_weight']:,.0f} kg")
                    ua4.write(f"**Value:** ${row['total_value']:,.0f}")
                    if row.get("notes"): st.caption(row["notes"])
                    st.markdown(t('assign_carrier_hdr'))
                    ac1,ac2,ac3 = st.columns(3)
                    carr_input = ac1.text_input(t('carrier_name_req'), key=f"carr_{row['id']}")
                    cont_input = ac2.selectbox(t('container_type'), ["FCL 20GP","FCL 40HC","LCL","Air Pallet","Bulk","Reefer FCL"],
                                               key=f"cont_{row['id']}")
                    inc_input  = ac3.selectbox(t('incoterms_field'), ["FOB","CIF","EXW","DDP","CFR","DAP"],
                                               key=f"inc_{row['id']}")
                    if st.button(t('assign_carrier_btn'), key=f"assgn_{row['id']}", type="primary"):
                        if not carr_input:
                            st.error(t('carrier_required'))
                        else:
                            db.assign_carrier(row["id"], carr_input, cont_input, inc_input)
                            st.success(f"Carrier assigned to {row['shipment_number']}.")
                            st.rerun()
        else:
            st.success(t('all_assigned'))

    with tab_active:
        if not df_all_a.empty:
            active = df_all_a[
                df_all_a["carrier_name"].notna() &
                (df_all_a["carrier_name"]!="") &
                (~df_all_a["status"].isin(["Delivered"]))
            ]
        else:
            active = pd.DataFrame()

        if not active.empty:
            st.caption(f"{len(active)} active assignment(s)")
            STATUS_ICON = {"In Transit":"🔵","Confirmed":"🟣","Customs":"🟡","Overdue":"🔴"}
            for _, row in active.iterrows():
                icon = STATUS_ICON.get(str(row["status"]), "⚪")
                with st.expander(f"{icon} {row['shipment_number']} — {row['origin']} → {row['destination']}  |  {row['status']}"):
                    ac1,ac2,ac3,ac4 = st.columns(4)
                    ac1.write(f"**Carrier:** {row['carrier_name']}")
                    ac2.write(f"**Container:** {row.get('container_type','—') or '—'}")
                    ac3.write(f"**Incoterms:** {row.get('incoterms','—') or '—'}")
                    ac4.write(f"**ETA:** {row['expected_arrival']}")
                    ac5,ac6,ac7 = st.columns(3)
                    ac5.write(f"**Mode:** {row['freight_mode']}")
                    ac6.write(f"**Weight:** {row['total_weight']:,.0f} kg")
                    ac7.write(f"**Value:** ${row['total_value']:,.0f}")

                    st.markdown("---")
                    b1,b2,b3 = st.columns(3)

                    if b1.button(t('mark_delivered'), key=f"delv_{row['id']}", type="primary"):
                        db.update_shipment_status_logistics(row["id"], "Delivered")
                        db.add_tracking_update_logistics(row["id"], row["destination"],
                            "Delivered", "Shipment delivered successfully. Assignment closed.",
                            st.session_state.get("user_id", 1))
                        st.success(f"{row['shipment_number']} marked as Delivered.")
                        st.rerun()

                    other_statuses = [s for s in ["Confirmed","In Transit","Customs","Overdue"] if s != row["status"]]
                    sel = b2.selectbox(t('change_status_lbl2'), other_statuses, key=f"act_sel_{row['id']}")
                    if b3.button(t('update_status_lbl'), key=f"act_upd_{row['id']}"):
                        db.update_shipment_status_logistics(row["id"], sel)
                        st.rerun()
        else:
            st.info(t('no_active_assignments'))

    with tab_done:
        if not df_all_a.empty:
            done = df_all_a[df_all_a["status"]=="Delivered"]
        else:
            done = pd.DataFrame()

        if not done.empty:
            st.caption(f"{len(done)} completed delivery(ies)")
            disp2 = done[["shipment_number","shipment_type","origin","destination",
                           "carrier_name","actual_arrival","total_weight","total_value"]].copy()
            disp2.columns = ["Shipment #","Type","Origin","Destination","Carrier","Delivered On","Weight(kg)","Value(USD)"]
            st.dataframe(disp2, use_container_width=True, hide_index=True)
        else:
            st.info(t('no_completed'))

elif page_matches(page, 'customer_service_dashboard'):
    st.header(t('customer_service_dashboard'))
    st.caption(f"{t('last_updated')}: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    st.markdown("---")

    summary = db.get_cs_summary()

    k1,k2,k3,k4,k5 = st.columns(5)
    k1.metric(t('total_tickets'),    summary["total"])
    k2.metric(t('open_status'),      summary["open_count"],
              delta=t('needs_attention') if summary["open_count"] > 0 else None,
              delta_color="inverse" if summary["open_count"] > 0 else "normal")
    k3.metric(t('in_progress'),      summary["in_progress"])
    k4.metric(t('resolved_today'),   summary["resolved_today"])
    k5.metric(t('avg_rating'),       f"⭐ {summary['avg_rating']:.1f} / 5.0")

    st.markdown("---")
    col_l, col_r = st.columns(2)

    with col_l:
        tkt_df = db.get_cs_tickets()
        if not tkt_df.empty:
            sc = tkt_df["status"].value_counts().reset_index()
            sc.columns = [t('status'), "Count"]
            color_map = {"Open":"#ef4444","In Progress":"#f59e0b","Resolved":"#22c55e","Closed":"#94a3b8"}
            fig_s = px.pie(sc, names=t('status'), values="Count", color=t('status'),
                           color_discrete_map=color_map, title=t('chart_tickets_status'), height=280)
            fig_s.update_layout(margin=dict(l=0,r=0,t=30,b=0))
            st.plotly_chart(fig_s, use_container_width=True)

    with col_r:
        tkt_df2 = db.get_cs_tickets()
        if not tkt_df2.empty:
            cat_c = tkt_df2["category"].value_counts().reset_index()
            cat_c.columns = [t('category_label'), "Count"]
            fig_c = px.bar(cat_c, x="Count", y=t('category_label'), orientation="h",
                           color_discrete_sequence=["#3b82f6"],
                           title=t('chart_tickets_category'), height=280)
            fig_c.update_layout(margin=dict(l=0,r=0,t=30,b=0), yaxis_title="")
            st.plotly_chart(fig_c, use_container_width=True)

    st.markdown("---")
    st.subheader(t('open_urgent_tickets'))
    urgent_df = db.get_cs_tickets(status_filter=None)
    if not urgent_df.empty:
        urgent_df = urgent_df[urgent_df["status"].isin(["Open","In Progress"])]
        if not urgent_df.empty:
            PRIORITY_ICON = {"Urgent":"🔴","High":"🟠","Medium":"🟡","Low":"🟢"}
            for _, row in urgent_df.head(5).iterrows():
                icon = PRIORITY_ICON.get(str(row["priority"]),"⚪")
                st.markdown(f"{icon} **{row['ticket_number']}** — {row['client_name']} — *{row['subject'][:70]}*  `{_topt(str(row['status']))}`")
        else:
            st.success(t('no_open_tickets'))

    if summary["unread_feedback"] > 0:
        st.info(f"📬 {t('unread_feedback_msg').format(n=summary['unread_feedback'])}")

elif page_matches(page, 'client_tickets'):
    st.header("🎫 " + t('client_tickets'))
    st.markdown("---")

    CATEGORIES = ["Shipment Inquiry","Document Request","Customs Issue","Complaint","Rate Request","General Inquiry"]
    PRIORITIES  = ["Low","Medium","High","Urgent"]
    STATUSES    = ["Open","In Progress","Resolved","Closed"]
    PRIORITY_ICON = {"Urgent":"🔴","High":"🟠","Medium":"🟡","Low":"🟢"}
    STATUS_ICON   = {"Open":"🔴","In Progress":"🟡","Resolved":"🟢","Closed":"⚫"}

    f1,f2,f3 = st.columns(3)
    sf  = f1.selectbox(t('status'),         ["All"]+STATUSES,    key="cs_sf", format_func=_topt)
    pf  = f2.selectbox(t('priority'),       ["All"]+PRIORITIES,  key="cs_pf", format_func=_topt)
    cf  = f3.selectbox(t('category_label'), ["All"]+CATEGORIES,  key="cs_cf", format_func=_topt)
    tkt_df = db.get_cs_tickets(
        sf  if sf  != "All" else None,
        pf  if pf  != "All" else None,
        cf  if cf  != "All" else None
    )
    if not tkt_df.empty:
        cs_unread_map = db.get_all_unread_counts("cs")
        st.caption(f"{len(tkt_df)} {t('tickets_found')}")
        for _, row in tkt_df.iterrows():
            tid    = int(row["id"])
            s_icon = STATUS_ICON.get(str(row["status"]),"⚪")
            p_icon = PRIORITY_ICON.get(str(row["priority"]),"⚪")
            cs_unread  = cs_unread_map.get(tid, 0)
            cs_notif   = f"  🔴 {cs_unread} new" if cs_unread > 0 else ""
            with st.expander(f"{s_icon} {row['ticket_number']}  |  {p_icon} {row['priority']}  |  {row['client_name']}  —  {row['subject'][:50]}{cs_notif}"):
                _edit_key = f"editing_ticket_{tid}"
                _del_key  = f"confirm_del_ticket_{tid}"

                # View mode
                if not st.session_state.get(_edit_key, False):
                    tc1,tc2,tc3 = st.columns(3)
                    tc1.write(f"**{t('client')}:** {row['client_name']}")
                    tc2.write(f"**{t('email_field')}:** {row.get('client_email','—') or '—'}")
                    tc3.write(f"**{t('shipment_ref')}:** {row.get('shipment_ref','—') or '—'}")
                    tc4,tc5,tc6 = st.columns(3)
                    tc4.write(f"**{t('category_label')}:** {_topt(str(row['category']))}")
                    tc5.write(f"**{t('priority')}:** {_topt(str(row['priority']))}")
                    tc6.write(f"**{t('created_lbl')}:** {str(row['created_at'])[:10]}")
                    st.markdown(f"**{t('description')}:**")
                    st.markdown(f"<div style='background:rgba(120,120,180,0.12);border-left:3px solid rgba(120,120,220,0.5);border-radius:6px;padding:10px 14px;margin:4px 0'>{row['description']}</div>", unsafe_allow_html=True)
                    if row.get("resolution"):
                        st.markdown(f"<div style='background:rgba(34,197,94,0.12);border-left:3px solid rgba(34,197,94,0.5);border-radius:6px;padding:10px 14px;margin:4px 0'><b>{t('resolution_lbl')}:</b> {row['resolution']}</div>", unsafe_allow_html=True)
                    st.markdown("---")

                    # ── Conversation toggle button ──
                    cs_conv_key = f"show_conv_cs_{tid}"
                    cs_btn_lbl  = (f"💬 {t('ticket_conversation')}  🔴 {cs_unread} {t('unread_count')}"
                                   if cs_unread > 0 else f"💬 {t('ticket_conversation')}")
                    if st.button(cs_btn_lbl, key=f"conv_btn_cs_{tid}"):
                        st.session_state[cs_conv_key] = not st.session_state.get(cs_conv_key, False)

                    if st.session_state.get(cs_conv_key, False):
                        if cs_unread > 0:
                            db.mark_ticket_replies_read(tid, "cs")
                        cs_replies_df = db.get_ticket_replies(tid)
                        if cs_replies_df.empty:
                            st.caption(t('no_replies_yet'))
                        else:
                            for _, rpl in cs_replies_df.iterrows():
                                is_cs  = str(rpl["sender_role"]) == "cs"
                                slbl   = t('reply_from_cs') if is_cs else t('reply_from_client')
                                bg     = "rgba(59,130,246,0.15)" if is_cs else "rgba(34,197,94,0.12)"
                                border = "rgba(59,130,246,0.45)" if is_cs else "rgba(34,197,94,0.4)"
                                align  = "right" if is_cs else "left"
                                st.markdown(
                                    f"<div style='background:{bg};border-left:3px solid {border};"
                                    f"border-radius:8px;padding:8px 12px;margin:4px 0;text-align:{align}'>"
                                    f"<small><b>{slbl}</b> · {str(rpl['created_at'])[:16]}</small><br>{rpl['message']}"
                                    f"</div>", unsafe_allow_html=True
                                )
                        cs_agent_email = user.get("email","") if user else ""
                        with st.form(f"cs_reply_form_{tid}", clear_on_submit=True):
                            cs_rpl_msg = st.text_area(t('type_reply'), height=70, label_visibility="collapsed", placeholder=t('type_reply'))
                            if st.form_submit_button(f"📤 {t('send_reply_btn')}", type="primary"):
                                if not cs_rpl_msg.strip():
                                    st.error(t('reply_required'))
                                else:
                                    db.add_ticket_reply(tid, "cs", cs_agent_email, cs_rpl_msg.strip())
                                    st.success(t('reply_sent'))
                                    st.rerun()

                    st.markdown("---")
                    if row["status"] not in ("Resolved","Closed"):
                        with st.form(f"resolve_form_{row['id']}", clear_on_submit=True):
                            new_status  = st.selectbox(t('update_status_btn'), [s for s in STATUSES if s != row["status"]], key=f"nst_{row['id']}", format_func=_topt)
                            resolution  = st.text_area(t('resolution_response'), height=80, key=f"res_{row['id']}")
                            if st.form_submit_button(t('update_ticket_btn'), type="primary"):
                                db.update_cs_ticket(row["id"], new_status, resolution or None)
                                st.success(t('ticket_updated_success'))
                                st.rerun()
                    else:
                        st.caption(f"{t('resolved_at_lbl')}: {str(row.get('resolved_at',''))[:16]}")

                    # Edit / Delete action buttons
                    btn_e, btn_d, _ = st.columns([1, 1, 4])
                    if btn_e.button(f"✏️ {t('edit_ticket_btn')}", key=f"edit_btn_{row['id']}"):
                        st.session_state[_edit_key] = True
                        st.rerun()
                    if btn_d.button(f"🗑️ {t('delete_ticket_btn')}", key=f"del_btn_{row['id']}", type="secondary"):
                        st.session_state[_del_key] = True
                        st.rerun()

                    if st.session_state.get(_del_key, False):
                        st.warning(t('confirm_delete_ticket'))
                        dc1, dc2, _ = st.columns([1, 1, 4])
                        if dc1.button(f"✅ {t('delete_ticket_btn')}", key=f"del_confirm_{row['id']}", type="primary"):
                            if db.delete_cs_ticket(row["id"]):
                                db.insert_activity_log("Ticket Deleted", f"CS Ticket #{row['id']} deleted", user.get("email",""))
                                st.session_state.pop(_del_key, None)
                                st.warning(t('cs_ticket_deleted_ok'))
                                st.rerun()
                        if dc2.button(f"❌ {t('cancel_btn')}", key=f"del_cancel_{row['id']}"):
                            st.session_state.pop(_del_key, None)
                            st.rerun()

                # Edit mode
                else:
                    with st.form(f"edit_ticket_form_{row['id']}", clear_on_submit=False):
                        et1, et2 = st.columns(2)
                        cat_idx = CATEGORIES.index(row["category"]) if row["category"] in CATEGORIES else 0
                        pri_idx = PRIORITIES.index(row["priority"]) if row["priority"] in PRIORITIES else 1
                        new_cat = et1.selectbox(t('category_label'), CATEGORIES, index=cat_idx, key=f"ecat_{row['id']}", format_func=_topt)
                        new_pri = et2.selectbox(t('priority'), PRIORITIES, index=pri_idx, key=f"epri_{row['id']}", format_func=_topt)
                        new_sub  = st.text_input(t('subject'), value=str(row["subject"]), key=f"esub_{row['id']}")
                        new_desc = st.text_area(t('description'), value=str(row["description"]), height=120, key=f"edesc_{row['id']}")
                        ef1, ef2 = st.columns(2)
                        if ef1.form_submit_button(f"💾 {t('save_ticket_btn')}", type="primary"):
                            if not new_sub or not new_desc:
                                st.error(t('client_subj_desc_required'))
                            else:
                                if db.edit_cs_ticket(row["id"], new_sub, new_desc, new_cat, new_pri):
                                    db.insert_activity_log("Ticket Updated", f"CS Ticket #{row['id']} edited", user.get("email",""))
                                    st.session_state.pop(_edit_key, None)
                                    st.success(t('ticket_edited_success'))
                                    st.rerun()
                        if ef2.form_submit_button(f"❌ {t('cancel_btn')}"):
                            st.session_state.pop(_edit_key, None)
                            st.rerun()
    else:
        st.info(t('no_tickets_found'))

elif page_matches(page, 'customer_feedback'):
    st.header("💬 " + t('customer_feedback'))
    st.markdown("---")

    f_unread = st.checkbox(t('show_unread_only'), value=False)
    fb_df = db.get_cs_feedback(unread_only=f_unread)
    if not fb_df.empty:
        avg_r = fb_df["rating"].mean()
        fa1,fa2,fa3 = st.columns(3)
        fa1.metric(t('total_feedback'), len(fb_df))
        fa2.metric(t('average_rating'), f"⭐ {avg_r:.1f} / 5.0")
        fa3.metric(t('unread_count'), int((fb_df["is_read"]==0).sum()))

        st.markdown("---")
        rc = fb_df["rating"].value_counts().sort_index().reset_index()
        _rating_col = t('avg_rating')
        rc.columns = [_rating_col, "Count"]
        rc[_rating_col] = rc[_rating_col].apply(lambda x: f"{'⭐'*int(x)} ({int(x)})")
        fig_r = px.bar(rc, x=_rating_col, y="Count", color_discrete_sequence=["#f59e0b"],
                       title=t('chart_rating_dist'), height=250)
        fig_r.update_layout(margin=dict(l=0,r=0,t=30,b=0))
        st.plotly_chart(fig_r, use_container_width=True)

        st.markdown("---")
        STAR = {5:"⭐⭐⭐⭐⭐",4:"⭐⭐⭐⭐",3:"⭐⭐⭐",2:"⭐⭐",1:"⭐"}
        for _, row in fb_df.iterrows():
            stars = STAR.get(int(row["rating"]),"⭐")
            unread_badge = " 🔵" if row["is_read"]==0 else ""
            with st.expander(f"{stars}  {row['client_name']}  |  {row['category']}  |  {str(row['created_at'])[:10]}{unread_badge}"):
                fb1,fb2,fb3 = st.columns(3)
                fb1.write(f"**{t('client')}:** {row['client_name']}")
                fb2.write(f"**{t('email')}:** {row.get('client_email','—') or '—'}")
                fb3.write(f"**{t('shipment_ref')}:** {row.get('shipment_ref','—') or '—'}")
                if row.get("comment"):
                    st.write(row["comment"])
                if row["is_read"] == 0:
                    if st.button(t('mark_as_read'), key=f"fbread_{row['id']}"):
                        db.mark_feedback_read(row["id"])
                        st.rerun()
    else:
        st.info(t('no_feedback_yet'))

elif page_matches(page, 'administration_dashboard'):
    st.header(t('administration_dashboard'))
    st.caption(f"{t('last_updated')}: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    st.markdown("---")

    s = db.get_admin_summary()
    k1,k2,k3,k4,k5 = st.columns(5)
    k1.metric(t('total_documents'),    s["active_docs"],    delta=f"{s['total_docs']} {t('total_word')}")
    k2.metric(t('active_contracts'),   s["active_contracts"])
    k3.metric(t('upcoming_meetings'),  s["upcoming_meetings"])
    k4.metric(t('expiring_30d'),       s["expiring_soon"],
              delta="Renew soon" if s["expiring_soon"] > 0 else None,
              delta_color="inverse" if s["expiring_soon"] > 0 else "normal")
    k5.metric(t('total_contracts_lbl'), s["total_contracts"])

    st.markdown("---")
    col_l, col_r = st.columns(2)

    with col_l:
        st.subheader(t('upcoming_meetings_hdr'))
        mtg_df = db.get_admin_meetings(upcoming_only=True)
        if not mtg_df.empty:
            TYPE_ICON = {"Internal":"🏢","With Carrier":"🚢","With Client":"🤝",
                         "With Customs":"🛂","Regulatory":"📋","Board":"👔"}
            for _, row in mtg_df.head(5).iterrows():
                icon = TYPE_ICON.get(str(row["meeting_type"]),"📅")
                st.markdown(f"{icon} **{row['meeting_date']}** {str(row.get('meeting_time',''))[:5]}  — {row['title']}")
                st.caption(f"&nbsp;&nbsp;&nbsp;&nbsp;{row['meeting_type']} · {row.get('location','')}")
        else:
            st.info(t('no_upcoming_meetings'))

    with col_r:
        st.subheader(t('contracts_expiring'))
        contr_df = db.get_admin_contracts(status="Active")
        if not contr_df.empty:
            contr_df["end_date"] = pd.to_datetime(contr_df["end_date"])
            soon = contr_df[contr_df["end_date"] <= pd.Timestamp.now() + pd.Timedelta(days=90)].sort_values("end_date")
            if not soon.empty:
                for _, row in soon.iterrows():
                    days_left = (row["end_date"] - pd.Timestamp.now()).days
                    color = "🔴" if days_left <= 30 else "🟡"
                    st.markdown(f"{color} **{row['title']}** — expires **{str(row['end_date'])[:10]}** ({days_left}d)")
            else:
                st.success(t('no_contracts_expiring'))

    st.markdown("---")
    st.subheader(t('recent_docs'))
    doc_df = db.get_admin_documents()
    if not doc_df.empty:
        disp = doc_df[["title","category","status","expiry_date"]].head(6).copy()
        disp.columns = ["Title","Category","Status","Expiry"]
        st.dataframe(disp, use_container_width=True, hide_index=True)

elif page_matches(page, 'documents'):
    st.header(t('documents'))
    st.markdown("---")
    tab_list, tab_add = st.tabs([t('tab_all_docs'), t('tab_add_doc')])

    DOC_CATS = ["Company Certificate","Carrier Agreement","Client Agreement",
                "Regulatory","SOP","Insurance Policy","Lease Agreement","Other"]

    with tab_list:
        dc1,dc2 = st.columns(2)
        cat_f = dc1.selectbox(t('category_lbl'), ["All"] + DOC_CATS, key="doc_cat")
        sta_f = dc2.selectbox(t('status'),        ["All","Active","Archived","Expired"], key="doc_sta", format_func=_topt)
        doc_df = db.get_admin_documents(
            cat_f if cat_f != "All" else None,
            sta_f if sta_f != "All" else None
        )
        if not doc_df.empty:
            st.caption(t('x_documents').format(n=len(doc_df)))
            STATUS_ICON = {"Active":"🟢","Archived":"⚫","Expired":"🔴"}
            for _, row in doc_df.iterrows():
                icon = STATUS_ICON.get(str(row["status"]),"⚪")
                with st.expander(f"{icon} {row['title']}  |  {row['category']}"):
                    dc1b,dc2b,dc3b = st.columns(3)
                    dc1b.write(f"**{t('category_lbl')}:** {row['category']}")
                    dc2b.write(f"**{t('status')}:** {row['status']}")
                    dc3b.write(f"**{t('expiry_lbl')}:** {str(row.get('expiry_date','—') or '—')[:10]}")
                    if row.get("description"): st.caption(row["description"])
                    if row.get("file_name"):   st.write(f"📄 `{row['file_name']}`")
                    st.markdown("---")
                    b1,b2,b3 = st.columns(3)
                    if row["status"] == "Active":
                        if b1.button(t('archive_btn'), key=f"arch_{row['id']}"):
                            db.update_document_status(row["id"], "Archived"); st.rerun()
                    else:
                        if b1.button(t('restore_btn'), key=f"rest_{row['id']}"):
                            db.update_document_status(row["id"], "Active"); st.rerun()
        else:
            st.info(t('no_documents_found'))

    with tab_add:
        st.subheader(t('add_new_doc'))
        with st.form("add_doc_form", clear_on_submit=True):
            ad1,ad2 = st.columns(2)
            doc_title   = ad1.text_input(t('doc_title'))
            doc_cat     = ad2.selectbox(t('category_field'), DOC_CATS)
            ad3,ad4 = st.columns(2)
            doc_status  = ad3.selectbox(t('status'), ["Active","Archived"], format_func=_topt)
            doc_expiry  = ad4.date_input(t('expiry_date_field'))
            doc_fname   = st.text_input(t('file_name_field'))
            doc_desc    = st.text_area(t('description'), height=80)
            use_expiry  = st.checkbox(t('set_expiry'))
            if st.form_submit_button(t('add_doc_btn'), type="primary"):
                if not doc_title:
                    st.error(t('meeting_title_required'))
                else:
                    uid = st.session_state.get("user_id", 1)
                    exp = str(doc_expiry) if use_expiry else None
                    ok  = db.add_document(doc_title, doc_cat, doc_desc, doc_fname, doc_status, exp, uid)
                    if ok: st.success(t('document_added_success'))
                    else:  st.error(t('failed_add_document_err'))

elif page_matches(page, 'meetings'):
    st.header(t('meetings'))
    st.markdown("---")
    tab_up, tab_past, tab_new = st.tabs([t('tab_upcoming'), t('tab_past_meetings'), t('tab_schedule')])

    MTG_TYPES = ["Internal","With Carrier","With Client","With Customs","Regulatory","Board"]
    TYPE_ICON = {"Internal":"🏢","With Carrier":"🚢","With Client":"🤝",
                 "With Customs":"🛂","Regulatory":"📋","Board":"👔"}

    all_mtg = db.get_admin_meetings()
    if not all_mtg.empty:
        all_mtg["meeting_date"] = pd.to_datetime(all_mtg["meeting_date"])
        upcoming_m = all_mtg[all_mtg["meeting_date"] >= pd.Timestamp.now().normalize()]
        past_m     = all_mtg[all_mtg["meeting_date"] <  pd.Timestamp.now().normalize()]
    else:
        upcoming_m = past_m = pd.DataFrame()

    with tab_up:
        if not upcoming_m.empty:
            for _, row in upcoming_m.iterrows():
                icon = TYPE_ICON.get(str(row["meeting_type"]),"📅")
                with st.expander(f"{icon} {str(row['meeting_date'])[:10]}  {str(row.get('meeting_time',''))[:5]}  — {row['title']}  |  `{row['status']}`"):
                    mc1,mc2,mc3 = st.columns(3)
                    mc1.write(f"**{t('type_lbl')}:** {row['meeting_type']}")
                    mc2.write(f"**{t('location_lbl')}:** {row.get('location','—') or '—'}")
                    mc3.write(f"**{t('duration_lbl')}:** {row.get('duration_min',60)} {t('min_lbl')}")
                    if row.get("attendees"): st.write(f"**{t('attendees_lbl')}:** {row['attendees']}")
                    if row.get("agenda"):
                        st.markdown(f"**{t('agenda_lbl')}:**")
                        st.info(row["agenda"])
                    if row["status"] == "Scheduled":
                        st.markdown("---")
                        with st.form(f"mtg_form_{row['id']}", clear_on_submit=True):
                            new_st  = st.selectbox(t('update_meeting'), ["Completed","Cancelled"], key=f"mst_{row['id']}", format_func=_topt)
                            minutes = st.text_area(t('meeting_minutes'), height=80, key=f"mmin_{row['id']}")
                            if st.form_submit_button(t('save_btn'), type="primary"):
                                db.update_meeting(row["id"], new_st, minutes or None)
                                st.rerun()
        else:
            st.info(t('no_upcoming_meetings'))

    with tab_past:
        if not past_m.empty:
            for _, row in past_m.sort_values("meeting_date", ascending=False).iterrows():
                icon = TYPE_ICON.get(str(row["meeting_type"]),"📅")
                status_icon = "✅" if row["status"]=="Completed" else "❌"
                with st.expander(f"{status_icon} {str(row['meeting_date'])[:10]} — {row['title']}"):
                    mc1,mc2 = st.columns(2)
                    mc1.write(f"**{t('type_lbl')}:** {row['meeting_type']}")
                    mc2.write(f"**{t('location_lbl')}:** {row.get('location','—') or '—'}")
                    if row.get("attendees"): st.write(f"**{t('attendees_lbl')}:** {row['attendees']}")
                    if row.get("minutes"):
                        st.markdown(f"**{t('minutes_lbl')}:**")
                        st.success(row["minutes"])
        else:
            st.info(t('no_past_meetings'))

    with tab_new:
        st.subheader(t('schedule_new_meeting'))
        with st.form("new_mtg_form", clear_on_submit=True):
            nm1,nm2 = st.columns(2)
            mtg_title = nm1.text_input(t('meeting_title'))
            mtg_type  = nm2.selectbox(t('meeting_type'), MTG_TYPES)
            nm3,nm4 = st.columns(2)
            mtg_date  = nm3.date_input(t('date_field2'))
            mtg_time  = nm4.time_input(t('time_field'))
            nm5,nm6 = st.columns(2)
            mtg_loc   = nm5.text_input(t('location_platform'))
            mtg_dur   = nm6.number_input(t('duration_min'), min_value=15, max_value=480, value=60, step=15)
            mtg_att   = st.text_input(t('attendees_field'))
            mtg_ag    = st.text_area(t('agenda_field'), height=100)
            if st.form_submit_button(t('schedule_meeting_btn'), type="primary"):
                if not mtg_title:
                    st.error(t('meeting_title_required'))
                else:
                    uid = st.session_state.get("user_id", 1)
                    ok  = db.add_meeting(mtg_title, mtg_type, mtg_att, mtg_loc,
                                         str(mtg_date), str(mtg_time), mtg_dur, mtg_ag, uid)
                    if ok: st.success(f"'{mtg_title}' {t('meeting_scheduled')}")
                    else:  st.error(t('failed_schedule_meeting_err'))

elif page_matches(page, 'contracts'):
    st.header(t('contracts'))
    st.markdown("---")
    tab_list, tab_new = st.tabs([t('tab_all_contracts'), t('tab_new_contract')])

    PARTY_TYPES  = ["Carrier","Client","Supplier","Government","Other"]
    CONTR_TYPES  = ["Service Agreement","Rate Agreement","Agency Agreement",
                    "Insurance Policy","Software License","Lease Agreement","NDA","Other"]

    with tab_list:
        cf1,cf2 = st.columns(2)
        sta_cf  = cf1.selectbox(t('status'),     ["All","Active","Expired","Pending","Terminated"], key="ct_sta", format_func=_topt)
        pty_cf  = cf2.selectbox(t('party_type_filter'), ["All"]+PARTY_TYPES, key="ct_pty", format_func=_topt)
        contr_df = db.get_admin_contracts(
            sta_cf if sta_cf != "All" else None,
            pty_cf if pty_cf != "All" else None
        )
        if not contr_df.empty:
            contr_df["end_date"] = pd.to_datetime(contr_df["end_date"])
            total_val = contr_df[contr_df["status"]=="Active"]["value"].sum()
            st.metric(t('active_contracts_value'), f"${total_val:,.0f}")
            st.markdown("---")
            STATUS_ICON = {"Active":"🟢","Expired":"🔴","Pending":"🟡","Terminated":"⚫"}
            for _, row in contr_df.iterrows():
                icon = STATUS_ICON.get(str(row["status"]),"⚪")
                days_left = (row["end_date"] - pd.Timestamp.now()).days
                expiry_txt = f"⚠️ {days_left}d left" if 0 < days_left <= 60 and row["status"]=="Active" else str(row["end_date"])[:10]
                with st.expander(f"{icon} {row['contract_number']} — {row['title']}  |  {row['party_name']}"):
                    cc1,cc2,cc3,cc4 = st.columns(4)
                    cc1.write(f"**{t('party_type_lbl')}:** {row['party_type']}")
                    cc2.write(f"**{t('contract_type_lbl')}:** {row['contract_type']}")
                    cc3.write(f"**{t('value_lbl')}:** ${row.get('value',0) or 0:,.0f} {row.get('currency','USD')}")
                    cc4.write(f"**{t('status')}:** {row['status']}")
                    cc5,cc6 = st.columns(2)
                    cc5.write(f"**{t('start_lbl')}:** {str(row['start_date'])[:10]}")
                    cc6.write(f"**{t('end_lbl')}:** {expiry_txt}")
                    if row.get("description"): st.caption(row["description"])
                    st.markdown("---")
                    b1,b2 = st.columns(2)
                    other = [s for s in ["Active","Expired","Terminated"] if s != row["status"]]
                    sel_s = b1.selectbox(t('change_status'), other, key=f"cst_{row['id']}", format_func=_topt)
                    if b2.button(t('update_btn_lbl'), key=f"cupd_{row['id']}"):
                        db.update_contract_status(row["id"], sel_s); st.rerun()
        else:
            st.info(t('no_contracts_found'))

    with tab_new:
        st.subheader(t('add_new_contract'))
        with st.form("new_contract_form", clear_on_submit=True):
            nc1,nc2 = st.columns(2)
            contr_title = nc1.text_input(t('contract_title'))
            party_name  = nc2.text_input(t('party_name'))
            nc3,nc4 = st.columns(2)
            party_type  = nc3.selectbox(t('party_type_field'), PARTY_TYPES, format_func=_topt)
            contr_type  = nc4.selectbox(t('contract_type_field'), CONTR_TYPES, format_func=_topt)
            nc5,nc6,nc7 = st.columns(3)
            contr_val   = nc5.number_input(t('value_field'), min_value=0.0, step=1000.0)
            contr_cur   = nc6.selectbox(t('currency_field'), ["USD","EUR","TRY","GBP"])
            nc8,nc9 = st.columns(2)
            start_d     = nc8.date_input(t('start_date_lbl'))
            end_d       = nc9.date_input(t('end_date_lbl'))
            contr_desc  = st.text_area(t('description'), height=80)
            if st.form_submit_button(t('add_contract_btn'), type="primary"):
                if not contr_title or not party_name:
                    st.error(t('title_party_required'))
                elif end_d <= start_d:
                    st.error(t('end_after_start'))
                else:
                    uid = st.session_state.get("user_id", 1)
                    cn  = f"CNTR-{datetime.now().strftime('%Y-%m%d%H%M%S')}"
                    ok  = db.add_contract(cn, contr_title, party_name, party_type, contr_type,
                                          contr_val, contr_cur, str(start_d), str(end_d), contr_desc, uid)
                    if ok: st.success(t('contract_added_success'))
                    else:  st.error(t('failed_add_contract_err'))

elif page_matches(page, 'sales_dashboard'):
    st.header(t('sales_dashboard'))
    st.caption(f"{t('last_updated')}: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    st.markdown("---")
    s = db.get_sales_summary()
    sc1,sc2,sc3,sc4,sc5,sc6 = st.columns(6)
    sc1.metric(t('total_leads'),    s["total_leads"])
    sc2.metric(t('active_leads'),   s["active_leads"])
    sc3.metric(t('open_deals'),     s["total_deals"] - s["won_deals"])
    sc4.metric(t('pipeline_value'), f"${s['pipeline_value']:,.0f}")
    sc5.metric(t('won_deals'),      s["won_deals"])
    sc6.metric(t('win_rate'),       f"{s['win_rate']}%")
    st.markdown("---")

    try:
        leads_df_dash = db.get_sales_leads()
        deals_df_dash = db.get_sales_deals()

        col_l, col_r = st.columns(2)
        with col_l:
            if leads_df_dash is not None and not leads_df_dash.empty:
                lsc = leads_df_dash["status"].value_counts().reset_index()
                lsc.columns = [t('status'), "Count"]
                fig_ls = px.bar(lsc, x=t('status'), y="Count", title=t('leads_by_status'),
                                color=t('status'), height=260)
                fig_ls.update_layout(margin=dict(l=0,r=0,t=30,b=0), showlegend=False)
                st.plotly_chart(fig_ls, use_container_width=True)

        with col_r:
            if deals_df_dash is not None and not deals_df_dash.empty:
                stage_order = ["Discovery","Proposal","Negotiation","Won","Lost"]
                dsc = deals_df_dash["stage"].value_counts().reindex(stage_order, fill_value=0).reset_index()
                _stage_lbl = t('stage_filter')
                dsc.columns = [_stage_lbl, "Count"]
                fig_ds = px.funnel(dsc, x="Count", y=_stage_lbl, title=t('deals_pipeline_chart'),
                                   height=260)
                fig_ds.update_layout(margin=dict(l=0,r=0,t=30,b=0))
                st.plotly_chart(fig_ds, use_container_width=True)

        st.subheader(t('recent_leads'))
        if leads_df_dash is not None and not leads_df_dash.empty:
            recent_l = leads_df_dash.head(8)[["lead_name","company_name","country","source","status","freight_interest","created_at"]].copy()
            recent_l.columns = [t('col_name'),t('col_company'),t('col_country'),t('col_source'),t('status'),t('col_freight'),t('col_created')]
            st.dataframe(recent_l, use_container_width=True, hide_index=True)
    except Exception as _ex:
        st.error(str(_ex))

elif page_matches(page, 'leads'):
    st.header(t('leads'))
    st.caption(f"{t('last_updated')}: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    st.markdown("---")

    LEAD_SOURCES   = ["Cold Call","Referral","LinkedIn","Trade Show","Website","Email Campaign","Other"]
    LEAD_STATUSES  = ["New","Contacted","Qualified","Proposal Sent","Won","Lost"]
    FREIGHT_TYPES  = ["Sea FCL","Sea LCL","Air Cargo","Customs","Road","Other"]
    COUNTRIES_LEAD = ["Afghanistan","Albania","Algeria","Argentina","Armenia","Australia","Austria",
                      "Azerbaijan","Bahrain","Bangladesh","Belarus","Belgium","Brazil","Bulgaria",
                      "Canada","Chile","China","Colombia","Croatia","Cyprus","Czech Republic",
                      "Denmark","Ecuador","Egypt","Estonia","Ethiopia","Finland","France",
                      "Georgia","Germany","Ghana","Greece","Hungary","India","Indonesia","Iran",
                      "Iraq","Ireland","Israel","Italy","Japan","Jordan","Kazakhstan","Kenya",
                      "Kuwait","Latvia","Lebanon","Libya","Lithuania","Luxembourg","Malaysia",
                      "Malta","Mexico","Morocco","Netherlands","New Zealand","Nigeria","Norway",
                      "Oman","Pakistan","Palestine","Peru","Philippines","Poland","Portugal",
                      "Qatar","Romania","Russia","Saudi Arabia","Senegal","Serbia","Singapore",
                      "Slovakia","Slovenia","South Africa","South Korea","Spain","Sri Lanka",
                      "Sudan","Sweden","Switzerland","Syria","Taiwan","Thailand","Tunisia",
                      "Turkey","Ukraine","United Arab Emirates","United Kingdom","United States",
                      "Uzbekistan","Venezuela","Vietnam","Yemen"]

    lf1,lf2,lf3 = st.columns(3)
    flt_sta = lf1.selectbox(t('status'),["All"]+LEAD_STATUSES, key="lead_flt_sta", format_func=_topt)
    flt_src = lf2.selectbox(t('source_filter'),["All"]+LEAD_SOURCES, key="lead_flt_src", format_func=_topt)
    flt_fi  = lf3.selectbox(t('freight_filter'),["All"]+FREIGHT_TYPES, key="lead_flt_fi", format_func=_topt)

    leads_df = db.get_sales_leads(status=flt_sta if flt_sta!="All" else None,
                                   source=flt_src if flt_src!="All" else None)
    if leads_df is not None and not leads_df.empty:
        if flt_fi != "All":
            leads_df = leads_df[leads_df["freight_interest"]==flt_fi]

    STATUS_COLOR = {"New":"🔵","Contacted":"🟡","Qualified":"🟠","Proposal Sent":"🟣","Won":"🟢","Lost":"🔴"}
    if leads_df is not None and not leads_df.empty:
        st.write(f"**{len(leads_df)} {t('leads_found')}**")
        for _, row in leads_df.iterrows():
            icon = STATUS_COLOR.get(str(row["status"]),"⚪")
            with st.expander(f"{icon} {row['lead_name']}  |  {row.get('company_name','—')}  |  {row.get('country','—')}"):
                lc1,lc2,lc3,lc4 = st.columns(4)
                lc1.write(f"**{t('email_field')}:** {row.get('email','—')}")
                lc2.write(f"**{t('phone_lbl')}:** {row.get('phone','—')}")
                lc3.write(f"**{t('col_source')}:** {row.get('source','—')}")
                lc4.write(f"**{t('col_freight')}:** {row.get('freight_interest','—')}")
                if row.get("notes"): st.caption(row["notes"])
                st.markdown("---")
                curr_status = str(row["status"])
                next_opts = [s for s in LEAD_STATUSES if s != curr_status]
                ns1,ns2 = st.columns(2)
                new_sta = ns1.selectbox(t('update_status_lead'), next_opts, key=f"lsta_{row['id']}", format_func=_topt)
                if ns2.button(t('update_btn'), key=f"lstabtn_{row['id']}", type="primary"):
                    if db.update_lead_status(int(row["id"]), new_sta):
                        st.success(t('status_updated'))
                        st.rerun()
    else:
        st.info(t('no_leads_filters_msg'))

    st.markdown("---")
    with st.expander(t('add_new_lead')):
        with st.form("add_lead_form", clear_on_submit=True):
            al1,al2 = st.columns(2)
            ln_name    = al1.text_input(t('lead_name'))
            ln_company = al2.text_input(t('company_name'))
            al3,al4 = st.columns(2)
            ln_email   = al3.text_input(t('email_field'))
            ln_phone   = al4.text_input(t('phone_field'))
            al5,al6 = st.columns(2)
            ln_country = al5.selectbox(t('country_field'), COUNTRIES_LEAD)
            ln_source  = al6.selectbox(t('source_field'), LEAD_SOURCES, format_func=_topt)
            ln_freight = st.selectbox(t('freight_interest'), FREIGHT_TYPES, format_func=_topt)
            ln_notes   = st.text_area(t('notes_desc'), height=80)
            if st.form_submit_button(t('add_lead_btn'), type="primary"):
                if not ln_name:
                    st.error(t('lead_name_required'))
                else:
                    uid = st.session_state.get("user",{}).get("id")
                    if db.add_lead(ln_name,ln_company,ln_email,ln_phone,ln_country,
                                   ln_source,ln_freight,ln_notes,uid):
                        st.success(t('lead_added_success'))
                        st.rerun()
                    else:
                        st.error(t('failed_add_lead_err'))

elif page_matches(page, 'deals'):
    st.header(t('deals'))
    st.caption(f"{t('last_updated')}: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    st.markdown("---")

    DEAL_STAGES   = ["Discovery","Proposal","Negotiation","Won","Lost"]
    FREIGHT_TYPES_D = ["Sea FCL","Sea LCL","Air Cargo","Customs","Road","Other"]
    STAGE_PROB    = {"Discovery":20,"Proposal":50,"Negotiation":75,"Won":100,"Lost":0}
    STAGE_ICON    = {"Discovery":"🔵","Proposal":"🟡","Negotiation":"🟠","Won":"🟢","Lost":"🔴"}

    df1,df2 = st.columns(2)
    flt_stg = df1.selectbox(t('stage_filter'), ["All"]+DEAL_STAGES, key="deal_flt_stg", format_func=_topt)
    flt_ft  = df2.selectbox(t('freight_type_filter'), ["All"]+FREIGHT_TYPES_D, key="deal_flt_ft", format_func=_topt)

    deals_df = db.get_sales_deals(stage=flt_stg if flt_stg!="All" else None)
    if deals_df is not None and not deals_df.empty:
        if flt_ft != "All":
            deals_df = deals_df[deals_df["freight_type"]==flt_ft]

    if deals_df is not None and not deals_df.empty:
        total_pipeline = deals_df[~deals_df["stage"].isin(["Won","Lost"])]["value"].sum()
        st.write(f"**{len(deals_df)} {t('deals_found')}** | {t('pipeline_lbl')}: **${total_pipeline:,.0f}**")

        for _, row in deals_df.iterrows():
            icon = STAGE_ICON.get(str(row["stage"]),"⚪")
            val_str = f"${float(row['value'] or 0):,.0f} {row.get('currency','USD')}" if row.get("value") else "TBD"
            with st.expander(f"{icon} {row['title']}  |  {row['client_name']}  |  {val_str}  |  {row['stage']}"):
                dc1,dc2,dc3,dc4 = st.columns(4)
                dc1.write(f"{t('freight_lbl')} {row.get('freight_type','—')}")
                dc2.write(f"{t('route_lbl')} {row.get('origin','—')} → {row.get('destination','—')}")
                dc3.write(f"{t('probability_lbl')} {row.get('probability',0)}%")
                dc4.write(f"{t('close_date_lbl')} {str(row.get('close_date','—'))[:10]}")
                if row.get("notes"): st.caption(row["notes"])
                if str(row["stage"]) not in ("Won","Lost"):
                    st.markdown("---")
                    ds1,ds2 = st.columns(2)
                    next_stages = [s for s in DEAL_STAGES if s != str(row["stage"])]
                    sel_stg = ds1.selectbox(t('move_to_stage'), next_stages, key=f"dstg_{row['id']}", format_func=_topt)
                    if ds2.button(t('update_stage_btn'), key=f"dstgbtn_{row['id']}", type="primary"):
                        prob = STAGE_PROB.get(sel_stg, 50)
                        if db.update_deal_stage(int(row["id"]), sel_stg, prob):
                            st.success(t('stage_updated_success'))
                            st.rerun()
    else:
        st.info(t('no_deals'))

    st.markdown("---")
    with st.expander(t('add_new_deal')):
        COUNTRIES_DEAL = ["Afghanistan","Albania","Algeria","Argentina","Armenia","Australia","Austria",
                          "Azerbaijan","Bahrain","Bangladesh","Belarus","Belgium","Brazil","Bulgaria",
                          "Canada","Chile","China","Colombia","Croatia","Cyprus","Czech Republic",
                          "Denmark","Ecuador","Egypt","Estonia","Ethiopia","Finland","France",
                          "Georgia","Germany","Ghana","Greece","Hungary","India","Indonesia","Iran",
                          "Iraq","Ireland","Israel","Italy","Japan","Jordan","Kazakhstan","Kenya",
                          "Kuwait","Latvia","Lebanon","Libya","Lithuania","Luxembourg","Malaysia",
                          "Malta","Mexico","Morocco","Netherlands","New Zealand","Nigeria","Norway",
                          "Oman","Pakistan","Palestine","Peru","Philippines","Poland","Portugal",
                          "Qatar","Romania","Russia","Saudi Arabia","Senegal","Serbia","Singapore",
                          "Slovakia","Slovenia","South Africa","South Korea","Spain","Sri Lanka",
                          "Sudan","Sweden","Switzerland","Syria","Taiwan","Thailand","Tunisia",
                          "Turkey","Ukraine","United Arab Emirates","United Kingdom","United States",
                          "Uzbekistan","Venezuela","Vietnam","Yemen"]
        with st.form("add_deal_form", clear_on_submit=True):
            ad1,ad2 = st.columns(2)
            dl_title  = ad1.text_input(t('deal_title_field'))
            dl_client = ad2.text_input(t('client_name_field'))
            ad3,ad4,ad5 = st.columns(3)
            dl_value  = ad3.number_input(t('value_usd_field'), min_value=0.0, step=1000.0)
            dl_stage  = ad4.selectbox(t('stage_filter'), DEAL_STAGES, format_func=_topt)
            dl_prob   = ad5.number_input(t('probability_pct'), min_value=0, max_value=100, value=STAGE_PROB.get(DEAL_STAGES[0],20))
            ad6,ad7 = st.columns(2)
            dl_ft     = ad6.selectbox(t('freight_type_filter'), FREIGHT_TYPES_D, format_func=_topt)
            dl_cd     = ad7.date_input(t('expected_close'))
            ad8,ad9 = st.columns(2)
            dl_orig   = ad8.selectbox(t('origin_lbl'), COUNTRIES_DEAL, index=COUNTRIES_DEAL.index("Turkey") if "Turkey" in COUNTRIES_DEAL else 0)
            dl_dst    = ad9.selectbox(t('destination_lbl'), COUNTRIES_DEAL)
            dl_notes  = st.text_area(t('notes_lbl'), height=60)
            if st.form_submit_button(t('add_deal_btn'), type="primary"):
                if not dl_title or not dl_client:
                    st.error(t('deal_title_client_req'))
                else:
                    uid = st.session_state.get("user",{}).get("id")
                    if db.add_deal(dl_title,dl_client,dl_value,"USD",dl_stage,dl_prob,
                                   str(dl_cd),dl_ft,dl_orig,dl_dst,dl_notes,None,uid):
                        st.success(t('deal_added_success'))
                        st.rerun()
                    else:
                        st.error(t('failed_add_deal'))

elif page_matches(page, 'offers'):
    st.header(t('offers'))
    st.caption(f"{t('last_updated')}: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    st.markdown("---")

    OFFER_STATUSES = ["Draft","Sent","Accepted","Rejected","Expired"]
    FREIGHT_TYPES_O = ["Sea FCL","Sea LCL","Air Cargo","Customs","Road","Other"]
    STATUS_ICON_O  = {"Draft":"⚪","Sent":"🔵","Accepted":"🟢","Rejected":"🔴","Expired":"⚫"}

    of1,of2 = st.columns(2)
    flt_osta = of1.selectbox(t('status'), ["All"]+OFFER_STATUSES, key="offer_flt_sta", format_func=_topt)
    flt_oft  = of2.selectbox(t('freight_filter_lbl'), ["All"]+FREIGHT_TYPES_O, key="offer_flt_ft", format_func=_topt)

    offers_df = db.get_sales_offers(status=flt_osta if flt_osta!="All" else None)
    if offers_df is not None and not offers_df.empty:
        if flt_oft != "All":
            offers_df = offers_df[offers_df["freight_type"]==flt_oft]

    if offers_df is not None and not offers_df.empty:
        total_sent = offers_df[offers_df["status"]=="Sent"]["total_value"].sum()
        accepted_val = offers_df[offers_df["status"]=="Accepted"]["total_value"].sum()
        om1,om2,om3 = st.columns(3)
        om1.metric(t('total_offers'), len(offers_df))
        om2.metric(t('pending_sent'), f"${float(total_sent or 0):,.0f}")
        om3.metric(t('accepted_value'), f"${float(accepted_val or 0):,.0f}")
        st.markdown("---")

        for _, row in offers_df.iterrows():
            icon = STATUS_ICON_O.get(str(row["status"]),"⚪")
            val_str = f"${float(row['total_value'] or 0):,.0f} {row.get('currency','USD')}" if row.get("total_value") else "TBD"
            with st.expander(f"{icon} {row['offer_number']}  |  {row['client_name']}  |  {val_str}  |  {row['status']}"):
                oc1,oc2,oc3,oc4 = st.columns(4)
                oc1.write(f"{t('freight_lbl')} {row.get('freight_type','—')}")
                oc2.write(f"{t('route_lbl')} {row.get('origin','—')} → {row.get('destination','—')}")
                oc3.write(f"{t('commodity_lbl')} {row.get('commodity','—')}")
                oc4.write(f"{t('valid_until_lbl')} {str(row.get('validity_date','—'))[:10]}")
                wt = row.get("weight_kg"); vol = row.get("volume_cbm")
                ow1,ow2 = st.columns(2)
                ow1.write(f"{t('weight_lbl')} {f'{float(wt):,.1f} kg' if wt else '—'}")
                ow2.write(f"{t('volume_lbl')} {f'{float(vol):,.2f} CBM' if vol else '—'}")
                if row.get("notes"): st.caption(row["notes"])
                if str(row["status"]) in ("Draft","Sent"):
                    st.markdown("---")
                    curr_s = str(row["status"])
                    next_s = [s for s in OFFER_STATUSES if s != curr_s]
                    os1,os2 = st.columns(2)
                    sel_os = os1.selectbox(t('update_status_offer'), next_s, key=f"osta_{row['id']}", format_func=_topt)
                    if os2.button(t('update_btn'), key=f"ostabtn_{row['id']}", type="primary"):
                        if db.update_offer_status(int(row["id"]), sel_os):
                            st.success(t('offer_status_updated'))
                            st.rerun()
    else:
        st.info(t('no_offers'))

    st.markdown("---")
    with st.expander(t('create_new_offer_exp')):
        COUNTRIES_OFFER = ["Afghanistan","Albania","Algeria","Argentina","Armenia","Australia","Austria",
                           "Azerbaijan","Bahrain","Bangladesh","Belarus","Belgium","Brazil","Bulgaria",
                           "Canada","Chile","China","Colombia","Croatia","Cyprus","Czech Republic",
                           "Denmark","Ecuador","Egypt","Estonia","Ethiopia","Finland","France",
                           "Georgia","Germany","Ghana","Greece","Hungary","India","Indonesia","Iran",
                           "Iraq","Ireland","Israel","Italy","Japan","Jordan","Kazakhstan","Kenya",
                           "Kuwait","Latvia","Lebanon","Libya","Lithuania","Luxembourg","Malaysia",
                           "Malta","Mexico","Morocco","Netherlands","New Zealand","Nigeria","Norway",
                           "Oman","Pakistan","Palestine","Peru","Philippines","Poland","Portugal",
                           "Qatar","Romania","Russia","Saudi Arabia","Senegal","Serbia","Singapore",
                           "Slovakia","Slovenia","South Africa","South Korea","Spain","Sri Lanka",
                           "Sudan","Sweden","Switzerland","Syria","Taiwan","Thailand","Tunisia",
                           "Turkey","Ukraine","United Arab Emirates","United Kingdom","United States",
                           "Uzbekistan","Venezuela","Vietnam","Yemen"]
        offer_clients_df = db.fetch_dataframe("SELECT id, email FROM users WHERE role='client' ORDER BY email")
        client_opts_offer = {r["email"]: r["id"] for _, r in offer_clients_df.iterrows()} if not offer_clients_df.empty else {}
        sel_offer_client = st.selectbox(t('client') + " *", ["— Select —"] + list(client_opts_offer.keys()), key="offer_client_sel")
        with st.form("add_offer_form", clear_on_submit=True):
            no1,no2 = st.columns(2)
            of_num    = no1.text_input(t('offer_num_field'), placeholder="OFR-2026-XXX")
            of_ft     = no2.selectbox(t('freight_type_filter'), FREIGHT_TYPES_O, format_func=_topt)
            no5,no6 = st.columns(2)
            of_orig   = no5.selectbox(t('origin_lbl'), COUNTRIES_OFFER, index=COUNTRIES_OFFER.index("Turkey") if "Turkey" in COUNTRIES_OFFER else 0)
            of_dst    = no6.selectbox(t('destination_lbl'), COUNTRIES_OFFER)
            of_comm   = st.text_input(t('commodity_desc_field'))
            no7,no8,no9 = st.columns(3)
            of_wt     = no7.number_input(t('weight_kg_field'), min_value=0.0, step=100.0)
            of_vol    = no8.number_input(t('volume_cbm_field'), min_value=0.0, step=0.5)
            of_val    = no9.number_input(t('total_value_usd'), min_value=0.0, step=500.0)
            of_valid  = st.date_input(t('valid_until_field'))
            of_notes  = st.text_area(t('notes_incl'), height=70)
            if st.form_submit_button(t('create_offer_btn'), type="primary"):
                of_email  = sel_offer_client if sel_offer_client != "— Select —" else ""
                of_client = of_email
                if not of_num or not of_email:
                    st.error(t('offer_num_client_req'))
                else:
                    uid = st.session_state.get("user",{}).get("id")
                    if db.add_offer(of_num,of_client,of_email,of_ft,of_orig,of_dst,
                                    of_comm,of_wt if of_wt>0 else None,
                                    of_vol if of_vol>0 else None,
                                    of_val if of_val>0 else None,
                                    "USD",str(of_valid),of_notes,uid):
                        st.success(t('offer_created_success'))
                        st.rerun()
                    else:
                        st.error(t('failed_create_offer'))

elif page_matches(page, 'training'):
    st.header(t('training'))
    st.caption(f"{t('last_updated')}: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    st.markdown("---")
    tab_dash, tab_programs, tab_new = st.tabs([t('tab_overview'), t('tab_all_programs'), t('tab_add_program')])

    TRAIN_CATS = ["Safety & Compliance","Customs & Regulations","IATA Cargo","Customer Service",
                  "IT & Systems","Leadership","Language","Operations","Other"]

    with tab_dash:
        try:
            prog_df = db.fetch_dataframe("SELECT * FROM admin_training_programs ORDER BY scheduled_date ASC")
            enrl_df = db.fetch_dataframe("SELECT * FROM admin_training_enrollments")
        except Exception:
            prog_df = enrl_df = pd.DataFrame()

        td1,td2,td3,td4 = st.columns(4)
        td1.metric(t('total_programs'),   len(prog_df) if not prog_df.empty else 0)
        td2.metric(t('upcoming_lbl'),     len(prog_df[prog_df["status"]=="Scheduled"]) if not prog_df.empty else 0)
        td3.metric(t('total_enrollments'),len(enrl_df) if not enrl_df.empty else 0)
        td4.metric(t('completions_lbl'),  len(enrl_df[enrl_df["status"]=="Completed"]) if not enrl_df.empty else 0)

        if not prog_df.empty:
            st.markdown("---")
            col_l, col_r = st.columns(2)
            with col_l:
                cat_c = prog_df["category"].value_counts().reset_index()
                cat_c.columns = ["Category","Count"]
                fig_t = px.pie(cat_c, names="Category", values="Count",
                               title=t('programs_by_cat'), height=280)
                fig_t.update_layout(margin=dict(l=0,r=0,t=30,b=0))
                st.plotly_chart(fig_t, use_container_width=True)
            with col_r:
                st.subheader(t('upcoming_training'))
                upcoming_t = prog_df[prog_df["status"]=="Scheduled"].head(5)
                for _, row in upcoming_t.iterrows():
                    st.markdown(f"📅 **{str(row['scheduled_date'])[:10]}** — {row['title']}")
                    st.caption(f"&nbsp;&nbsp;&nbsp;&nbsp;{row['category']} · {row.get('duration_hours',0)}h · Trainer: {row.get('trainer','TBD')}")

    with tab_programs:
        try:
            prog_df2 = db.fetch_dataframe("SELECT * FROM admin_training_programs ORDER BY scheduled_date DESC")
            enrl_df2 = db.fetch_dataframe("SELECT * FROM admin_training_enrollments")
        except Exception:
            prog_df2 = enrl_df2 = pd.DataFrame()

        pf1,pf2 = st.columns(2)
        cat_pf = pf1.selectbox(t('cat_filter_lbl'), ["All"]+TRAIN_CATS, key="tp_cat")
        sta_pf = pf2.selectbox(t('status'), ["All","Scheduled","Ongoing","Completed","Cancelled"], key="tp_sta", format_func=_topt)

        if not prog_df2.empty:
            disp_p = prog_df2.copy()
            if cat_pf != "All": disp_p = disp_p[disp_p["category"]==cat_pf]
            if sta_pf != "All": disp_p = disp_p[disp_p["status"]==sta_pf]

            STATUS_ICON = {"Scheduled":"🔵","Ongoing":"🟡","Completed":"🟢","Cancelled":"🔴"}
            for _, row in disp_p.iterrows():
                icon = STATUS_ICON.get(str(row["status"]),"⚪")
                enrolled_count = len(enrl_df2[enrl_df2["program_id"]==row["id"]]) if not enrl_df2.empty else 0
                with st.expander(f"{icon} {row['title']}  |  {row['category']}  |  {str(row.get('scheduled_date',''))[:10]}"):
                    pc1,pc2,pc3,pc4 = st.columns(4)
                    pc1.write(f"**Trainer:** {row.get('trainer','TBD') or 'TBD'}")
                    pc2.write(f"**Duration:** {row.get('duration_hours',0)}h")
                    pc3.write(f"**Enrolled:** {enrolled_count}/{row.get('max_participants','∞') or '∞'}")
                    pc4.write(f"**Status:** {row['status']}")
                    if row.get("description"): st.caption(row["description"])

                    # Enrollments
                    if not enrl_df2.empty:
                        prog_enrl = enrl_df2[enrl_df2["program_id"]==row["id"]]
                        if not prog_enrl.empty:
                            st.markdown(t('enrolled_participants'))
                            disp_e = prog_enrl[["employee_name","status","score","completed_at"]].copy()
                            disp_e.columns = ["Employee","Status","Score","Completed"]
                            st.dataframe(disp_e, use_container_width=True, hide_index=True)

                    st.markdown("---")
                    ep1,ep2 = st.columns(2)
                    # Enroll employee
                    emp_df = db.fetch_dataframe("SELECT id, email FROM users WHERE role='employee' ORDER BY email")
                    if not emp_df.empty:
                        sel_emp = ep1.selectbox(t('enroll_employee'), emp_df["email"].tolist(), key=f"enrl_emp_{row['id']}")
                        if ep2.button(t('enroll_btn'), key=f"enrl_btn_{row['id']}", type="primary"):
                            emp_row = emp_df[emp_df["email"]==sel_emp].iloc[0]
                            try:
                                with db.get_connection() as conn:
                                    conn.execute(text("""
                                        INSERT IGNORE INTO admin_training_enrollments
                                            (program_id,employee_name,employee_email,user_id,status)
                                        VALUES (:pid,:name,:email,:uid,'Enrolled')
                                    """), dict(pid=int(row["id"]),name=sel_emp,email=sel_emp,uid=int(emp_row["id"])))
                                    conn.commit()
                                st.success(f"{sel_emp} enrolled.")
                                st.rerun()
                            except Exception as ex:
                                st.error(str(ex))

                    if row["status"] not in ("Completed","Cancelled"):
                        sta_opts = [s for s in ["Scheduled","Ongoing","Completed","Cancelled"] if s != row["status"]]
                        sel_sta = st.selectbox(t('update_status_prog'), sta_opts, key=f"psta_{row['id']}")
                        if st.button(t('update_prog_status'), key=f"pstabtn_{row['id']}"):
                            try:
                                with db.get_connection() as conn:
                                    conn.execute(text("UPDATE admin_training_programs SET status=:s WHERE id=:id"),
                                                 {"s":sel_sta,"id":int(row["id"])})
                                    conn.commit()
                                st.rerun()
                            except Exception as ex:
                                st.error(str(ex))
        else:
            st.info(t('no_programs'))

    with tab_new:
        st.subheader(t('add_program_hdr'))
        with st.form("new_training_form", clear_on_submit=True):
            nt1,nt2 = st.columns(2)
            tr_title    = nt1.text_input(t('program_title'))
            tr_cat      = nt2.selectbox(t('category_field'), TRAIN_CATS)
            nt3,nt4 = st.columns(2)
            tr_date     = nt3.date_input(t('scheduled_date'))
            tr_dur      = nt4.number_input(t('duration_hours'), min_value=1, max_value=40, value=8)
            nt5,nt6 = st.columns(2)
            tr_trainer  = nt5.text_input(t('trainer_instructor'))
            tr_max      = nt6.number_input(t('max_participants'), min_value=1, max_value=100, value=20)
            tr_desc     = st.text_area(t('description_area'), height=80)
            if st.form_submit_button(t('add_program_btn'), type="primary"):
                if not tr_title:
                    st.error(t('title_required_err'))
                else:
                    try:
                        with db.get_connection() as conn:
                            conn.execute(text("""
                                INSERT INTO admin_training_programs
                                    (title,category,description,trainer,duration_hours,max_participants,scheduled_date,status)
                                VALUES (:t,:cat,:desc,:tr,:dur,:mx,:dt,'Scheduled')
                            """), dict(t=tr_title,cat=tr_cat,desc=tr_desc,tr=tr_trainer,
                                       dur=tr_dur,mx=tr_max,dt=str(tr_date)))
                            conn.commit()
                        st.success(f"Program '{tr_title}' added.")
                    except Exception as ex:
                        st.error(str(ex))

elif page_matches(page, 'profile'):
    try:
        current_user = st.session_state.get('user')
        if not current_user:
            st.error(t('please_login'))
            st.stop()

        _is_dark_p = st.session_state.get('theme', 'light') == 'dark'

        users_df = db.get_all_users()
        if users_df is None or users_df.empty:
            st.info(t('no_users_avail'))
        else:
            emails = users_df['email'].tolist()

            # Manager: pick any user; others: own profile
            if current_user.get('role') == 'manager':
                try:
                    qp = st.query_params
                    if qp and 'profile_selected_email' in qp:
                        val = qp['profile_selected_email'][0] if isinstance(qp['profile_selected_email'], (list, tuple)) else qp['profile_selected_email']
                        if val and val in emails:
                            st.session_state['profile_selected_email'] = val
                except Exception:
                    pass
                _my_email = current_user.get('email')

                # If "My Profile" was clicked last run, apply the override before widget creation
                if st.session_state.pop('_goto_my_profile', False):
                    st.session_state['_profile_sel_idx'] = emails.index(_my_email) if _my_email in emails else 0

                _sel_idx = st.session_state.get('_profile_sel_idx', emails.index(_my_email) if _my_email in emails else 0)

                col_sel, col_btn = st.columns([4, 1])
                with col_sel:
                    selected_email = st.selectbox(
                        t('select_user_view'),
                        options=emails,
                        index=_sel_idx
                    )
                    st.session_state['_profile_sel_idx'] = emails.index(selected_email)
                with col_btn:
                    st.write("")
                    if st.button(t('my_profile_btn'), disabled=(selected_email == _my_email)):
                        st.session_state['_goto_my_profile'] = True
                        _safe_rerun()
            else:
                selected_email = current_user.get('email')

            selected = db.get_user_by_email(selected_email)
            if not selected:
                st.error(t('user_not_found'))
            else:
                # ── Avatar card ──────────────────────────────────────────────
                _role = selected.get('role', 'employee')
                _role_colors = {'manager': '#7c3aed', 'employee': '#1d4ed8', 'client': '#0369a1', 'admin': '#b91c1c'}
                _role_color = _role_colors.get(_role, '#1d4ed8')
                _created = selected.get('created_at')
                _join_str = ''
                if _created:
                    try:
                        from datetime import datetime as _dt
                        _cd = _created if isinstance(_created, _dt) else _dt.strptime(str(_created)[:19], '%Y-%m-%d %H:%M:%S')
                        _months = (_dt.now().year - _cd.year) * 12 + (_dt.now().month - _cd.month)
                        _join_str = f"Joined {_cd.strftime('%b %Y')} &nbsp;·&nbsp; {_months} months ago"
                    except Exception:
                        _join_str = str(_created)[:10]

                # pull name/department/position/phone from company_records for all roles
                _dept = ''
                _record_name = ''
                _position = ''
                _phone_rec = ''
                try:
                    _recs = get_cached_records()
                    _match = _recs[_recs['email'].astype(str).str.strip().str.lower() == str(selected_email).strip().lower()]
                    if not _match.empty:
                        _rec_row = _match.iloc[0]
                        _record_name = str(_rec_row.get('employee_name', '') or '')
                        if _record_name in ('None', 'nan'):
                            _record_name = ''
                        # phone available for all roles
                        _raw_ph = str(_rec_row.get('phone', '') or '')
                        _phone_rec = _raw_ph if _raw_ph not in ('', 'None', 'nan') else ''
                        # dept and position only meaningful for employees
                        if _role == 'employee':
                            _raw_dept = str(_rec_row.get('department', '') or '')
                            _dept = _raw_dept if _raw_dept not in ('', 'None', 'nan') else ''
                            _raw_pos = str(_rec_row.get('position', '') or '')
                            _position = _raw_pos if _raw_pos not in ('', 'None', 'nan') else ''
                except Exception:
                    pass

                # Display name: users.full_name > company_records.employee_name > email
                _display_name = selected.get('full_name') or _record_name or selected_email
                _avatar_letter = _display_name[0].upper()

                _bg  = '#1e293b' if _is_dark_p else '#f0f4f8'
                _txt = '#e2e8f0' if _is_dark_p else '#1e2a3a'
                _sub = '#94a3b8' if _is_dark_p else '#64748b'
                _bdr = '#334155' if _is_dark_p else '#cbd5e1'

                # ── Pre-compute conditional HTML snippets ─────────────────────
                _position_html = f"<div style='font-size:0.82rem;color:{_sub};margin-top:2px;'>{_position}</div>" if _position and _position not in ('', 'None', 'nan') else ""
                _phone_val = _phone_rec or selected.get('phone', '') or ''
                _phone_html = f"<div style='font-size:0.82rem;color:{_sub};margin-top:2px;'>&#128222; {_phone_val}</div>" if _phone_val and _phone_val not in ('', 'None', 'nan') else ""
                _dept_html = f"<span style='background:#0f766e;color:#fff;padding:2px 12px;border-radius:20px;font-size:0.78rem;font-weight:600;'>{_dept}</span>" if _dept else ""

                # ── Avatar HTML (image or letter) ─────────────────────────────
                _avatar_path = (selected.get('avatar') or '').strip()
                _can_edit_avatar = (selected_email == current_user.get('email')) or (current_user.get('role') == 'manager')

                if _avatar_path and os.path.exists(_avatar_path):
                    import base64 as _b64
                    with open(_avatar_path, 'rb') as _af:
                        _ext = _avatar_path.rsplit('.', 1)[-1].lower()
                        _mime = 'image/png' if _ext == 'png' else 'image/jpeg'
                        _img_b64 = _b64.b64encode(_af.read()).decode()
                    _avatar_html = f"<img src='data:{_mime};base64,{_img_b64}' style='width:72px;height:72px;border-radius:50%;object-fit:cover;border:2px solid {_role_color};flex-shrink:0;'/>"
                else:
                    _avatar_html = f"<div style='width:72px;height:72px;border-radius:50%;background:{_role_color};display:flex;align-items:center;justify-content:center;font-size:2rem;font-weight:700;color:#fff;flex-shrink:0;'>{_avatar_letter}</div>"

                _badges = (f"<span style='background:{_role_color};color:#fff;padding:2px 12px;border-radius:20px;font-size:0.78rem;font-weight:600;'>{_role.capitalize()}</span>"
                           + _dept_html
                           + f"<span style='color:{_sub};font-size:0.8rem;margin-left:4px;'>{_join_str}</span>")
                _card_html = (
                    f"<div style='background:{_bg};border:1px solid {_bdr};border-radius:14px;padding:24px 28px;display:flex;align-items:center;gap:24px;margin-bottom:20px;'>"
                    + _avatar_html
                    + f"<div><div style='font-size:1.1rem;font-weight:700;color:{_txt};'>{_display_name}</div>"
                    + f"<div style='font-size:0.85rem;color:{_sub};margin-top:2px;'>{selected_email}</div>"
                    + _position_html
                    + _phone_html
                    + f"<div style='margin-top:8px;display:flex;gap:8px;flex-wrap:wrap;align-items:center;'>{_badges}</div>"
                    + "</div></div>"
                )
                st.markdown(_card_html, unsafe_allow_html=True)

                # ── Upload / change avatar ────────────────────────────────────
                if _can_edit_avatar:
                    with st.expander(t('change_profile_pic')):
                        if not IS_DESKTOP:
                            st.info("🖼️ Profile picture upload is only available in the desktop application.")
                        else:
                            _up_col, _del_col = st.columns([3, 1])
                            with _up_col:
                                _avatar_file = st.file_uploader(t('upload_photo'), type=['jpg', 'jpeg', 'png'], key='avatar_upload')
                                if _avatar_file is not None:
                                    if _avatar_file.size > 2 * 1024 * 1024:
                                        st.error(t('file_too_large'))
                                    else:
                                        if st.button(t('save_photo')):
                                            try:
                                                _av_dir = os.path.join('uploads', 'avatars')
                                                os.makedirs(_av_dir, exist_ok=True)
                                                _ext = _avatar_file.name.rsplit('.', 1)[-1].lower()
                                                _av_name = f"avatar_{selected.get('id')}.{_ext}"
                                                _av_path = os.path.join(_av_dir, _av_name)
                                                with open(_av_path, 'wb') as _f:
                                                    _f.write(_avatar_file.getbuffer())
                                                db.update_user_avatar(selected.get('id'), _av_path)
                                                db.insert_activity_log("avatar_update", "Updated profile picture", current_user.get('email',''))
                                                st.success(t('profile_pic_updated'))
                                                _safe_rerun()
                                            except Exception as _e:
                                                st.error(f"Failed to save: {_e}")
                            with _del_col:
                                if _avatar_path and os.path.exists(_avatar_path):
                                    if st.button(t('remove_photo')):
                                        try:
                                            os.remove(_avatar_path)
                                        except Exception:
                                            pass
                                    db.update_user_avatar(selected.get('id'), '')
                                    st.success(t('photo_removed'))
                                    _safe_rerun()

                # ── Stats (per role) ─────────────────────────────────────────
                m1, m2, m3, m4 = st.columns(4)
                try:
                    if _role == 'employee':
                        _lr_df    = db.get_leave_requests_by_user(selected.get('id'))
                        _total    = len(_lr_df)
                        _approved = len(_lr_df[_lr_df['status'].str.lower() == 'approved']) if _total else 0
                        _pending  = len(_lr_df[_lr_df['status'].str.lower() == 'pending'])  if _total else 0
                        _rejected = len(_lr_df[_lr_df['status'].str.lower() == 'rejected']) if _total else 0
                        m1.metric(t('leave_requests_lbl'), _total)
                        m2.metric(t('approved'), _approved)
                        m3.metric(t('pending'), _pending)
                        m4.metric(t('rejected'), _rejected)

                    elif _role == 'manager':
                        _all_users  = db.get_all_users()
                        _employees  = len(_all_users[_all_users['role'] == 'employee']) if not _all_users.empty else 0
                        _clients    = len(_all_users[_all_users['role'] == 'client'])   if not _all_users.empty else 0
                        _leave_df   = db.get_all_leave_requests()
                        _pending_lv = len(_leave_df[_leave_df['status'].str.lower() == 'pending']) if not _leave_df.empty else 0
                        _total_users = len(_all_users) if not _all_users.empty else 0
                        m1.metric(t('employees_lbl'), _employees)
                        m2.metric(t('clients_lbl'), _clients)
                        m3.metric(t('leave_req_pending'), _pending_lv)
                        m4.metric(t('total_users'), _total_users)

                    elif _role == 'client':
                        _ship_df   = db.get_shipments_by_client(selected.get('id'))
                        _total_sh  = len(_ship_df)
                        _active    = len(_ship_df[_ship_df['status'].str.lower().isin(['in transit', 'processing', 'pending'])]) if _total_sh else 0
                        _delivered = len(_ship_df[_ship_df['status'].str.lower() == 'delivered']) if _total_sh else 0
                        _customs   = len(_ship_df[_ship_df['customs_cleared'].astype(str) == '1']) if _total_sh else 0
                        m1.metric(t('total_shipments'), _total_sh)
                        m2.metric(t('active_lbl'), _active)
                        m3.metric(t('delivered_lbl'), _delivered)
                        m4.metric(t('customs_cleared'), _customs)
                except Exception:
                    pass

                st.markdown("---")

                # ── Tabs ─────────────────────────────────────────────────────
                tab_info, tab_security, tab_activity = st.tabs([f"👤 {t('tab_info')}", f"🔒 {t('tab_security')}", f"📋 {t('tab_activity')}"])

                # ── Tab: Info ────────────────────────────────────────────────
                with tab_info:
                    with st.form('profile_info_form'):
                        # users table first, fall back to company_records
                        _fn = selected.get('full_name') or _record_name or ''
                        _ph = selected.get('phone') or _phone_rec or ''
                        new_full_name = st.text_input(t('full_name_lbl'), value=_fn, placeholder='Enter your full name')
                        new_phone     = st.text_input(t('phone_lbl'), value=_ph, placeholder='+1 234 567 8900')
                        if current_user.get('role') == 'manager':
                            role_options = ['client', 'employee', 'manager']
                            cur_role = selected.get('role') or 'employee'
                            new_role_sel = st.selectbox(t('role_lbl'), role_options, index=role_options.index(cur_role) if cur_role in role_options else 1)
                        else:
                            new_role_sel = selected.get('role')
                        if st.form_submit_button(t('save_info')):
                            try:
                                db.update_user_profile(selected.get('id'), new_full_name, new_phone)
                                db.insert_activity_log("profile_update", f"Updated profile info (name/phone)", current_user.get('email',''))
                                if current_user.get('role') == 'manager' and new_role_sel != selected.get('role'):
                                    db.update_user_role(selected.get('id'), new_role_sel)
                                    db.insert_activity_log("role_update", f"Changed role of {selected_email} → {new_role_sel}", current_user.get('email',''))
                                st.success(t('profile_updated'))
                                _safe_rerun()
                            except Exception as e:
                                st.error(f'Failed to save: {e}')

                # ── Tab: Security ────────────────────────────────────────────
                with tab_security:
                    if current_user.get('role') == 'manager' and selected.get('id') != current_user.get('id'):
                        with st.form('mgr_reset_pw'):
                            st.write(f"Reset password for **{selected_email}**")
                            new_pw_mgr = st.text_input(t('new_pass_auto'), type='password')
                            if st.form_submit_button(t('reset_password_btn')):
                                try:
                                    _new_pw = secrets.token_urlsafe(8) if not new_pw_mgr else new_pw_mgr
                                    _salt = _generate_salt()
                                    _hashed = _hash_password(_new_pw, _salt)
                                    if db.update_user_password(selected.get('id'), _hashed, _salt):
                                        st.success(t('password_reset_done'))
                                        st.code(f'New password: {_new_pw}')
                                    else:
                                        st.error(t('failed_short'))
                                except Exception as e:
                                    st.error(f'Error: {e}')
                    else:
                        with st.form('user_change_pw'):
                            cur_pw_in  = st.text_input(t('current_password'), type='password')
                            new_pw_in  = st.text_input(t('new_password'), type='password')
                            conf_pw_in = st.text_input(t('confirm_new_password'), type='password')
                            if st.form_submit_button(t('update_password_btn')):
                                if not cur_pw_in or not new_pw_in:
                                    st.error(t('fill_all_fields'))
                                elif new_pw_in != conf_pw_in:
                                    st.error(t('pass_no_match_new'))
                                else:
                                    try:
                                        _ur = db.get_user_by_email(selected.get('email'))
                                        if _hash_password(cur_pw_in, _ur.get('salt')) != _ur.get('password_hash'):
                                            st.error(t('current_pass_incorrect'))
                                        else:
                                            _salt = _generate_salt()
                                            _hashed = _hash_password(new_pw_in, _salt)
                                            if db.update_user_password(selected.get('id'), _hashed, _salt):
                                                st.success(t('password_updated'))
                                            else:
                                                st.error(t('failed_update_pass'))
                                    except Exception as e:
                                        st.error(f'Error: {e}')

                # ── Tab: My Activity ─────────────────────────────────────────
                with tab_activity:
                    _pcolors = {'Urgent': '#ef4444', 'High': '#f97316', 'Medium': '#3b82f6', 'Low': '#6b7280'}
                    _scolors = {'approved': '#16a34a', 'rejected': '#dc2626', 'pending': '#d97706',
                                'delivered': '#16a34a', 'in transit': '#3b82f6', 'processing': '#d97706'}

                    def _act_card(left_top, left_sub, badge1_text, badge1_color, badge2_text=None, badge2_color=None):
                        _b2 = f"<span style='background:{badge2_color};color:#fff;padding:2px 10px;border-radius:12px;font-size:0.75rem;font-weight:600;'>{badge2_text}</span>" if badge2_text else ""
                        _html = (
                            f"<div style='background:{_bg};border:1px solid {_bdr};border-radius:10px;padding:14px 18px;margin-bottom:10px;display:flex;justify-content:space-between;align-items:center;'>"
                            f"<div><div style='color:{_txt};font-weight:600;'>{left_top}</div>"
                            f"<div style='color:{_sub};font-size:0.82rem;margin-top:4px;'>{left_sub}</div></div>"
                            f"<div style='display:flex;gap:6px;flex-direction:column;align-items:flex-end;'>"
                            f"<span style='background:{badge1_color};color:#fff;padding:2px 10px;border-radius:12px;font-size:0.75rem;font-weight:600;'>{badge1_text}</span>"
                            f"{_b2}</div></div>"
                        )
                        st.markdown(_html, unsafe_allow_html=True)

                    # ── icon + color map per action type ─────────────────────
                    _action_icons = {
                        'login_success':     ('🔑', '#16a34a'),
                        'logout':            ('🚪', '#6b7280'),
                        'login_failed':      ('⛔', '#dc2626'),
                        'leave_request':     ('📋', '#3b82f6'),
                        'leave_status_update': ('✅', '#7c3aed'),
                        'add_record':        ('➕', '#0891b2'),
                        'edit_record':       ('✏️', '#d97706'),
                        'delete_record':     ('🗑️', '#dc2626'),
                        'role_update':       ('🔄', '#7c3aed'),
                        'profile_update':    ('👤', '#0369a1'),
                        'avatar_update':     ('🖼️', '#0369a1'),
                        'password_change':   ('🔒', '#b91c1c'),
                        'Ticket Created':    ('🎫', '#0891b2'),
                        'Ticket Updated':    ('🎫', '#d97706'),
                        'Ticket Deleted':    ('🎫', '#dc2626'),
                    }
                    _action_labels = {
                        'login_success':     t('act_login_success'),
                        'logout':            t('act_logout'),
                        'login_failed':      t('act_login_failed'),
                        'leave_request':     t('act_leave_request'),
                        'leave_status_update': t('act_leave_updated'),
                        'add_record':        t('act_add_record'),
                        'edit_record':       t('act_edit_record'),
                        'delete_record':     t('act_delete_record'),
                        'role_update':       t('act_role_update'),
                        'profile_update':    t('act_profile_update'),
                        'avatar_update':     t('act_avatar_update'),
                        'password_change':   t('act_password_change'),
                        'Ticket Created':    t('act_ticket_created'),
                        'Ticket Updated':    t('act_ticket_updated'),
                        'Ticket Deleted':    t('act_ticket_deleted'),
                    }

                    try:
                        _log_df = db.get_user_activity_log(selected_email, limit=50)
                        if _log_df.empty:
                            st.info(t('no_activity_yet'))
                        else:
                            st.caption(f"{len(_log_df)} {t('recent_activities')}")
                            for _, _r in _log_df.iterrows():
                                _etype = str(_r.get('event_type', '') or '')
                                _icon, _color = _action_icons.get(_etype, ('⚡', '#6b7280'))
                                _badge_label = _action_labels.get(_etype, _etype.replace('_', ' ').title())
                                _ts = str(_r.get('timestamp', ''))[:16]
                                _desc = str(_r.get('description', '') or '')
                                _html = (
                                    f"<div style='background:{_bg};border:1px solid {_bdr};border-radius:10px;"
                                    f"padding:12px 16px;margin-bottom:8px;display:flex;align-items:center;gap:14px;'>"
                                    f"<div style='font-size:1.3rem;'>{_icon}</div>"
                                    f"<div style='flex:1;'>"
                                    f"<div style='color:{_txt};font-weight:600;font-size:0.88rem;'>{_desc}</div>"
                                    f"<div style='color:{_sub};font-size:0.75rem;margin-top:2px;'>{_ts}</div>"
                                    f"</div>"
                                    f"<span style='background:{_color};color:#fff;padding:2px 10px;border-radius:12px;"
                                    f"font-size:0.72rem;font-weight:600;white-space:nowrap;'>{_badge_label}</span>"
                                    f"</div>"
                                )
                                st.markdown(_html, unsafe_allow_html=True)
                    except Exception as e:
                        st.error(f'Could not load activity: {e}')

    except Exception as e:
        st.error(f"Error loading profile: {str(e)}")

elif page_matches(page, 'messages'):
    st.header(t('messages'))
    
    try:
        current_user = st.session_state['user']
        user_id = current_user['id']
        user_role = current_user.get('role', 'client')
        
        # Create tabs for Inbox and Send Message
        tab1, tab2 = st.tabs([f"📥 {t('inbox')}", f"✉️ {t('compose')}"])
        
        with tab1:
            # Display messages
            st.markdown(f"#### {t('your_messages')}")

            messages_df = db.get_user_messages(user_id)

            if len(messages_df) == 0:
                st.info(t('no_messages'))
            else:
                # Show unread count
                unread_count = db.get_unread_count(user_id)
                if unread_count > 0:
                    st.warning(f"📬 {t('unread_messages_warn').format(count=unread_count)}")
                
                for idx, msg in messages_df.iterrows():
                    # Determine if this is sent or received
                    is_received = msg['to_user_id'] == user_id
                    
                    # Create expandable message
                    with st.expander(
                        f"{'📬' if is_received and msg['is_read'] == 0 else '📭'} "
                        f"{t('from_lbl') if is_received else t('to_lbl')}: {msg['from_email'] if is_received else msg['to_email']} - "
                        f"{msg['subject']} ({msg['created_at']})"
                    ):
                        col1, col2 = st.columns([3, 1])

                        with col1:
                            st.markdown(f"{t('subject_lbl')} {msg['subject']}")
                            st.markdown(f"{t('from_lbl2')} {msg['from_email']}")
                            st.markdown(f"{t('to_lbl2')} {msg['to_email']}")
                            if msg['shipment_number']:
                                st.markdown(f"{t('shipment_lbl')} {msg['shipment_number']}")
                            st.markdown(f"{t('date_lbl')} {msg['created_at']}")
                            st.markdown("---")
                            st.markdown(t('message_lbl'))
                            st.write(msg['content'])

                        with col2:
                            if is_received and msg['is_read'] == 0:
                                if st.button(t('mark_as_read_btn'), key=f"read_{msg['id']}"):
                                    db.mark_message_read(msg['id'])
                                    _safe_rerun()

                            # Reply button
                            if st.button(t('reply_btn'), key=f"reply_{msg['id']}"):
                                st.session_state['reply_to'] = {
                                    'email': msg['from_email'] if is_received else msg['to_email'],
                                    'subject': f"Re: {msg['subject']}",
                                    'shipment_id': msg.get('shipment_id')
                                }
                                _safe_rerun()

                # If user clicked reply, show the send form inline here so they don't have to switch tabs
                if 'reply_to' in st.session_state:
                    st.markdown("---")
                    st.markdown(f"#### {t('reply_hdr')}")
                    reply_data = st.session_state['reply_to']
                    # Build minimal recipient list to validate email -> id mapping
                    if user_role == 'client':
                        with db.get_connection() as conn:
                            res = conn.execute(text("""
                                SELECT u.id, u.email FROM users u
                                JOIN company_records cr ON cr.email = u.email
                                WHERE cr.department = 'Customer Service'
                            """))
                            users = [(r[0], r[1]) for r in res.fetchall()]
                    else:
                        with db.get_connection() as conn:
                            res = conn.execute(text("SELECT id, email FROM users WHERE role='client'"))
                            users = [(r[0], r[1]) for r in res.fetchall()]

                    with st.form("send_message_form_reply"):
                        to_email = st.text_input(t('to_field'), value=reply_data['email'], disabled=True)
                        to_user = next((u for u in users if u[1] == reply_data['email']), None)
                        to_user_id = to_user[0] if to_user else (users[0][0] if users else None)
                        subject = st.text_input(t('subject_field'), value=reply_data['subject'])
                        shipment_id = reply_data.get('shipment_id')
                        content = st.text_area(t('message_field'), height=200)
                        submit = st.form_submit_button(t('send_reply_btn'))
                        if submit:
                            if not subject or not content:
                                st.error(t('fill_subject_msg'))
                            else:
                                if to_user_id is None:
                                    st.error(t('recipient_not_found'))
                                else:
                                    if db.send_message(user_id, to_user_id, subject, content, shipment_id):
                                        st.success(t('reply_sent'))
                                        # clear reply state and refresh
                                        try:
                                            del st.session_state['reply_to']
                                        except Exception:
                                            pass
                                        time.sleep(1)
                                        _safe_rerun()
                                    else:
                                        st.error(t('error_sending_reply'))
        
        with tab2:
            st.markdown(f"#### {t('send_new_msg')}")

            # Get list of users to send to
            user_department = current_user.get('department', '')
            if user_role == 'client':
                with db.get_connection() as conn:
                    res = conn.execute(text("""
                        SELECT u.id, u.email FROM users u
                        JOIN company_records cr ON cr.email = u.email
                        WHERE cr.department = 'Customer Service'
                    """))
                    users = [(r[0], r[1]) for r in res.fetchall()]
            elif user_role == 'manager' or user_department in ('Customer Service', 'Sales'):
                with db.get_connection() as conn:
                    res = conn.execute(text("SELECT id, email FROM users WHERE id != :uid"), {"uid": user_id})
                    users = [(r[0], r[1]) for r in res.fetchall()]
            else:
                with db.get_connection() as conn:
                    res = conn.execute(text("SELECT id, email FROM users WHERE id != :uid AND role != 'client'"), {"uid": user_id})
                    users = [(r[0], r[1]) for r in res.fetchall()]

            if len(users) == 0:
                st.warning(t('no_users_send_msg'))
            else:
                with st.form("send_message_form"):
                    # Check if replying
                    if 'reply_to' in st.session_state:
                        reply_data = st.session_state['reply_to']
                        to_email = st.text_input(t('to_field'), value=reply_data['email'], disabled=True)
                        # Find user ID from email
                        to_user = next((u for u in users if u[1] == reply_data['email']), None)
                        to_user_id = to_user[0] if to_user else users[0][0]
                        subject = st.text_input(t('subject_field'), value=reply_data['subject'])

                        # Clear reply state after form
                        del st.session_state['reply_to']
                    else:
                        # Allow broadcast option for non-client roles
                        user_emails = [u[1] for u in users]
                        can_contact_clients = user_role == 'manager' or user_department in ('Customer Service', 'Sales')
                        if user_role != 'client':
                            if can_contact_clients:
                                group_opts = ['All Users', 'All Managers', 'All Employees', 'All Clients']
                            else:
                                group_opts = ['All Managers', 'All Employees']
                            user_emails = group_opts + user_emails
                            to_email = st.selectbox(t('to_field'), user_emails, index=len(group_opts))
                        else:
                            to_email = st.selectbox(t('to_field'), user_emails)
                        # pre-resolve to_user_id when an individual email is selected
                        if to_email not in ('All Users', 'All Managers', 'All Employees', 'All Clients'):
                            to_user_id = next((u[0] for u in users if u[1] == to_email), users[0][0])
                        else:
                            to_user_id = None
                        subject = st.text_input(t('subject_field'))

                    # Optional: Select shipment
                    if user_role == 'client':
                        client_shipments = db.get_shipments_by_client(user_id)
                        if len(client_shipments) > 0:
                            shipment_options = ["None"] + client_shipments['shipment_number'].tolist()
                            selected_shipment = st.selectbox(t('related_shipment'), shipment_options)
                            shipment_id = None
                            if selected_shipment != "None":
                                shipment_id = client_shipments[client_shipments['shipment_number'] == selected_shipment]['id'].iloc[0]
                        else:
                            shipment_id = None
                    else:
                        shipment_id = None

                    content = st.text_area(t('message_field'), height=200)

                    submit = st.form_submit_button(t('send_msg_btn'))

                    if submit:
                        if not subject or not content:
                            st.error(t('fill_subject_msg'))
                        else:
                            try:
                                # Broadcast to groups if a group option selected
                                sent = 0
                                failed = 0
                                target_clause = None
                                if to_email == 'All Users':
                                    target_clause = ("", {"uid": user_id})
                                elif to_email == 'All Managers':
                                    target_clause = ("WHERE role = 'manager' AND id != :uid", {"uid": user_id})
                                elif to_email == 'All Employees':
                                    target_clause = ("WHERE role = 'employee' AND id != :uid", {"uid": user_id})
                                elif to_email == 'All Clients':
                                    target_clause = ("WHERE role = 'client' AND id != :uid", {"uid": user_id})

                                if target_clause is not None:
                                    where_sql, params = target_clause
                                    with db.get_connection() as conn:
                                        if where_sql:
                                            q = text(f"SELECT id FROM users {where_sql}")
                                            res = conn.execute(q, params)
                                        else:
                                            # All Users except sender
                                            q = text("SELECT id FROM users WHERE id != :uid")
                                            res = conn.execute(q, params)
                                        all_user_ids = [r[0] for r in res.fetchall()]
                                    for uid in all_user_ids:
                                        ok = db.send_message(user_id, uid, subject, content, shipment_id)
                                        if ok:
                                            sent += 1
                                        else:
                                            failed += 1
                                    st.success(f"✅ Broadcast complete — sent: {sent}, failed: {failed}")
                                    time.sleep(1)
                                    _safe_rerun()
                                else:
                                    # single recipient
                                    if to_user_id is None:
                                        to_user_id = next((u[0] for u in users if u[1] == to_email), None)
                                    if to_user_id is None:
                                        st.error(t('recipient_not_found'))
                                    else:
                                        if db.send_message(user_id, to_user_id, subject, content, shipment_id):
                                            st.success("✅ Message sent successfully!")
                                            time.sleep(1)
                                            _safe_rerun()
                                        else:
                                            st.error(t('error_sending_msg'))
                            except Exception as e:
                                logger.exception("Error sending message")
                                st.error(f"❌ Error sending message: {str(e)}")
    
    except Exception as e:
        st.error(f"Error: {str(e)}")

st.markdown("---")
st.markdown(
    f"""
    <div style='text-align: center; color: #666;'>
        <p>{t('footer_text')}</p>
    </div>
    """,
    unsafe_allow_html=True
)

